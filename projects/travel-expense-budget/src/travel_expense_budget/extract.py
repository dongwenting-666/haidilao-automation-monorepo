"""Data extraction from KSB1 and QBI xlsx files."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from travel_expense_budget.config import (
    TRAVEL_GL_PREFIXES,
    KSB1_COL_COST_CENTER,
    KSB1_COL_GL_CODE,
    KSB1_COL_CAD,
    KSB1_COL_CURRENCY,
    DEFAULT_CAD_TO_USD,
    QBI_COL_STORE,
    QBI_COL_REVENUE,
    QBI_STORE_MAP,
)

log = logging.getLogger(__name__)


def extract_travel_expenses(
    ksb1_path: Path,
    cad_to_usd: float = DEFAULT_CAD_TO_USD,
) -> dict[str, float]:
    """Extract travel expenses from a KSB1 xlsx file, grouped by cost center.

    Returns {cost_center_code: total_cad}.
    Detects transaction currency per row: CAD rows use rate 1.0, USD rows are
    converted back to CAD using the configured CAD→USD exchange rate.
    Filters for GL accounts starting with 5101160x (travel expense codes).
    """
    wb = openpyxl.load_workbook(ksb1_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        totals: dict[str, float] = {}
        row_count = 0
        usd_to_cad = 1 / cad_to_usd if cad_to_usd else 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= KSB1_COL_CURRENCY:
                continue
            gl_code = str(row[KSB1_COL_GL_CODE]) if row[KSB1_COL_GL_CODE] else ""
            if not any(gl_code.startswith(p) for p in TRAVEL_GL_PREFIXES):
                continue

            cc = str(row[KSB1_COL_COST_CENTER]) if row[KSB1_COL_COST_CENTER] else ""
            try:
                amount = float(row[KSB1_COL_CAD] or 0)
            except (ValueError, TypeError):
                continue
            currency = str(row[KSB1_COL_CURRENCY]) if row[KSB1_COL_CURRENCY] else "CAD"
            cad = amount * usd_to_cad if currency == "USD" else amount

            if cc:
                totals[cc] = totals.get(cc, 0) + cad
                row_count += 1

        log.info("Extracted %d travel expense rows from %s → %d cost centers (CAD output, rate=%.4f)",
                 row_count, ksb1_path.name, len(totals), cad_to_usd)
        return totals
    finally:
        wb.close()


def extract_travel_from_multiple(
    ksb1_paths: list[Path],
    cad_to_usd: float = DEFAULT_CAD_TO_USD,
) -> dict[str, float]:
    """Aggregate travel expenses across multiple KSB1 files."""
    combined: dict[str, float] = {}
    for path in ksb1_paths:
        if not path.exists():
            log.warning("KSB1 file not found, skipping: %s", path)
            continue
        for cc, total in extract_travel_expenses(path, cad_to_usd=cad_to_usd).items():
            combined[cc] = combined.get(cc, 0) + total
    return combined


def extract_store_revenue(qbi_path: Path) -> dict[str, float]:
    """Extract total revenue per store from a QBI daily report xlsx.

    Returns {store_name: total_revenue} where store_name matches config entities.
    Sums 营业收入(不含税) across all dates in the file.
    """
    wb = openpyxl.load_workbook(qbi_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        totals: dict[str, float] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= QBI_COL_REVENUE:
                continue
            store_raw = str(row[QBI_COL_STORE]) if row[QBI_COL_STORE] else ""
            try:
                revenue = float(row[QBI_COL_REVENUE] or 0)
            except (ValueError, TypeError):
                continue

            entity_name = QBI_STORE_MAP.get(store_raw)
            if entity_name:
                totals[entity_name] = totals.get(entity_name, 0) + revenue

        log.info("Extracted revenue for %d stores from %s",
                 len(totals), qbi_path.name)
        return totals
    finally:
        wb.close()


def extract_revenue_from_multiple(qbi_paths: list[Path]) -> dict[str, float]:
    """Aggregate store revenue across multiple QBI files."""
    combined: dict[str, float] = {}
    for path in qbi_paths:
        if not path.exists():
            log.warning("QBI file not found, skipping: %s", path)
            continue
        for store, total in extract_store_revenue(path).items():
            combined[store] = combined.get(store, 0) + total
    return combined
