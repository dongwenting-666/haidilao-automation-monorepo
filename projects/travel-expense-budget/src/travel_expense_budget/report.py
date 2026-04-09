"""Travel expense budget report generation."""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from travel_expense_budget.config import ALL_ENTITIES, DEPARTMENTS, STORES, Entity

log = logging.getLogger(__name__)


@dataclass
class EntityRow:
    """Computed data for one row in the report."""
    name: str
    section: str
    prev_year_revenue: float
    prev_year_travel: float
    prev_year_ratio: float
    q1_revenue: float         # 26年Q1实际收入
    remaining_target: float   # 26年4-12月目标收入
    q1_budget: float          # Q1差旅费额度 = Q1 revenue × ratio
    full_year_budget: float   # 全年预算 = full target × ratio
    ytd_actual: float         # YTD已发生差旅费


def compute_report(
    curr_year_travel: dict[str, float],
    db_data: dict[str, dict],
    report_month: int,
    prev_year: int = 2025,
    curr_year: int = 2026,
) -> list[EntityRow]:
    """Compute the travel expense budget report rows.

    Uses coworker methodology: budget split into Q1 (actual revenue) + remaining (target).

    Args:
        curr_year_travel: {cost_center: total_usd} for current year YTD (from KSB1)
        db_data: {store_name: {target_revenue, prev_year_revenue, prev_year_travel, q1_revenue}}
        report_month: Month number (1-12) that is the last month included in YTD
    """
    # Gather store-level totals from DB
    total_prev_revenue = sum(
        db_data.get(e.name, {}).get("prev_year_revenue", 0) for e in STORES
        if db_data.get(e.name, {}).get("prev_year_revenue", 0) > 0
    )
    total_store_travel = sum(
        db_data.get(e.name, {}).get("prev_year_travel", 0) for e in STORES
    )
    total_target_revenue = sum(
        db_data.get(e.name, {}).get("target_revenue", 0) for e in STORES
    )
    total_q1_revenue = sum(
        db_data.get(e.name, {}).get("q1_revenue", 0) for e in STORES
    )

    # Stores: uniform ratio (rounded to 4 dp)
    store_ratio = round(total_store_travel / total_prev_revenue, 4) if total_prev_revenue else 0

    # Dept ratio = dept-only travel / total store revenue
    total_dept_travel = sum(
        db_data.get(e.name, {}).get("prev_year_travel", 0) for e in DEPARTMENTS
    )
    dept_ratio = total_dept_travel / total_prev_revenue if total_prev_revenue else 0

    # Growth factor for department budget scaling
    growth_factor = total_target_revenue / total_prev_revenue if total_prev_revenue else 1

    rows: list[EntityRow] = []

    for entity in ALL_ENTITIES:
        d = db_data.get(entity.name, {})
        cy_travel = sum(curr_year_travel.get(cc, 0) for cc in entity.cost_centers)

        if entity.section == "营业门店":
            py_rev = d.get("prev_year_revenue", 0)
            py_travel = d.get("prev_year_travel", 0)
            q1_rev = d.get("q1_revenue", 0)
            tgt_rev = d.get("target_revenue", 0)
            remaining_tgt = tgt_rev - q1_rev if tgt_rev > q1_rev else 0
            ratio = store_ratio
            q1_budget = q1_rev * store_ratio
            # "全年预算" = remaining months budget (target minus Q1 actual)
            remaining_budget = remaining_tgt * store_ratio
        else:
            py_rev = 0
            py_travel = d.get("prev_year_travel", 0)
            q1_rev = total_q1_revenue
            tgt_rev = total_target_revenue
            remaining_tgt = tgt_rev - q1_rev if tgt_rev > q1_rev else 0
            ratio = dept_ratio
            full_annual = py_travel * growth_factor
            # Depts: Q1 budget = proportional by month count (no individual revenue)
            q1_budget = full_annual * (report_month / 12)
            # Depts: "全年预算" column shows full annual budget
            remaining_budget = full_annual

        rows.append(EntityRow(
            name=entity.name,
            section=entity.section,
            prev_year_revenue=py_rev,
            prev_year_travel=py_travel,
            prev_year_ratio=ratio,
            q1_revenue=q1_rev,
            remaining_target=remaining_tgt,
            q1_budget=q1_budget,
            full_year_budget=remaining_budget,
            ytd_actual=cy_travel,
        ))

    return rows


def generate_excel(
    rows: list[EntityRow],
    output_path: Path,
    report_month: int,
    prev_year: int = 2025,
    curr_year: int = 2026,
) -> Path:
    """Generate the Excel report matching the coworker's format.

    Columns:
    A: section label (merged vertically)
    B: entity name
    C: prev year revenue
    D: prev year travel
    E: ratio (merged vertically per section)
    F: Q1 actual revenue
    G: 4-12月 target revenue
    H: Q1 travel budget
    I: full year budget
    J: YTD actual travel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差旅费预算明细"

    # Styles
    font_normal = Font(name="宋体", size=11)
    font_bold = Font(name="宋体", size=11, bold=True)
    font_title = Font(name="宋体", size=14, bold=True)
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    num_fmt = "#,##0.00"
    pct_fmt = "0.00%"

    # ── Row 1: Title + currency ──────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = f"加拿大{curr_year}年差旅费预算明细"
    ws["A1"].font = font_title
    ws.cell(row=1, column=10, value="货币单位：美元")
    ws.cell(row=1, column=10).font = font_normal
    ws.cell(row=1, column=10).alignment = Alignment(horizontal="right")

    # ── Row 2: Headers ───────────────────────────────────────────────────
    headers = {
        1: "部门", 2: "",
        3: f"{prev_year}年收入",
        4: f"{prev_year}年差旅费",
        5: f"{prev_year}年差旅费占比",
        6: f"{curr_year}年1-{report_month}月收入",
        7: f"{curr_year}年{report_month + 1}-12月目标收入" if report_month < 12 else f"{curr_year}年目标收入",
        8: f"{curr_year}年1-{report_month}月份差旅费额度",
        9: f"{curr_year}年差旅费预算（按{prev_year}年占比测算）",
        10: f"{curr_year}年1-{report_month}月份已发生差旅费",
    }
    ws.merge_cells("A2:B2")
    for col, text in headers.items():
        if col == 2:
            continue
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = font_bold
        cell.fill = header_fill
        cell.border = border_all
    ws.cell(row=2, column=2).font = font_bold
    ws.cell(row=2, column=2).fill = header_fill
    ws.cell(row=2, column=2).border = border_all

    # ── Data rows ────────────────────────────────────────────────────────
    data_start = 3
    data_row = data_start
    section_ranges: dict[str, tuple[int, int]] = {}
    current_section = None
    section_start = data_start

    store_rows = [r for r in rows if r.section == "营业门店"]

    for entity_row in rows:
        if entity_row.section != current_section:
            if current_section:
                section_ranges[current_section] = (section_start, data_row - 1)
            current_section = entity_row.section
            section_start = data_row

        is_store = entity_row.section == "营业门店"
        is_first_dept = (entity_row.section == "职能部门" and entity_row.name == DEPARTMENTS[0].name)

        ws.cell(row=data_row, column=2, value=entity_row.name)

        # Col C: prev year revenue (stores only)
        if is_store and entity_row.prev_year_revenue:
            ws.cell(row=data_row, column=3, value=entity_row.prev_year_revenue)

        # Col D: prev year travel
        ws.cell(row=data_row, column=4, value=entity_row.prev_year_travel)

        # Col F: Q1 revenue (stores only)
        if is_store and entity_row.q1_revenue:
            ws.cell(row=data_row, column=6, value=entity_row.q1_revenue)

        # Col G: remaining target (stores + first dept)
        if is_store:
            ws.cell(row=data_row, column=7, value=entity_row.remaining_target)
        elif is_first_dept:
            ws.cell(row=data_row, column=7, value=entity_row.remaining_target)

        # Col H: Q1 budget
        ws.cell(row=data_row, column=8, value=entity_row.q1_budget)

        # Col I: full year budget
        ws.cell(row=data_row, column=9, value=entity_row.full_year_budget)

        # Col J: YTD actual
        ws.cell(row=data_row, column=10, value=entity_row.ytd_actual)

        data_row += 1

    if current_section:
        section_ranges[current_section] = (section_start, data_row - 1)

    # ── Merged cells: section labels (A), ratio (E), dept columns ────────
    for section_name, (first, last) in section_ranges.items():
        ws.cell(row=first, column=1, value=section_name)
        ws.cell(row=first, column=1).font = font_bold
        ws.cell(row=first, column=1).alignment = Alignment(vertical="center")
        if last > first:
            ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)

        # Ratio merged per section
        if section_name == "营业门店":
            ratio_val = rows[0].prev_year_ratio
        else:
            ratio_val = rows[len(store_rows)].prev_year_ratio
        ws.cell(row=first, column=5, value=ratio_val)
        ws.cell(row=first, column=5).alignment = Alignment(vertical="center")
        if last > first:
            ws.merge_cells(start_row=first, start_column=5, end_row=last, end_column=5)

        # Dept section: merge remaining target column (same total for all)
        if section_name == "职能部门" and last > first:
            ws.merge_cells(start_row=first, start_column=7, end_row=last, end_column=7)

    # ── Total row ────────────────────────────────────────────────────────
    total_row = data_row
    ws.cell(row=total_row, column=2, value="加拿大大区")
    ws.cell(row=total_row, column=2).font = font_bold

    # Totals: store budgets + dept budgets (summed separately to avoid double-counting)
    stores = [r for r in rows if r.section == "营业门店"]
    depts = [r for r in rows if r.section == "职能部门"]
    total_prev_rev = sum(r.prev_year_revenue for r in stores if r.prev_year_revenue)

    ws.cell(row=total_row, column=3, value=total_prev_rev)
    ws.cell(row=total_row, column=4, value=sum(r.prev_year_travel for r in rows))
    ws.cell(row=total_row, column=5, value=sum(r.prev_year_travel for r in rows) / total_prev_rev if total_prev_rev else 0)
    ws.cell(row=total_row, column=6, value=sum(r.q1_revenue for r in stores))
    ws.cell(row=total_row, column=7, value=sum(r.remaining_target for r in stores))
    ws.cell(row=total_row, column=8, value=sum(r.q1_budget for r in stores) + sum(r.q1_budget for r in depts))
    ws.cell(row=total_row, column=9, value=sum(r.full_year_budget for r in stores) + sum(r.full_year_budget for r in depts))
    ws.cell(row=total_row, column=10, value=sum(r.ytd_actual for r in rows))

    # ── Apply borders, fonts, number formats ─────────────────────────────
    for row_idx in range(data_start, total_row + 1):
        for col in range(1, 11):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = font_normal
            cell.border = border_all
            if col == 5:
                cell.number_format = pct_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col >= 3:
                cell.number_format = num_fmt

    double_top = Border(left=thin, right=thin, top=Side(style="double"), bottom=thin)
    for col in range(1, 11):
        cell = ws.cell(row=total_row, column=col)
        cell.border = double_top
        cell.font = font_bold

    # ── Column widths ────────────────────────────────────────────────────
    widths = {"A": 12, "B": 27, "C": 15, "D": 18, "E": 16, "F": 20, "G": 24, "H": 28, "I": 37, "J": 28}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Report saved to %s", output_path)
    return output_path
