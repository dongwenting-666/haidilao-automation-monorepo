"""Store working-hour data collection.

Daily at 6:30 AM Vancouver time:
1. Check if a monthly spreadsheet exists for the current month in the target folder.
   If not, copy the template to create one.
2. Check columns D (翻台率) and E (总桌数) for all dates from day 1 to T-2.
   For each missing date, load from the generated daily report XLSX and fill in.
3. Check which blue columns (F–K: staffing data) are still empty for past dates.
   Alert stores (store_hours chat) if any are unfilled; completely silent if everything is fine.

Template: https://haidilao.feishu.cn/sheets/SbTns7kTxhxn5TtLMyccOrqRnqe
Folder:   https://haidilao.feishu.cn/drive/folder/AVt8fGZLHl5PzJd2gw3cNa10ntd

Notification routing (configured in server/notify.toml [store-hours-collect]):
    chat       = "hongming"    # data-fill summary → admin
    alert_chat = "store_hours" # unfilled alert    → store group
    all stores filled → silent

    Env vars HOURS_NOTIFY_CHAT_ID / HOURS_ALERT_CHAT_ID override toml for ad-hoc runs.

Environment variables:
    LARK_APP_ID / LARK_APP_SECRET   Feishu bot credentials
    HOURS_NOTIFY_CHAT_ID            (optional) override for data-fill summary chat
    HOURS_ALERT_CHAT_ID             (optional) override for unfilled-store alert chat
    HOURS_TEMPLATE_TOKEN            Template spreadsheet token (default provided)
    HOURS_FOLDER_TOKEN              Target folder token (default provided)
"""

from __future__ import annotations

import calendar
import logging
import os
import sys
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv

from lark_client import LarkClient, chat_id_for, command_chat_for

logger = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────

TEMPLATE_TOKEN = "SbTns7kTxhxn5TtLMyccOrqRnqe"
FOLDER_TOKEN   = "AVt8fGZLHl5PzJd2gw3cNa10ntd"

STORES = [
    "加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店",
    "加拿大五店", "加拿大六店", "加拿大七店", "加拿大八店",
]

# Layout: Row 4-5 = headers, Row 6 = day 1 of month
_HEADER_ROWS = 5


# ── Helpers ───────────────────────────────────────────────────────────────────

def _excel_serial(d: date) -> int:
    return (d - date(1900, 1, 1)).days + 2


def _weekday_cn(d: date) -> str:
    return ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"][d.weekday()]


def _month_file_name(year: int, month: int) -> str:
    return f"加拿大门店用工数据跟踪-{year}{month:02d}"


def _output_dir() -> Path:
    """Resolve the output directory for daily reports."""
    p = Path(__file__).resolve().parent
    while p != p.parent:
        if (p / "pyproject.toml").exists() and "[tool.uv.workspace]" in (p / "pyproject.toml").read_text():
            return p / "output" / "daily-report"
        p = p.parent
    return Path.cwd() / "output" / "daily-report"


# ── Load data from daily report XLSX ──────────────────────────────────────────

def load_daily_data(report_date: date) -> tuple[dict[str, float], dict[str, float]]:
    """Load turnover rate and table count from the generated daily report XLSX.

    Returns (turnover_by_store, tables_by_store). Empty dicts if file not found.
    """
    import openpyxl

    report_path = _output_dir() / f"database_report_{report_date.year}_{report_date.month:02d}_{report_date.day:02d}.xlsx"
    if not report_path.exists():
        logger.warning("Daily report not found: %s", report_path)
        return {}, {}

    wb = openpyxl.load_workbook(report_path, data_only=True)
    ws = wb.worksheets[0]  # 对比上月表

    # Row 2 columns C-J = store names in order
    stores = [ws.cell(2, c).value for c in range(3, 11)]

    # Row 3: 今日总桌数 (columns C-J, same order as stores)
    tables: dict[str, float] = {}
    for i, store in enumerate(stores):
        val = ws.cell(3, 3 + i).value
        if store and val is not None:
            tables[store] = float(val)

    # Row 25: 翻台率排名店铺 (store names, ranked)
    # Row 26: 翻台率排名 (values, ranked)
    ranked_stores = [ws.cell(25, c).value for c in range(3, 11)]
    ranked_values = [ws.cell(26, c).value for c in range(3, 11)]
    turnover: dict[str, float] = {}
    for store, val in zip(ranked_stores, ranked_values):
        if store and val is not None:
            turnover[store] = float(val)

    wb.close()
    logger.info("Loaded daily data from %s: %d stores", report_path.name, len(turnover))
    return turnover, tables


def ensure_daily_report(report_date: date) -> bool:
    """Ensure the daily report XLSX exists for the given date.

    If not cached on disk, generates it directly via subprocess (NOT via the
    server queue, to avoid deadlocking the serial execution queue).
    Returns True if the file is available after generation.
    """
    import subprocess

    report_path = _output_dir() / f"database_report_{report_date.year}_{report_date.month:02d}_{report_date.day:02d}.xlsx"
    if report_path.exists():
        return True

    logger.info("Generating daily report for %s...", report_date)
    repo_root = _output_dir().parent.parent
    try:
        result = subprocess.run(
            [
                "uv", "run",
                "--project", str(repo_root / "projects" / "daily-store-operation-report"),
                "python", "-m", "daily_store_operation_report.main",
                report_date.isoformat(),
            ],
            capture_output=True, text=True, timeout=300, cwd=str(repo_root),
        )
        if result.returncode != 0:
            logger.error("Daily report generation failed for %s:\n%s", report_date, result.stdout[-500:])
            return False
        logger.info("Daily report generated for %s", report_date)
        return report_path.exists()
    except subprocess.TimeoutExpired:
        logger.error("Daily report generation timed out for %s", report_date)
        return False
    except Exception as e:
        logger.error("Failed to generate daily report for %s: %s", report_date, e)
        return False


# ── Find or create monthly spreadsheet ────────────────────────────────────────

def find_monthly_sheet(client: LarkClient, folder_token: str, year: int, month: int) -> str | None:
    target_name = _month_file_name(year, month)
    for f in client.list_folder(folder_token):
        if target_name in f.get("name", ""):
            logger.info("Found existing sheet: %s (token=%s)", f["name"], f["token"])
            return f["token"]
    return None


def create_monthly_sheet(client: LarkClient, template_token: str, folder_token: str,
                         year: int, month: int) -> str:
    title = _month_file_name(year, month)
    logger.info("Creating monthly sheet: %s", title)

    data = client._post(f"/drive/v1/files/{template_token}/copy",
        {"name": title, "type": "sheet", "folder_token": folder_token})
    new_token = data["data"]["file"]["token"]
    logger.info("Created %s (token=%s)", title, new_token)

    # Get sheet tab IDs
    sheets_data = client._get(f"/sheets/v3/spreadsheets/{new_token}/sheets/query").json()
    if sheets_data.get("code") != 0:
        raise RuntimeError(f"Lark API error: {sheets_data.get('code')} {sheets_data.get('msg')}")
    new_sheets = {s["title"]: s["sheet_id"] for s in sheets_data["data"]["sheets"]}

    # Fill dates for each store tab
    days_in_month = calendar.monthrange(year, month)[1]
    for store in STORES:
        sheet_id = new_sheets.get(store)
        if not sheet_id:
            continue
        values = []
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            values.append([_excel_serial(d), _weekday_cn(d)])
        last_row = _HEADER_ROWS + days_in_month
        client._put(f"/sheets/v2/spreadsheets/{new_token}/values",
            {"valueRange": {"range": f"{sheet_id}!B{_HEADER_ROWS+1}:C{last_row}", "values": values}})
        logger.info("  %s: wrote %d date rows", store, days_in_month)

    return new_token


def get_sheet_tabs(client: LarkClient, sheet_token: str) -> dict[str, str]:
    """Get {store_name: sheet_id} mapping for a spreadsheet."""
    data = client._get(f"/sheets/v3/spreadsheets/{sheet_token}/sheets/query").json()
    if data.get("code") != 0:
        raise RuntimeError(f"Lark API error: {data.get('code')} {data.get('msg')}")
    return {s["title"]: s["sheet_id"] for s in data["data"]["sheets"]}


def _sheet_values(client: LarkClient, sheet_token: str, range_str: str) -> list:
    """Read a sheet range and return the values list (may be empty)."""
    data = client._get(f"/sheets/v2/spreadsheets/{sheet_token}/values/{range_str}").json()
    if data.get("code") != 0:
        raise RuntimeError(f"Lark API error: {data.get('code')} {data.get('msg')}")
    return data.get("data", {}).get("valueRange", {}).get("values", [])


# ── Fill turnover and table data ──────────────────────────────────────────────

def fill_missing_data(client: LarkClient, sheet_token: str, year: int, month: int,
                      up_to_date: date) -> list[date]:
    """Check and fill columns D/E for all dates from day 1 up to *up_to_date*.

    Returns the list of dates that were newly filled.
    """
    store_sheets = get_sheet_tabs(client, sheet_token)
    filled_dates: list[date] = []

    # Read existing D/E data for the first store to find which days are already filled
    first_store = STORES[0]
    first_id = store_sheets.get(first_store)
    if not first_id:
        return []

    last_day = up_to_date.day
    existing_rows = _sheet_values(
        client, sheet_token, f"{first_id}!D{_HEADER_ROWS+1}:E{_HEADER_ROWS+last_day}"
    )

    for day in range(1, last_day + 1):
        d = date(year, month, day)
        row_idx = day - 1  # 0-based in the values list

        # Check if this day already has data
        if row_idx < len(existing_rows):
            row = existing_rows[row_idx]
            if row and len(row) >= 2 and row[0] is not None and row[0] != "":
                continue  # already filled

        # Need to fill this date — ensure daily report exists
        if not ensure_daily_report(d):
            logger.warning("Skipping %s — daily report unavailable", d)
            continue

        turnover, tables = load_daily_data(d)
        if not turnover:
            logger.warning("Skipping %s — no data in report", d)
            continue

        # Write to all store tabs
        row_num = _HEADER_ROWS + day
        for store in STORES:
            sid = store_sheets.get(store)
            if not sid:
                continue
            t = turnover.get(store, 0)
            tb = tables.get(store, 0)
            client._put(f"/sheets/v2/spreadsheets/{sheet_token}/values",
                {"valueRange": {"range": f"{sid}!D{row_num}:E{row_num}", "values": [[t, tb]]}})

        filled_dates.append(d)
        logger.info("Filled %s: %d stores", d, len(turnover))

    return filled_dates


# ── Check unfilled blue columns ───────────────────────────────────────────────

def check_unfilled(client: LarkClient, sheet_token: str, year: int, month: int,
                   up_to_date: date) -> dict[str, list[date]]:
    """Check which stores have unfilled blue columns (F-K) for past dates."""
    store_sheets = get_sheet_tabs(client, sheet_token)
    unfilled: dict[str, list[date]] = {}
    last_day = up_to_date.day

    for store in STORES:
        sid = store_sheets.get(store)
        if not sid:
            continue

        rows = _sheet_values(
            client, sheet_token, f"{sid}!F{_HEADER_ROWS+1}:K{_HEADER_ROWS+last_day}"
        )

        missing = []
        for day_offset in range(last_day):
            row = rows[day_offset] if day_offset < len(rows) else []
            all_empty = all((c is None or c == "" or c == 0) for c in (row if row else []))
            if all_empty:
                missing.append(date(year, month, day_offset + 1))

        if missing:
            unfilled[store] = missing

    return unfilled


# ── Notifications ─────────────────────────────────────────────────────────────

def send_data_summary(client: LarkClient, chat_id: str, year: int, month: int,
                      filled_dates: list[date], turnover: dict[str, float],
                      tables: dict[str, float]) -> None:
    lines = [f"**✅ {year}年{month}月 用工表 翻台率/总桌数已自动填入**\n"]
    lines.append(f"新填入日期：{', '.join(d.strftime('%m/%d') for d in filled_dates)}\n")
    if turnover:
        for store in STORES:
            t = turnover.get(store, 0)
            tb = tables.get(store, 0)
            lines.append(f"▸ {store}：翻台率 {t:.2f}　总桌数 {tb:.0f}")
    client.send_card(
        title="📊 用工表数据已更新",
        content="\n".join(lines),
        chat_id=chat_id,
        color="green",
    )


def send_unfilled_alert(client: LarkClient, chat_id: str, year: int, month: int,
                        unfilled: dict[str, list[date]], sheet_url: str) -> None:
    lines = [f"**📋 {year}年{month}月 用工数据表 — 以下门店有未填写数据：**\n"]
    for store, dates in sorted(unfilled.items()):
        date_strs = ", ".join(d.strftime("%m/%d") for d in dates[:7])
        extra = f" 等{len(dates)}天" if len(dates) > 7 else ""
        lines.append(f"▸ **{store}**：{date_strs}{extra}")
    lines.append(f"\n[👉 点击填写]({sheet_url})")
    lines.append("\n> 请门店尽快完成蓝色列数据填写")
    client.send_card(
        title=f"⚠️ {month}月用工数据未填写提醒",
        content="\n".join(lines),
        chat_id=chat_id,
        color="yellow",
    )
    logger.info("Sent unfilled alert for %d stores", len(unfilled))


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    import argparse

    load_dotenv()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    parser = argparse.ArgumentParser(description="Store working-hour data collection")
    parser.add_argument("--date", type=str, default=None, help="Target date YYYY-MM-DD (default: T-2)")
    args = parser.parse_args()

    app_id = os.environ.get("LARK_APP_ID", "")
    app_secret = os.environ.get("LARK_APP_SECRET", "")
    template_token = os.environ.get("HOURS_TEMPLATE_TOKEN", TEMPLATE_TOKEN)
    folder_token = os.environ.get("HOURS_FOLDER_TOKEN", FOLDER_TOKEN)

    # Chat routing — configured in server/notify.toml [store-hours-collect]
    # Env vars override toml for ad-hoc testing; toml is the source of truth for production.
    summary_chat_id = os.environ.get("HOURS_NOTIFY_CHAT_ID") or command_chat_for("store-hours-collect", "chat")        or ""
    alert_chat_id   = os.environ.get("HOURS_ALERT_CHAT_ID")  or command_chat_for("store-hours-collect", "alert_chat")  or ""

    if not app_id or not app_secret:
        logger.error("LARK_APP_ID and LARK_APP_SECRET must be set")
        sys.exit(1)
    if not summary_chat_id:
        logger.error("notify.toml [store-hours-collect] 'chat' not set and HOURS_NOTIFY_CHAT_ID not provided")
        sys.exit(1)
    if not alert_chat_id:
        logger.warning("notify.toml [store-hours-collect] 'alert_chat' not set and HOURS_ALERT_CHAT_ID not provided; unfilled alerts will be suppressed")

    # T-2: if today is 2026-03-18, target date = 2026-03-16
    target_date = date.fromisoformat(args.date) if args.date else date.today() - timedelta(days=2)
    year, month = target_date.year, target_date.month
    logger.info("Target date: %s (month: %04d-%02d)", target_date, year, month)

    with LarkClient(app_id=app_id, app_secret=app_secret) as client:
        # Step 1: Find or create monthly spreadsheet
        sheet_token = find_monthly_sheet(client, folder_token, year, month)
        if not sheet_token:
            sheet_token = create_monthly_sheet(client, template_token, folder_token, year, month)

        sheet_url = f"https://haidilao.feishu.cn/sheets/{sheet_token}"

        # Step 2: Fill missing D/E data for all dates from day 1 to target_date
        filled_dates = fill_missing_data(client, sheet_token, year, month, target_date)
        if filled_dates:
            turnover, tables = load_daily_data(filled_dates[-1])
            send_data_summary(client, summary_chat_id, year, month, filled_dates, turnover, tables)
        else:
            logger.info("All dates already filled for D/E columns")

        # Step 3: Check unfilled blue columns; alert admin only, silent if all good
        unfilled = check_unfilled(client, sheet_token, year, month, target_date)
        if unfilled:
            if alert_chat_id:
                send_unfilled_alert(client, alert_chat_id, year, month, unfilled, sheet_url)
            else:
                logger.warning("Unfilled stores detected (%d) but alert_chat_id not configured — skipping", len(unfilled))
        else:
            logger.info("All stores have filled staffing data up to %s ✓", target_date)

    logger.info("Done")


if __name__ == "__main__":
    main()
