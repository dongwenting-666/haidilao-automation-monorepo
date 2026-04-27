from __future__ import annotations

import argparse
from copy import copy
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

SHEET_NAME = "加拿大片区假想敌外卖收入对比"


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[4]


def _daily_output_dir() -> Path:
    return _repo_root() / "output" / "daily-report"


def _default_report_date() -> date:
    return date.today() - timedelta(days=2)


def _default_source(report_date: date | None) -> Path:
    daily_dir = _daily_output_dir()
    effective_date = report_date or _default_report_date()
    candidate = daily_dir / f"database_report_{effective_date.year}_{effective_date.month:02d}_{effective_date.day:02d}.xlsx"
    if candidate.exists():
        return candidate
    raise FileNotFoundError(f"Daily report not found for {effective_date.isoformat()}: {candidate}")


def _copy_sheet(source_ws, target_ws) -> None:
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    for key, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dim.width
        target_ws.column_dimensions[key].hidden = dim.hidden

    for key, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key].height = dim.height
        target_ws.row_dimensions[key].hidden = dim.hidden

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.freeze_panes = source_ws.freeze_panes


def export_competitor_takeout_sheet(source_path: Path, output_dir: Path) -> Path:
    wb = load_workbook(source_path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_NAME}")

    src_ws = wb[SHEET_NAME]
    out_wb = Workbook()
    default_ws = out_wb.active
    out_wb.remove(default_ws)
    out_ws = out_wb.create_sheet(SHEET_NAME)
    _copy_sheet(src_ws, out_ws)

    output_dir.mkdir(parents=True, exist_ok=True)
    report_date = _report_date_from_filename(source_path)
    filename = f"加拿大片区假想敌外卖收入对比_{report_date.year}-{report_date.month:02d}-{report_date.day:02d}.xlsx"
    output_path = output_dir / filename
    out_wb.save(output_path)
    return output_path


def _report_date_from_filename(path: Path) -> date:
    stem = path.stem
    parts = stem.rsplit("_", 3)
    if len(parts) >= 4:
        try:
            return date(int(parts[-3]), int(parts[-2]), int(parts[-1]))
        except ValueError:
            pass
    return date.today()


def main() -> None:
    parser = argparse.ArgumentParser(description="Export competitor takeout comparison sheet from daily report")
    parser.add_argument("date", nargs="?", help="Report date in YYYY-MM-DD; defaults to T-2")
    parser.add_argument("--source", help="Explicit source daily report path")
    parser.add_argument("--output-dir", help="Output directory for the exported xlsx")
    args = parser.parse_args()

    report_date = date.fromisoformat(args.date) if args.date else None
    source_path = Path(args.source) if args.source else _default_source(report_date)
    output_dir = Path(args.output_dir) if args.output_dir else (_daily_output_dir() / "competitor-takeout")

    output_path = export_competitor_takeout_sheet(source_path, output_dir)
    print(f"Report saved to {output_path}")


if __name__ == "__main__":
    main()
