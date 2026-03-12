"""All openpyxl fill/font/border/alignment constants and helpers."""

from __future__ import annotations

import functools

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

# ── Shared ────────────────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

BOLD = Font(bold=True, size=11)
BOLD_TITLE = Font(bold=True, size=12)
BOLD_LARGE = Font(bold=True, size=20, color="FFFF0000")
RED_FONT = Font(color="FFFF0000", size=11)
WHITE_BOLD = Font(color="FFFFFFFF", bold=True, size=11)
WHITE_BOLD_TITLE = Font(color="FFFFFFFF", bold=True, size=12)

# ── Sheet 1 & 2: Gold/MoM theme ──────────────────────────────────────────────
GOLD_FILL = PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF99", end_color="FFFFFF99", fill_type="solid")
BRIGHT_YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
LIGHT_BLUE_FILL_GOLD = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")

# ── Sheet 3: Blue/YoY theme ──────────────────────────────────────────────────
DARK_BLUE_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
MED_BLUE_FILL = PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="FFB4C7E7", end_color="FFB4C7E7", fill_type="solid")
PALE_BLUE_FILL = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")

# ── Sheet 4: Per-store colors ─────────────────────────────────────────────────
STORE_COLORS: dict[str, str] = {
    "加拿大一店": "FFFFFF99",
    "加拿大二店": "FFE6F3FF",
    "加拿大三店": "FFFFE6E6",
    "加拿大四店": "FFE6FFE6",
    "加拿大五店": "FFFFE6CC",
    "加拿大六店": "FFF0E6FF",
    "加拿大七店": "FFE6FFFF",
    "加拿大八店": "FFFFFF99",
}

GRAY_FILL = PatternFill(start_color="FFD0D0D0", end_color="FFD0D0D0", fill_type="solid")
NAVY_FILL = PatternFill(start_color="FF000080", end_color="FF000080", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


@functools.lru_cache(maxsize=None)
def store_fill(store: str) -> PatternFill:
    """Get the fill color for a store in Sheet 4."""
    color = STORE_COLORS.get(store, "FFFFFFFF")
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def apply_border(ws: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply thin borders to a rectangular range."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def apply_fill_row(ws: Worksheet, row: int, fill: PatternFill, min_col: int = 1, max_col: int | None = None) -> None:
    """Apply a fill to all cells in a row."""
    if max_col is None:
        max_col = ws.max_column
    for col in range(min_col, max_col + 1):
        ws.cell(row=row, column=col).fill = fill


def apply_fill_range(ws: Worksheet, min_row: int, max_row: int, fill: PatternFill, min_col: int = 1, max_col: int | None = None) -> None:
    """Apply a fill to a rectangular range."""
    if max_col is None:
        max_col = ws.max_column
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).fill = fill
