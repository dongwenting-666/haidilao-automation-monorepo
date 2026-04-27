"""Competitor comparison sheets.

Layout
------
Turnover sheet:
    Row 1 : Title spanning all 6 columns
    Row 2 : Column headers
    Rows 3–10 : One row per store (8 stores total)
    Row 11 : Footer note

Takeout revenue sheet:
    Row 1 : Title spanning all 6 columns
    Row 2 : Column headers
    Row 3 : 七店子店 Snappy vs 四店外卖 BI
    Row 4 : Footer note

Columns
-------
A  门店                      Store name
B  假想敌                    Competitor store name
C  2月份翻台率差异            Prev-month full-month turnover delta (store − competitor)
D  3月截止目前门店翻台率       MTD turnover rate (current store)
E  3月截止目前假想敌翻台率     MTD turnover rate (competitor store)
F  差异对比                  MTD delta (D − E)
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from daily_store_operation_report.constants import STORES, WAN_DIVISOR
from daily_store_operation_report.sheets.styles import (
    BOLD,
    CENTER,
    WHITE_BOLD,
    apply_border,
)
from daily_store_operation_report.transform import ReportData

# Header fill — teal to match the screenshot's style
_HEADER_FILL = PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid")
_TITLE_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
_ALT_FILL = PatternFill(start_color="FFDEEAF6", end_color="FFDEEAF6", fill_type="solid")

_COL_WIDTHS = [16, 16, 20, 22, 22, 12]
_TURNOVER_SHEET_NAME = "加拿大片区假想敌翻台率对比"
_TAKEOUT_SHEET_NAME = "加拿大片区假想敌外卖收入对比"


def _write_placeholder(ws) -> None:
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "假想敌配置未设置 — 请前往管理后台配置：/admin/competitors"
    cell.font = Font(italic=True, color="FF888888", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30


def _apply_table_header(ws, title: str, headers: list[str]) -> None:
    for i, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color="FFFFFFFF")
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = CENTER
    apply_border(ws, 1, 1, 1, 6)

    ws.row_dimensions[2].height = 36
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = WHITE_BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = CENTER
    apply_border(ws, 2, 2, 1, 6)


def build_competitor_sheet(wb: Workbook, data: ReportData) -> None:
    """Add the competitor turnover comparison sheet to *wb*.

    If ``data.competitor`` is empty (DB not configured), the sheet is still
    created but shows a placeholder message instead of data rows.
    """
    ws = wb.create_sheet(_TURNOVER_SHEET_NAME)

    # If no competitor config, show a placeholder and return early.
    if not data.competitor:
        _write_placeholder(ws)
        return

    dates = data.dates
    prev_month = dates.prev_end.month
    cur_month = dates.report_date.month

    headers = [
        "门店",
        "假想敌",
        f"{prev_month}月份翻台率差异",
        f"{cur_month}月截止目前门店翻台率",
        f"{cur_month}月截止目前假想敌翻台率",
        "差异对比",
    ]
    _apply_table_header(ws, _TURNOVER_SHEET_NAME, headers)

    # ── Rows 3–10: Data ──────────────────────────────────────────────────────
    competitor_map = data.competitor

    for row_idx, store in enumerate(STORES, start=3):
        ws.row_dimensions[row_idx].height = 20
        competitor = competitor_map.get(store, "—")

        # Pull turnover values from the pre-computed store metrics
        store_metrics = data.stores.get(store)
        comp_metrics = data.stores.get(competitor)

        # Previous month full-month average turnover rate
        prev_store = store_metrics.prev_mtd_turnover_rate if store_metrics else 0.0
        prev_comp = comp_metrics.prev_mtd_turnover_rate if comp_metrics else 0.0
        prev_diff = prev_store - prev_comp

        # MTD current month
        mtd_store = store_metrics.mtd_turnover_rate if store_metrics else 0.0
        mtd_comp = comp_metrics.mtd_turnover_rate if comp_metrics else 0.0
        mtd_diff = mtd_store - mtd_comp

        row_values = [store, competitor, prev_diff, mtd_store, mtd_comp, mtd_diff]
        fill = _ALT_FILL if row_idx % 2 == 0 else None

        for col, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = CENTER
            cell.font = BOLD
            if fill:
                cell.fill = fill
            # Colour negative differences red
            if col in (3, 6) and isinstance(val, float) and val < 0:
                cell.font = Font(bold=True, color="FFFF0000", size=11)

        apply_border(ws, row_idx, row_idx, 1, 6)

    # ── Footer note ──────────────────────────────────────────────────────────
    footer_row = len(STORES) + 3
    ws.row_dimensions[footer_row].height = 18
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    note = ws[f"A{footer_row}"]
    note.value = "每周一公布截止到周五的，比如22号公布1-20号"
    note.font = Font(italic=True, size=10, color="FF595959")
    note.alignment = Alignment(horizontal="left", vertical="center")
    apply_border(ws, footer_row, footer_row, 1, 6)


def build_competitor_takeout_sheet(wb: Workbook, data: ReportData) -> None:
    """Add the takeout revenue comparison sheet to *wb*."""
    ws = wb.create_sheet(_TAKEOUT_SHEET_NAME)

    if not data.competitor:
        _write_placeholder(ws)
        return

    dates = data.dates
    prev_month = dates.prev_end.month
    cur_month = dates.report_date.month

    headers = [
        "门店",
        "假想敌",
        f"{prev_month}月份收入差异",
        f"{cur_month}月截止目前门店收入",
        f"{cur_month}月截止目前假想敌收入",
        "差异对比",
    ]
    _apply_table_header(ws, _TAKEOUT_SHEET_NAME, headers)

    ws.row_dimensions[3].height = 20
    snappy = data.snappy
    store4_metrics = data.stores.get("加拿大四店")

    snappy_cur_mtd_wan = snappy.mtd_net_sales / WAN_DIVISOR
    snappy_prev_full_wan = snappy.prev_full_net_sales / WAN_DIVISOR
    store4_cur_takeout_wan = store4_metrics.mtd_takeout_wan if store4_metrics else 0.0
    store4_prev_takeout_wan = store4_metrics.prev_full_takeout_wan if store4_metrics else 0.0

    row_values = [
        "加拿大七店子店麻辣烫收入",
        "加拿大四店外卖收入",
        snappy_prev_full_wan - store4_prev_takeout_wan,
        snappy_cur_mtd_wan,
        store4_cur_takeout_wan,
        snappy_cur_mtd_wan - store4_cur_takeout_wan,
    ]

    for col, val in enumerate(row_values, start=1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.alignment = CENTER
        cell.font = BOLD
        cell.fill = _ALT_FILL
        if col in (3, 6) and isinstance(val, float) and val < 0:
            cell.font = Font(bold=True, color="FFFF0000", size=11)
    apply_border(ws, 3, 3, 1, 6)

    ws.row_dimensions[4].height = 18
    ws.merge_cells("A4:F4")
    note = ws["A4"]
    note.value = "七店子店麻辣烫收入来自 Snappy；四店外卖收入来自 BI（营业收入(外卖)+营业收入(外送)）。"
    note.font = Font(italic=True, size=10, color="FF595959")
    note.alignment = Alignment(horizontal="left", vertical="center")
    apply_border(ws, 4, 4, 1, 6)
