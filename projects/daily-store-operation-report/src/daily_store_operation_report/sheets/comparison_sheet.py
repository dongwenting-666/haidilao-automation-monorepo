"""Shared builder for comparison sheets (MoM Sheet 1, YoY Sheet 3).

Eliminates ~85% code duplication between mom.py and yoy_detail.py by
parameterizing the comparison period and color theme.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Callable

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from daily_store_operation_report.constants import REGION_LABEL, STORES, WAN_DIVISOR, WEEKDAY_NAMES
from daily_store_operation_report.sheets.styles import (
    BOLD,
    BOLD_LARGE,
    CENTER,
    WHITE_FILL,
    apply_border,
    apply_fill_range,
    apply_fill_row,
)
from daily_store_operation_report.transform import ReportData, StoreMetrics
from daily_store_operation_report.utils import comp_text, div_or_zero, pct_str

_NCOLS = 11  # A..K


@dataclass
class SheetTheme:
    """Color/font theme for a comparison sheet."""

    title_font: Font
    header_font: Font
    header_fill: PatternFill
    section_a_fill: PatternFill  # 桌数 + 单桌消费 sections
    section_b_fill: PatternFill  # 收入 section + turnover alternating
    highlight_fill: PatternFill  # diff/target rows + turnover alternating


@dataclass
class ComparisonConfig:
    """What to compare against (previous month or previous year)."""

    sheet_name: str
    comp_type: str  # "环比" or "同比"
    comp_label: str  # "上月" or "上年"
    get_comp_tables: Callable[[StoreMetrics], float]
    get_comp_raw_tables: Callable[[StoreMetrics], float]
    get_comp_revenue_wan: Callable[[StoreMetrics], float]
    get_comp_per_table: Callable[[StoreMetrics], float]
    get_comp_dine_in_wan: Callable[[StoreMetrics], float]
    get_comp_takeout_wan: Callable[[StoreMetrics], float]
    theme: SheetTheme


def _write_row(
    ws: Worksheet, row: int, label: str, values: list, *, region_sum: bool = True
) -> None:
    """Write a label + per-store values row, with optional region sum in column K."""
    ws.cell(row=row, column=2, value=label)
    for i, v in enumerate(values):
        ws.cell(row=row, column=3 + i, value=v)
    if region_sum and len(values) == len(STORES):
        ws.cell(
            row=row,
            column=_NCOLS,
            value=sum(v for v in values if isinstance(v, (int, float))),
        )


def build_comparison_sheet(wb: Workbook, data: ReportData, config: ComparisonConfig) -> Worksheet:
    """Build a comparison sheet (MoM or YoY detail)."""
    ws: Worksheet = wb.create_sheet(config.sheet_name)
    dates = data.dates
    d = dates.report_date
    weekday = WEEKDAY_NAMES[d.weekday()]
    month = d.month
    day = d.day
    stores = STORES
    theme = config.theme
    comp = config.comp_label
    comp_type = config.comp_type

    # Row 1: Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"加拿大-各门店{d.year}年{month:02d}月{day:02d}日{comp_type}数据-{weekday}"
    ws["A1"].font = theme.title_font
    ws["A1"].alignment = CENTER

    # Row 2: Headers
    headers = ["项目", "内容"] + stores + [REGION_LABEL]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = theme.header_font
        c.alignment = CENTER

    # ── Section 1: 桌数(考核) — rows 3-8 ──
    ws.merge_cells("A3:A8")
    ws["A3"] = "桌数\n(考核)"
    ws["A3"].font = BOLD
    ws["A3"].alignment = CENTER

    _write_row(ws, 3, "今日总桌数", [data.stores[s].today_tables for s in stores])
    _write_row(ws, 4, "今日外卖桌数", [data.stores[s].today_takeout_tables for s in stores])
    _write_row(ws, 5, "今日未计入考核桌数", [data.stores[s].today_non_assessed_tables for s in stores])
    _write_row(ws, 6, f"{month}月总桌数", [data.stores[s].mtd_tables for s in stores])
    _write_row(
        ws, 7, f"{comp}同期总桌数", [config.get_comp_tables(data.stores[s]) for s in stores]
    )

    # Row 8: comparison diff
    diffs = [data.stores[s].mtd_tables - config.get_comp_tables(data.stores[s]) for s in stores]
    _write_row(ws, 8, f"对比{comp}同期总桌数", [comp_text(diff) for diff in diffs], region_sum=False)
    ws.cell(row=8, column=_NCOLS, value=comp_text(sum(diffs)))

    # ── Section 2: 收入(不含税-万加元) ──
    # MoM (环比): rows 9-23 (15 rows)
    # YoY (同比): rows 9-25 (17 rows) — 2 extra rows for 上年堂食/外卖
    is_yoy = comp_type == "同比"
    r = 9  # current row counter

    ws["A9"] = "收入\n(不含税-万加元)"
    ws["A9"].font = BOLD
    ws["A9"].alignment = CENTER

    _write_row(ws, r, "今日营业收入(万)", [data.stores[s].today_revenue_wan for s in stores]); r += 1

    if is_yoy:
        _write_row(ws, r, "今日堂食营业收入(万)", [data.stores[s].today_dine_in_wan for s in stores]); r += 1
        _write_row(ws, r, "今日外卖营业收入(万)", [data.stores[s].today_takeout_wan for s in stores]); r += 1
    else:
        _write_row(ws, r, "外卖收入", [data.stores[s].today_takeout_wan for s in stores]); r += 1
        _write_row(ws, r, "堂食收入", [data.stores[s].today_dine_in_wan for s in stores]); r += 1

    _write_row(ws, r, "本月截止目前营业收入(万)", [data.stores[s].mtd_revenue_wan for s in stores]); r += 1

    if is_yoy:
        _write_row(ws, r, "本月截止目前堂食营业收入(万)", [data.stores[s].mtd_dine_in_wan for s in stores]); r += 1
        _write_row(ws, r, "本月截止目前外卖营业收入(万)", [data.stores[s].mtd_takeout_wan for s in stores]); r += 1
    else:
        _write_row(ws, r, "本月截止目前堂食收入(万)", [data.stores[s].mtd_dine_in_wan for s in stores]); r += 1
        _write_row(ws, r, "本月截止目前外卖收入(万)", [data.stores[s].mtd_takeout_wan for s in stores]); r += 1

    _write_row(ws, r, f"{comp}截止目前营业收入(万)",
               [config.get_comp_revenue_wan(data.stores[s]) for s in stores]); r += 1

    if is_yoy:
        _write_row(ws, r, f"{comp}截止目前堂食营业收入(万)",
                   [config.get_comp_dine_in_wan(data.stores[s]) for s in stores]); r += 1
        _write_row(ws, r, f"{comp}截止目前外卖营业收入(万)",
                   [config.get_comp_takeout_wan(data.stores[s]) for s in stores]); r += 1

    _write_row(ws, r, f"{comp_type}营业收入变化(万)",
               [data.stores[s].mtd_revenue_wan - config.get_comp_revenue_wan(data.stores[s]) for s in stores]); r += 1

    if not is_yoy:
        # MoM: show takeout/dine-in change breakdown
        _write_row(ws, r, f"{comp_type}营业外卖收入变化(万)",
                   [data.stores[s].mtd_takeout_wan - config.get_comp_takeout_wan(data.stores[s]) for s in stores]); r += 1
        _write_row(ws, r, f"{comp_type}营业堂食收入变化(万)",
                   [data.stores[s].mtd_dine_in_wan - config.get_comp_dine_in_wan(data.stores[s]) for s in stores]); r += 1

    _write_row(ws, r, "本月营业收入目标(万)", [data.stores[s].revenue_target for s in stores]); r += 1

    # Target completion %
    vals_pct = [
        pct_str(div_or_zero(data.stores[s].mtd_revenue_wan, data.stores[s].revenue_target) * 100)
        for s in stores
    ]
    _write_row(ws, r, "本月截止目标完成率", vals_pct, region_sum=False)
    region_mtd = sum(data.stores[s].mtd_revenue_wan for s in stores)
    region_target = sum(data.stores[s].revenue_target for s in stores)
    ws.cell(row=r, column=_NCOLS, value=pct_str(div_or_zero(region_mtd, region_target) * 100)); r += 1

    # Standard time progress
    tp = pct_str(dates.time_progress * 100)
    ws.cell(row=r, column=2, value="标准时间进度")
    for col in range(3, _NCOLS + 1):
        ws.cell(row=r, column=col, value=tp)
    r += 1

    # Discounts
    _write_row(ws, r, "当月累计优惠总金额(万)", [data.stores[s].mtd_discount_wan for s in stores]); r += 1
    vals_disc = [pct_str(data.stores[s].mtd_discount_pct) for s in stores]
    _write_row(ws, r, "当月累计优惠占比", vals_disc, region_sum=False)
    region_disc = sum(data.stores[s].mtd_discount_wan for s in stores)
    ws.cell(row=r, column=_NCOLS, value=pct_str(div_or_zero(region_disc, region_mtd) * 100)); r += 1

    # Merge revenue section label (A9 through end of discount row)
    revenue_end = r - 1
    ws.merge_cells(f"A9:A{revenue_end}")

    # ── Section 3: 单桌消费(不含税) ──
    sec3_start = r
    ws.cell(row=r, column=1, value="单桌消费\n(不含税)")
    ws.cell(row=r, column=1).font = BOLD
    ws.cell(row=r, column=1).alignment = CENTER

    region_rev_today = sum(data.stores[s].today_revenue_wan * WAN_DIVISOR for s in stores)
    region_cust_today = sum(data.stores[s].today_customers for s in stores)
    _write_row(ws, r, "今日人均消费", [data.stores[s].today_per_capita for s in stores], region_sum=False)
    ws.cell(row=r, column=_NCOLS, value=div_or_zero(region_rev_today, region_cust_today)); r += 1

    _write_row(ws, r, "今日消费客数", [data.stores[s].today_customers for s in stores]); r += 1

    region_raw_tables_today = sum(data.stores[s].today_raw_tables for s in stores)
    _write_row(ws, r, "今日单桌消费", [data.stores[s].today_per_table for s in stores], region_sum=False)
    ws.cell(row=r, column=_NCOLS, value=div_or_zero(region_rev_today, region_raw_tables_today)); r += 1

    region_mtd_raw_rev = sum(data.stores[s].mtd_revenue_wan * WAN_DIVISOR for s in stores)
    region_mtd_raw_tables = sum(data.stores[s].mtd_raw_tables for s in stores)
    _write_row(ws, r, "截止今日单桌消费", [data.stores[s].mtd_per_table for s in stores], region_sum=False)
    ws.cell(row=r, column=_NCOLS, value=div_or_zero(region_mtd_raw_rev, region_mtd_raw_tables)); r += 1

    region_comp_rev = sum(config.get_comp_revenue_wan(data.stores[s]) * WAN_DIVISOR for s in stores)
    region_comp_raw_tables = sum(config.get_comp_raw_tables(data.stores[s]) for s in stores)
    _write_row(ws, r, f"{comp}单桌消费",
               [config.get_comp_per_table(data.stores[s]) for s in stores], region_sum=False)
    ws.cell(row=r, column=_NCOLS, value=div_or_zero(region_comp_rev, region_comp_raw_tables)); r += 1

    _write_row(ws, r, f"{comp_type}{comp}变化",
               [data.stores[s].mtd_per_table - config.get_comp_per_table(data.stores[s]) for s in stores],
               region_sum=False)
    ws.cell(row=r, column=_NCOLS,
            value=div_or_zero(region_mtd_raw_rev, region_mtd_raw_tables) - div_or_zero(region_comp_rev, region_comp_raw_tables)); r += 1

    ws.merge_cells(f"A{sec3_start}:A{r - 1}")

    # ── Section 4: 翻台率 ──
    sec4_start = r
    ws.cell(row=r, column=1, value="翻台率")
    ws.cell(row=r, column=1).font = BOLD
    ws.cell(row=r, column=1).alignment = CENTER

    ws.cell(row=r, column=2, value="名次")
    for i in range(len(stores)):
        ws.cell(row=r, column=3 + i, value=f"第{i + 1}名")
    ws.cell(row=r, column=_NCOLS, value="当月累计平均翻台率"); r += 1

    # Today ranking
    today_tr = [(s, data.stores[s].today_turnover_rate) for s in stores]
    today_tr.sort(key=lambda x: x[1], reverse=True)
    tr_start = r
    ws.cell(row=r, column=2, value=f"{month}月{day}日翻台率排名店铺"); r += 1
    ws.cell(row=r, column=2, value=f"{month}月{day}日翻台率排名")
    for i, (store, tr) in enumerate(today_tr):
        ws.cell(row=r - 1, column=3 + i, value=store)
        ws.cell(row=r, column=3 + i, value=tr)
    r += 1

    # MTD ranking
    mtd_tr = [(s, data.stores[s].mtd_turnover_rate) for s in stores]
    mtd_tr.sort(key=lambda x: x[1], reverse=True)
    ws.cell(row=r, column=2, value=f"{month}月平均翻台率排名店铺"); r += 1
    ws.cell(row=r, column=2, value=f"{month}月平均翻台率排名")
    for i, (store, tr) in enumerate(mtd_tr):
        ws.cell(row=r - 1, column=3 + i, value=store)
        ws.cell(row=r, column=3 + i, value=tr)

    ws.merge_cells(f"A{sec4_start}:A{r}")

    # Region weighted-average turnover rate (merged in K)
    ws.merge_cells(f"K{tr_start}:K{r}")
    region_mtd_tables_total = sum(data.stores[s].mtd_tables for s in stores)
    region_seats_total = sum(data.stores[s].seats for s in stores if data.stores[s].seats > 0)
    num_days = dates.day_of_month
    region_avg_tr = div_or_zero(region_mtd_tables_total, region_seats_total * num_days) if region_seats_total else 0
    ws.cell(row=tr_start, column=_NCOLS, value=region_avg_tr)
    ws.cell(row=tr_start, column=_NCOLS).font = BOLD_LARGE
    ws.cell(row=tr_start, column=_NCOLS).alignment = CENTER

    last_row = r

    # ── Styling (dynamic row ranges) ──
    apply_fill_row(ws, 1, theme.header_fill, 1, _NCOLS)
    apply_fill_row(ws, 2, theme.header_fill, 1, _NCOLS)

    # Section 1 (桌数): rows 3-7, row 8 highlight
    apply_fill_range(ws, 3, 7, theme.section_a_fill, 1, _NCOLS)
    apply_fill_row(ws, 8, theme.highlight_fill, 1, _NCOLS)

    # Section 2 (收入): 9 to revenue_end → section_b
    apply_fill_range(ws, 9, revenue_end, theme.section_b_fill, 1, _NCOLS)

    # Section 3 (单桌消费)
    apply_fill_range(ws, sec3_start, sec4_start - 1, theme.section_a_fill, 1, _NCOLS)

    # Section 4 (翻台率): alternating
    for i in range(sec4_start, last_row + 1):
        fill = theme.highlight_fill if (i - sec4_start) % 2 == 0 else theme.section_b_fill
        apply_fill_row(ws, i, fill, 1, _NCOLS)

    # K region avg white fill override
    ws.cell(row=tr_start, column=_NCOLS).fill = WHITE_FILL

    # Borders
    apply_border(ws, 1, last_row, 1, _NCOLS)

    # Alignment
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=_NCOLS):
        for cell in row:
            cell.alignment = CENTER

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    for col in range(3, _NCOLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    return ws
