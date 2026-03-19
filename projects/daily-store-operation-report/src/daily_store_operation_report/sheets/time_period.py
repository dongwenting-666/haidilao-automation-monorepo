"""Sheet 4: 分时段-上报 (Time-period breakdown with per-store colors)."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from daily_store_operation_report.constants import (
    STORES,
    TIME_SLOTS,
    WEEKDAY_NAMES,
)
from daily_store_operation_report.sheets.styles import (
    BOLD,
    BOLD_TITLE,
    CENTER,
    GOLD_FILL,
    GRAY_FILL,
    NAVY_FILL,
    RED_FONT,
    WHITE_BOLD,
    apply_border,
    apply_fill_row,
    store_fill,
)
from daily_store_operation_report.transform import ReportData, StoreMetrics

_NCOLS = 15  # A..O

# Columns that represent turnover rates or rate diffs (averaged across stores).
# Stores with no data (all-zero subtotals) are excluded from the average
# to avoid diluting results with new stores that lack historical data.
# Col 3=今年翻台率, 4=去年翻台率, 5=本月目标, 6=目标差异, 7=同比差异,
# Col 8=当日翻台率, 10=去年同周同日翻台率, 12=翻台率同比差异
_AVG_COLS = {3, 4, 5, 6, 7, 8, 10, 12}


def _store_slot_totals(m: StoreMetrics) -> dict[int, float]:
    """Sum all time-slot values into per-column totals for a single store.

    Used for both the per-store subtotal row and the region total row,
    ensuring they stay in sync.

    Turnover rates (c3, c4, c5, c8, c10) are summed across slots because
    daily turnover rate = sum of per-slot rates (each slot's rate is
    independent: tables_in_slot / total_seats).
    """
    c3 = sum(m.tp_turnover_cur.get(s, 0) for s in TIME_SLOTS)
    c4 = sum(m.tp_turnover_yoy.get(s, 0) for s in TIME_SLOTS)
    c5 = sum(m.tp_turnover_target.get(s, 0) for s in TIME_SLOTS)
    c8 = sum(m.tp_turnover_today.get(s, 0) for s in TIME_SLOTS)
    c9 = sum(m.tp_tables_today.get(s, 0) for s in TIME_SLOTS)
    c10 = sum(m.tp_turnover_yoy_weekday.get(s, 0) for s in TIME_SLOTS)
    c11 = sum(m.tp_tables_yoy_weekday.get(s, 0) for s in TIME_SLOTS)
    c13 = sum(m.tp_mtd_tables_cur.get(s, 0) for s in TIME_SLOTS)
    c14 = sum(m.tp_mtd_tables_yoy.get(s, 0) for s in TIME_SLOTS)
    return {
        3: c3, 4: c4, 5: c5, 6: c3 - c5, 7: c3 - c4,
        8: c8, 9: c9, 10: c10, 11: c11, 12: c8 - c10,
        13: c13, 14: c14, 15: c13 - c14,
    }


def build_time_period_sheet(wb: Workbook, data: ReportData) -> Worksheet:
    """Build the 分时段-上报 sheet."""
    ws: Worksheet = wb.create_sheet("分时段-上报")
    dates = data.dates
    d = dates.report_date
    weekday = WEEKDAY_NAMES[d.weekday()]
    month = d.month
    day = d.day
    yoy_year = d.year - 1

    # Row 1: Title (merged A1:O1)
    ws.merge_cells("A1:O1")
    ws["A1"] = (
        f"门店分时段营业数据{d.year}年{month}月vs{yoy_year}年{month}月"
        f"截至{day}日-{weekday}（考核）"
    )
    ws["A1"].font = BOLD_TITLE
    ws["A1"].alignment = CENTER

    # Row 2: Group headers
    ws.merge_cells("A2:A3")
    ws["A2"] = "门店名称"
    ws["A2"].font = BOLD
    ws["A2"].alignment = CENTER

    ws.merge_cells("B2:B3")
    ws["B2"] = "分时段"
    ws["B2"].font = BOLD
    ws["B2"].alignment = CENTER

    ws.merge_cells("C2:G2")
    ws["C2"] = "翻台率（考核）"
    ws["C2"].font = BOLD
    ws["C2"].alignment = CENTER

    ws.merge_cells("H2:I2")
    ws["H2"] = f"{day:02d}/{month:02d}/{d.year}"
    ws["H2"].font = BOLD
    ws["H2"].alignment = CENTER

    ws.merge_cells("J2:L2")
    ws["J2"] = "去年同周同日"
    ws["J2"].font = BOLD
    ws["J2"].alignment = CENTER

    ws.merge_cells("M2:O2")
    ws["M2"] = "本月截止目前桌数"
    ws["M2"].font = BOLD
    ws["M2"].alignment = CENTER

    # Row 3: Sub-headers
    sub_headers = {
        3: "今年", 4: "去年", 5: "本月目标", 6: "目标差异", 7: "同比差异",
        8: "翻台率（考核）", 9: "桌数（考核）",
        10: "翻台率", 11: "桌数", 12: "翻台率同比差异",
        13: "今年", 14: "去年", 15: "同比差异",
    }
    for col, label in sub_headers.items():
        c = ws.cell(row=3, column=col, value=label)
        c.font = BOLD
        c.alignment = CENTER

    # ── Data rows ──
    current_row = 4

    for store in STORES:
        m = data.stores[store]
        sfill = store_fill(store)
        store_start_row = current_row

        for slot_idx, slot in enumerate(TIME_SLOTS):
            r = current_row

            if slot_idx == 0:
                ws.cell(row=r, column=1, value=store)
            ws.cell(row=r, column=2, value=slot)

            cur_tr = m.tp_turnover_cur.get(slot, 0)
            yoy_tr = m.tp_turnover_yoy.get(slot, 0)
            target_tr = m.tp_turnover_target.get(slot, 0)
            today_tr = m.tp_turnover_today.get(slot, 0)
            today_tables = m.tp_tables_today.get(slot, 0)
            yoy_wd_tr = m.tp_turnover_yoy_weekday.get(slot, 0)
            yoy_wd_tables = m.tp_tables_yoy_weekday.get(slot, 0)
            mtd_tables = m.tp_mtd_tables_cur.get(slot, 0)
            yoy_mtd_tables = m.tp_mtd_tables_yoy.get(slot, 0)

            ws.cell(row=r, column=3, value=cur_tr)
            ws.cell(row=r, column=4, value=yoy_tr)
            ws.cell(row=r, column=5, value=target_tr)

            target_diff = cur_tr - target_tr
            cell_f = ws.cell(row=r, column=6, value=target_diff)
            if target_diff < 0:
                cell_f.font = RED_FONT

            yoy_diff = cur_tr - yoy_tr
            cell_g = ws.cell(row=r, column=7, value=yoy_diff)
            if yoy_diff < 0:
                cell_g.font = RED_FONT

            ws.cell(row=r, column=8, value=today_tr)
            ws.cell(row=r, column=9, value=today_tables)
            ws.cell(row=r, column=10, value=yoy_wd_tr)
            ws.cell(row=r, column=11, value=yoy_wd_tables)

            tr_diff_wd = today_tr - yoy_wd_tr
            cell_l = ws.cell(row=r, column=12, value=tr_diff_wd)
            if tr_diff_wd < 0:
                cell_l.font = RED_FONT

            ws.cell(row=r, column=13, value=mtd_tables)
            ws.cell(row=r, column=14, value=yoy_mtd_tables)

            tables_diff = mtd_tables - yoy_mtd_tables
            cell_o = ws.cell(row=r, column=15, value=tables_diff)
            if tables_diff < 0:
                cell_o.font = RED_FONT

            for col in range(1, _NCOLS + 1):
                ws.cell(row=r, column=col).fill = sfill

            current_row += 1

        # Merge store name cells
        if len(TIME_SLOTS) > 1:
            ws.merge_cells(
                start_row=store_start_row, start_column=1,
                end_row=store_start_row + len(TIME_SLOTS) - 1, end_column=1,
            )

        # Store subtotal row (gray)
        r = current_row
        ws.cell(row=r, column=2, value=f"{store}汇总")

        totals = _store_slot_totals(m)
        for col_idx, val in totals.items():
            ws.cell(row=r, column=col_idx, value=val)

        # Gray fill + bold for subtotal
        for col in range(1, _NCOLS + 1):
            c = ws.cell(row=r, column=col)
            c.fill = GRAY_FILL
            c.font = BOLD

        # Red font for negative diffs in subtotal
        for col in (6, 7, 12, 15):
            c = ws.cell(row=r, column=col)
            if isinstance(c.value, (int, float)) and c.value < 0:
                c.font = Font(bold=True, color="FFFF0000")

        current_row += 1

    # ── Region total row (navy) — computed from ReportData ──
    r = current_row
    ws.cell(row=r, column=1, value="区域整体")

    store_subtotals = [_store_slot_totals(data.stores[s]) for s in STORES]

    for col_idx in range(3, _NCOLS + 1):
        values = [st[col_idx] for st in store_subtotals]
        total = sum(values)
        if col_idx in _AVG_COLS:
            nonzero = sum(1 for v in values if v != 0)
            total = total / (nonzero or 1)
        ws.cell(row=r, column=col_idx, value=total)

    # Navy fill + white bold
    for col in range(1, _NCOLS + 1):
        c = ws.cell(row=r, column=col)
        c.fill = NAVY_FILL
        c.font = WHITE_BOLD

    # ── Styling ──
    apply_fill_row(ws, 1, GOLD_FILL, 1, _NCOLS)
    apply_fill_row(ws, 2, GOLD_FILL, 1, _NCOLS)
    apply_fill_row(ws, 3, GOLD_FILL, 1, _NCOLS)

    apply_border(ws, 1, current_row, 1, _NCOLS)

    for row_cells in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=_NCOLS):
        for cell in row_cells:
            cell.alignment = CENTER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    for col in range(3, _NCOLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return ws
