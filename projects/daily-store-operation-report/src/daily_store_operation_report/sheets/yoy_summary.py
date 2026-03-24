"""Sheet 2: 同比数据 (YoY summary, region-grouped with gold theme)."""

from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from daily_store_operation_report.constants import (
    EAST_STORES,
    REGION_LABEL,
    WAN_DIVISOR,
    WEEKDAY_NAMES,
    WEST_STORES,
)
from daily_store_operation_report.sheets.styles import (
    BOLD,
    BOLD_TITLE,
    BRIGHT_YELLOW_FILL,
    CENTER,
    GOLD_FILL,
    LIGHT_BLUE_FILL_GOLD,
    YELLOW_FILL,
    apply_border,
    apply_fill_range,
    apply_fill_row,
)
from daily_store_operation_report.transform import ReportData
from daily_store_operation_report.utils import div_or_zero, pct_str, region_turnover_rate


_NCOLS = 11  # A..K

def build_yoy_summary_sheet(wb: Workbook, data: ReportData) -> Worksheet:
    """Build the 同比数据 sheet."""
    ws: Worksheet = wb.create_sheet("同比数据")
    dates = data.dates
    d = dates.report_date
    weekday = WEEKDAY_NAMES[d.weekday()]
    month = d.month
    day = d.day

    # Column order: west stores then east stores
    ordered_stores = WEST_STORES + EAST_STORES

    # Row 1: Title
    ws.merge_cells("A1:K1")
    ws["A1"] = f"加拿大-各门店{d.year}年{month}月{day}日同比数据-{weekday}"
    ws["A1"].font = BOLD_TITLE
    ws["A1"].alignment = CENTER

    # Row 2: Region headers
    ws.merge_cells("A2:B2")
    ws["A2"] = "分类"
    ws["A2"].font = BOLD
    ws["A2"].alignment = CENTER
    ws.merge_cells("C2:E2")
    ws["C2"] = "西部"
    ws["C2"].font = BOLD
    ws["C2"].alignment = CENTER
    ws.merge_cells("F2:I2")
    ws["F2"] = "东部"
    ws["F2"].font = BOLD
    ws["F2"].alignment = CENTER
    region_cell = ws.cell(row=2, column=_NCOLS, value=REGION_LABEL)
    region_cell.font = BOLD
    region_cell.alignment = CENTER

    # Row 3: Store name headers
    ws.cell(row=3, column=1, value="项目").font = BOLD
    ws.cell(row=3, column=2, value="内容").font = BOLD
    for i, store in enumerate(ordered_stores):
        ws.cell(row=3, column=3 + i, value=store).font = BOLD
    ws.cell(row=3, column=_NCOLS, value=REGION_LABEL).font = BOLD

    def _write_row(row: int, label: str, values: list, region_val=None):
        ws.cell(row=row, column=2, value=label)
        for i, v in enumerate(values):
            ws.cell(row=row, column=3 + i, value=v)
        if region_val is not None:
            ws.cell(row=row, column=_NCOLS, value=region_val)

    # ── Section 1: 桌数对比同期数据 — rows 4-7 ──
    ws.merge_cells("A4:A7")
    ws["A4"] = "桌数\n对比同期数据"
    ws["A4"].font = BOLD
    ws["A4"].alignment = CENTER

    mtd_vals = [data.stores[s].mtd_tables for s in ordered_stores]
    _write_row(4, "本月截止目前", mtd_vals, sum(mtd_vals))

    yoy_vals = [data.stores[s].yoy_mtd_tables for s in ordered_stores]
    _write_row(5, "去年截止同期", yoy_vals, sum(yoy_vals))

    diff_vals = [data.stores[s].mtd_tables - data.stores[s].yoy_mtd_tables for s in ordered_stores]
    _write_row(6, "对比去年同期", diff_vals, sum(diff_vals))

    growth_vals = []
    for s in ordered_stores:
        yoy_t = data.stores[s].yoy_mtd_tables
        diff = data.stores[s].mtd_tables - yoy_t
        growth_vals.append(pct_str(div_or_zero(diff, yoy_t) * 100, 1))
    region_mtd_t = sum(data.stores[s].mtd_tables for s in ordered_stores)
    region_yoy_t = sum(data.stores[s].yoy_mtd_tables for s in ordered_stores)
    _write_row(7, "桌数增长率", growth_vals, pct_str(div_or_zero(region_mtd_t - region_yoy_t, region_yoy_t) * 100, 1))

    # ── Section 2: 翻台率对比同期数据 — rows 8-11 ──
    ws.merge_cells("A8:A11")
    ws["A8"] = "翻台率\n对比同期数据"
    ws["A8"].font = BOLD
    ws["A8"].alignment = CENTER

    tr_cur = [data.stores[s].mtd_turnover_rate for s in ordered_stores]
    num_days = dates.day_of_month
    region_avg_tr = region_turnover_rate(
        data.stores, ordered_stores, tables_attr="mtd_tables", num_days=num_days,
    )
    _write_row(8, "本月截止目前", tr_cur, region_avg_tr)

    tr_yoy = [data.stores[s].yoy_mtd_turnover_rate for s in ordered_stores]
    region_avg_tr_yoy = region_turnover_rate(
        data.stores, ordered_stores, tables_attr="yoy_mtd_tables", num_days=num_days,
    )
    _write_row(9, "去年截止同期", tr_yoy, region_avg_tr_yoy)

    tr_diff = [data.stores[s].mtd_turnover_rate - data.stores[s].yoy_mtd_turnover_rate for s in ordered_stores]
    _write_row(10, "对比去年同期", tr_diff, region_avg_tr - region_avg_tr_yoy)

    tr_growth = []
    for s in ordered_stores:
        yoy_tr = data.stores[s].yoy_mtd_turnover_rate
        diff = data.stores[s].mtd_turnover_rate - yoy_tr
        tr_growth.append(pct_str(div_or_zero(diff, yoy_tr) * 100, 1))
    _write_row(11, "翻台率增长率", tr_growth,
               pct_str(div_or_zero(region_avg_tr - region_avg_tr_yoy, region_avg_tr_yoy) * 100, 1))

    # ── Section 3: 营业收入(不含税-万加元) — rows 12-15 ──
    ws.merge_cells("A12:A15")
    ws["A12"] = "营业收入\n(不含税-万加元)"
    ws["A12"].font = BOLD
    ws["A12"].alignment = CENTER

    rev_cur = [data.stores[s].mtd_revenue_wan for s in ordered_stores]
    _write_row(12, "本月截止目前", rev_cur, sum(rev_cur))

    rev_yoy = [data.stores[s].yoy_mtd_revenue_wan for s in ordered_stores]
    _write_row(13, "去年截止同期", rev_yoy, sum(rev_yoy))

    rev_diff = [data.stores[s].mtd_revenue_wan - data.stores[s].yoy_mtd_revenue_wan for s in ordered_stores]
    _write_row(14, "对比去年同期", rev_diff, sum(rev_diff))

    rev_growth = []
    for s in ordered_stores:
        yoy_r = data.stores[s].yoy_mtd_revenue_wan
        diff = data.stores[s].mtd_revenue_wan - yoy_r
        rev_growth.append(pct_str(div_or_zero(diff, yoy_r) * 100, 1))
    region_cur_r = sum(rev_cur)
    region_yoy_r = sum(rev_yoy)
    _write_row(15, "收入增长率", rev_growth,
               pct_str(div_or_zero(region_cur_r - region_yoy_r, region_yoy_r) * 100, 1))

    # ── Section 4: 单桌消费对比同期数据 — rows 16-19 ──
    ws.merge_cells("A16:A19")
    ws["A16"] = "单桌消费\n对比同期数据"
    ws["A16"].font = BOLD
    ws["A16"].alignment = CENTER

    pt_cur = [data.stores[s].mtd_per_table for s in ordered_stores]
    region_pt_cur = div_or_zero(
        sum(data.stores[s].mtd_revenue_wan * WAN_DIVISOR for s in ordered_stores),
        sum(data.stores[s].mtd_raw_tables for s in ordered_stores),
    )
    _write_row(16, "本月截止目前", pt_cur, region_pt_cur)

    pt_yoy = [data.stores[s].yoy_mtd_per_table for s in ordered_stores]
    region_pt_yoy = div_or_zero(
        sum(data.stores[s].yoy_mtd_revenue_wan * WAN_DIVISOR for s in ordered_stores),
        sum(data.stores[s].yoy_mtd_raw_tables for s in ordered_stores),
    )
    _write_row(17, "去年截止同期", pt_yoy, region_pt_yoy)

    pt_diff = [data.stores[s].mtd_per_table - data.stores[s].yoy_mtd_per_table for s in ordered_stores]
    _write_row(18, "对比去年同期", pt_diff, region_pt_cur - region_pt_yoy)

    pt_growth = []
    for s in ordered_stores:
        yoy_pt = data.stores[s].yoy_mtd_per_table
        diff = data.stores[s].mtd_per_table - yoy_pt
        pt_growth.append(pct_str(div_or_zero(diff, yoy_pt) * 100, 1))
    _write_row(19, "单桌消费增长率", pt_growth,
               pct_str(div_or_zero(region_pt_cur - region_pt_yoy, region_pt_yoy) * 100, 1))

    # ── Styling ──
    apply_fill_row(ws, 1, GOLD_FILL, 1, _NCOLS)
    apply_fill_row(ws, 2, GOLD_FILL, 1, _NCOLS)
    apply_fill_row(ws, 3, GOLD_FILL, 1, _NCOLS)

    apply_fill_range(ws, 4, 6, YELLOW_FILL, 1, _NCOLS)
    apply_fill_range(ws, 12, 14, YELLOW_FILL, 1, _NCOLS)

    apply_fill_row(ws, 7, BRIGHT_YELLOW_FILL, 1, _NCOLS)
    apply_fill_row(ws, 11, BRIGHT_YELLOW_FILL, 1, _NCOLS)
    apply_fill_row(ws, 15, BRIGHT_YELLOW_FILL, 1, _NCOLS)
    apply_fill_row(ws, 19, BRIGHT_YELLOW_FILL, 1, _NCOLS)

    apply_fill_range(ws, 8, 10, LIGHT_BLUE_FILL_GOLD, 1, _NCOLS)
    apply_fill_range(ws, 16, 18, LIGHT_BLUE_FILL_GOLD, 1, _NCOLS)

    # Red font for negative growth rates
    for row_num in (7, 11, 15, 19):
        for col in range(3, _NCOLS + 1):
            cell = ws.cell(row=row_num, column=col)
            val = str(cell.value or "")
            if val.startswith("-"):
                cell.font = Font(color="FFFF0000")

    apply_border(ws, 1, 19, 1, _NCOLS)

    for row in ws.iter_rows(min_row=1, max_row=19, min_col=1, max_col=_NCOLS):
        for cell in row:
            cell.alignment = CENTER

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    for col in range(3, _NCOLS + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    return ws
