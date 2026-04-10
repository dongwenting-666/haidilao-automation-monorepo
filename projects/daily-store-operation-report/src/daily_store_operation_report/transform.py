"""Raw QBI data loading and transformation into report-ready dataclass."""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import TypedDict

from excel_utils import load_data_rows

from daily_store_operation_report.validation import (
    validate_daily_rows,
    validate_no_all_zero_columns,
    validate_store_coverage,
    validate_time_period_rows,
)
from daily_store_operation_report.constants import (
    COL_CUSTOMERS,
    COL_DATE,
    COL_DISCOUNT,
    COL_REVENUE,
    COL_REVENUE_DINE_IN,
    COL_REVENUE_TAKEOUT,
    COL_REVENUE_DELIVERY,
    COL_SEATS,
    COL_STORE,
    COL_TABLES_ASSESSED,
    COL_TABLES_RAW,
    COL_TABLES_TAKEOUT,
    COL_TIME_SLOT,
    COL_TURNOVER,
    QBI_SHEET_DAILY,
    QBI_SHEET_TIME_PERIOD,
    STORES,
    TIME_SLOTS,
    WAN_DIVISOR,
)
from daily_store_operation_report.dates import ReportDates
from daily_store_operation_report.download import DownloadedFiles
from daily_store_operation_report.utils import div_or_zero

logger = logging.getLogger(__name__)


# ── Raw data loading ──────────────────────────────────────────────────────────


def _normalize_date(val: object) -> str:
    """Normalize a date value from Excel to YYYYMMDD string.

    Handles datetime objects (from openpyxl), date objects, and raw strings.
    Note: datetime must be checked before date since datetime is a subclass of date.
    """
    from datetime import date, datetime

    if isinstance(val, datetime):
        return val.strftime("%Y%m%d")
    if isinstance(val, date):
        return val.strftime("%Y%m%d")
    return str(val)


def _load_daily(path: Path) -> list[dict]:
    """Load rows from the 不含税 sheet of a daily report.

    Validates columns and row counts after loading.
    """
    try:
        rows = load_data_rows(path, sheet_name=QBI_SHEET_DAILY)
    except KeyError as e:
        # openpyxl raises KeyError when sheet name doesn't exist
        import openpyxl as _opx
        try:
            _wb = _opx.load_workbook(path, read_only=True, data_only=True)
            available = _wb.sheetnames
            _wb.close()
        except Exception:
            available = ["(could not open file)"]
        raise ValueError(
            f"Sheet {QBI_SHEET_DAILY!r} not found in {path.name}.\n"
            f"  Available sheets: {available}\n"
            f"  Is this the correct daily report file?"
        ) from e
    except Exception as e:
        raise ValueError(f"Failed to load daily report from {path}: {e}") from e

    validate_daily_rows(rows, path)
    return rows


def _load_time_period(path: Path) -> list[dict]:
    """Load rows from the 不含税 sheet of a time-period report.

    Validates columns and row counts after loading.
    """
    try:
        rows = load_data_rows(path, sheet_name=QBI_SHEET_TIME_PERIOD)
    except KeyError as e:
        import openpyxl as _opx
        try:
            _wb = _opx.load_workbook(path, read_only=True, data_only=True)
            available = _wb.sheetnames
            _wb.close()
        except Exception:
            available = ["(could not open file)"]
        raise ValueError(
            f"Sheet {QBI_SHEET_TIME_PERIOD!r} not found in {path.name}.\n"
            f"  Available sheets: {available}\n"
            f"  Is this the correct time-period report file?"
        ) from e
    except Exception as e:
        raise ValueError(f"Failed to load time-period report from {path}: {e}") from e

    validate_time_period_rows(rows, path)
    return rows


def _safe_float(val: object, *, context: str = "") -> float:
    """Coerce *val* to float, returning 0.0 on None/NaN/non-numeric values.

    Logs a debug message when a non-None non-numeric value is silently replaced.
    """
    if val is None:
        return 0.0
    try:
        result = float(val)  # type: ignore[arg-type]
        import math
        if math.isnan(result) or math.isinf(result):
            return 0.0
        return result
    except (TypeError, ValueError):
        if context:
            logger.debug("Non-numeric value in column %s: %r — treating as 0", context, val)
        return 0.0


def _sum_by_store(rows: list[dict], col: str) -> dict[str, float]:
    """Sum a column grouped by store name.

    Handles NaN, None, and non-numeric values gracefully.
    """
    totals: dict[str, float] = {}
    for row in rows:
        store = str(row.get(COL_STORE) or "").strip()
        if not store:
            continue
        val = _safe_float(row.get(col), context=col)
        totals[store] = totals.get(store, 0) + val
    return totals


def _sum_two_cols_by_store(rows: list[dict], col1: str, col2: str) -> dict[str, float]:
    """Sum two columns per store (e.g., 外卖+外送 = total non-dine-in)."""
    totals: dict[str, float] = {}
    for row in rows:
        store = str(row.get(COL_STORE) or "").strip()
        if not store:
            continue
        val = _safe_float(row.get(col1), context=col1) + _safe_float(row.get(col2), context=col2)
        totals[store] = totals.get(store, 0) + val
    return totals


def _last_day_two_cols_by_store(rows: list[dict], report_date_str: str, col1: str, col2: str) -> dict[str, float]:
    """Get last day value for two columns combined per store."""
    totals: dict[str, float] = {}
    for row in rows:
        store = str(row.get(COL_STORE) or "").strip()
        if not store or _normalize_date(row.get(COL_DATE, "")) != report_date_str:
            continue
        val = _safe_float(row.get(col1), context=col1) + _safe_float(row.get(col2), context=col2)
        totals[store] = totals.get(store, 0) + val
    return totals


def _last_day_by_store(rows: list[dict], report_date_str: str, col: str) -> dict[str, float]:
    """Get a column value for a specific day, summed per store."""
    totals: dict[str, float] = {}
    for row in rows:
        if _normalize_date(row.get(COL_DATE, "")) == report_date_str:
            store = str(row.get(COL_STORE) or "").strip()
            if not store:
                continue
            val = _safe_float(row.get(col), context=col)
            totals[store] = totals.get(store, 0) + val
    return totals


def _sum_time_period(rows: list[dict], col: str) -> dict[str, dict[str, float]]:
    """Sum a column grouped by (store, time slot). Returns {store: {slot: val}}."""
    result: dict[str, dict[str, float]] = {}
    for row in rows:
        store = str(row.get(COL_STORE) or "").strip()
        slot = str(row.get(COL_TIME_SLOT) or "").strip()
        if not store or not slot or slot == "-":
            continue
        val = _safe_float(row.get(col), context=col)
        result.setdefault(store, {})
        result[store][slot] = result[store].get(slot, 0) + val
    return result


def _avg_time_period(rows: list[dict], col: str) -> dict[str, dict[str, float]]:
    """Average a column grouped by (store, time slot)."""
    sums: dict[str, dict[str, float]] = {}
    counts: dict[str, dict[str, int]] = {}
    for row in rows:
        store = str(row.get(COL_STORE) or "").strip()
        slot = str(row.get(COL_TIME_SLOT) or "").strip()
        if not store or not slot or slot == "-":
            continue
        val = _safe_float(row.get(col), context=col)
        sums.setdefault(store, {})
        counts.setdefault(store, {})
        sums[store][slot] = sums[store].get(slot, 0) + val
        counts[store][slot] = counts[store].get(slot, 0) + 1
    result: dict[str, dict[str, float]] = {}
    for store in sums:
        result[store] = {}
        for slot in sums[store]:
            c = counts[store][slot]
            result[store][slot] = sums[store][slot] / c if c else 0.0
    return result


def _last_day_time_period(rows: list[dict], date_str: str, col: str) -> dict[str, dict[str, float]]:
    """Get column value for a specific day grouped by (store, slot)."""
    result: dict[str, dict[str, float]] = {}
    for row in rows:
        if _normalize_date(row.get(COL_DATE, "")) != date_str:
            continue
        store = str(row.get(COL_STORE) or "").strip()
        slot = str(row.get(COL_TIME_SLOT) or "").strip()
        if not store or not slot or slot == "-":
            continue
        val = _safe_float(row.get(col), context=col)
        result.setdefault(store, {})
        result[store][slot] = result[store].get(slot, 0) + val
    return result


def _avg_turnover_by_store(rows: list[dict], col: str = COL_TURNOVER) -> dict[str, float]:
    """Pre-aggregate average turnover rate per store (single pass).

    Only stores with at least one non-zero row are included to avoid
    diluting averages for stores with partial data.
    """
    sums: dict[str, float] = {}
    counts: dict[str, int] = {}
    for row in rows:
        store = str(row.get(COL_STORE) or "").strip()
        if not store:
            continue
        val = _safe_float(row.get(col), context=col)
        sums[store] = sums.get(store, 0) + val
        counts[store] = counts.get(store, 0) + 1
    return {s: sums[s] / counts[s] for s in sums if counts[s]}


# ── Targets loading ───────────────────────────────────────────────────────────


class Targets(TypedDict):
    """Structure of per-month targets from targets.json."""

    revenue: dict[str, float]
    turnover_rate: dict[str, dict[str, float]]


def load_competitor(path: Path) -> dict[str, str]:
    """Load the competitor (假想敌) mapping from competitor.json.

    Returns a dict mapping store name → competitor store name.
    Returns an empty dict (and logs a warning) if the file is missing.
    Raises ValueError on malformed JSON.
    """
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        logger.warning("Competitor file not found: %s — competitor sheet will be empty", path)
        return {}
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in competitor file {path}: {e}") from e
    if not isinstance(data, dict):
        raise ValueError(f"competitor.json must be a flat dict, got {type(data)}")
    return data


def load_targets(path: Path, month_key: str) -> Targets:
    """Load targets for a given month from targets.json."""
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        logger.warning("Targets file not found: %s — using defaults", path)
        return {"revenue": {}, "turnover_rate": {}}
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in targets file {path}: {e}") from e
    result = data.get(month_key)
    if result is None:
        logger.warning("No targets found for %s in %s — using defaults", month_key, path)
        return {"revenue": {}, "turnover_rate": {}}
    if not isinstance(result.get("revenue"), dict) or not isinstance(result.get("turnover_rate"), dict):
        raise ValueError(
            f"Invalid target structure for {month_key} in {path}: "
            "expected 'revenue' and 'turnover_rate' dicts"
        )
    return result


def _filter_rows_by_date(rows: list[dict], max_date_str: str) -> list[dict]:
    """Filter rows to include only dates <= max_date_str (YYYYMMDD format)."""
    return [r for r in rows if _normalize_date(r.get(COL_DATE, "")) <= max_date_str]


# ── Report data ───────────────────────────────────────────────────────────────


@dataclass
class StoreMetrics:
    """Computed metrics for a single store.

    Values are stored at full precision; rounding to 2 decimals
    happens at display time (in report.py's _format_numbers).
    """

    # Daily (report date)
    today_tables: float = 0
    today_raw_tables: float = 0
    today_takeout_tables: float = 0
    today_non_assessed_tables: float = 0
    today_revenue_wan: float = 0
    today_per_capita: float = 0
    today_customers: int = 0
    today_per_table: float = 0
    today_dine_in_wan: float = 0
    today_takeout_wan: float = 0
    today_turnover_rate: float = 0
    seats: int = 0  # 所有餐位数 (total seat count, constant per store)

    # MTD current month
    mtd_tables: float = 0
    mtd_raw_tables: float = 0
    mtd_revenue_wan: float = 0
    mtd_dine_in_wan: float = 0
    mtd_takeout_wan: float = 0
    mtd_turnover_rate: float = 0
    mtd_per_table: float = 0
    mtd_discount_wan: float = 0
    mtd_discount_pct: float = 0

    # MTD previous month (same period)
    prev_mtd_tables: float = 0
    prev_mtd_raw_tables: float = 0
    prev_mtd_revenue_wan: float = 0
    prev_mtd_dine_in_wan: float = 0
    prev_mtd_takeout_wan: float = 0
    prev_mtd_per_table: float = 0
    prev_mtd_turnover_rate: float = 0   # full prev-month avg turnover (for competitor sheet)

    # MTD previous year (same period)
    yoy_mtd_tables: float = 0
    yoy_mtd_raw_tables: float = 0
    yoy_mtd_revenue_wan: float = 0
    yoy_mtd_dine_in_wan: float = 0
    yoy_mtd_takeout_wan: float = 0
    yoy_mtd_turnover_rate: float = 0
    yoy_mtd_per_table: float = 0

    # Targets
    revenue_target: float = 0

    # Time period data: {slot: value}
    tp_turnover_cur: dict[str, float] = field(default_factory=dict)
    tp_turnover_yoy: dict[str, float] = field(default_factory=dict)
    tp_turnover_target: dict[str, float] = field(default_factory=dict)
    tp_turnover_target_total: float = 0  # pre-computed total from DB (avoids slot rounding drift)
    tp_tables_today: dict[str, float] = field(default_factory=dict)
    tp_tables_yoy_weekday: dict[str, float] = field(default_factory=dict)
    tp_turnover_today: dict[str, float] = field(default_factory=dict)
    tp_turnover_yoy_weekday: dict[str, float] = field(default_factory=dict)
    tp_mtd_tables_cur: dict[str, float] = field(default_factory=dict)
    tp_mtd_tables_yoy: dict[str, float] = field(default_factory=dict)


@dataclass
class ReportData:
    """All computed data needed to build the report."""

    dates: ReportDates
    stores: dict[str, StoreMetrics] = field(default_factory=dict)
    competitor: dict[str, str] = field(default_factory=dict)  # store → competitor store


# ── Metric computation ────────────────────────────────────────────────────────


@dataclass
class RawData:
    """Pre-aggregated raw data from all 5 QBI files."""

    # Today (report date)
    today_tables: dict[str, float]
    today_raw_tables: dict[str, float]
    today_takeout: dict[str, float]
    today_revenue: dict[str, float]
    today_dine_in: dict[str, float]
    today_takeout_rev: dict[str, float]
    today_customers: dict[str, float]
    today_turnover: dict[str, float]
    # MTD current
    mtd_tables: dict[str, float]
    mtd_raw_tables: dict[str, float]
    mtd_revenue: dict[str, float]
    mtd_dine_in: dict[str, float]
    mtd_takeout_rev: dict[str, float]
    mtd_discount: dict[str, float]
    mtd_avg_turnover: dict[str, float]
    # Previous month
    prev_tables: dict[str, float]
    prev_raw_tables: dict[str, float]
    prev_revenue: dict[str, float]
    prev_dine_in: dict[str, float]
    prev_takeout_rev: dict[str, float]
    prev_avg_turnover: dict[str, float]   # full prev-month average turnover rate
    # YoY
    yoy_tables: dict[str, float]
    yoy_raw_tables: dict[str, float]
    yoy_revenue: dict[str, float]
    yoy_dine_in: dict[str, float]
    yoy_takeout_rev: dict[str, float]
    yoy_avg_turnover: dict[str, float]
    # Time-period: {store: {slot: value}}
    tp_turnover_cur: dict[str, dict[str, float]]
    tp_turnover_yoy: dict[str, dict[str, float]]
    tp_tables_today: dict[str, dict[str, float]]
    tp_turnover_today: dict[str, dict[str, float]]
    tp_tables_yoy_weekday: dict[str, dict[str, float]]
    tp_turnover_yoy_weekday: dict[str, dict[str, float]]
    tp_mtd_tables_cur: dict[str, dict[str, float]]
    tp_mtd_tables_yoy: dict[str, dict[str, float]]
    # Seats per store (constant, taken from any row)
    seats: dict[str, int]


def _seats_by_store(rows: list[dict]) -> dict[str, int]:
    """Extract seat count per store (takes the first non-zero value found)."""
    result: dict[str, int] = {}
    for row in rows:
        store = row.get(COL_STORE, "")
        if store and store not in result:
            val = row.get(COL_SEATS, 0) or 0
            if val:
                result[store] = int(val)
    return result


def _load_all_raw_data(dates: ReportDates, files: DownloadedFiles) -> RawData:
    """Load and aggregate all raw QBI data into a typed bundle."""
    cur_rows = _load_daily(files.cur_daily)
    prev_rows = _load_daily(files.prev_daily)
    yoy_rows_all = _load_daily(files.yoy_daily)
    cur_tp_rows = _load_time_period(files.cur_time_period)
    yoy_tp_rows_all = _load_time_period(files.yoy_time_period)

    # Filter YoY rows to MTD end date (extra days kept for same-weekday lookups)
    yoy_end_str = dates.yoy_end.strftime("%Y%m%d")
    yoy_rows = _filter_rows_by_date(yoy_rows_all, yoy_end_str)
    yoy_tp_rows = _filter_rows_by_date(yoy_tp_rows_all, yoy_end_str)

    date_str = dates.report_date.strftime("%Y%m%d")
    yoy_weekday_str = dates.yoy_same_weekday.strftime("%Y%m%d")

    return RawData(
        # Today
        today_tables=_last_day_by_store(cur_rows, date_str, COL_TABLES_ASSESSED),
        today_raw_tables=_last_day_by_store(cur_rows, date_str, COL_TABLES_RAW),
        today_takeout=_last_day_by_store(cur_rows, date_str, COL_TABLES_TAKEOUT),
        today_revenue=_last_day_by_store(cur_rows, date_str, COL_REVENUE),
        today_dine_in=_last_day_by_store(cur_rows, date_str, COL_REVENUE_DINE_IN),
        today_takeout_rev=_last_day_two_cols_by_store(cur_rows, date_str, COL_REVENUE_TAKEOUT, COL_REVENUE_DELIVERY),
        today_customers=_last_day_by_store(cur_rows, date_str, COL_CUSTOMERS),
        today_turnover=_last_day_by_store(cur_rows, date_str, COL_TURNOVER),
        # MTD current
        mtd_tables=_sum_by_store(cur_rows, COL_TABLES_ASSESSED),
        mtd_raw_tables=_sum_by_store(cur_rows, COL_TABLES_RAW),
        mtd_revenue=_sum_by_store(cur_rows, COL_REVENUE),
        mtd_dine_in=_sum_by_store(cur_rows, COL_REVENUE_DINE_IN),
        mtd_takeout_rev=_sum_two_cols_by_store(cur_rows, COL_REVENUE_TAKEOUT, COL_REVENUE_DELIVERY),
        mtd_discount=_sum_by_store(cur_rows, COL_DISCOUNT),
        mtd_avg_turnover=_avg_turnover_by_store(cur_rows),
        # Previous month
        prev_tables=_sum_by_store(prev_rows, COL_TABLES_ASSESSED),
        prev_raw_tables=_sum_by_store(prev_rows, COL_TABLES_RAW),
        prev_revenue=_sum_by_store(prev_rows, COL_REVENUE),
        prev_dine_in=_sum_by_store(prev_rows, COL_REVENUE_DINE_IN),
        prev_takeout_rev=_sum_two_cols_by_store(prev_rows, COL_REVENUE_TAKEOUT, COL_REVENUE_DELIVERY),
        prev_avg_turnover=_avg_turnover_by_store(prev_rows),
        # YoY
        yoy_tables=_sum_by_store(yoy_rows, COL_TABLES_ASSESSED),
        yoy_raw_tables=_sum_by_store(yoy_rows, COL_TABLES_RAW),
        yoy_revenue=_sum_by_store(yoy_rows, COL_REVENUE),
        yoy_dine_in=_sum_by_store(yoy_rows, COL_REVENUE_DINE_IN),
        yoy_takeout_rev=_sum_two_cols_by_store(yoy_rows, COL_REVENUE_TAKEOUT, COL_REVENUE_DELIVERY),
        yoy_avg_turnover=_avg_turnover_by_store(yoy_rows),
        # Time-period
        tp_turnover_cur=_avg_time_period(cur_tp_rows, COL_TURNOVER),
        tp_turnover_yoy=_avg_time_period(yoy_tp_rows, COL_TURNOVER),
        tp_tables_today=_last_day_time_period(cur_tp_rows, date_str, COL_TABLES_ASSESSED),
        tp_turnover_today=_last_day_time_period(cur_tp_rows, date_str, COL_TURNOVER),
        tp_tables_yoy_weekday=_last_day_time_period(yoy_tp_rows_all, yoy_weekday_str, COL_TABLES_ASSESSED),
        tp_turnover_yoy_weekday=_last_day_time_period(yoy_tp_rows_all, yoy_weekday_str, COL_TURNOVER),
        tp_mtd_tables_cur=_sum_time_period(cur_tp_rows, COL_TABLES_ASSESSED),
        tp_mtd_tables_yoy=_sum_time_period(yoy_tp_rows, COL_TABLES_ASSESSED),
        seats=_seats_by_store(cur_rows),
    )


def _build_store_metrics(store: str, raw: RawData, targets: Targets) -> StoreMetrics:
    """Compute all metrics for a single store from raw aggregated data."""
    m = StoreMetrics()
    rev_targets = targets.get("revenue", {})
    tr_targets = targets.get("turnover_rate", {})

    # Today
    m.today_tables = raw.today_tables.get(store, 0)
    m.today_raw_tables = raw.today_raw_tables.get(store, 0)
    m.today_takeout_tables = raw.today_takeout.get(store, 0)
    m.today_non_assessed_tables = m.today_raw_tables - m.today_tables
    rev_today = raw.today_revenue.get(store, 0)
    m.today_revenue_wan = rev_today / WAN_DIVISOR
    m.today_dine_in_wan = raw.today_dine_in.get(store, 0) / WAN_DIVISOR
    m.today_takeout_wan = raw.today_takeout_rev.get(store, 0) / WAN_DIVISOR
    cust_today = int(raw.today_customers.get(store, 0))
    m.today_customers = cust_today
    m.today_per_capita = div_or_zero(rev_today, cust_today)
    m.today_per_table = div_or_zero(rev_today, m.today_raw_tables)
    m.today_turnover_rate = raw.today_turnover.get(store, 0)
    m.seats = raw.seats.get(store, 0)

    # MTD current
    mtd_r = raw.mtd_revenue.get(store, 0)
    m.mtd_tables = raw.mtd_tables.get(store, 0)
    m.mtd_raw_tables = raw.mtd_raw_tables.get(store, 0)
    m.mtd_revenue_wan = mtd_r / WAN_DIVISOR
    m.mtd_dine_in_wan = raw.mtd_dine_in.get(store, 0) / WAN_DIVISOR
    m.mtd_takeout_wan = raw.mtd_takeout_rev.get(store, 0) / WAN_DIVISOR
    m.mtd_turnover_rate = raw.mtd_avg_turnover.get(store, 0)
    m.mtd_per_table = div_or_zero(mtd_r, m.mtd_raw_tables)
    m.mtd_discount_wan = raw.mtd_discount.get(store, 0) / WAN_DIVISOR
    m.mtd_discount_pct = div_or_zero(raw.mtd_discount.get(store, 0), mtd_r) * 100

    # Prev month
    prev_r = raw.prev_revenue.get(store, 0)
    m.prev_mtd_tables = raw.prev_tables.get(store, 0)
    m.prev_mtd_raw_tables = raw.prev_raw_tables.get(store, 0)
    m.prev_mtd_revenue_wan = prev_r / WAN_DIVISOR
    m.prev_mtd_dine_in_wan = raw.prev_dine_in.get(store, 0) / WAN_DIVISOR
    m.prev_mtd_takeout_wan = raw.prev_takeout_rev.get(store, 0) / WAN_DIVISOR
    m.prev_mtd_per_table = div_or_zero(prev_r, m.prev_mtd_raw_tables)
    m.prev_mtd_turnover_rate = raw.prev_avg_turnover.get(store, 0)

    # YoY
    yoy_r = raw.yoy_revenue.get(store, 0)
    m.yoy_mtd_tables = raw.yoy_tables.get(store, 0)
    m.yoy_mtd_raw_tables = raw.yoy_raw_tables.get(store, 0)
    m.yoy_mtd_revenue_wan = yoy_r / WAN_DIVISOR
    m.yoy_mtd_dine_in_wan = raw.yoy_dine_in.get(store, 0) / WAN_DIVISOR
    m.yoy_mtd_takeout_wan = raw.yoy_takeout_rev.get(store, 0) / WAN_DIVISOR
    m.yoy_mtd_turnover_rate = raw.yoy_avg_turnover.get(store, 0)
    m.yoy_mtd_per_table = div_or_zero(yoy_r, m.yoy_mtd_raw_tables)

    # Targets
    m.revenue_target = rev_targets.get(store, 0)
    store_tr_target = tr_targets.get(store, {})
    m.tp_turnover_target = {slot: store_tr_target.get(slot, 0) for slot in TIME_SLOTS}
    m.tp_turnover_target_total = store_tr_target.get("total", 0)

    # Time-period data
    m.tp_turnover_cur = raw.tp_turnover_cur.get(store, {})
    m.tp_turnover_yoy = raw.tp_turnover_yoy.get(store, {})
    m.tp_tables_today = raw.tp_tables_today.get(store, {})
    m.tp_tables_yoy_weekday = raw.tp_tables_yoy_weekday.get(store, {})
    m.tp_turnover_today = raw.tp_turnover_today.get(store, {})
    m.tp_turnover_yoy_weekday = raw.tp_turnover_yoy_weekday.get(store, {})
    m.tp_mtd_tables_cur = raw.tp_mtd_tables_cur.get(store, {})
    m.tp_mtd_tables_yoy = raw.tp_mtd_tables_yoy.get(store, {})

    return m


def compute_metrics(
    dates: ReportDates,
    files: DownloadedFiles,
) -> ReportData:
    """Load all raw data and compute every metric for the report.

    Targets and competitor mapping are loaded from the database.
    Use the admin UI at /admin to configure them.
    """
    from server.db import get_competitor_for_report, get_targets_for_report

    try:
        raw = _load_all_raw_data(dates, files)
    except ValueError as e:
        raise ValueError(f"Data loading failed: {e}") from e
    except Exception as e:
        raise RuntimeError(
            f"Unexpected error while loading QBI data: {type(e).__name__}: {e}\n"
            "Check that all 5 QBI files are valid and not corrupted."
        ) from e

    try:
        targets: Targets = get_targets_for_report(dates.month_key)
    except Exception as e:
        logger.warning("Failed to load targets for %s: %s — using empty targets", dates.month_key, e)
        targets = {"revenue": {}, "turnover_rate": {}}

    try:
        competitor = get_competitor_for_report()
    except Exception as e:
        logger.warning("Failed to load competitor config: %s — competitor sheet will be empty", e)
        competitor = {}

    data = ReportData(dates=dates, competitor=competitor)
    for store in STORES:
        try:
            data.stores[store] = _build_store_metrics(store, raw, targets)
        except Exception as e:
            raise RuntimeError(
                f"Failed to build metrics for store {store!r}: {type(e).__name__}: {e}"
            ) from e

    # Validate that computed metrics look sane
    validate_store_coverage(data.stores, label="post-transform")
    validate_no_all_zero_columns(data.stores)

    return data
