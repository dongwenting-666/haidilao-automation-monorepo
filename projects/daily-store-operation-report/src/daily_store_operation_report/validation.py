"""Validation helpers for the daily store operation report pipeline.

Called at multiple stages to catch bad data early:
  1. After file resolution  → validate_file_exists_and_readable, validate_file_timestamps
  2. After loading xlsx     → validate_daily_rows, validate_time_period_rows
  3. After transform        → validate_store_coverage, validate_no_all_zero_columns
  4. After report generation → validate_report_output (post-gen self-test)
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from daily_store_operation_report.constants import (
    COL_DATE,
    COL_REVENUE,
    COL_STORE,
    COL_TABLES_ASSESSED,
    COL_TIME_SLOT,
    COL_TURNOVER,
    STORES,
)

if TYPE_CHECKING:
    from daily_store_operation_report.transform import StoreMetrics

logger = logging.getLogger(__name__)


# ── File-level validation ─────────────────────────────────────────────────────


def validate_file_exists_and_readable(path: Path, label: str = "") -> None:
    """Verify a file exists, is non-empty, and is a valid xlsx.

    Raises FileNotFoundError or ValueError on failure.
    """
    name = label or path.name
    if not path.exists():
        raise FileNotFoundError(f"[Validation] {name}: file not found: {path}")
    size = path.stat().st_size
    if size < 100:
        raise ValueError(
            f"[Validation] {name}: file appears too small ({size} bytes), may be corrupted or empty"
        )
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        logger.debug("[Validation] %s: valid xlsx with sheets: %s", name, sheet_names)
    except Exception as e:
        raise ValueError(
            f"[Validation] {name}: cannot open as xlsx: {e}\n"
            f"  Path: {path}"
        ) from e


def validate_xlsx_has_sheet(path: Path, sheet_name: str) -> None:
    """Verify a specific sheet exists in an xlsx file, with a clear error listing available sheets."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        available = wb.sheetnames
        wb.close()
    except Exception as e:
        raise ValueError(f"[Validation] Cannot open {path.name}: {e}") from e

    if sheet_name not in available:
        raise ValueError(
            f"[Validation] Sheet {sheet_name!r} not found in {path.name}.\n"
            f"  Available sheets: {available}\n"
            f"  Check that the correct file is being used."
        )


# ── Row-level validation ──────────────────────────────────────────────────────


def validate_daily_rows(rows: list[dict], path: Path) -> None:
    """Validate that daily rows contain expected columns and non-trivial data.

    Raises ValueError on fatal problems; logs warnings for soft issues.
    """
    if not rows:
        raise ValueError(
            f"[Validation] Daily file {path.name}: no data rows found — "
            "file may be empty or the sheet structure has changed"
        )

    required_cols = {COL_STORE, COL_DATE, COL_REVENUE, COL_TABLES_ASSESSED}
    actual_cols = set(rows[0].keys())
    missing = required_cols - actual_cols
    if missing:
        raise ValueError(
            f"[Validation] Daily file {path.name}: missing expected columns {missing}.\n"
            f"  Actual columns ({len(actual_cols)}): {sorted(actual_cols)}\n"
            f"  This usually means the QBI export format changed or the wrong sheet was loaded."
        )

    stores_found = {r.get(COL_STORE) for r in rows if r.get(COL_STORE)}
    stores_found_clean = {s for s in stores_found if s}
    if not stores_found_clean:
        raise ValueError(
            f"[Validation] Daily file {path.name}: column {COL_STORE!r} is empty — "
            "no store names found in any row"
        )

    # Warn if expected stores are absent
    expected_stores = set(STORES)
    absent = expected_stores - stores_found_clean
    if absent:
        logger.warning(
            "[Validation] Daily file %s: %d expected store(s) not found: %s",
            path.name, len(absent), sorted(absent),
        )

    logger.info(
        "[Validation] Daily file %s OK: %d rows, %d stores found",
        path.name, len(rows), len(stores_found_clean),
    )


def validate_time_period_rows(rows: list[dict], path: Path) -> None:
    """Validate that time-period rows contain expected columns."""
    if not rows:
        raise ValueError(
            f"[Validation] Time-period file {path.name}: no data rows found — "
            "file may be empty or the sheet structure has changed"
        )

    required_cols = {COL_STORE, COL_DATE, COL_TIME_SLOT, COL_TURNOVER}
    actual_cols = set(rows[0].keys())
    missing = required_cols - actual_cols
    if missing:
        raise ValueError(
            f"[Validation] Time-period file {path.name}: missing expected columns {missing}.\n"
            f"  Actual columns ({len(actual_cols)}): {sorted(actual_cols)}\n"
            f"  This usually means the QBI export format changed or the wrong sheet was loaded."
        )

    slots_found = {
        r.get(COL_TIME_SLOT)
        for r in rows
        if r.get(COL_TIME_SLOT) and r.get(COL_TIME_SLOT) != "-"
    }
    logger.info(
        "[Validation] Time-period file %s OK: %d rows, time slots found: %s",
        path.name, len(rows), sorted(s for s in slots_found if s),
    )


# ── Transform-level validation ────────────────────────────────────────────────


def validate_store_coverage(metrics: dict[str, "StoreMetrics"], label: str = "") -> None:
    """Warn if any expected store has zero MTD revenue AND zero tables."""
    tag = f"{label}: " if label else ""
    missing_data = []
    for store in STORES:
        m = metrics.get(store)
        if m is None:
            missing_data.append(f"{store} (no metrics object)")
            continue
        if m.mtd_revenue_wan == 0 and m.mtd_tables == 0:
            missing_data.append(f"{store} (revenue=0, tables=0)")

    if missing_data:
        logger.warning(
            "[Validation] %sstores with no MTD data: %s",
            tag, missing_data,
        )
    else:
        logger.info("[Validation] %sAll %d stores have non-zero MTD data ✓", tag, len(STORES))


def validate_no_all_zero_columns(metrics: dict[str, "StoreMetrics"]) -> None:
    """Raise if ALL stores have zero for a critical period (signals file-ordering bug).

    A file ordering bug (like the one fixed in commit 5274ed0) causes all
    stores in a period to be zero because wrong-year data gets filtered out.
    """
    checks = [
        ("MTD revenue", [metrics[s].mtd_revenue_wan for s in STORES if s in metrics]),
        ("YoY MTD revenue", [metrics[s].yoy_mtd_revenue_wan for s in STORES if s in metrics]),
        ("prev-month MTD revenue", [metrics[s].prev_mtd_revenue_wan for s in STORES if s in metrics]),
        ("MTD tables", [metrics[s].mtd_tables for s in STORES if s in metrics]),
        ("YoY MTD tables", [metrics[s].yoy_mtd_tables for s in STORES if s in metrics]),
    ]

    errors = []
    warnings = []
    for label, values in checks:
        if not values:
            continue
        nonzero = [v for v in values if v != 0]
        if not nonzero:
            # MTD current-month zeros are a hard error (report is useless)
            if "MTD revenue" in label and "YoY" not in label and "prev" not in label:
                errors.append(label)
            else:
                warnings.append(label)

    for label in warnings:
        logger.warning(
            "[Validation] ALL stores have zero %s — check the corresponding QBI file "
            "(this may indicate a file-ordering bug or new store with no history)",
            label,
        )

    if errors:
        raise ValueError(
            f"[Validation] CRITICAL: ALL stores have zero {', '.join(errors)}.\n"
            "This strongly indicates a file-ordering bug or the wrong file was loaded.\n"
            "Run with --cur-daily / --prev-daily / --yoy-daily to provide explicit files."
        )


# ── Session/timestamp validation ──────────────────────────────────────────────


def _parse_file_timestamp(sort_key: str) -> datetime | None:
    """Parse a download datetime from a sort key like '20260319_2001' or '20260319_2002_2'."""
    m = re.match(r"^(\d{8})_(\d{4})", sort_key)
    if not m:
        return None
    try:
        return datetime.strptime(f"{m.group(1)}{m.group(2)}", "%Y%m%d%H%M")
    except ValueError:
        return None


def validate_file_timestamps(
    files_with_keys: list[tuple[Path, str]],
    max_gap_minutes: int = 15,
) -> None:
    """Warn if the 5 resolved files have download timestamps more than *max_gap_minutes* apart.

    Files from the same download session should have timestamps within a few minutes.
    Large gaps suggest files were accidentally mixed from different sessions.
    """
    parsed: list[tuple[str, datetime]] = []
    for path, sort_key in files_with_keys:
        ts = _parse_file_timestamp(sort_key)
        if ts is not None:
            parsed.append((path.name, ts))

    if len(parsed) < 2:
        logger.debug("[Validation] Not enough timestamped files to check session consistency")
        return

    parsed.sort(key=lambda x: x[1])
    min_name, min_ts = parsed[0]
    max_name, max_ts = parsed[-1]
    gap_minutes = (max_ts - min_ts).total_seconds() / 60

    if gap_minutes > max_gap_minutes:
        logger.warning(
            "[Validation] ⚠️  Files may be from DIFFERENT download sessions!\n"
            "  Timestamp spread: %.1f minutes (threshold: %d min)\n"
            "  Earliest: %s at %s\n"
            "  Latest:   %s at %s\n"
            "  Consider using --cur-daily / --prev-daily / --yoy-daily / --cur-tp / --yoy-tp "
            "to explicitly specify which files to use.",
            gap_minutes,
            max_gap_minutes,
            min_name,
            min_ts.strftime("%Y-%m-%d %H:%M"),
            max_name,
            max_ts.strftime("%Y-%m-%d %H:%M"),
        )
    else:
        logger.info(
            "[Validation] File session OK — all timestamps within %.1f minutes ✓",
            gap_minutes,
        )


# ── Post-generation self-test ─────────────────────────────────────────────────


def validate_report_output(output_path: Path, stores: list[str] | None = None) -> None:
    """Post-generation self-test: open the output xlsx and verify key data integrity.

    Checks performed:
    - All expected sheets are present
    - 对比上年表 rows for MTD tables and MTD revenue are non-zero for each store
    - Region total (column K) ≈ sum of store columns (within 1%)
    - 分时段-上报 subtotal rows have non-zero time-period data for each store

    Logs warnings for soft failures; raises ValueError for hard failures.
    """
    if stores is None:
        stores = STORES

    logger.info("[Self-test] Validating generated report: %s", output_path.name)

    try:
        wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(
            f"[Self-test] Cannot open generated report {output_path}: {e}"
        ) from e

    actual_sheets = set(wb.sheetnames)
    errors: list[str] = []
    warnings: list[str] = []

    # ── 1. Check all expected sheets exist ───────────────────────────────────
    for sheet in [
        "对比上年表",
        "分时段-上报",
        "同比数据",
        "对比上月表",
        "加拿大片区假想敌翻台率对比",
        "加拿大片区假想敌外卖收入对比",
    ]:
        if sheet not in actual_sheets:
            errors.append(f"Missing sheet: {sheet!r}")

    if errors:
        wb.close()
        raise ValueError(f"[Self-test] Missing sheets: {'; '.join(errors)}")

    n_stores = len(stores)
    store_cols = list(range(3, 3 + n_stores))  # columns C through J (1-indexed)
    region_col = 11  # column K

    # ── 2. Validate 对比上年表 ────────────────────────────────────────────────
    ws_yoy = wb["对比上年表"]

    #  Row 6  = "{month}月总桌数"
    row6_vals = [ws_yoy.cell(row=6, column=c).value for c in store_cols]
    row6_numeric = [v for v in row6_vals if isinstance(v, (int, float))]
    row6_nonzero = [v for v in row6_numeric if v != 0]

    if not row6_nonzero:
        errors.append("对比上年表 row 6 (月总桌数): all stores are zero — data loading likely failed")
    elif len(row6_nonzero) < n_stores:
        zero_stores = [stores[i] for i, v in enumerate(row6_vals) if not (isinstance(v, (int, float)) and v != 0)]
        warnings.append(f"对比上年表 row 6 (月总桌数): {len(zero_stores)} store(s) are zero: {zero_stores}")

    #  Row 12 = "本月截止目前营业收入(万)" (shifted +2 from old row 10 by dine-in/takeout rows)
    row12_vals = [ws_yoy.cell(row=12, column=c).value for c in store_cols]
    row12_numeric = [v for v in row12_vals if isinstance(v, (int, float))]
    row12_nonzero = [v for v in row12_numeric if v != 0]

    if not row12_nonzero:
        errors.append("对比上年表 row 12 (本月截止目前营业收入): all stores are zero")
    elif len(row12_nonzero) < n_stores:
        zero_stores = [stores[i] for i, v in enumerate(row12_vals) if not (isinstance(v, (int, float)) and v != 0)]
        warnings.append(f"对比上年表 row 12 (营业收入): {len(zero_stores)} store(s) are zero: {zero_stores}")

    # Region total check for row 12
    region_val = ws_yoy.cell(row=12, column=region_col).value
    store_sum = sum(v for v in row12_numeric if v != 0)
    if isinstance(region_val, (int, float)) and store_sum > 0:
        ratio = abs(region_val - store_sum) / store_sum
        if ratio > 0.01:
            warnings.append(
                f"对比上年表 row 12: region total ({region_val:.4f}) "
                f"≠ sum of stores ({store_sum:.4f}), diff={ratio:.2%} — "
                "region total may be computed incorrectly"
            )
        else:
            logger.info(
                "[Self-test] 对比上年表 row 12: region total matches store sum (diff=%.4f%%) ✓",
                ratio * 100,
            )

    #  Row 15 = "上年截止目前营业收入(万)" (shifted +4 from old row 11 by dine-in/takeout rows)
    row15_vals = [ws_yoy.cell(row=15, column=c).value for c in store_cols]
    row15_nonzero = [v for v in row15_vals if isinstance(v, (int, float)) and v != 0]
    if not row15_nonzero:
        warnings.append(
            "对比上年表 row 15 (上年营业收入): all stores are zero — "
            "check that yoy_daily file contains data for the comparison period"
        )

    # ── 3. Validate 分时段-上报 ───────────────────────────────────────────────
    ws_tp = wb["分时段-上报"]
    n_slots = 4  # TIME_SLOTS has 4 entries
    rows_per_store = n_slots + 1  # 4 data rows + 1 subtotal
    stores_zero_tp: list[str] = []

    for store_idx, store in enumerate(stores):
        subtotal_row = 4 + store_idx * rows_per_store + n_slots
        # col 3 = 今年翻台率 (MTD avg), col 13 = 本月截止目前今年桌数
        c3 = ws_tp.cell(row=subtotal_row, column=3).value
        c13 = ws_tp.cell(row=subtotal_row, column=13).value
        c3_ok = isinstance(c3, (int, float)) and c3 != 0
        c13_ok = isinstance(c13, (int, float)) and c13 != 0
        if not c3_ok and not c13_ok:
            stores_zero_tp.append(store)

    if stores_zero_tp:
        warnings.append(
            f"分时段-上报: {len(stores_zero_tp)} store(s) have all-zero time-period subtotals: "
            f"{stores_zero_tp}"
        )
    else:
        logger.info("[Self-test] 分时段-上报: all stores have non-zero time-period data ✓")

    wb.close()

    # ── Summarise ────────────────────────────────────────────────────────────
    if errors:
        msg = "\n".join(f"  ❌ {e}" for e in errors)
        raise ValueError(f"[Self-test] Report validation FAILED:\n{msg}")

    if warnings:
        for w in warnings:
            logger.warning("[Self-test] ⚠️  %s", w)
        logger.info(
            "[Self-test] Report validation passed with %d warning(s) — review above",
            len(warnings),
        )
    else:
        logger.info(
            "[Self-test] ✅ Report validation PASSED — "
            "all key metrics are non-zero and region totals match"
        )
