"""KSB1 accounting check: compare last month vs this month by 科目 per store."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from ksb1_accounting_check.rules import NOTE_CURR_ONLY, NOTE_PREV_ONLY, analyze_store

log = logging.getLogger(__name__)

MAPPING_FILE = Path(__file__).resolve().parent / "报表科目.xlsx"

BOLD_FONT = Font(bold=True)


def load_cost_element_mapping(path: Path = MAPPING_FILE) -> dict[str, str]:
    """Load 成本要素 → 报表科目 mapping from the mapping spreadsheet.

    Returns:
        dict mapping cost element code (str) to 报表科目 name.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, _name, category = row[0], row[1], row[2]
        if code is not None and category is not None:
            mapping[str(int(code) if isinstance(code, (int, float)) else code)] = str(category)
    wb.close()
    log.info("Loaded %d cost element mappings", len(mapping))
    return mapping


def load_ksb1_data(path: Path) -> list[dict]:
    """Load raw KSB1 export into a list of row dicts.

    Expected columns from SAP KSB1 export:
        过账日期, 公司代码, 成本中心, CO对象名称, 成本要素名称, 成本要素,
        名称, AuxAcctAs1, 业务货币值, 交易货币, 对象货币值, 参考凭证号码,
        物料描述, 凭证抬头文本, 用户名, 物料, 报表货币值, 贸易伙伴, 期间
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        # Skip empty/summary rows
        if data.get("过账日期") is None:
            continue
        rows.append(data)
    wb.close()
    log.info("Loaded %d rows from %s", len(rows), path.name)
    return rows


def enrich_rows(rows: list[dict], mapping: dict[str, str]) -> None:
    """Add 月份 and 科目 columns to each row (mutates in place)."""
    unmapped = set()
    for row in rows:
        # Extract month from 过账日期
        post_date = row["过账日期"]
        if hasattr(post_date, "month"):
            row["月份"] = post_date.month
        else:
            row["月份"] = None

        # Map 成本要素 → 科目
        code = str(row.get("成本要素", "")).strip()
        if code in mapping:
            row["科目"] = mapping[code]
        else:
            row["科目"] = None
            if code:
                unmapped.add(code)

    if unmapped:
        log.warning("Unmapped cost elements: %s", ", ".join(sorted(unmapped)))


def split_by_month(rows: list[dict], prev_month: int, curr_month: int) -> tuple[list[dict], list[dict]]:
    """Split rows into previous month and current month."""
    prev = [r for r in rows if r.get("月份") == prev_month]
    curr = [r for r in rows if r.get("月份") == curr_month]
    log.info("Split: %d rows in month %d, %d rows in month %d", len(prev), prev_month, len(curr), curr_month)
    return prev, curr


def build_store_data(
    prev_rows: list[dict],
    curr_rows: list[dict],
    store: str,
) -> tuple[list[dict], dict[str, list[dict]]]:
    """Build per-store 科目 summary and grouped detail rows.

    Returns:
        kemu_summary: list of dicts (科目, 上月金额, 本月金额, 差异, 备注, 明细)
        kemu_rows: dict mapping 科目 → list of all transaction rows (both months)
    """
    store_prev = [r for r in prev_rows if r.get("CO对象名称") == store]
    store_curr = [r for r in curr_rows if r.get("CO对象名称") == store]

    def collect(rows: list[dict]) -> dict[str, list[dict]]:
        groups: dict[str, list[dict]] = {}
        for r in rows:
            kemu = r.get("科目")
            if not kemu:
                continue
            groups.setdefault(kemu, []).append(r)
        return groups

    def total(rows: list[dict]) -> float:
        return sum(r.get("对象货币值") or 0 for r in rows if isinstance(r.get("对象货币值"), (int, float)))

    def build_sub_detail(p_rows: list[dict], c_rows: list[dict]) -> list[dict]:
        """Build 成本要素名称-level breakdown within a 科目."""
        def group_by_name(rows):
            groups = {}
            for r in rows:
                name = r.get("成本要素名称") or "未知"
                groups.setdefault(name, []).append(r)
            return groups

        prev_by_name = group_by_name(p_rows)
        curr_by_name = group_by_name(c_rows)
        all_names = sorted(set(prev_by_name) | set(curr_by_name))

        details = []
        for name in all_names:
            p = prev_by_name.get(name, [])
            c = curr_by_name.get(name, [])
            p_amt = round(total(p), 2)
            c_amt = round(total(c), 2)
            d = round(c_amt - p_amt, 2)
            note = ""
            if p and not c:
                note = NOTE_PREV_ONLY
            elif c and not p:
                note = NOTE_CURR_ONLY
            details.append({
                "成本要素名称": name,
                "上月金额": p_amt,
                "本月金额": c_amt,
                "差异": d,
                "备注": note,
            })
        return details

    prev_groups = collect(store_prev)
    curr_groups = collect(store_curr)
    all_kemu = sorted(set(prev_groups) | set(curr_groups))

    kemu_summary = []
    kemu_rows: dict[str, list[dict]] = {}
    for kemu in all_kemu:
        p_rows = prev_groups.get(kemu, [])
        c_rows = curr_groups.get(kemu, [])
        prev_amt = round(total(p_rows), 2)
        curr_amt = round(total(c_rows), 2)
        diff = round(curr_amt - prev_amt, 2)

        note = ""
        if p_rows and not c_rows:
            note = NOTE_PREV_ONLY
        elif c_rows and not p_rows:
            note = NOTE_CURR_ONLY

        sub_detail = build_sub_detail(p_rows, c_rows)

        kemu_summary.append({
            "科目": kemu,
            "上月金额": prev_amt,
            "本月金额": curr_amt,
            "差异": diff,
            "备注": note,
            "明细": sub_detail,
        })
        kemu_rows[kemu] = p_rows + c_rows

    return kemu_summary, kemu_rows


# Default stores to analyze — 销售公共组 1-8
DEFAULT_STORE_KEYWORDS = ["一店销售公共组", "二店销售公共组", "三店销售公共组", "四店销售公共组",
                          "五店销售公共组", "六店销售公共组", "七店销售公共组", "八店销售公共组"]


def generate_report(
    ksb1_path: Path,
    output_path: Path,
    target_month: int,
    mapping_path: Path = MAPPING_FILE,
    store_keywords: list[str] | None = None,
    model: str | None = None,
    prompt_path: Path | None = None,
) -> Path:
    """Generate the full KSB1 accounting check report.

    Args:
        ksb1_path: Path to raw KSB1 export (XLSX from SAP).
        output_path: Path for the output report.
        target_month: The month being checked (1-12).
        mapping_path: Path to 报表科目.xlsx mapping file.
        store_keywords: List of keywords to filter stores. A store is included
            if any keyword is found in its name. Defaults to 销售公共组 1-8.
        model: Ollama model name for LLM enhancement. None = rules only.
        prompt_path: Override path to LLM prompt file. None = use default.

    Returns:
        Path to the generated report.
    """
    prev_month = 12 if target_month == 1 else target_month - 1
    keywords = store_keywords if store_keywords is not None else DEFAULT_STORE_KEYWORDS

    # Initialize LLM client if requested
    llm_client = None
    if model:
        from ksb1_accounting_check.llm import create_client, set_prompt_path
        log.info("Initializing LLM (%s) for observation enhancement...", model)
        llm_client = create_client(model=model)
        if prompt_path is not None:
            set_prompt_path(prompt_path)

    # Load data
    mapping = load_cost_element_mapping(mapping_path)
    rows = load_ksb1_data(ksb1_path)
    enrich_rows(rows, mapping)
    prev_rows, curr_rows = split_by_month(rows, prev_month, target_month)

    # Get unique stores, filtered by keywords
    all_stores = sorted({r["CO对象名称"] for r in rows if r.get("CO对象名称")})
    stores = [s for s in all_stores if any(kw in s for kw in keywords)]
    log.info("Found %d stores matching filter (of %d total)", len(stores), len(all_stores))

    # Build output workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create store comparison sheets
    for store in stores:
        kemu_summary, kemu_rows = build_store_data(prev_rows, curr_rows, store)
        if not kemu_summary:
            continue

        log.info("Analyzing %s...", store)
        # Collect all rows for this store (both months) for cross-store detection
        store_all_rows = [r for r in rows if r.get("CO对象名称") == store]
        findings = analyze_store(store, prev_month, target_month, kemu_summary, all_rows=store_all_rows)

        # Optional: enhance observations with LLM
        if llm_client and findings:
            from ksb1_accounting_check.llm import enhance_findings
            findings = enhance_findings(
                llm_client, store, prev_month, target_month, findings, kemu_rows,
            )

        sheet_name = _short_store_name(store)
        ws = wb.create_sheet(title=sheet_name)

        _write_findings_sheet(ws, findings, kemu_rows, prev_month, target_month)

    # Raw data sheet
    _write_raw_data_sheet(wb, rows, prev_month, target_month)

    # Copy mapping sheet
    _write_mapping_sheet(wb, mapping_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Report saved to %s", output_path)
    return output_path


def _build_detail_rows(raw_rows: list[dict]) -> list[dict]:
    """Build detail rows for display under a finding.

    If few rows per month (≤ threshold): show individual transactions.
    Only aggregate rows that share the exact same 名称 within a month.
    """
    # Group by (月份, 名称)
    groups: dict[tuple, list[dict]] = {}
    for r in raw_rows:
        m = r.get("月份")
        if m is None:
            continue
        name = r.get("名称") or ""
        key = (m, name)
        groups.setdefault(key, []).append(r)

    result = []
    for (month, name), rows in sorted(groups.items(), key=lambda x: (x[0][0], x[0][1])):
        total = sum(
            r.get("对象货币值") or 0
            for r in rows
            if isinstance(r.get("对象货币值"), (int, float))
        )
        if len(rows) == 1:
            display_name = name
        else:
            display_name = f"{name[:30]}等{len(rows)}笔"
        result.append({"月份": month, "对象货币值": round(total, 2), "名称": display_name})

    return result


def _write_findings_sheet(
    ws: Worksheet,
    findings: list[dict],
    kemu_rows: dict[str, list[dict]],
    prev_month: int,
    curr_month: int,
) -> None:
    """Write a store sheet: concise notes only, like the manual report.

    Output format (matching manual style):
        Row 1: "说明"
        Row 3: "12月多计提电费5K"
        Row 5: "12月多530 左右保险费"
        ...
    No detail tables — just one-line observations per finding, spaced out.
    Detail data lives in the 原数据 sheet for anyone who needs to drill down.
    """
    ws.cell(row=1, column=1, value="说明").font = BOLD_FONT
    row_num = 2

    for finding in findings:
        observation = finding["observation"]
        # Write observation as a clean one-liner
        ws.cell(row=row_num, column=1, value=observation)
        row_num += 2  # blank row between findings for readability

    # Auto-width column A
    max_len = 10
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions["A"].width = min(max_len + 4, 80)


def _short_store_name(name: str) -> str:
    """Shorten store name to match manual format: '一店', '二店', etc."""
    # "加拿大一店销售公共组" → "一店"
    name = name.replace("加拿大", "").replace("销售公共组", "").strip()
    return name[:31]


def _write_raw_data_sheet(wb: Workbook, rows: list[dict], prev_month: int, curr_month: int) -> None:
    """Write the 原数据 sheet with all enriched rows."""
    ws = wb.create_sheet(title=f"原数据（{prev_month}月&{curr_month}月）")

    headers = [
        "公司代码", "CO对象名称", "过账日期", "月份", "科目",
        "成本要素", "成本要素名称", "对象货币值", "报表货币值",
        "名称", "物料描述", "凭证抬头文本",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = BOLD_FONT

    for i, row in enumerate(rows, 2):
        for col, key in enumerate(headers, 1):
            ws.cell(row=i, column=col, value=row.get(key))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15


def _write_mapping_sheet(wb: Workbook, mapping_path: Path) -> None:
    """Copy the 报表科目 mapping sheet into the output workbook."""
    src_wb = load_workbook(mapping_path, read_only=True, data_only=True)
    src_ws = src_wb.active

    ws = wb.create_sheet(title="报表科目")
    for row in src_ws.iter_rows(values_only=True):
        # Only copy first 3 meaningful columns
        ws.append([row[0], row[1], row[2]])

    src_wb.close()
