"""End-to-end tests for the 毛利相关分析指标 orchestrator."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from zfi0049_report.gross_margin_analysis import (
    GrossMarginInputs,
    build_workbook,
)
from zfi0049_report.static_meta import STORE_META
from zfi0049_report.table1_dish import PosSale


MANUAL_WB = Path(
    "/Users/hongming-claw/Downloads/附件3-毛利相关分析指标-2603.xlsx"
)
ZFI0156_MARCH = Path(
    "/Users/hongming-claw/haidilao-automation-monorepo/output/zfi0156/zfi0156-202603.xlsx"
)
MB5B_MARCH = Path(
    "/Users/hongming-claw/haidilao-automation-monorepo/output/mb5b/mb5b202603.xls"
)


EXPECTED_SHEETS = [
    "填写说明",
    "细分毛利率表 (2)",
    "毛利率连续对比表",
    "毛利率环比",
    "毛利率同比",
    "表1-菜品价格变动及菜品损耗表 (模板) ",
    "表2-原材料成本变动表",
    "表3-打折优惠表",
    "基础数据",
]


def test_build_workbook_minimal_inputs_produces_all_9_sheets(tmp_path):
    """Empty inputs still produce a viewable workbook with all sheets."""
    inputs = GrossMarginInputs(year=2026, month=3)
    out = tmp_path / "minimal.xlsx"
    build_workbook(inputs, out)
    assert out.exists()
    wb = openpyxl.load_workbook(out, read_only=True)
    assert wb.sheetnames == EXPECTED_SHEETS
    wb.close()


def test_build_workbook_basic_data_covers_all_stores(tmp_path):
    """基础数据 sheet should have one row per store in STORE_META even when
    P&L data is missing (zero-filled)."""
    inputs = GrossMarginInputs(year=2026, month=3)
    out = tmp_path / "stores.xlsx"
    build_workbook(inputs, out)
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb["基础数据"]
    # 1 header row + 9 stores
    assert ws.max_row == 1 + len(STORE_META)
    wb.close()


def test_build_workbook_with_mom_data(tmp_path):
    """Provide cur+prev P&L → 毛利率环比 sheet gets populated."""
    cur_pnl = {
        "加拿大一店": {
            "三、毛利率": 0.6974,
            "1、销售净收入": 783542.0,
            "优惠总金额（不含税）": 48724.21,
        },
    }
    prev_pnl = {
        "加拿大一店": {
            "三、毛利率": 0.700639,
            "1、销售净收入": 700000.0,
            "优惠总金额（不含税）": 68148.38,
        },
    }
    inputs = GrossMarginInputs(
        year=2026, month=3,
        cur_pnl=cur_pnl, prev_pnl=prev_pnl,
    )
    out = tmp_path / "mom.xlsx"
    build_workbook(inputs, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["毛利率环比"]
    # Row 2 should be 加拿大一店 (only store with P&L data).
    assert ws.cell(2, 2).value == "加拿大一店"
    # 本月毛利率
    assert ws.cell(2, 3).value == pytest.approx(0.6974)
    wb.close()


def test_build_workbook_with_table3_discount(tmp_path):
    cur_pnl = {
        "加拿大一店": {
            "1、销售净收入": 1075411.75,
            "优惠总金额（不含税）": 48724.21,
        },
    }
    inputs = GrossMarginInputs(year=2026, month=3, cur_pnl=cur_pnl)
    out = tmp_path / "discount.xlsx"
    build_workbook(inputs, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["表3-打折优惠表"]
    # Header + 1 store
    assert ws.max_row == 2
    # 优惠占比 (col 4): 48724.21 / (1075411.75 + 48724.21) = 0.04335
    assert ws.cell(2, 4).value == pytest.approx(0.04335, abs=1e-4)
    wb.close()


@pytest.mark.skipif(
    not (ZFI0156_MARCH.exists() and MB5B_MARCH.exists()),
    reason="ZFI0156 + MB5B March 2026 not on this machine",
)
def test_build_workbook_with_real_zfi_and_mb5b(tmp_path):
    """E2E with real material data + synthetic POS for one store."""
    from zfi0049_report.table1_dish import load_mb5b_prices, load_zfi0156
    inputs = GrossMarginInputs(
        year=2026, month=3,
        zfi_cur=load_zfi0156(ZFI0156_MARCH),
        mb5b_cur=load_mb5b_prices(MB5B_MARCH),
    )
    out = tmp_path / "real_materials.xlsx"
    build_workbook(inputs, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["表2-原材料成本变动表"]
    # Real data should have >100 material rows in 表2
    assert ws.max_row > 100
    # Spot check: row should reference a real material
    found_a_material = False
    for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
        if row[3] and isinstance(row[3], int):  # 物料代码
            found_a_material = True
            break
    assert found_a_material
    wb.close()


@pytest.mark.skipif(
    not MANUAL_WB.exists(), reason="manual workbook not on this machine"
)
def test_sheet_names_match_manual():
    """Our orchestrator produces the same 9 sheet names as the manual."""
    wb = openpyxl.load_workbook(MANUAL_WB, read_only=True)
    manual_sheets = wb.sheetnames
    wb.close()
    assert manual_sheets == EXPECTED_SHEETS, (
        f"manual sheets: {manual_sheets}\nours: {EXPECTED_SHEETS}"
    )
