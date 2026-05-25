"""Tests for the 毛利率 derivative sheets (毛利率环比/连续对比/同比)."""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from zfi0049_report.derivative_sheets import (
    DEFAULT_MONTH_SERIALS,
    MOM_HEADERS_ROW_4,
    MomRow,
    TREND_HEADERS_ROW_5,
    _flag_consecutive_decline,
    _flag_two_month_anomaly,
    build_mom_sheet,
    build_trend_rows,
    build_trend_sheet,
    build_yoy_sheet,
    mom_row_to_excel,
    restored_gp_discount,
    restored_gp_dish,
    restored_gp_material,
)


# ── Formula tests verified against manual workbook ──


def test_restored_gp_dish_matches_manual_canada_one():
    """加拿大一店 March 2026: cur_gp=0.6974, revenue=1075411.75, Δdish=-1127.27
    → manual restored = 0.697716859766."""
    r = restored_gp_dish(cur_gp=0.6974, revenue=1075411.75, delta_dish=-1127.27)
    assert r == pytest.approx(0.69772, abs=1e-4)


def test_restored_gp_material_matches_manual_canada_one():
    """加拿大一店 March 2026: cur_gp=0.6974, revenue=1075411.75, Δmat=-2093
    → manual restored = 0.695453768707."""
    r = restored_gp_material(cur_gp=0.6974, revenue=1075411.75, delta_material=-2093)
    assert r == pytest.approx(0.69545, abs=1e-4)


def test_restored_gp_dish_handles_zero_revenue():
    assert restored_gp_dish(0.5, 0, 100.0) == 0.5


def test_restored_gp_material_handles_zero_revenue():
    assert restored_gp_material(0.5, 0, 100.0) == 0.5


def test_restored_gp_discount_matches_manual_canada_one_march():
    """加拿大一店 March 2026: cur_gp=0.6974, discount_pct=0.0433
    → manual 本月还原毛利率 = 0.7105."""
    r = restored_gp_discount(cur_gp=0.6974, discount_pct=0.0433)
    assert r == pytest.approx(0.7105, abs=1e-4)


def test_restored_gp_discount_matches_manual_canada_one_feb():
    """加拿大一店 上月: cur_gp=0.7006, discount_pct=0.0634
    → manual 上月还原毛利率 = 0.7196."""
    r = restored_gp_discount(cur_gp=0.7006, discount_pct=0.0634)
    assert r == pytest.approx(0.7196, abs=1e-4)


def test_restored_gp_discount_no_discount_returns_cur_gp():
    """When discount = 0 the restored margin equals the input margin."""
    assert restored_gp_discount(0.6974, 0.0) == pytest.approx(0.6974)


# ── MomRow → 23-col Excel projection ──


def test_mom_row_to_excel_length_and_layout():
    r = MomRow(
        store="加拿大一店", cur_gp=0.6974, prev_gp=0.700639573585,
        cur_revenue=1075411.75, prev_revenue=1006336.12,
        delta_dish=-1127.27, delta_material=-2093,
        cur_discount_pct=0.043343698, prev_discount_pct=0.063424256,
    )
    row = mom_row_to_excel(r)
    assert len(row) == 23
    assert row[1] == "加拿大一店"
    # Manual cur_gp - prev_gp = -0.003239573...
    assert row[4] == pytest.approx(-0.003240, abs=1e-4)
    # 还原毛利率 (dish, col 7 — 0-indexed [6])
    assert row[6] == pytest.approx(0.69772, abs=1e-4)
    # 毛利率影响 (dish, col 8) = cur - restored = manual -0.00031685976
    assert row[7] == pytest.approx(-0.000317, abs=1e-4)
    # 还原毛利率 (material, col 10)
    assert row[9] == pytest.approx(0.69545, abs=1e-4)
    # 毛利率影响 (material, col 11) = manual 0.001946231292
    assert row[10] == pytest.approx(0.001946, abs=1e-4)
    # Revenue MoM (col 23 — 0-indexed [22])
    assert row[22] == pytest.approx(0.06865, abs=1e-3)


def test_mom_row_revenue_mom_none_when_prev_zero():
    r = MomRow(
        store="x", cur_gp=0.5, prev_gp=0.5,
        cur_revenue=1000, prev_revenue=0,
        delta_dish=0, delta_material=0,
        cur_discount_pct=0, prev_discount_pct=0,
    )
    row = mom_row_to_excel(r)
    assert row[22] is None


def test_build_mom_sheet_smoke():
    wb = Workbook()
    wb.remove(wb.active)
    rows = [
        MomRow(store="加拿大一店", cur_gp=0.6974, prev_gp=0.7006,
               cur_revenue=1075412, prev_revenue=1006336,
               delta_dish=-1127.27, delta_material=-2093,
               cur_discount_pct=0.0433, prev_discount_pct=0.0634),
        MomRow(store="加拿大二店", cur_gp=0.6988, prev_gp=0.7202,
               cur_revenue=653250, prev_revenue=572025,
               delta_dish=-732.99, delta_material=-1128,
               cur_discount_pct=0.0652, prev_discount_pct=0.0682),
    ]
    ws = build_mom_sheet(wb, rows)
    assert ws.title == "毛利率环比"
    assert ws.cell(1, 2).value == "门店名称"
    assert ws.cell(2, 2).value == "加拿大一店"
    assert ws.max_row == 3
    assert ws.max_column == 23


# ── Trend (7-month) sheet ──


def test_flag_consecutive_decline():
    # 3 consecutive declines: most recent should be down → down → down
    assert _flag_consecutive_decline([0.7, 0.72, 0.74, 0.76], 3) == "Y"
    # Only 2 declines
    assert _flag_consecutive_decline([0.7, 0.72, 0.74, 0.73], 3) == "N"
    # 2 declines OK for k=2
    assert _flag_consecutive_decline([0.7, 0.72, 0.74, 0.73], 2) == "Y"


def test_flag_two_month_anomaly():
    # |delta| > 2% twice
    assert _flag_two_month_anomaly([-0.03, 0.025]) == "Y"
    # Only one exceeds
    assert _flag_two_month_anomaly([-0.03, 0.01]) == "N"


def test_build_trend_rows_canada_one_match_manual():
    """Manual 加拿大一店 7-month gross margins from the screenshot:
    0.6974, 0.700640, 0.693248, 0.724327, 0.718356, 0.696945, 0.694533
    Expected flags: 是否连续3个月下降=N, 是否连续2个月下降=N, anomaly=N,
    drop>2%=N, below60%=N."""
    monthly_gp = {
        "加拿大一店": [0.6974, 0.700639573585016, 0.69324773666432,
                       0.724327369726264, 0.718356127001213,
                       0.696945251618259, 0.694533436007962],
        "加拿大二店": [0.6988, 0.720222155748908, 0.708552555997,
                       0.719513169551, 0.734392971946, 0.726357251750,
                       0.730793081799],
    }
    rows = build_trend_rows(monthly_gp=monthly_gp)
    assert len(rows) == 2
    r1 = rows[0]
    # row layout: [_, idx, region, store, m1..m7, Δ1..Δ4, region_avg, diff, 5 flags]
    assert r1[3] == "加拿大一店"
    assert r1[4] == pytest.approx(0.6974)
    # Δ1 (m1-m2) = 0.6974 - 0.700640 = -0.003240
    assert r1[11] == pytest.approx(-0.003240, abs=1e-4)
    # Flags
    assert r1[17] == "N"  # 3-month decline
    assert r1[18] == "N"  # 2-month decline
    assert r1[19] == "N"  # 2-month anomaly
    assert r1[20] == "N"  # drop > 2%
    assert r1[21] == "N"  # below 60%


def test_build_trend_rows_below_60_percent_flag():
    monthly_gp = {"加拿大十店": [0.55] + [0.6] * 6}
    rows = build_trend_rows(monthly_gp=monthly_gp)
    assert rows[0][21] == "Y"


def test_build_trend_sheet_smoke():
    wb = Workbook()
    wb.remove(wb.active)
    monthly_gp = {
        "加拿大一店": [0.7] + [0.69] * 6,
    }
    rows = build_trend_rows(monthly_gp=monthly_gp)
    ws = build_trend_sheet(wb, rows)
    assert ws.title == "毛利率连续对比表"
    # Header should have replaced M1-M7 placeholders with default serials
    assert ws.cell(1, 5).value == DEFAULT_MONTH_SERIALS[0]
    assert ws.cell(2, 4).value == "加拿大一店"


# ── YoY sheet (same layout as MoM with different labels) ──


def test_build_yoy_sheet_relabels_prev_columns():
    wb = Workbook()
    wb.remove(wb.active)
    rows = [MomRow(
        store="加拿大一店", cur_gp=0.6974, prev_gp=0.683,
        cur_revenue=1075412, prev_revenue=1100000,
        delta_dish=6424, delta_material=-2635,
        cur_discount_pct=0.0433, prev_discount_pct=0.0671,
    )]
    ws = build_yoy_sheet(wb, rows)
    assert ws.title == "毛利率同比"
    # Header relabels
    assert ws.cell(1, 4).value == "去年同期毛利率"
    assert ws.cell(1, 13).value == "去年同期优惠占比"
    assert ws.cell(1, 22).value == "去年同期收入"
