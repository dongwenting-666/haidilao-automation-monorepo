"""Tests for 表3 (打折优惠表) builder."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from zfi0049_report.table3_discount import (
    HEADERS_LINE_1,
    PNL_DISCOUNT_KEY,
    PNL_REVENUE_KEY,
    Table3Row,
    build_rows,
    to_row,
    write_sheet,
)

MANUAL_WB = Path(
    "/Users/hongming-claw/Downloads/附件3-毛利相关分析指标-2603.xlsx"
)


def test_header_count_15():
    assert len(HEADERS_LINE_1) == 15


def test_row_discount_pct_formula():
    r = Table3Row(
        region="加拿大", store="加拿大一店",
        cur_revenue=1075411.75, cur_discount=48724.21,
        prev_revenue=1006336.12, prev_discount=68148.38,
        yoy_revenue=1106125.93, yoy_discount=74283.62,
        is_comparable=True,
    )
    # Match manual: 加拿大一店 cur 优惠占比 = 0.0433437
    assert r.cur_discount_pct == pytest.approx(0.04334, abs=1e-4)
    # prev 优惠占比 = 0.0634
    assert r.prev_discount_pct == pytest.approx(0.06342, abs=1e-4)


def test_row_deltas_match_manual():
    """Manual 加拿大一店: cur=0.0433, prev=0.0634 → 环比=-0.0201."""
    r = Table3Row(
        region="加拿大", store="加拿大一店",
        cur_revenue=1075411.75, cur_discount=48724.21,
        prev_revenue=1006336.12, prev_discount=68148.38,
        yoy_revenue=1106125.93, yoy_discount=74283.62,
        is_comparable=True,
    )
    assert r.mom_delta == pytest.approx(-0.0201, abs=1e-3)


def test_row_handles_missing_prev_yoy():
    r = Table3Row(
        region="加拿大", store="加拿大八店",
        cur_revenue=500000.0, cur_discount=20000.0,
        prev_revenue=None, prev_discount=None,
        yoy_revenue=None, yoy_discount=None,
        is_comparable=False,
    )
    # 20000 / (500000 + 20000) = 0.03846
    assert r.cur_discount_pct == pytest.approx(0.03846, abs=1e-4)
    assert r.prev_discount_pct is None
    assert r.yoy_discount_pct is None
    assert r.mom_delta is None
    assert r.yoy_delta is None


def test_to_row_layout():
    r = Table3Row(
        region="加拿大", store="加拿大一店",
        cur_revenue=1075411.75, cur_discount=48724.21,
        prev_revenue=1006336.12, prev_discount=68148.38,
        yoy_revenue=1106125.93, yoy_discount=74283.62,
        is_comparable=True,
    )
    row = to_row(r)
    assert len(row) == 15
    assert row[0] is None
    assert row[1] == "加拿大"
    assert row[2] == "加拿大一店"
    assert row[4] == 1075411.75       # col 5 (cur 收入)
    assert row[5] == 48724.21          # col 6 (cur 优惠总金额)
    assert row[14] == "是"             # col 15 (是否同比店)


def test_to_row_marks_non_comparable_as_no():
    r = Table3Row(
        region="加拿大", store="加拿大八店",
        cur_revenue=500000.0, cur_discount=20000.0,
        prev_revenue=None, prev_discount=None,
        yoy_revenue=None, yoy_discount=None,
        is_comparable=False,
    )
    row = to_row(r)
    assert row[14] == "否"


def test_build_rows_skips_inactive_stores():
    cur = {
        "加拿大一店": {PNL_REVENUE_KEY: 100.0, PNL_DISCOUNT_KEY: 5.0},
        "加拿大二店": {PNL_REVENUE_KEY: 0, PNL_DISCOUNT_KEY: 0},
        "加拿大九店": {PNL_REVENUE_KEY: 0, PNL_DISCOUNT_KEY: 0},
    }
    rows = build_rows(cur_pnl=cur)
    assert {r.store for r in rows} == {"加拿大一店"}


def test_build_rows_marks_classification_from_static_meta():
    """加拿大八店 is classified as '2025年新开店' (not 可比店) → is_comparable=False."""
    cur = {
        "加拿大一店": {PNL_REVENUE_KEY: 100.0, PNL_DISCOUNT_KEY: 5.0},
        "加拿大八店": {PNL_REVENUE_KEY: 100.0, PNL_DISCOUNT_KEY: 5.0},
    }
    rows = build_rows(cur_pnl=cur)
    by_store = {r.store: r for r in rows}
    assert by_store["加拿大一店"].is_comparable is True
    assert by_store["加拿大八店"].is_comparable is False


def test_write_sheet_smoke(tmp_path):
    cur = {"加拿大一店": {PNL_REVENUE_KEY: 1000.0, PNL_DISCOUNT_KEY: 50.0}}
    rows = build_rows(cur_pnl=cur)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表3"
    n = write_sheet(ws, rows)
    assert n == 1
    out = tmp_path / "table3.xlsx"
    wb.save(out)
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["表3"]
    assert ws2.cell(1, 2).value == "区域"
    assert ws2.cell(2, 3).value == "加拿大一店"


@pytest.mark.skipif(
    not MANUAL_WB.exists(), reason="manual workbook not on this machine"
)
def test_manual_table3_column_count():
    wb = openpyxl.load_workbook(MANUAL_WB, data_only=True, read_only=True)
    ws = wb["表3-打折优惠表"]
    assert ws.max_column == 15 == len(HEADERS_LINE_1)
    wb.close()
