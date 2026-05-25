"""Tests for 细分毛利率表 (per-category gross margin) builder."""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from zfi0049_report.dish_category import (
    OTHER,
    REPORT_CATEGORIES,
    map_pos_to_report_category,
)
from zfi0049_report.subdivided_gp_sheet import (
    build_subdivided_gp_sheet,
    compute_category_gp,
)
from zfi0049_report.table1_dish import Table1Row


# ── Category mapping ──


def test_report_categories_layout():
    assert REPORT_CATEGORIES == [
        "锅底类", "荤菜类", "素菜类", "酒水类",
        "小料台类-酱料水果", "小吃类", "其他类（如有）",
    ]


@pytest.mark.parametrize("dalei,expected", [
    ("锅底类", "锅底类"),
    ("荤菜", "荤菜类"),
    ("素菜", "素菜类"),
    ("酒水饮料", "酒水类"),
    ("小吃甜品", "小吃类"),
    ("火锅周边", "小料台类-酱料水果"),
    ("套餐", OTHER),
    ("赠送及其他类", OTHER),
    ("经典火锅菜", OTHER),
    (None, OTHER),
    ("", OTHER),
])
def test_map_pos_to_report_category(dalei, expected):
    assert map_pos_to_report_category(dalei) == expected


# ── GP aggregation ──


def _mk(store, dish, spec, mat, *, sales, portion, price, mat_price,
        category, loss=1.0):
    return Table1Row(
        region="加拿大", store=store, dish_code=dish, dish_short_code=None,
        dish_name="X", portion=portion, dish_unit=None, spec=spec,
        sales_qty=sales, loss_factor=loss,
        cur_dish_price=price,
        material_code=mat, material_unit_price=mat_price,
        category=category,
    )


def test_compute_category_gp_single_dish_single_material():
    """Revenue 100×10 = 1000, cost 100×0.5×1.0×2 = 100, gm = 0.9."""
    rows = [_mk("加拿大一店", 1, "整份", 100,
                sales=100, portion=0.5, price=10.0, mat_price=2.0,
                category="荤菜类")]
    gp = compute_category_gp(rows)
    assert gp[("加拿大一店", "荤菜类")] == pytest.approx((1000.0, 100.0, 0.9))


def test_compute_category_gp_revenue_deduped_across_materials():
    """One dish_spec, two materials → revenue counted once, costs summed."""
    rows = [
        _mk("加拿大一店", 1, "整份", 100,
            sales=100, portion=0.5, price=10.0, mat_price=2.0,
            category="荤菜类"),
        _mk("加拿大一店", 1, "整份", 200,
            sales=100, portion=0.2, price=10.0, mat_price=3.0,
            category="荤菜类"),
    ]
    gp = compute_category_gp(rows)
    rev, cost, gm = gp[("加拿大一店", "荤菜类")]
    # rev = 100 × 10 (counted once across the two material rows)
    assert rev == pytest.approx(1000.0)
    # cost = 100×0.5×2 + 100×0.2×3 = 100 + 60 = 160
    assert cost == pytest.approx(160.0)
    assert gm == pytest.approx(0.84)


def test_compute_category_gp_groups_per_store_per_category():
    """Two stores × two categories → 4 separate aggregates."""
    rows = [
        _mk("加拿大一店", 1, "整份", 100,
            sales=10, portion=1.0, price=5.0, mat_price=1.0,
            category="荤菜类"),
        _mk("加拿大一店", 2, "整份", 100,
            sales=20, portion=1.0, price=5.0, mat_price=1.0,
            category="素菜类"),
        _mk("加拿大二店", 1, "整份", 100,
            sales=30, portion=1.0, price=5.0, mat_price=1.0,
            category="荤菜类"),
    ]
    gp = compute_category_gp(rows)
    assert ("加拿大一店", "荤菜类") in gp
    assert ("加拿大一店", "素菜类") in gp
    assert ("加拿大二店", "荤菜类") in gp


def test_compute_category_gp_uses_other_when_no_category():
    rows = [_mk("加拿大一店", 1, "整份", 100,
                sales=10, portion=1.0, price=5.0, mat_price=1.0,
                category=None)]
    gp = compute_category_gp(rows)
    assert (("加拿大一店", OTHER)) in gp


def test_compute_category_gp_gm_none_when_no_revenue():
    """Cost-only rows produce gm=None instead of dividing by zero."""
    rows = [_mk("加拿大一店", 1, "整份", 100,
                sales=10, portion=1.0, price=None, mat_price=1.0,
                category="荤菜类")]
    gp = compute_category_gp(rows)
    rev, cost, gm = gp[("加拿大一店", "荤菜类")]
    assert rev == 0
    assert cost == pytest.approx(10.0)
    assert gm is None


# ── Sheet writer ──


def test_build_subdivided_gp_sheet_layout():
    wb = Workbook()
    wb.remove(wb.active)
    cur = {
        ("加拿大一店", "锅底类"): (1000.0, 350.0, 0.65),
        ("加拿大一店", "荤菜类"): (2000.0, 800.0, 0.60),
    }
    prev = {
        ("加拿大一店", "锅底类"): (900.0, 290.0, 0.6778),
    }
    ws = build_subdivided_gp_sheet(
        wb, cur_gp=cur, prev_gp=prev,
        year=2026, month=3,
    )
    assert ws.title == "细分毛利率表 (2)"
    # Title
    assert ws.cell(1, 2).value == "2026年3月细分毛利率环比表"
    # Row 2 category header positions: 锅底类 at col 3, 荤菜类 at col 6, ...
    assert ws.cell(2, 3).value == "锅底类"
    assert ws.cell(2, 6).value == "荤菜类"
    assert ws.cell(2, 21).value == "其他类（如有）"
    # Row 4: 加拿大一店 — cur GM for 锅底类 at col 3, prev at col 4, MoM at col 5
    assert ws.cell(4, 2).value == "加拿大一店"
    assert ws.cell(4, 3).value == pytest.approx(0.65)
    assert ws.cell(4, 4).value == pytest.approx(0.6778)
    assert ws.cell(4, 5).value == pytest.approx(0.65 - 0.6778)
    # 荤菜类 cur GM at col 6
    assert ws.cell(4, 6).value == pytest.approx(0.60)
    # prev/MoM blank for 荤菜类 (not in prev dict)
    assert ws.cell(4, 7).value is None
    assert ws.cell(4, 8).value is None


def test_build_subdivided_gp_sheet_sorts_stores():
    wb = Workbook()
    wb.remove(wb.active)
    cur = {
        ("加拿大三店", "锅底类"): (1000.0, 500.0, 0.5),
        ("加拿大一店", "锅底类"): (1000.0, 500.0, 0.5),
    }
    ws = build_subdivided_gp_sheet(
        wb, cur_gp=cur, year=2026, month=3,
    )
    # Row 4 = first store sorted, row 5 = second
    assert ws.cell(4, 2).value == "加拿大一店"
    assert ws.cell(5, 2).value == "加拿大三店"
