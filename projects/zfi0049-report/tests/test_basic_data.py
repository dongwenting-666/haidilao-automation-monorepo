"""Tests for the 基础数据 sheet builder.

Two flavours:
  - Structural: header layout, column count, computed-formula correctness
    against hand-picked values. Always runs.
  - Reference-diff: loads the actual manual workbook (附件3-...xlsx) when
    present and verifies static metadata + structure match. Skipped if the
    file is absent (so CI stays green on machines without it).
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
import pytest

from zfi0049_report.basic_data import (
    DERIVED_HEADERS,
    HEADERS,
    META_HEADERS,
    OPS_HEADERS,
    PNL_ITEMS,
    StoreMonthRecord,
    build_row,
    compute_derived,
    write_basic_data_sheet,
)
from zfi0049_report.static_meta import STORE_META


REFERENCE_WORKBOOK = Path(
    "/Users/hongming-claw/Downloads/附件3-毛利相关分析指标-2603.xlsx"
)


def test_total_column_count_is_126():
    assert len(HEADERS) == 126


def test_pnl_row_count_matches():
    # Cols 7–76 must be the 70 P&L line items from canada_pnl.ROWS.
    assert len(PNL_ITEMS) == 70
    assert HEADERS[6:76] == PNL_ITEMS


def test_meta_block_position():
    # Cols 86–92 are static metadata.
    assert HEADERS[85:92] == META_HEADERS
    assert HEADERS[85] == "地区"
    assert HEADERS[91] == "门店分类"


def test_static_meta_covers_all_canadian_stores():
    for n in range(1, 10):
        store = f"加拿大{['一','二','三','四','五','六','七','八','九'][n-1]}店"
        assert store in STORE_META, f"missing meta for {store}"


def test_derived_hotpot_profit_formula_matches_canada_one_march_2026():
    # Numbers pulled from manual workbook 加拿大一店 March 2026.
    rec = StoreMonthRecord(
        store="加拿大一店", year=2026, month=3, period_serial=46082,
        pnl={
            "一、主营业务收入": 783542.0,
            '九、利润总额(亏损以"-"号表示)': 125967.92,
            "十、所得税费用": 0.0,
            '十一、净利润(亏损以"-"号表示)': 125967.92,
        },
        audit_adjustment=-1509.5538,
        functional_fees=33120.5627,
    )
    d = compute_derived(rec)
    # 火锅经营净利润 = 利润总额 - 审计调整 - 职能费用
    assert d["火锅经营净利润"] == pytest.approx(94356.91, abs=0.05)
    assert d["火锅经营净利润（所得税前）"] == pytest.approx(94356.91, abs=0.05)
    assert d["火锅经营净利润率（所得税前）"] == pytest.approx(0.1204, abs=0.0005)


def test_derived_cash_flow_adds_back_non_cash_pnl_items():
    """经营性现金流 = 火锅经营净利润 + 资产折旧费 + 装修费摊销 + 资产减值损失.

    Verified against 加拿大三店 March 2026: 火锅经营净利润 = -4044.87,
    经营性现金流 = 16428.91 → implied depreciation+amortization ≈ 20,473.
    Here we test the formula mechanically with synthetic numbers.
    """
    rec = StoreMonthRecord(
        store="加拿大三店", year=2026, month=3, period_serial=46082,
        pnl={
            "一、主营业务收入": 617235.0,
            '九、利润总额(亏损以"-"号表示)': 27548.73,
            "十、所得税费用": 0.0,
            '十一、净利润(亏损以"-"号表示)': 27548.73,
            "3.3、资产折旧费": 18000.0,
            "3.5、装修费摊销": 2473.0,
            "3.7、资产减值损失": 0.0,
        },
        audit_adjustment=-1525.4675,
        functional_fees=33119.0658,
    )
    d = compute_derived(rec)
    # 火锅经营净利润 = 27548.73 - (-1525.47) - 33119.07 = -4044.87
    assert d["火锅经营净利润"] == pytest.approx(-4044.87, abs=0.05)
    # cash flow = -4044.87 + 18000 + 2473 + 0 = 16428.13 ≈ manual 16428.91
    assert d["经营性现金流"] == pytest.approx(16428.13, abs=1.0)
    assert d["经营性现金流（所得税前）"] == d["经营性现金流"]


def test_derived_cash_flow_zero_when_no_depreciation_data():
    """When the P&L lacks depreciation lines, cash flow falls back to
    just 火锅经营净利润 (no add-backs)."""
    rec = StoreMonthRecord(
        store="加拿大一店", year=2026, month=3, period_serial=46082,
        pnl={'九、利润总额(亏损以"-"号表示)': 100.0},
        audit_adjustment=0.0,
        functional_fees=0.0,
    )
    d = compute_derived(rec)
    assert d["经营性现金流"] == pytest.approx(100.0)


def test_build_row_shape():
    rec = StoreMonthRecord(
        store="加拿大一店", year=2026, month=3, period_serial=46082,
    )
    row = build_row(rec)
    assert len(row) == 126
    # Identifiers
    assert row[0] == "2026年"
    assert row[1] == "3月"
    assert row[2] == 46082
    assert row[5] == "加拿大一店"
    # Static meta
    assert row[85] == "北美"  # 地区
    assert row[87] == "温哥华"  # 城市
    assert row[90] == "营业"   # 营业状态
    # Col 125 == 门店名称 dup of col 6
    assert row[124] == "加拿大一店"


def test_write_basic_data_sheet_smoke(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础数据"
    records = [
        StoreMonthRecord(store=s, year=2026, month=3, period_serial=46082)
        for s in STORE_META
    ]
    n = write_basic_data_sheet(ws, records)
    assert n == len(STORE_META)
    out = tmp_path / "basic_data.xlsx"
    wb.save(out)
    assert out.exists()

    # Re-read and sanity-check
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["基础数据"]
    assert ws2.cell(1, 1).value == "年份"
    assert ws2.cell(1, 126).value == "利润档位"
    assert ws2.max_row == 1 + len(STORE_META)


# ── Reference-diff tests (require manual workbook present) ──────────────────


@pytest.mark.skipif(
    not REFERENCE_WORKBOOK.exists(),
    reason="manual workbook not on this machine",
)
def test_header_layout_matches_manual():
    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True, read_only=True)
    ws = wb["基础数据"]
    manual_headers = [c for c in next(ws.iter_rows(values_only=True))]
    wb.close()
    assert len(manual_headers) == 126
    # Headers should match position-by-position.
    mismatches = [
        (i + 1, ours, theirs)
        for i, (ours, theirs) in enumerate(zip(HEADERS, manual_headers))
        if ours != theirs
    ]
    assert not mismatches, f"header mismatches: {mismatches[:10]}"


@pytest.mark.skipif(
    not REFERENCE_WORKBOOK.exists(),
    reason="manual workbook not on this machine",
)
def test_static_metadata_matches_manual_for_march_2026():
    wb = openpyxl.load_workbook(REFERENCE_WORKBOOK, data_only=True, read_only=True)
    ws = wb["基础数据"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    march = [r for r in rows[1:] if r[0] == "2026年" and r[1] == "3月"]
    by_store = {r[5]: r for r in march}

    mismatches = []
    for store, meta in STORE_META.items():
        if store not in by_store:
            continue  # store didn't exist in March 2026 manual
        m = by_store[store]
        # Cols 86–92 (idx 85–91)
        expected_open = (meta.open_date.strftime("%Y-%m-%d 00:00:00")
                         if meta.open_date else "工程")
        actual_open = m[88]
        actual_open_str = (actual_open.strftime("%Y-%m-%d 00:00:00")
                           if hasattr(actual_open, "strftime")
                           else str(actual_open))
        checks = [
            ("地区", meta.region, m[85]),
            ("国家", meta.country, m[86]),
            ("城市", meta.city, m[87]),
            ("开业日期", expected_open, actual_open_str),
            ("门店级别", meta.level, m[89]),
            ("营业状态", meta.status, m[90]),
            ("门店分类", meta.classification, m[91]),
        ]
        for label, ours, theirs in checks:
            if ours != theirs:
                mismatches.append((store, label, ours, theirs))

    assert not mismatches, f"static-meta mismatches: {mismatches}"
