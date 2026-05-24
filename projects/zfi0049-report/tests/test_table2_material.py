"""Tests for 表2 (原材料成本变动) builder."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from zfi0049_report.table1_dish import MaterialUsage
from zfi0049_report.table2_material import (
    HEADERS,
    Table2Row,
    build_rows,
    build_rows_from_paths,
    to_row,
    write_sheet,
)

MANUAL_WB = Path(
    "/Users/hongming-claw/Downloads/附件3-毛利相关分析指标-2603.xlsx"
)
ZFI0156_MARCH = Path(
    "/Users/hongming-claw/haidilao-automation-monorepo/output/zfi0156/zfi0156-202603.xlsx"
)
MB5B_MARCH = Path(
    "/Users/hongming-claw/haidilao-automation-monorepo/output/mb5b/mb5b202603.xls"
)


def test_header_count_29():
    assert len(HEADERS) == 29
    # Used 14 cols + 15 trailing blanks
    assert HEADERS[0] == "区域"
    assert HEADERS[13] == "同比影响成本"


def test_row_code_layout():
    r = Table2Row(
        region="加拿大", store="加拿大一店", werks="CA01",
        matnr=1000049, matxt="金标生抽", unit="升",
        cur_price=2.53, prev_price=2.41, yoy_price=2.35,
        period_usage=191.1,
    )
    # Matches manual: "加拿大一店1000049升"
    assert r.code == "加拿大一店1000049升"


def test_row_price_deltas_and_cost_impacts_match_manual():
    """Manual 加拿大一店 / 1000049: cur=2.53, prev=2.41, yoy=2.35, usage=191.1
    → 环比变动 0.12, 同比变动 0.18, 环比影响成本 22.93, 同比影响成本 34.40."""
    r = Table2Row(
        region="加拿大", store="加拿大一店", werks="CA01",
        matnr=1000049, matxt="金标生抽", unit="升",
        cur_price=2.53, prev_price=2.41, yoy_price=2.35,
        period_usage=191.1,
    )
    assert r.mom_price_delta == pytest.approx(0.12)
    assert r.yoy_price_delta == pytest.approx(0.18)
    assert r.mom_cost_impact == pytest.approx(22.93, abs=0.01)
    assert r.yoy_cost_impact == pytest.approx(34.40, abs=0.01)


def test_row_deltas_none_when_data_missing():
    r = Table2Row(
        region="加拿大", store="加拿大一店", werks="CA01",
        matnr=1, matxt=None, unit=None,
        cur_price=2.53, prev_price=None, yoy_price=None,
        period_usage=100.0,
    )
    assert r.mom_price_delta is None
    assert r.yoy_price_delta is None
    assert r.mom_cost_impact is None
    assert r.yoy_cost_impact is None


def test_to_row_layout():
    r = Table2Row(
        region="加拿大", store="加拿大一店", werks="CA01",
        matnr=1000049, matxt="金标生抽", unit="升",
        cur_price=2.53, prev_price=2.41, yoy_price=2.35,
        period_usage=191.1,
    )
    row = to_row(r)
    assert len(row) == 29
    assert row[0] == "加拿大"
    assert row[1] == "加拿大一店1000049升"
    assert row[2] == "加拿大一店"
    assert row[3] == 1000049
    assert row[6] == 2.53
    assert row[11] == 191.1
    assert row[12] == pytest.approx(22.93, abs=0.01)
    assert row[13] == pytest.approx(34.40, abs=0.01)
    # Trailing cols blank
    assert row[14] is None
    assert row[28] is None


def test_build_rows_skips_zero_usage_materials():
    zfi = {
        ("CA01", 1): MaterialUsage(werks="CA01", matnr=1, matxt="X",
                                    unit="kg", unit_price=1.0,
                                    quantity=10.0, amount=10.0),
    }
    # MB5B has another material with no usage — should be excluded
    mb5b = {("CA01", 1): 1.0, ("CA01", 999): 5.0}
    rows = build_rows(zfi_cur=zfi, mb5b_cur=mb5b)
    assert len(rows) == 1
    assert rows[0].matnr == 1


def test_build_rows_orders_by_werks_then_matnr():
    zfi = {
        ("CA02", 100): MaterialUsage(werks="CA02", matnr=100, matxt="B",
                                      unit="kg", unit_price=1.0,
                                      quantity=1.0, amount=1.0),
        ("CA01", 200): MaterialUsage(werks="CA01", matnr=200, matxt="A",
                                      unit="kg", unit_price=1.0,
                                      quantity=1.0, amount=1.0),
        ("CA01", 100): MaterialUsage(werks="CA01", matnr=100, matxt="C",
                                      unit="kg", unit_price=1.0,
                                      quantity=1.0, amount=1.0),
    }
    mb5b = {k: 1.0 for k in zfi}
    rows = build_rows(zfi_cur=zfi, mb5b_cur=mb5b)
    assert [(r.werks, r.matnr) for r in rows] == [
        ("CA01", 100), ("CA01", 200), ("CA02", 100),
    ]


def test_write_sheet_smoke(tmp_path):
    rows = [
        Table2Row(region="加拿大", store="加拿大一店", werks="CA01",
                  matnr=1000049, matxt="金标生抽", unit="升",
                  cur_price=2.53, prev_price=2.41, yoy_price=2.35,
                  period_usage=191.1),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表2"
    n = write_sheet(ws, rows)
    assert n == 1
    out = tmp_path / "table2.xlsx"
    wb.save(out)
    # Re-read
    wb2 = openpyxl.load_workbook(out)
    ws2 = wb2["表2"]
    assert ws2.cell(1, 1).value == "区域"
    assert ws2.cell(2, 2).value == "加拿大一店1000049升"


# ── Integration tests ──


@pytest.mark.skipif(
    not (ZFI0156_MARCH.exists() and MB5B_MARCH.exists()),
    reason="ZFI0156 + MB5B March 2026 not on this machine",
)
def test_build_rows_from_paths_real_data():
    rows = build_rows_from_paths(
        zfi_cur_path=ZFI0156_MARCH,
        mb5b_cur_path=MB5B_MARCH,
    )
    assert len(rows) > 100  # many materials × 8 stores

    # Spot check 加拿大一店 / 1000049 (金标生抽)
    target = [r for r in rows if r.werks == "CA01" and r.matnr == 1000049]
    assert target, "expected CA01/1000049 in build_rows result"
    r = target[0]
    assert r.cur_price == pytest.approx(2.53, abs=0.01)
    assert r.period_usage > 0
    # prev/yoy unset (we didn't pass those paths) — deltas should be None
    assert r.mom_price_delta is None
    assert r.yoy_price_delta is None


@pytest.mark.skipif(
    not MANUAL_WB.exists(), reason="manual workbook not on this machine"
)
def test_manual_table2_layout_compatibility():
    """Manual 表2 has 29 cols; our HEADERS list matches that count."""
    wb = openpyxl.load_workbook(MANUAL_WB, data_only=True, read_only=True)
    ws = wb["表2-原材料成本变动表"]
    manual_cols = ws.max_column
    wb.close()
    assert manual_cols == 29 == len(HEADERS)
