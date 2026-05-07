from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from zfi0049_report.canada_pnl import (
    EXCLUDED_COST_CENTERS,
    STORE_ORDER,
    load_account_map,
    map_store,
)

TARGET_ITEMS = (
    "1、销售净收入",
    "2、仓储服务收入",
    "1、产品销售成本",
    "2、仓储服务成本",
)

ITEM_LABELS = {
    "1、销售净收入": "销售净收入",
    "2、仓储服务收入": "仓储服务收入",
    "1、产品销售成本": "产品销售成本",
    "2、仓储服务成本": "仓储服务成本",
}


def _normalize_code(value: object) -> str:
    return str(value).zfill(10)[-8:]


def collect_gross_margin_rows(source_path: Path, mapping_path: Path, store_name: str = ""):
    account_map = load_account_map(mapping_path)
    wb = load_workbook(source_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    stores = [store_name] if store_name else [s for s in STORE_ORDER if s != "加拿大九店"]
    summary = {store: defaultdict(float) for store in stores}
    detail_rows: list[dict[str, object]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        company_code, cost_element, amount = row[0], row[1], row[2]
        profit_center, profit_desc = row[3], row[4]
        cost_center, cost_desc = row[5], row[6]
        cost_element_desc, currency = row[7], row[8]

        if cost_element is None or amount is None:
            continue

        code8 = _normalize_code(cost_element)
        mapped = account_map.get(code8)
        if not mapped:
            continue

        report_item, data_source, sign = mapped
        if report_item not in TARGET_ITEMS:
            continue

        cost_desc_text = str(cost_desc or "").strip()
        if cost_desc_text in EXCLUDED_COST_CENTERS:
            continue

        store = map_store(profit_desc) if data_source == "利润中心" else map_store(cost_desc, profit_desc)
        if store not in summary:
            continue

        signed_amount = float(amount) * float(sign)
        summary[store][report_item] += signed_amount
        detail_rows.append(
            {
                "门店": store,
                "毛利项目": ITEM_LABELS[report_item],
                "报表项目": report_item,
                "公司代码": str(company_code or "").strip(),
                "成本要素": code8,
                "成本要素描述": str(cost_element_desc or "").strip(),
                "金额": signed_amount,
                "原始金额": float(amount),
                "取数维度": data_source,
                "利润中心": str(profit_center or "").strip(),
                "利润中心描述": str(profit_desc or "").strip(),
                "成本中心": str(cost_center or "").strip(),
                "成本中心描述": cost_desc_text,
                "货币码": str(currency or "").strip(),
            }
        )

    wb.close()
    return stores, summary, detail_rows


def _write_summary_sheet(ws, stores: list[str], summary: dict[str, defaultdict[str, float]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)

    headers = ["项目"] + stores
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ("销售净收入", "1、销售净收入", "amount"),
        ("仓储服务收入", "2、仓储服务收入", "amount"),
        ("主营业务收入", None, "amount"),
        ("产品销售成本", "1、产品销售成本", "amount"),
        ("仓储服务成本", "2、仓储服务成本", "amount"),
        ("主营业务成本", None, "amount"),
        ("产品销售毛利额", None, "amount"),
        ("产品销售毛利率", None, "ratio"),
        ("仓储服务毛利额", None, "amount"),
        ("仓储服务毛利率", None, "ratio"),
        ("综合毛利额", None, "amount"),
        ("综合毛利率", None, "ratio"),
    ]

    for row_idx, (label, key, kind) in enumerate(rows, start=2):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 1).font = bold if label in {"主营业务收入", "主营业务成本", "综合毛利额", "综合毛利率"} else Font(bold=False)
        if label in {"主营业务收入", "主营业务成本", "综合毛利额", "综合毛利率"}:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = section_fill
        for col_idx, store in enumerate(stores, start=2):
            sales = summary[store]["1、销售净收入"]
            ware_rev = summary[store]["2、仓储服务收入"]
            product_cost = summary[store]["1、产品销售成本"]
            ware_cost = summary[store]["2、仓储服务成本"]
            total_rev = sales + ware_rev
            total_cost = product_cost + ware_cost
            product_gp = sales - product_cost
            ware_gp = ware_rev - ware_cost
            total_gp = total_rev - total_cost

            values = {
                "销售净收入": sales,
                "仓储服务收入": ware_rev,
                "主营业务收入": total_rev,
                "产品销售成本": product_cost,
                "仓储服务成本": ware_cost,
                "主营业务成本": total_cost,
                "产品销售毛利额": product_gp,
                "产品销售毛利率": 0.0 if sales == 0 else product_gp / sales,
                "仓储服务毛利额": ware_gp,
                "仓储服务毛利率": 0.0 if ware_rev == 0 else ware_gp / ware_rev,
                "综合毛利额": total_gp,
                "综合毛利率": 0.0 if total_rev == 0 else total_gp / total_rev,
            }

            cell = ws.cell(row_idx, col_idx, values[label])
            cell.number_format = "0.00%" if kind == "ratio" else "#,##0.00;[Red]-#,##0.00"

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 18
    for idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16


def _write_detail_sheet(ws, detail_rows: list[dict[str, object]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    headers = [
        "门店",
        "毛利项目",
        "报表项目",
        "公司代码",
        "成本要素",
        "成本要素描述",
        "金额",
        "原始金额",
        "取数维度",
        "利润中心",
        "利润中心描述",
        "成本中心",
        "成本中心描述",
        "货币码",
    ]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in detail_rows:
        ws.append([row[h] for h in headers])

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 7).number_format = "#,##0.00;[Red]-#,##0.00"
        ws.cell(row_idx, 8).number_format = "#,##0.00;[Red]-#,##0.00"

    ws.freeze_panes = "A2"
    widths = {
        "A": 14,
        "B": 14,
        "C": 18,
        "D": 10,
        "E": 12,
        "F": 26,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 14,
        "K": 20,
        "L": 14,
        "M": 24,
        "N": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def generate_gross_margin_workbook(
    source_path: Path,
    mapping_path: Path,
    output_path: Path,
    store_name: str = "",
) -> Path:
    stores, summary, detail_rows = collect_gross_margin_rows(source_path, mapping_path, store_name=store_name)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    _write_summary_sheet(ws_summary, stores, summary)

    ws_detail = wb.create_sheet("底层明细")
    _write_detail_sheet(ws_detail, detail_rows)

    wb.save(output_path)
    return output_path
