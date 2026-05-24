"""Build the four 毛利率 derivative sheets:
  - 毛利率连续对比表 (7-month rolling trend + anomaly flags)
  - 毛利率环比 (MoM decomposition: dish-price / material-cost / discount impact)
  - 毛利率同比 (YoY equivalent)
  - 细分毛利率表 (per-category gross margins — TODO, needs dish→category map)

Formulas verified against the manual workbook 附件3-毛利相关分析指标-2603.xlsx
for 加拿大一店 March 2026:
  - 还原毛利率(dish)  = (cur_gp · revenue − Δdish) / (revenue − Δdish)
    where Δdish = 菜品涨价金额 (Σ over 表1 col 24 per store)
  - 还原毛利率(material) = cur_gp + Δcost / revenue
    where Δcost = 原材料价格变动金额 (Σ over 表2 col 13 per store)
  - 毛利率影响 = cur_gp − 还原毛利率
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from zfi0049_report.table1_dish import Table1Row
from zfi0049_report.table2_material import Table2Row
from zfi0049_report.table3_discount import Table3Row


# ── Formulas ────────────────────────────────────────────────────────────────


def restored_gp_dish(cur_gp: float, revenue: float, delta_dish: float) -> float:
    """Restored gross margin if dish prices hadn't changed.

    The "restoration" removes the price-change effect: revenue would have
    been (revenue − Δdish), cost is unchanged. So:
      restored = (revenue − Δdish − cost) / (revenue − Δdish)
    With cost = revenue · (1 − cur_gp):
      restored = (cur_gp · revenue − Δdish) / (revenue − Δdish)

    Falls back to ``cur_gp`` when revenue is 0 (no baseline to restore from).
    """
    if not revenue:
        return cur_gp
    denom = revenue - delta_dish
    if denom == 0:
        return cur_gp
    return (cur_gp * revenue - delta_dish) / denom


def restored_gp_material(cur_gp: float, revenue: float,
                         delta_material: float) -> float:
    """Restored gross margin if material prices hadn't changed.

    Cost would have been (cost − Δmaterial), revenue is unchanged. So:
      restored = (revenue − (cost − Δmaterial)) / revenue
             = cur_gp + Δmaterial / revenue
    """
    if revenue == 0:
        return cur_gp
    return cur_gp + delta_material / revenue


# ── 毛利率环比 ───────────────────────────────────────────────────────────────


# 23-col header (rows 3-4 of manual, flattened).
MOM_HEADERS_ROW_4 = [
    None,                      # 1
    "门店名称",                # 2
    "本月毛利率",              # 3 (manual r4 col 3 = month-end serial; we use literal label)
    "上月毛利率",              # 4
    "环比",                    # 5
    "菜品涨价金额",            # 6
    "还原毛利率",              # 7
    "毛利率影响",              # 8
    "原材料价格变动金额",      # 9
    "还原毛利率",              # 10
    "毛利率影响",              # 11
    "本月优惠占比",            # 12
    "上月优惠占比",            # 13
    "环比",                    # 14
    "本月还原毛利率",          # 15
    "上月还原毛利率",          # 16
    "毛利率影响",              # 17
    "菜品损耗环比变动金额",    # 18
    "还原毛利率",              # 19
    "毛利率影响",              # 20
    "本月收入",                # 21
    "上月收入",                # 22
    "收入环比",                # 23
]
assert len(MOM_HEADERS_ROW_4) == 23


@dataclass
class MomRow:
    """One row in 毛利率环比 (per store)."""

    store: str
    cur_gp: float            # 本月毛利率 (from basic_data 三、毛利率)
    prev_gp: float           # 上月毛利率
    cur_revenue: float       # 本月收入 (优惠后 from 表3 cur_revenue)
    prev_revenue: float      # 上月收入
    delta_dish: float        # Σ 表1 col 24
    delta_material: float    # Σ 表2 col 13
    cur_discount_pct: float  # 表3 col 4
    prev_discount_pct: float  # 表3 col 7
    delta_loss: float = 0.0  # 表1 col 34 sum (TODO when loss-impact cols built)


def _impact_columns(cur_gp: float, restored: float) -> tuple[float, float]:
    """Return (restored_gp, impact_on_gp = cur − restored)."""
    return restored, cur_gp - restored


def mom_row_to_excel(r: MomRow) -> list[Any]:
    """Project MomRow → 23-col row matching MOM_HEADERS_ROW_4 layout."""
    mom_gp = r.cur_gp - r.prev_gp

    # Dish impact (cols 7-8)
    restored_dish = restored_gp_dish(r.cur_gp, r.cur_revenue, r.delta_dish)
    impact_dish = r.cur_gp - restored_dish

    # Material impact (cols 10-11)
    restored_mat = restored_gp_material(r.cur_gp, r.cur_revenue, r.delta_material)
    impact_mat = r.cur_gp - restored_mat

    # Discount impact (cols 14-17)
    discount_mom = r.cur_discount_pct - r.prev_discount_pct
    # Per-discount "还原毛利率": gp + cur_discount_pct (rough but matches manual
    # for 加拿大一店: 0.6974 + 0.0433 = 0.7407 vs manual 0.7105 — formula needs
    # refinement). For now provide the column with a documented placeholder.
    # TODO(iter-8): refine 还原毛利率 (discount) formula against manual numbers.
    restored_disc_cur = r.cur_gp + r.cur_discount_pct
    restored_disc_prev = r.prev_gp + r.prev_discount_pct
    impact_disc = r.cur_gp - restored_disc_cur

    # Loss impact (cols 19-20)
    restored_loss = restored_gp_material(r.cur_gp, r.cur_revenue, r.delta_loss)
    impact_loss = r.cur_gp - restored_loss

    # Revenue MoM (col 23)
    rev_mom = ((r.cur_revenue - r.prev_revenue) / r.prev_revenue
               if r.prev_revenue else None)

    return [
        None,                  # 1
        r.store,               # 2
        r.cur_gp,              # 3
        r.prev_gp,             # 4
        mom_gp,                # 5
        r.delta_dish,          # 6
        restored_dish,         # 7
        impact_dish,           # 8
        r.delta_material,      # 9
        restored_mat,          # 10
        impact_mat,            # 11
        r.cur_discount_pct,    # 12
        r.prev_discount_pct,   # 13
        discount_mom,          # 14
        restored_disc_cur,     # 15
        restored_disc_prev,    # 16
        impact_disc,           # 17
        r.delta_loss,          # 18
        restored_loss,         # 19
        impact_loss,           # 20
        r.cur_revenue,         # 21
        r.prev_revenue,        # 22
        rev_mom,               # 23
    ]


def build_mom_rows(
    *,
    cur_pnl: dict[str, dict[str, float]],
    prev_pnl: dict[str, dict[str, float]],
    table1_rows: list[Table1Row],
    table2_rows: list[Table2Row],
    table3_rows: list[Table3Row],
) -> list[MomRow]:
    """Aggregate 表1/表2/表3 per-row data into per-store MoM rows.

    cur_pnl[store]["三、毛利率"] is the gross margin from 基础数据.
    """
    # Sum per-store impacts from 表1 (col 24) and 表2 (col 13)
    dish_impact_sum: dict[str, float] = defaultdict(float)
    for t1 in table1_rows:
        if t1.mom_revenue_impact is not None:
            dish_impact_sum[t1.store] += t1.mom_revenue_impact

    mat_impact_sum: dict[str, float] = defaultdict(float)
    for t2 in table2_rows:
        v = t2.mom_cost_impact
        if v is not None:
            mat_impact_sum[t2.store] += v

    # Index 表3 by store
    t3_by_store = {t.store: t for t in table3_rows}

    out: list[MomRow] = []
    for store, cur in cur_pnl.items():
        prev = prev_pnl.get(store, {})
        cur_gp = float(cur.get("三、毛利率", 0.0))
        prev_gp = float(prev.get("三、毛利率", 0.0))
        t3 = t3_by_store.get(store)
        cur_rev = t3.cur_revenue if t3 else 0.0
        prev_rev = t3.prev_revenue if t3 and t3.prev_revenue else 0.0
        cur_disc_pct = (t3.cur_discount_pct if t3 and t3.cur_discount_pct
                        is not None else 0.0)
        prev_disc_pct = (t3.prev_discount_pct if t3 and t3.prev_discount_pct
                         is not None else 0.0)
        out.append(MomRow(
            store=store, cur_gp=cur_gp, prev_gp=prev_gp,
            cur_revenue=cur_rev, prev_revenue=prev_rev,
            delta_dish=dish_impact_sum.get(store, 0.0),
            delta_material=mat_impact_sum.get(store, 0.0),
            cur_discount_pct=cur_disc_pct, prev_discount_pct=prev_disc_pct,
        ))
    return out


def build_mom_sheet(wb: Workbook, mom_rows: list[MomRow]) -> Worksheet:
    """Create the 毛利率环比 sheet on ``wb`` and populate."""
    ws = wb.create_sheet("毛利率环比")
    ws.append(MOM_HEADERS_ROW_4)
    for r in mom_rows:
        ws.append(mom_row_to_excel(r))
    return ws


# ── 毛利率连续对比表 (7-month rolling trend) ───────────────────────────────


# Month-end Excel serials in the manual (most-recent first).
DEFAULT_MONTH_SERIALS = [46082, 46054, 46023, 45992, 45962, 45931, 45901]

TREND_HEADERS_ROW_5 = [
    None,
    "序号", "区域", "分店名称",
    "M1", "M2", "M3", "M4", "M5", "M6", "M7",  # placeholders — caller fills serials
    "M1-M2", "M2-M3", "M3-M4", "M4-M5",
    "区域平均水平", "差异",
    "是否连续3个月下降", "是否连续2个月下降",
    "毛利率连续波2个月动异常",
    "毛利率环比下降变动超过2%",
    "毛利率是否低于60%",
]
assert len(TREND_HEADERS_ROW_5) == 22


def _flag_consecutive_decline(values: list[float], k: int) -> str:
    """'Y' iff the most-recent k MoM deltas are all negative."""
    deltas = [values[i] - values[i + 1] for i in range(len(values) - 1)]
    return "Y" if len(deltas) >= k and all(d < 0 for d in deltas[:k]) else "N"


def _flag_two_month_anomaly(deltas: list[float]) -> str:
    """'Y' iff two consecutive MoM deltas exceed 2% in absolute terms."""
    if len(deltas) < 2:
        return "N"
    return "Y" if abs(deltas[0]) > 0.02 and abs(deltas[1]) > 0.02 else "N"


def build_trend_rows(
    *,
    monthly_gp: dict[str, list[float]],
    month_serials: list[int] | None = None,
) -> list[list[Any]]:
    """Build per-store rows for 毛利率连续对比表.

    ``monthly_gp[store]`` is a list of 7 floats: most-recent-month-first.
    Returns a list of 22-element rows ready to ``ws.append``.
    """
    month_serials = month_serials or DEFAULT_MONTH_SERIALS
    stores = list(monthly_gp.keys())
    # Region average for current month (col 16)
    cur_values = [monthly_gp[s][0] for s in stores
                  if monthly_gp[s] and monthly_gp[s][0]]
    region_avg = sum(cur_values) / len(cur_values) if cur_values else 0.0

    rows: list[list[Any]] = []
    for idx, store in enumerate(stores, start=1):
        gps = monthly_gp[store]
        # Pad to 7 if fewer months
        while len(gps) < 7:
            gps.append(0.0)
        deltas = [gps[i] - gps[i + 1] for i in range(min(4, len(gps) - 1))]
        while len(deltas) < 4:
            deltas.append(0.0)
        diff = gps[0] - region_avg
        flag_3down = _flag_consecutive_decline(gps, 3)
        flag_2down = _flag_consecutive_decline(gps, 2)
        flag_2anom = _flag_two_month_anomaly(deltas)
        flag_drop2pct = "Y" if deltas[0] < -0.02 else "N"
        flag_below60 = "Y" if gps[0] < 0.60 else "N"
        rows.append([
            None,
            idx, "加拿大", store,
            gps[0], gps[1], gps[2], gps[3], gps[4], gps[5], gps[6],
            deltas[0], deltas[1], deltas[2], deltas[3],
            region_avg, diff,
            flag_3down, flag_2down, flag_2anom,
            flag_drop2pct, flag_below60,
        ])
    return rows


def build_trend_sheet(
    wb: Workbook,
    rows: list[list[Any]],
    month_serials: list[int] | None = None,
) -> Worksheet:
    """Create the 毛利率连续对比表 sheet on ``wb`` and populate."""
    month_serials = month_serials or DEFAULT_MONTH_SERIALS
    ws = wb.create_sheet("毛利率连续对比表")
    # Replace M1-M7 placeholders with actual serials.
    header = list(TREND_HEADERS_ROW_5)
    for i, s in enumerate(month_serials[:7]):
        header[4 + i] = s
    ws.append(header)
    for r in rows:
        ws.append(r)
    return ws


# ── 毛利率同比 (YoY) — mirrors MoM with yoy substitutions ───────────────────


def build_yoy_sheet(wb: Workbook, mom_rows_yoy: list[MomRow]) -> Worksheet:
    """Same layout as 毛利率环比 but the 'prev' fields hold YoY values.

    Caller is responsible for constructing the MomRow with YoY data
    (yoy_pnl in place of prev_pnl, delta_material derived from 表2
    yoy_cost_impact, etc.).
    """
    ws = wb.create_sheet("毛利率同比")
    headers = list(MOM_HEADERS_ROW_4)
    headers[3] = "去年同期毛利率"  # col 4 label
    headers[12] = "去年同期优惠占比"  # col 13 label
    headers[21] = "去年同期收入"     # col 22 label
    ws.append(headers)
    for r in mom_rows_yoy:
        ws.append(mom_row_to_excel(r))
    return ws
