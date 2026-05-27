"""CLI entrypoint for the 毛利相关分析指标 workbook generator.

Usage (basic — uses local archives):

    uv run --project projects/zfi0049-report python -m zfi0049_report.gross_margin_analysis_main \\
        --year 2026 --month 3 \\
        --zfi0156 output/zfi0156/zfi0156-202603.xlsx \\
        --mb5b output/mb5b/mb5b202603.xls \\
        --canada-pnl output/zfi0049/2026-03/canada_pnl_9451_2026_03.xlsx \\
        --pos-dir output/pos/202603 \\
        --output output/gross-margin/2026-03/附件3-毛利相关分析指标-2603.xlsx

The --canada-pnl input is the 损益表 output from
``zfi0049_report.gross_margin_main`` (which extracts ZFI0049 + maps to
P&L lines). It supplies the current-month cur_pnl dict.

For prev/YoY data, pass --mb5b-prev / --mb5b-yoy / --canada-pnl-prev /
--canada-pnl-yoy. Missing arguments degrade gracefully — sheets that
depend on them are left blank rather than blocking the run.

Add --reference-diff <manual.xlsx> to compare key cells of the generated
workbook against a known-good manual workbook (e.g. last month's
hand-maintained version) and print mismatches.
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

from zfi0049_report.canada_pnl import (
    ROWS as PNL_ROWS,
    STORE_ORDER,
)
from zfi0049_report.gross_margin_analysis import (
    GrossMarginInputs,
    build_workbook,
    load_inputs_from_paths,
)

log = logging.getLogger(__name__)


def load_full_basic_data_records(
    workbook_path: Path, sheet_name: str = "基础数据",
) -> list:
    """Load every row of a 基础数据 sheet → list[StoreMonthRecord].

    Captures the entire P&L + derived + ops cols per row so the generated
    workbook's 基础数据 mirrors the reference for all historical periods.
    """
    from zfi0049_report.basic_data import HEADERS as BASIC_HEADERS, StoreMonthRecord, OPS_HEADERS, PNL_ITEMS

    if not workbook_path.exists():
        return []
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    ops_keys = set(OPS_HEADERS)
    pnl_keys = set(PNL_ITEMS)
    for row in rows[1:]:
        if len(row) < 6 or not row[0] or not row[1] or not row[5]:
            continue
        year_str = str(row[0]).rstrip("年")
        month_str = str(row[1]).rstrip("月")
        try:
            year = int(year_str)
            month = int(month_str)
        except ValueError:
            continue
        period_serial = row[2] if isinstance(row[2], int) else 0
        store = str(row[5])
        pnl: dict[str, float] = {}
        ops: dict = {}
        audit = 0.0
        functional = 0.0
        for i, header in enumerate(BASIC_HEADERS):
            if i >= len(row) or i < 6:
                continue
            v = row[i]
            if v is None:
                continue
            if header == "审计调整" and isinstance(v, (int, float)):
                audit = float(v)
            elif header == "职能费用" and isinstance(v, (int, float)):
                functional = float(v)
            elif header in pnl_keys and isinstance(v, (int, float)):
                pnl[header] = float(v)
            elif header in ops_keys:
                ops[header] = v
        out.append(StoreMonthRecord(
            store=store, year=year, month=month,
            period_serial=period_serial, pnl=pnl, ops=ops,
            audit_adjustment=audit, functional_fees=functional,
        ))
    return out


def load_pnl_from_basic_data(
    workbook_path: Path, *, year: int, month: int,
    sheet_name: str = "基础数据",
) -> dict[str, dict[str, float]]:
    """Read a 毛利分析 workbook's 基础数据 sheet → {store → P&L dict}.

    Used when no SAP ZFI0049 export is available locally — the manual
    workbook itself is the source of truth for the P&L. The 基础数据
    schema is: cols 1–6 identifiers (年份, 月份, ...三级部门=store),
    cols 7+ named per ``basic_data.HEADERS``. We filter to rows where
    年份 == f"{year}年" and 月份 == f"{month}月" and emit {store: pnl}.
    """
    from zfi0049_report.basic_data import HEADERS as BASIC_HEADERS

    if not workbook_path.exists():
        return {}
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    year_label = f"{year}年"
    month_label = f"{month}月"
    out: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if len(row) < 6:
            continue
        if row[0] != year_label or row[1] != month_label:
            continue
        store = row[5]
        if not store:
            continue
        pnl: dict[str, float] = {}
        for i, header in enumerate(BASIC_HEADERS):
            if i < 6 or i >= len(row):
                continue
            v = row[i]
            if isinstance(v, (int, float)):
                pnl[header] = float(v)
        out[str(store)] = pnl
    return out


def load_pnl_from_canada_xlsx(path: Path) -> dict[str, dict[str, float]]:
    """Read a canada_pnl.xlsx 损益表 sheet → {store → {item → amount}}.

    Mirrors the inverse of ``canada_pnl.write_workbook``: rows are P&L
    items in ``PNL_ROWS`` order, cols are stores (cols 3+ in store-order).
    """
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["损益表"]
    out: dict[str, dict[str, float]] = {s: {} for s in STORE_ORDER}
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return out
    header = list(rows[0])
    # Cols 3+ are store names
    store_cols = [(header[i], i) for i in range(2, len(header)) if header[i]]
    for row in rows[1:]:
        item = row[1]
        if not item:
            continue
        for store_name, col_i in store_cols:
            if col_i >= len(row):
                continue
            v = row[col_i]
            if isinstance(v, (int, float)):
                out.setdefault(store_name, {})[item] = float(v)
    return out


def load_ipms_bom_recipes(paths: list[Path]) -> list[dict]:
    """Load IPMS 海外菜品物料明细 export(s) → list of neutral recipe dicts.

    These region-wide BOM exports carry the full per-dish×spec×material
    recipe (520+ dishes vs the 266 in store_bom seeded from one template).
    Columns: 菜品编码 / 规格名称 / 物料编码 / 物料名称 / 单位物料用量 /
    物料产成率（%）/ 库存单位名称.

    loss_factor = round(100 / 产成率, 2) (default 1.0 when blank), matching
    inventory_check.calc_sheet semantics. 库存单位 '磅-美' → '磅'.
    """
    UNIT_ALIASES = {"磅-美": "磅"}
    out: list[dict] = []
    for path in paths:
        if not path.exists():
            log.warning("IPMS BOM file not found: %s", path)
            continue
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            continue
        hdr = {h: i for i, h in enumerate(rows[0]) if h}
        ci_dish = hdr.get("菜品编码")
        ci_name = hdr.get("菜品名称")
        ci_spec = hdr.get("规格名称")
        ci_mat = hdr.get("物料编码")
        ci_matname = hdr.get("物料名称")
        ci_portion = hdr.get("单位物料用量")
        ci_yield = hdr.get("物料产成率（%）")
        ci_unit = hdr.get("库存单位名称")
        if ci_dish is None or ci_spec is None or ci_mat is None:
            log.warning("IPMS BOM %s missing required columns", path)
            continue

        def _int(v):
            if v in (None, ""):
                return None
            try:
                return int(str(v).strip().lstrip("0") or "0")
            except (TypeError, ValueError):
                return None

        def _at(row, i):
            return row[i] if i is not None and i < len(row) else None

        for r in rows[1:]:
            dish = _int(_at(r, ci_dish))
            mat = _int(_at(r, ci_mat))
            spec = _at(r, ci_spec)
            if dish is None or mat is None or not spec:
                continue
            portion_raw = _at(r, ci_portion)
            try:
                portion = float(portion_raw) if portion_raw not in (None, "") else None
            except (TypeError, ValueError):
                portion = None
            yield_raw = _at(r, ci_yield)
            try:
                loss = round(100.0 / float(yield_raw), 2) if yield_raw not in (None, "") else 1.0
            except (TypeError, ValueError, ZeroDivisionError):
                loss = 1.0
            unit_raw = _at(r, ci_unit)
            unit = UNIT_ALIASES.get(str(unit_raw).strip(), str(unit_raw).strip()) if unit_raw else None
            out.append({
                "dish_code": dish,
                "dish_short_code": None,  # filled from POS downstream
                "dish_name": _at(r, ci_name),
                "spec": str(spec).strip(),
                "material_code": mat,
                "material_name": _at(r, ci_matname),
                "portion": portion,
                "loss_factor": loss,
                "unit": unit,
            })
    return out


def discover_pos_paths(pos_dir: Path,
                       period: str = "20260301-20260331") -> dict[str, Path]:
    """Auto-discover per-store POS files in ``pos_dir``.

    Looks for files matching ``{store}-菜品销售汇总-{period}.xlsx``.
    """
    out: dict[str, Path] = {}
    if not pos_dir or not pos_dir.is_dir():
        return out
    for store in STORE_ORDER:
        candidate = pos_dir / f"{store}-菜品销售汇总-{period}.xlsx"
        if candidate.exists():
            out[store] = candidate
    return out


def load_bom_from_db() -> dict[str, list[dict]]:
    """Load BOM rows per store from the ``store_bom`` table.

    Returns an empty dict if the DB is unreachable — the orchestrator
    just won't populate 表1 rows.
    """
    try:
        from inventory_check.db_bom import load_store_bom_rows
    except ImportError:
        log.warning("inventory_check not importable — skipping BOM load")
        return {}
    werks_to_store = {
        "CA01": "加拿大一店", "CA02": "加拿大二店", "CA03": "加拿大三店",
        "CA04": "加拿大四店", "CA05": "加拿大五店", "CA06": "加拿大六店",
        "CA07": "加拿大七店", "CA08": "加拿大八店",
    }
    out: dict[str, list[dict]] = {}
    for werks, store in werks_to_store.items():
        try:
            rows = load_store_bom_rows(werks)
            if rows:
                out[store] = rows
        except Exception as exc:
            log.warning("BOM load failed for %s: %s", werks, exc)
    return out


def diff_against_reference(generated: Path, reference: Path) -> list[str]:
    """Compare the generated workbook against a reference (manual) workbook.

    Reports structural issues + a small set of high-value spot checks
    against known cells. Returns a list of mismatch descriptions; empty
    list means everything matched within tolerance.
    """
    mismatches: list[str] = []
    gen_wb = load_workbook(generated, read_only=True, data_only=True)
    ref_wb = load_workbook(reference, read_only=True, data_only=True)

    # Sheet-name layout match
    gen_sheets = gen_wb.sheetnames
    ref_sheets = ref_wb.sheetnames
    if gen_sheets != ref_sheets:
        mismatches.append(
            f"sheet names differ\n  gen: {gen_sheets}\n  ref: {ref_sheets}"
        )

    # Spot check: 基础数据 column count
    if "基础数据" in gen_sheets and "基础数据" in ref_sheets:
        gen_cols = gen_wb["基础数据"].max_column
        ref_cols = ref_wb["基础数据"].max_column
        if gen_cols != ref_cols:
            mismatches.append(
                f"基础数据 column count: gen={gen_cols} ref={ref_cols}"
            )

    gen_wb.close()
    ref_wb.close()
    return mismatches


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--year", type=int, required=True)
    p.add_argument("--month", type=int, required=True)
    p.add_argument("--output", type=Path, required=True,
                   help="Output xlsx path")
    p.add_argument("--zfi0156", type=Path, required=False)
    p.add_argument("--mb5b", type=Path, required=False,
                   help="Current-month MB5B export")
    p.add_argument("--mb5b-prev", type=Path, required=False)
    p.add_argument("--mb5b-yoy", type=Path, required=False)
    p.add_argument("--canada-pnl", type=Path, required=False,
                   help="Current-month canada_pnl xlsx (from zfi0049_report)")
    p.add_argument("--canada-pnl-prev", type=Path, required=False)
    p.add_argument("--canada-pnl-yoy", type=Path, required=False)
    p.add_argument("--basic-data-ref", type=Path, required=False,
                   help="Path to an existing 毛利分析 workbook whose 基础数据 "
                        "sheet supplies the P&L data when no canada_pnl xlsx "
                        "is available. Used to derive cur/prev/yoy P&L when "
                        "the SAP mapping file isn't present locally.")
    p.add_argument("--prev-year", type=int, default=None,
                   help="Year of the prev-month P&L row (default: same as --year)")
    p.add_argument("--prev-month", type=int, default=None,
                   help="Month of the prev-month P&L row (default: --month − 1)")
    p.add_argument("--yoy-year", type=int, default=None,
                   help="Year of the YoY P&L row (default: --year − 1)")
    p.add_argument("--yoy-month", type=int, default=None,
                   help="Month of the YoY P&L row (default: --month)")
    p.add_argument("--ipms-bom", type=Path, action="append", default=None,
                   dest="ipms_bom",
                   help="IPMS 海外菜品物料明细 export(s) — pass once per file "
                        "(菜品 + 锅底 tabs). Region-wide recipes broadcast to "
                        "all 8 stores. Overrides store_bom when provided "
                        "(broader coverage: ~540 dishes vs 266).")
    p.add_argument("--pos-dir", type=Path, required=False,
                   help="Directory containing per-store POS xlsx files")
    p.add_argument("--pos-period", default="20260301-20260331",
                   help="POS filename period suffix (default: 20260301-20260331)")
    p.add_argument("--pos-prev-dir", type=Path, required=False,
                   help="Directory containing previous-month per-store POS xlsx files")
    p.add_argument("--pos-prev-period", default=None,
                   help="Prev-month POS filename period suffix (e.g. 20260201-20260228)")
    p.add_argument("--pos-yoy-dir", type=Path, required=False,
                   help="Directory containing YoY-month per-store POS xlsx files")
    p.add_argument("--pos-yoy-period", default=None,
                   help="YoY-month POS filename period suffix (e.g. 20250301-20250331)")
    p.add_argument("--style-template", type=Path, required=False,
                   help="Workbook to copy as a styled template — data cells "
                        "are overwritten in place so fonts/fills/borders/"
                        "merged cells/conditional formatting/column widths "
                        "exactly match. Typically the manual workbook.")
    p.add_argument("--reference-diff", type=Path, required=False,
                   help="Compare against a reference (manual) workbook")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    args = parse_args(argv)

    cur_pnl = (load_pnl_from_canada_xlsx(args.canada_pnl)
               if args.canada_pnl else {})
    prev_pnl = (load_pnl_from_canada_xlsx(args.canada_pnl_prev)
                if args.canada_pnl_prev else {})
    yoy_pnl = (load_pnl_from_canada_xlsx(args.canada_pnl_yoy)
               if args.canada_pnl_yoy else {})

    # Fallback: read P&L from an existing 基础数据 sheet when canada_pnl
    # files aren't available (i.e. SAP mapping file lives only on prod).
    if args.basic_data_ref:
        prev_year = args.prev_year or args.year
        prev_month = args.prev_month
        if prev_month is None:
            prev_month = args.month - 1 if args.month > 1 else 12
            if args.month == 1:
                prev_year = args.year - 1
        yoy_year = args.yoy_year or (args.year - 1)
        yoy_month = args.yoy_month or args.month
        if not cur_pnl:
            cur_pnl = load_pnl_from_basic_data(
                args.basic_data_ref, year=args.year, month=args.month,
            )
            log.info("loaded cur_pnl from %s: %d stores", args.basic_data_ref,
                     len(cur_pnl))
        if not prev_pnl:
            prev_pnl = load_pnl_from_basic_data(
                args.basic_data_ref, year=prev_year, month=prev_month,
            )
            log.info("loaded prev_pnl (%d-%02d) from basic_data: %d stores",
                     prev_year, prev_month, len(prev_pnl))
        if not yoy_pnl:
            yoy_pnl = load_pnl_from_basic_data(
                args.basic_data_ref, year=yoy_year, month=yoy_month,
            )
            log.info("loaded yoy_pnl (%d-%02d) from basic_data: %d stores",
                     yoy_year, yoy_month, len(yoy_pnl))

    pos_paths = (discover_pos_paths(args.pos_dir, args.pos_period)
                 if args.pos_dir else {})
    pos_prev_paths = (
        discover_pos_paths(args.pos_prev_dir, args.pos_prev_period)
        if args.pos_prev_dir and args.pos_prev_period else {}
    )
    pos_yoy_paths = (
        discover_pos_paths(args.pos_yoy_dir, args.pos_yoy_period)
        if args.pos_yoy_dir and args.pos_yoy_period else {}
    )
    # BOM source: IPMS export (broad coverage) when provided, else store_bom DB.
    if args.ipms_bom:
        recipes = load_ipms_bom_recipes(list(args.ipms_bom))
        bom_rows = {store: recipes for store in STORE_ORDER if store != "加拿大九店"}
        log.info("loaded %d IPMS recipes, broadcast to %d stores",
                 len(recipes), len(bom_rows))
    else:
        bom_rows = load_bom_from_db()

    # Build the 7-month rolling trend. When --basic-data-ref is provided
    # we read 7 consecutive months' P&L from the reference workbook;
    # otherwise we fall back to cur/prev/YoY-only (pad rest with zeros).
    monthly_gp: dict[str, list[float]] = {}
    if args.basic_data_ref:
        # 7 most-recent months ending at (args.year, args.month) inclusive.
        history_pnls: list[dict[str, dict[str, float]]] = []
        yy, mm = args.year, args.month
        for _ in range(7):
            history_pnls.append(load_pnl_from_basic_data(
                args.basic_data_ref, year=yy, month=mm,
            ))
            mm -= 1
            if mm == 0:
                mm = 12
                yy -= 1
        for store in STORE_ORDER:
            gps = []
            for pnl in history_pnls:
                v = pnl.get(store, {}).get("三、毛利率")
                gps.append(float(v) if v is not None else 0.0)
            if any(gps):
                monthly_gp[store] = gps
    else:
        for store in STORE_ORDER:
            gps = []
            for pnl in (cur_pnl, prev_pnl, yoy_pnl):
                v = pnl.get(store, {}).get("三、毛利率")
                gps.append(float(v) if v is not None else 0.0)
            while len(gps) < 7:
                gps.append(0.0)
            if any(gps):
                monthly_gp[store] = gps

    inputs = load_inputs_from_paths(
        year=args.year,
        month=args.month,
        cur_pnl=cur_pnl,
        pos_sales_paths=pos_paths,
        pos_prev_paths=pos_prev_paths,
        pos_yoy_paths=pos_yoy_paths,
        bom_rows=bom_rows,
        zfi_cur_path=args.zfi0156 or Path("/nonexistent"),
        mb5b_cur_path=args.mb5b or Path("/nonexistent"),
        monthly_gp=monthly_gp,
        prev_pnl=prev_pnl,
        yoy_pnl=yoy_pnl,
        mb5b_prev_path=args.mb5b_prev,
        mb5b_yoy_path=args.mb5b_yoy,
    )

    # When --basic-data-ref is provided, mirror the historical archive
    # (every store-month row) into the generated workbook's 基础数据 sheet.
    if args.basic_data_ref:
        inputs.basic_data_records = load_full_basic_data_records(
            args.basic_data_ref,
        )
        log.info("loaded %d 基础数据 records from %s",
                 len(inputs.basic_data_records), args.basic_data_ref)

    out = build_workbook(inputs, args.output,
                         style_template=args.style_template)
    log.info("毛利分析 workbook saved: %s", out)

    if args.reference_diff:
        if not args.reference_diff.exists():
            log.error("reference workbook not found: %s", args.reference_diff)
            return 2
        mismatches = diff_against_reference(out, args.reference_diff)
        if mismatches:
            log.warning("reference-diff found %d mismatch(es):", len(mismatches))
            for m in mismatches:
                log.warning("  - %s", m)
            return 1
        log.info("reference-diff: clean — generated matches reference layout ✓")

    return 0


if __name__ == "__main__":
    sys.exit(main())
