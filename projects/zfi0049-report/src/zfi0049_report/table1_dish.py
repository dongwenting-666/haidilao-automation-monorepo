"""Build 表1-菜品价格变动及菜品损耗表 for the 毛利分析 workbook.

One row per (store × dish × spec × material) with sales, theoretical
material usage, prices, and loss-impact analysis. Joins POS sales
(红火台销售汇总) with store_bom (per-store recipes) and downstream
material usage + price data.

This module covers the structural skeleton + cols 1–12, 18–19 (the
parts derivable from POS + store_bom alone). Cols 13–17 (price history),
20–21 (material usage/price), 22–36 (套餐 / derived) are populated by
``enrich_with_materials`` / ``enrich_with_prices`` once those inputs are
wired (see iter 3+).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Full 36-column header layout, matching the manual workbook.
HEADERS: list[str] = [
    "区域", "门店名称", "唯一编码", "套餐唯一编码", "菜品编码",
    "菜品短编码", "菜品名称", "出品份量", "菜品单位", "规格",
    "本期销量", "理论耗用量", "本期单价", "上期单价", "去年同期单价",
    "环比价格变动", "同比价格变动", "对应物料代码", "对应物料名称（对应zfi0156每个门店V）",
    "单价/KG（对应zfi0156每个门店V）", "本期耗用量(对应zfi0156每个门店V）",
    "套餐、拼盘用量", "套餐销量", "环比影响收入", "同比影响收入",
    "调价日期", "点击率-按门店V汇总", "实际毛利率", "理论毛利率",
    "理论耗用量", "实际耗用量", "本月损耗影响成本金额", "上月损耗影响成本金额",
    "损耗环比变动金额", "可比期间损耗影响成本金额", "损耗同比变动金额",
]
assert len(HEADERS) == 36


@dataclass
class Table1Row:
    """One row in 表1 — per store × dish × spec × material."""

    # ── Identifiers (cols 1–10) ──
    region: str            # 区域 = '加拿大'
    store: str             # 门店名称
    dish_code: int         # 菜品编码
    dish_short_code: int | None  # 菜品短编码 (from POS)
    dish_name: str
    portion: float | None  # 出品份量 (from store_bom)
    dish_unit: str | None  # 菜品单位 (e.g. '锅', '整份') — from POS or store_bom
    spec: str              # 规格

    # ── Sales + theoretical usage (cols 11–12) ──
    sales_qty: float = 0           # 本期销量 (POS 实际出品数据)
    loss_factor: float = 1.0       # from store_bom

    # ── Prices (cols 13–17) — filled by enrich_with_prices ──
    cur_dish_price: float | None = None  # 本期单价
    prev_dish_price: float | None = None  # 上期单价
    yoy_dish_price: float | None = None   # 去年同期单价

    # ── Material join (cols 18–21) ──
    material_code: int | None = None    # 对应物料代码
    material_name: str | None = None    # 对应物料名称
    material_unit_price: float | None = None  # 单价/KG (MB5B)
    material_period_usage: float | None = None  # 本期耗用量 (ZFI0156)

    # ── Set meal (cols 22–23) — filled by enrich_with_set_meals ──
    set_meal_usage: float | None = None  # 套餐、拼盘用量
    set_meal_qty: float | None = None    # 套餐销量

    # ── Revenue impact (cols 24–25) — filled by compute_revenue_impact ──
    mom_revenue_impact: float | None = None  # 环比影响收入
    yoy_revenue_impact: float | None = None  # 同比影响收入

    # ── Report category (for 细分毛利率表 aggregation) ──
    category: str | None = None  # e.g. '锅底类', '荤菜类', '其他类（如有）'

    # ── Pricing date (col 26) ──
    price_change_date: Any = None  # 调价日期 — manual fill currently

    # ── Per-store-aggregate metrics (cols 27–29) ──
    click_rate: float | None = None  # 点击率 — store-level rollup
    actual_gp_rate: float | None = None  # 实际毛利率
    theoretical_gp_rate: float | None = None  # 理论毛利率

    # ── Actual usage (col 31) ──
    actual_usage: float | None = None  # 实际耗用量 (ZFI0156)

    # ── Loss-impact (cols 32–36) — filled by compute_loss_impact ──
    loss_cost_cur: float | None = None
    loss_cost_prev: float | None = None
    loss_cost_mom_delta: float | None = None
    loss_cost_comparable: float | None = None
    loss_cost_yoy_delta: float | None = None

    @property
    def theoretical_usage(self) -> float:
        """理论耗用量 = sales × portion × loss_factor."""
        if self.portion is None or self.sales_qty in (None, 0):
            return 0.0
        return self.sales_qty * self.portion * self.loss_factor

    @property
    def unique_code(self) -> str:
        """唯一编码 = store + dish_code + short_code + spec."""
        short = self.dish_short_code if self.dish_short_code is not None else ""
        return f"{self.store}{self.dish_code}{short}{self.spec}"

    @property
    def set_meal_unique_code(self) -> str:
        """套餐唯一编码 = store + dish_code + dish_name + spec."""
        return f"{self.store}{self.dish_code}{self.dish_name}{self.spec}"

    @property
    def mom_price_delta(self) -> float | None:
        if self.cur_dish_price is None or self.prev_dish_price is None:
            return None
        return self.cur_dish_price - self.prev_dish_price

    @property
    def yoy_price_delta(self) -> float | None:
        if self.cur_dish_price is None or self.yoy_dish_price is None:
            return None
        return self.cur_dish_price - self.yoy_dish_price


def to_row(r: Table1Row) -> list[Any]:
    """Project Table1Row → 36-column list aligned with HEADERS."""
    return [
        r.region,                          # 1
        r.store,                            # 2
        r.unique_code,                      # 3
        r.set_meal_unique_code,             # 4
        r.dish_code,                        # 5
        r.dish_short_code,                  # 6
        r.dish_name,                        # 7
        r.portion,                          # 8
        r.dish_unit,                        # 9
        r.spec,                             # 10
        r.sales_qty,                        # 11
        r.theoretical_usage,                # 12
        r.cur_dish_price,                   # 13
        r.prev_dish_price,                  # 14
        r.yoy_dish_price,                   # 15
        r.mom_price_delta,                  # 16
        r.yoy_price_delta,                  # 17
        r.material_code,                    # 18
        r.material_name,                    # 19
        r.material_unit_price,              # 20
        r.material_period_usage,            # 21
        r.set_meal_usage,                   # 22
        r.set_meal_qty,                     # 23
        r.mom_revenue_impact,               # 24
        r.yoy_revenue_impact,               # 25
        r.price_change_date,                # 26
        r.click_rate,                       # 27
        r.actual_gp_rate,                   # 28
        r.theoretical_gp_rate,              # 29
        r.theoretical_usage,                # 30 — dup of col 12
        r.actual_usage,                     # 31
        r.loss_cost_cur,                    # 32
        r.loss_cost_prev,                   # 33
        r.loss_cost_mom_delta,              # 34
        r.loss_cost_comparable,             # 35
        r.loss_cost_yoy_delta,              # 36
    ]


# ── POS loader ───────────────────────────────────────────────────────────────


@dataclass
class PosSale:
    store: str
    dish_code: int
    dish_short_code: int | None
    dish_name: str
    spec: str
    sales_qty: float
    dish_unit: str | None = None  # not in older POS files
    category: str | None = None   # report category (from POS 大类 → mapping)


def load_pos_sales(path: Path) -> list[PosSale]:
    """Read a 红火台销售汇总 xlsx → list of PosSale.

    The POS schema (output of pos-crawler):
      检索 / 门店名称 / 编码 / 菜品编码 / 菜品短编码 / 菜品名称 /
      规格 / 出品数量 / 退菜数量 / 实际出品数据 / 大类名称 / 子类名称

    实际出品数据 = 出品数量 - 退菜数量 (this is 本期销量 for 表1).
    """
    from zfi0049_report.dish_category import map_pos_to_report_category
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = list(rows[0])
    idx = {h: i for i, h in enumerate(headers) if h}
    try:
        store_i = idx["门店名称"]
        dish_i = idx["菜品编码"]
        short_i = idx["菜品短编码"]
        name_i = idx["菜品名称"]
        spec_i = idx["规格"]
        net_i = idx["实际出品数据（出品数量-退菜数量）"]
    except KeyError as e:
        raise ValueError(f"POS file {path} missing expected column: {e}")
    dalei_i = idx.get("大类名称")
    zilei_i = idx.get("子类名称")
    unit_i = idx.get("菜品单位")

    out = []
    for vals in rows[1:]:
        store = vals[store_i]
        dish = vals[dish_i]
        if not store or not dish:
            continue
        try:
            dish_int = int(str(dish).strip().lstrip("0") or "0")
        except (TypeError, ValueError):
            continue
        short = vals[short_i]
        short_int = None
        if short is not None and short != "":
            try:
                short_int = int(str(short).strip().lstrip("0") or "0")
            except (TypeError, ValueError):
                short_int = None
        sales = vals[net_i]
        try:
            sales_f = float(sales) if sales not in (None, "") else 0.0
        except (TypeError, ValueError):
            sales_f = 0.0
        dalei = vals[dalei_i] if dalei_i is not None and dalei_i < len(vals) else None
        zilei = vals[zilei_i] if zilei_i is not None and zilei_i < len(vals) else None
        unit = vals[unit_i] if unit_i is not None and unit_i < len(vals) else None
        out.append(PosSale(
            store=str(store).strip(),
            dish_code=dish_int,
            dish_short_code=short_int,
            dish_name=str(vals[name_i]).strip() if vals[name_i] else "",
            spec=str(vals[spec_i]).strip() if vals[spec_i] else "",
            sales_qty=sales_f,
            dish_unit=str(unit).strip() if unit else None,
            category=map_pos_to_report_category(dalei, zilei),
        ))
    return out


# ── Builder ──────────────────────────────────────────────────────────────────


def build_rows_for_store(
    store: str,
    *,
    pos_sales: list[PosSale],
    bom_rows: list[dict[str, Any]],
    region: str = "加拿大",
    sold_only: bool = False,
) -> list[Table1Row]:
    """Join POS sales with store_bom rows for a single store.

    Each BOM row (dish × spec × material) produces one Table1Row. Sales
    for the (dish, spec) come from ``pos_sales``; rows with no matching
    sale get sales_qty = 0.

    When ``sold_only`` is True, only BOM rows whose (dish, spec) appears
    in this store's POS sales are emitted — this matches the manual 表1
    which lists dishes actually sold per store (not the full region BOM).
    """
    # Index POS by (dish_code, spec) for fast lookup
    sales_idx: dict[tuple[int, str], PosSale] = {}
    for s in pos_sales:
        sales_idx[(s.dish_code, s.spec)] = s

    out = []
    for b in bom_rows:
        dish = int(b["dish_code"])
        spec = str(b["spec"]).strip()
        sale = sales_idx.get((dish, spec))
        if sold_only and sale is None:
            continue
        out.append(Table1Row(
            region=region,
            store=store,
            dish_code=dish,
            dish_short_code=(b.get("dish_short_code")
                             if b.get("dish_short_code") is not None
                             else (sale.dish_short_code if sale else None)),
            dish_name=str(b.get("dish_name") or (sale.dish_name if sale else "")),
            portion=float(b["portion"]) if b.get("portion") is not None else None,
            dish_unit=sale.dish_unit if sale else None,
            spec=spec,
            sales_qty=sale.sales_qty if sale else 0,
            loss_factor=(float(b["loss_factor"])
                         if b.get("loss_factor") is not None else 1.0),
            material_code=(int(b["material_code"])
                           if b.get("material_code") is not None else None),
            material_name=b.get("material_name"),
            category=sale.category if sale else None,
        ))
    return out


def write_sheet(ws: Worksheet, records: list[Table1Row]) -> int:
    """Write records to ``ws`` with the standard 表1 header. Returns row count."""
    ws.append(HEADERS)
    written = 0
    for r in records:
        ws.append(to_row(r))
        written += 1
    return written


# ── ZFI0156 material usage loader ────────────────────────────────────────────


@dataclass(frozen=True)
class MaterialUsage:
    """Per-(plant × material) summary from ZFI0156."""

    werks: str
    matnr: int
    matxt: str | None
    unit: str | None       # 单位描述 (e.g. '公斤', '升')
    unit_price: float | None  # 系统发出单价
    quantity: float        # 数量 (summed across 大类 rows)
    amount: float          # 系统发出金额 (summed)


def _coerce_matnr(v: Any) -> int | None:
    if v in (None, ""):
        return None
    try:
        return int(str(v).strip().lstrip("0") or "0")
    except (TypeError, ValueError):
        return None


def load_zfi0156(path: Path) -> dict[tuple[str, int], MaterialUsage]:
    """Read a ZFI0156 export → {(werks, matnr) → MaterialUsage}.

    ZFI0156 rows can repeat the same (plant, material) under different
    大类 ('费用类' / '成本类'). We sum 数量 + 金额 across rows; metadata
    (unit_price / matxt / unit) takes the last non-empty value (they're
    identical across the duplicate rows in practice).
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = list(rows[0])

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    werks_i = col("工厂")
    mat_i = col("物料")
    matxt_i = col("物料描述")
    unit_i = col("单位描述")
    price_i = col("系统发出单价")
    qty_i = col("数量")
    amt_i = col("系统发出金额")
    if werks_i is None or mat_i is None or qty_i is None:
        raise ValueError(f"ZFI0156 {path} missing required columns")

    acc: dict[tuple[str, int], dict[str, Any]] = {}
    for vals in rows[1:]:
        if not vals or vals[werks_i] is None or vals[mat_i] is None:
            continue
        werks = str(vals[werks_i]).strip()
        matnr = _coerce_matnr(vals[mat_i])
        if matnr is None:
            continue
        key = (werks, matnr)
        rec = acc.setdefault(key, {
            "matxt": None, "unit": None, "unit_price": None,
            "quantity": 0.0, "amount": 0.0,
        })
        qty = vals[qty_i] if qty_i is not None else None
        try:
            rec["quantity"] += float(qty) if qty not in (None, "") else 0.0
        except (TypeError, ValueError):
            pass
        if amt_i is not None:
            amt = vals[amt_i]
            try:
                rec["amount"] += float(amt) if amt not in (None, "") else 0.0
            except (TypeError, ValueError):
                pass
        for src_i, dst in [(matxt_i, "matxt"), (unit_i, "unit"),
                            (price_i, "unit_price")]:
            if src_i is None:
                continue
            v = vals[src_i]
            if v not in (None, ""):
                if dst == "unit_price":
                    try:
                        rec[dst] = float(v)
                    except (TypeError, ValueError):
                        pass
                else:
                    rec[dst] = str(v).strip()

    return {
        k: MaterialUsage(
            werks=k[0], matnr=k[1],
            matxt=v["matxt"], unit=v["unit"], unit_price=v["unit_price"],
            quantity=v["quantity"], amount=v["amount"],
        )
        for k, v in acc.items()
    }


# ── MB5B price loader (wraps inventory-check.mb5b_parse) ────────────────────


def load_mb5b_prices(path: Path) -> dict[tuple[str, int], float]:
    """Read an MB5B export → {(werks, matnr) → moving-average unit price}.

    MB5B doesn't expose a UnitPrice column in the raw 14-col export; we
    derive it from ClosingAmt / ClosingQty (the SAP moving-average price).
    Falls back to OpeningAmt / OpeningQty when closing is zero (newly-
    received materials with no consumption yet). Skips non-store werks
    rows (company-code headers like '9451').
    """
    from inventory_check.mb5b_parse import parse_mb5b_file
    rows = parse_mb5b_file(path)
    out: dict[tuple[str, int], float] = {}
    for r in rows:
        werks = r.get("Werks")
        if not isinstance(werks, str) or not werks.startswith("CA"):
            continue
        matnr = _coerce_matnr(r.get("Matnr"))
        if matnr is None:
            continue
        # Try explicit UnitPrice first, then ClosingAmt/Qty, then OpeningAmt/Qty.
        price = r.get("UnitPrice")
        if isinstance(price, (int, float)) and price > 0:
            out[(werks, matnr)] = float(price)
            continue
        for amt_key, qty_key in [("ClosingAmt", "ClosingQty"),
                                  ("OpeningAmt", "OpeningQty")]:
            amt = r.get(amt_key)
            qty = r.get(qty_key)
            if (isinstance(amt, (int, float)) and isinstance(qty, (int, float))
                    and qty > 0):
                out[(werks, matnr)] = float(amt) / float(qty)
                break
    return out


# ── POS price loader (newer crawler outputs only) ───────────────────────────


def load_pos_prices(path: Path) -> dict[tuple[int, str], tuple[float, str | None]]:
    """Read a POS xlsx → {(dish_code, spec) → (dish_price, dish_unit)}.

    Newer pos-crawler outputs include 菜品单价 + 菜品单位 columns (from the
    listDishPotSale API). Older outputs don't — we return an empty dict
    rather than failing so the caller can degrade gracefully.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = list(rows[0])

    def col(name: str) -> int | None:
        try:
            return header.index(name)
        except ValueError:
            return None

    dish_i = col("菜品编码")
    spec_i = col("规格")
    price_i = col("菜品单价")
    unit_i = col("菜品单位")
    if dish_i is None or spec_i is None or price_i is None:
        return {}  # older format, no price column

    out: dict[tuple[int, str], tuple[float, str | None]] = {}
    for vals in rows[1:]:
        dish = vals[dish_i] if dish_i < len(vals) else None
        spec = vals[spec_i] if spec_i < len(vals) else None
        price = vals[price_i] if price_i < len(vals) else None
        unit = vals[unit_i] if unit_i is not None and unit_i < len(vals) else None
        if dish is None or spec is None or price in (None, ""):
            continue
        try:
            dish_int = int(str(dish).strip().lstrip("0") or "0")
        except (TypeError, ValueError):
            continue
        try:
            price_f = float(price)
        except (TypeError, ValueError):
            continue
        key = (dish_int, str(spec).strip())
        if key not in out:
            out[key] = (price_f, str(unit).strip() if unit else None)
    return out


# ── Enrichment ──────────────────────────────────────────────────────────────


# Map werks → store_name (in step with canada_pnl.STORE_ORDER)
WERKS_TO_STORE = {
    "CA01": "加拿大一店", "CA02": "加拿大二店", "CA03": "加拿大三店",
    "CA04": "加拿大四店", "CA05": "加拿大五店", "CA06": "加拿大六店",
    "CA07": "加拿大七店", "CA08": "加拿大八店",
}


def enrich_with_materials(
    rows: list[Table1Row],
    *,
    werks: str,
    zfi: dict[tuple[str, int], MaterialUsage],
    mb5b_prices: dict[tuple[str, int], float] | None = None,
) -> None:
    """Populate cols 19, 20, 21, 31 (material name / unit price / theoretical
    usage / actual usage).

    - Col 19 (对应物料名称): from ZFI0156 (matxt) — overrides any value
      already on the row from store_bom.
    - Col 20 (单价/KG): prefer MB5B moving-avg price (matches manual
      workbook observed values); fall back to ZFI0156 系统发出单价.
    - Col 21 (本期耗用量): **theoretical** material consumption, computed
      as Σ(sales × portion × loss_factor) across all dish×spec rows
      that share this material in the store. Despite the manual's "对应
      zfi0156每个门店V" header note, observed numbers do not match raw
      ZFI0156 quantity; they match the theoretical sum.
    - Col 31 (实际耗用量): actual issued quantity from ZFI0156 数量.
    """
    mb5b_prices = mb5b_prices or {}

    # 1. Pre-compute theoretical usage per material (col 21 aggregate).
    theoretical_by_mat: dict[int, float] = {}
    for r in rows:
        if r.material_code is None:
            continue
        theoretical_by_mat[r.material_code] = (
            theoretical_by_mat.get(r.material_code, 0.0) + r.theoretical_usage
        )

    for r in rows:
        if r.material_code is None:
            continue
        usage = zfi.get((werks, r.material_code))
        if usage:
            if not r.material_name:
                r.material_name = usage.matxt
            r.actual_usage = usage.quantity  # col 31
        # Theoretical material usage at the material level (col 21)
        r.material_period_usage = theoretical_by_mat.get(r.material_code)
        # Price (col 20): prefer MB5B (matches manual workbook)
        mb_p = mb5b_prices.get((werks, r.material_code))
        if mb_p is not None:
            r.material_unit_price = mb_p
        elif usage and usage.unit_price is not None:
            r.material_unit_price = usage.unit_price


def enrich_with_prices(
    rows: list[Table1Row],
    *,
    cur_prices: dict[tuple[int, str], tuple[float, str | None]],
    prev_prices: dict[tuple[int, str], tuple[float, str | None]] | None = None,
    yoy_prices: dict[tuple[int, str], tuple[float, str | None]] | None = None,
) -> None:
    """Populate cols 13–17 (current / prev / YoY price + deltas).

    ``cur_prices`` / ``prev_prices`` / ``yoy_prices`` are dicts from
    ``load_pos_prices``. Older POS exports lack the 菜品单价 column —
    they yield empty dicts and rows stay None (consistent with how the
    manual workbook leaves cells blank for stores that opened recently).
    """
    prev_prices = prev_prices or {}
    yoy_prices = yoy_prices or {}
    for r in rows:
        key = (r.dish_code, r.spec)
        cur = cur_prices.get(key)
        if cur is not None:
            r.cur_dish_price = cur[0]
            if r.dish_unit is None:
                r.dish_unit = cur[1]
        prev = prev_prices.get(key)
        if prev is not None:
            r.prev_dish_price = prev[0]
        yoy = yoy_prices.get(key)
        if yoy is not None:
            r.yoy_dish_price = yoy[0]


# ── Canonical-row detection + blanking ──────────────────────────────────────


def find_canonical_rows(rows: list[Table1Row]) -> set[int]:
    """For each (store × dish × material) group, return the index of the
    canonical row — the row with the **largest** portion.

    Reasoning (verified against manual workbook, 加拿大一店 dish 1060066):
    when a single dish has multiple specs (单锅 / 拼锅 / 四宫格) all
    consuming the same material, the SAP report keeps material-level
    aggregates (本期耗用量, 实际耗用量, loss-impact) on a single row to
    avoid double-counting via SUMIF. The 单锅 (largest portion) row is
    chosen because that's the unsplit "main" spec.
    """
    groups: dict[tuple[str, int, int], list[tuple[int, float]]] = {}
    for idx, r in enumerate(rows):
        if r.material_code is None:
            continue
        portion = r.portion if r.portion is not None else 0.0
        groups.setdefault(
            (r.store, r.dish_code, r.material_code), []
        ).append((idx, portion))
    canonical: set[int] = set()
    for items in groups.values():
        # Largest portion wins; ties broken by row order (stable).
        items.sort(key=lambda t: (-t[1], t[0]))
        canonical.add(items[0][0])
    return canonical


def apply_canonical_blanking(rows: list[Table1Row],
                             canonical: set[int] | None = None) -> None:
    """Blank the canonical-only columns on non-canonical rows.

    Cols 21 (本期耗用量), 31 (实际耗用量), 32, 33, 35 (loss impact
    cur/prev/comparable) are mirrored only on the canonical row of each
    (store × dish × material) group. Cols 34 / 36 are deltas that
    consumers compute from the canonical-row values, so we leave them as
    derived values on non-canonical rows (they evaluate to 0 since their
    inputs are blank).
    """
    if canonical is None:
        canonical = find_canonical_rows(rows)
    for idx, r in enumerate(rows):
        if idx in canonical:
            continue
        r.material_period_usage = None  # col 21
        r.actual_usage = None           # col 31
        r.loss_cost_cur = None          # col 32
        r.loss_cost_prev = None         # col 33
        r.loss_cost_comparable = None   # col 35


# ── Revenue impact (cols 24-25) ─────────────────────────────────────────────


def compute_revenue_impact(rows: list[Table1Row]) -> None:
    """Populate cols 24 (环比影响收入) and 25 (同比影响收入).

    Formulas (verified against manual 加拿大一店 1060066 拼锅 — MoM=0,
    YoY=2: yoy_revenue_impact = 1275 × 2 = 2550 ✓):
      环比影响收入 = (本期单价 − 上期单价) × 本期销量
      同比影响收入 = (本期单价 − 去年同期单价) × 本期销量

    Stored back into the row as new attributes accessible to ``to_row``.
    """
    for r in rows:
        if r.mom_price_delta is not None:
            r.mom_revenue_impact = r.mom_price_delta * r.sales_qty
        if r.yoy_price_delta is not None:
            r.yoy_revenue_impact = r.yoy_price_delta * r.sales_qty
