"""Build 细分毛利率表 (per-category gross margin breakdown).

Per (store × category): current-month gross margin, previous-month gross
margin, and the MoM delta. Categories come from
``dish_category.REPORT_CATEGORIES``.

Layout (matches the manual workbook 细分毛利率表 (2)):
  Row 1: title (e.g. '2026年3月细分毛利率环比表')
  Row 2: category headers — col 2 '门店名称', cols 3-5 锅底类, 6-8 荤菜类,
         9-11 素菜类, 12-14 酒水类, 15-17 小料台类, 18-20 小吃类, 21-23 其他类
  Row 3: sub-headers — for each category three sub-cols (cur / prev / 环比)
  Rows 4+: per-store data
"""
from __future__ import annotations

from collections import defaultdict

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from zfi0049_report.dish_category import OTHER, REPORT_CATEGORIES
from zfi0049_report.table1_dish import Table1Row


# ── Aggregation ─────────────────────────────────────────────────────────────


def compute_category_gp(
    rows: list[Table1Row],
) -> dict[tuple[str, str], tuple[float, float, float | None]]:
    """Return {(store, category) → (revenue, cost, gross_margin)}.

    Revenue is unique per (store, dish_code, spec): the row's
    sales_qty × cur_dish_price counted once across the rows that share
    the same dish_spec but different materials.

    Cost aggregates every (store, dish, spec, material) row:
      cost = sales_qty × portion × loss_factor × material_unit_price

    Rows missing prices/portions contribute 0 to that side. Stores with
    zero revenue across all categories are excluded from the output.
    """
    revenue: dict[tuple[str, str], float] = defaultdict(float)
    cost: dict[tuple[str, str], float] = defaultdict(float)
    seen_dish_spec: set[tuple[str, int, str]] = set()

    for r in rows:
        cat = r.category or OTHER
        key = (r.store, cat)

        # Revenue: only count each (store, dish, spec) once
        dish_spec_key = (r.store, r.dish_code, r.spec)
        if dish_spec_key not in seen_dish_spec:
            seen_dish_spec.add(dish_spec_key)
            if r.sales_qty and r.cur_dish_price is not None:
                revenue[key] += r.sales_qty * r.cur_dish_price

        # Cost: per-row contribution
        if (r.sales_qty and r.portion is not None
                and r.material_unit_price is not None):
            cost[key] += (
                r.sales_qty * r.portion * r.loss_factor
                * r.material_unit_price
            )

    out: dict[tuple[str, str], tuple[float, float, float | None]] = {}
    all_keys = set(revenue) | set(cost)
    for key in all_keys:
        rev = revenue[key]
        c = cost[key]
        gm = ((rev - c) / rev) if rev else None
        out[key] = (rev, c, gm)
    return out


# ── Sheet writer ────────────────────────────────────────────────────────────


def _column_block_start(idx: int) -> int:
    """Cols 3-5 for cat 0, 6-8 for cat 1, etc. (3 cols per category)."""
    return 3 + idx * 3


def build_subdivided_gp_sheet(
    wb: Workbook,
    *,
    cur_gp: dict[tuple[str, str], tuple[float, float, float | None]],
    prev_gp: dict[tuple[str, str], tuple[float, float, float | None]] | None = None,
    year: int,
    month: int,
    stores: list[str] | None = None,
) -> Worksheet:
    """Create the 细分毛利率表 (2) sheet on ``wb`` and populate.

    ``stores`` defaults to the set of stores that appear in either
    ``cur_gp`` or ``prev_gp`` (sorted by store-name).
    """
    prev_gp = prev_gp or {}
    if stores is None:
        store_set = {k[0] for k in cur_gp} | {k[0] for k in prev_gp}
        stores = sorted(store_set)

    ws = wb.create_sheet("细分毛利率表 (2)")

    # Row 1: title
    ws.cell(row=1, column=2, value=f"{year}年{month}月细分毛利率环比表")

    # Row 2: category headers spanning 3 cols each
    ws.cell(row=2, column=2, value="门店名称")
    for cat_idx, cat in enumerate(REPORT_CATEGORIES):
        col = _column_block_start(cat_idx)
        ws.cell(row=2, column=col, value=cat)

    # Row 3: sub-headers (cur / prev / 环比) under each category
    for cat_idx in range(len(REPORT_CATEGORIES)):
        col = _column_block_start(cat_idx)
        ws.cell(row=3, column=col, value=f"{year}年{month}月")
        ws.cell(row=3, column=col + 1, value="上月")
        ws.cell(row=3, column=col + 2, value="环比")

    # Rows 4+: per-store data
    for r_idx, store in enumerate(stores, start=4):
        ws.cell(row=r_idx, column=2, value=store)
        for cat_idx, cat in enumerate(REPORT_CATEGORIES):
            col = _column_block_start(cat_idx)
            cur = cur_gp.get((store, cat))
            prev = prev_gp.get((store, cat))
            cur_gm = cur[2] if cur and cur[2] is not None else None
            prev_gm = prev[2] if prev and prev[2] is not None else None
            ws.cell(row=r_idx, column=col, value=cur_gm)
            ws.cell(row=r_idx, column=col + 1, value=prev_gm)
            mom = (cur_gm - prev_gm) if (cur_gm is not None
                                         and prev_gm is not None) else None
            ws.cell(row=r_idx, column=col + 2, value=mom)

    return ws
