"""Fill a styled template workbook with computed data, preserving styles.

The manual 附件3-毛利相关分析指标 workbook carries all the formatting we
need (fonts, fills, borders, merged cells, conditional formatting,
column widths, number formats). Rather than rebuild that styling from
scratch in openpyxl, we copy the template and overwrite only the data
cells in place — openpyxl keeps a cell's style when you set just
``cell.value``.

Per-sheet fill functions write computed rows into the template's data
region starting at the known data-start row. When our row count exceeds
the template's, the style of the last template data row is copied down
to the new rows so they stay formatted.
"""
from __future__ import annotations

import logging
import shutil
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# Data-start row (1-based) per sheet — the first row that holds store/dish data.
DATA_START_ROW = {
    "表1-菜品价格变动及菜品损耗表 (模板) ": 4,
    "表2-原材料成本变动表": 3,
    "基础数据": 2,
}


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, ncols: int) -> None:
    """Copy cell styles (not values) from src_row to dst_row."""
    for c in range(1, ncols + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        if s.has_style:
            d.font = copy(s.font)
            d.fill = copy(s.fill)
            d.border = copy(s.border)
            d.alignment = copy(s.alignment)
            d.number_format = s.number_format
            d.protection = copy(s.protection)


def _write_rows_in_place(
    ws: Worksheet, rows: list[list[Any]], start_row: int, ncols: int,
    *, clear_to_row: int | None = None,
) -> None:
    """Overwrite values from start_row downward, preserving styles.

    Styles for rows beyond the template's existing styled region are
    copied from the template's first data row (start_row). If
    ``clear_to_row`` is set, any leftover template data rows below the
    new data are cleared (values only).
    """
    proto = start_row
    for i, row in enumerate(rows):
        r = start_row + i
        if r > ws.max_row:
            _copy_row_style(ws, proto, r, ncols)
        for j, val in enumerate(row):
            if j >= ncols:
                break
            cell = ws.cell(row=r, column=j + 1)
            # Skip cells inside a merged range anchor that isn't top-left.
            cell.value = val
    # Clear leftover template rows below our data.
    if clear_to_row is not None:
        last = start_row + len(rows)
        for r in range(last, clear_to_row + 1):
            for c in range(1, ncols + 1):
                ws.cell(row=r, column=c).value = None


def open_template(template_path: Path, out_path: Path):
    """Copy template → out_path and open it for editing (styles preserved)."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, out_path)
    return load_workbook(out_path)


def fill_table_sheet(
    wb, sheet_name: str, rows: list[list[Any]], ncols: int,
) -> None:
    """Generic filler for the flat data tables (表1 / 表2 / 基础数据)."""
    if sheet_name not in wb.sheetnames:
        logger.warning("template missing sheet %s", sheet_name)
        return
    ws = wb[sheet_name]
    start = DATA_START_ROW[sheet_name]
    template_last = ws.max_row
    _write_rows_in_place(ws, rows, start, ncols, clear_to_row=template_last)
    logger.info("filled %s: %d rows from row %d", sheet_name, len(rows), start)


def fill_positioned_rows(
    wb, sheet_name: str, rows_by_store: dict[str, list[Any]],
    *, store_col: int, start_row: int, end_row: int,
    value_cols: list[int], strict: bool = True,
) -> None:
    """Fill a sheet where each store occupies a fixed row in [start,end].

    Reads the store name from ``store_col`` on each existing template row
    and writes the matching row's values into ``value_cols``.

    When ``strict=True`` (default), every cell at (row, value_col) for
    rows in [start_row, end_row] is FIRST CLEARED so anything that ends
    up in the workbook comes from our pipeline — never inherited from
    the template. Cells where we have no computed value remain None
    (blank), which is honest about coverage.
    """
    if sheet_name not in wb.sheetnames:
        logger.warning("template missing sheet %s", sheet_name)
        return
    ws = wb[sheet_name]
    if strict:
        for r in range(start_row, end_row + 1):
            for col in value_cols:
                cell = ws.cell(row=r, column=col)
                cur = cell.value
                # Preserve formulas (Excel recomputes them on open). Only
                # clear literal values so any literal that survives must
                # come from our write step below.
                if isinstance(cur, str) and cur.startswith("="):
                    continue
                safe_set(ws, r, col, None)
    written = 0
    for r in range(start_row, end_row + 1):
        store = ws.cell(row=r, column=store_col).value
        if not store or store not in rows_by_store:
            continue
        full = rows_by_store[store]
        for col in value_cols:
            if col - 1 < len(full):
                v = full[col - 1]
                if v is not None:
                    safe_set(ws, r, col, v)
        written += 1
    logger.info("filled %s: %d store rows (%d..%d, strict=%s)",
                sheet_name, written, start_row, end_row, strict)


def safe_set(ws: Worksheet, row: int, col: int, value: Any) -> bool:
    """Write a cell value, skipping cells inside a merge that aren't the anchor.

    openpyxl raises 'MergedCell object attribute value is read-only' when
    you try to write to a non-anchor cell of a merged range. Returns True
    if the write succeeded, False if skipped.
    """
    try:
        ws.cell(row=row, column=col).value = value
        return True
    except AttributeError:
        return False


def fill_specific_row(
    wb, sheet_name: str, row: int, values: list[Any],
    *, value_cols: list[int] | None = None,
) -> None:
    """Write a single row's values into specific columns.

    Used for total / aggregate / header rows that don't fit the
    per-store positioned pattern. ``values`` is a 1-based column → value
    list (index col-1). When ``value_cols`` is None, writes all non-None
    values from ``values``; otherwise writes only the listed columns.
    """
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    cols_to_write = value_cols if value_cols is not None else list(range(1, len(values) + 1))
    for col in cols_to_write:
        if col - 1 < len(values):
            v = values[col - 1]
            if v is not None:
                ws.cell(row=row, column=col).value = v


def clear_cells(wb, sheet_name: str, ranges: list[tuple[int, int, int, int]]) -> None:
    """Blank cell values in arbitrary rectangular ranges (row1, col1, row2, col2)
    so any value present afterwards came from our pipeline."""
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    for r1, c1, r2, c2 in ranges:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).value = None


def fill_instructions_sheet(wb, year: int, month: int) -> None:
    """Strict-overwrite the 填写说明 sheet from source (no template-inherited text)."""
    if "填写说明" not in wb.sheetnames:
        return
    ws = wb["填写说明"]
    # Clear existing content (skip merged-cell anchors that openpyxl
    # raises on; safe_set swallows MergedCell read-only errors).
    last = max(ws.max_row, 12)
    for r in range(1, last + 1):
        for c in range(1, max(ws.max_column, 3) + 1):
            safe_set(ws, r, c, None)
    safe_set(ws, 1, 1, "填表说明：涉及金额的填写为本币金额")
    safe_set(ws, 2, 1, "1、先更新《表1-菜品价格变动及菜品损耗表》、《表2-原材料成本变动表》、《表3-打折优惠表》")
    safe_set(ws, 3, 1, "2、其次填写《细分毛利率表》、《毛利率连续对比表》")
    safe_set(ws, 4, 1, "3、再填写《毛利率环比》和《毛利率同比》中贴数部分")
    safe_set(ws, 5, 1,
        "4、通过本表数据对比分析后，需识别毛利率相关分析问题在经营分析报告上描述即可"
        "（描述内容需包含如下图示例中问题类型，内容建议等）"
    )
    safe_set(ws, 7, 1, f"{year}年{month}月毛利率相关问题")
    safe_set(ws, 8, 1, "序号")
    safe_set(ws, 8, 2, "问题内容")
    safe_set(ws, 8, 3, "问题描述及建议")
    safe_set(ws, 9, 1, 1)
    safe_set(ws, 9, 2, "毛利率环比下降异常")
    safe_set(ws, 9, 3,
        "(1) 问题描述：本月XX店毛利率环比下降超过3%，且毛利率低于60%；\n"
        "(2) 原因：如：本月库存盘点有误，影响成本虚增3万元，影响毛利率下降0.5%；"
        "原材料成本环比增加8万元，影响毛利率下降XXX；"
    )
    safe_set(ws, 10, 1, 2)
    safe_set(ws, 10, 2, "低毛利低点击率产品")
    safe_set(ws, 10, 3,
        "（1）问题描述：通过数据发现XXX店和XXX店销售的牛蛙（现杀）属于负毛利、低点击率产品"
        "（毛利率为-7.8%，点击率在1.5%左右），反馈至门店以及大区，由门店及大区评估是否下架或"
        "通过其他如调价等措施提升毛利率\n"
        "（2）建议：反馈至门店以及大区，由门店及大区评估是否下架或通过其他如调价等措施提升毛利率"
    )
    safe_set(ws, 11, 1, 3)
    safe_set(ws, 11, 2, "单品锅底毛利率异常")
    safe_set(ws, 11, 3,
        "(1) 问题描述：XXX店的番茄锅底及白玉锅底的毛利率较低损耗较大，该店番茄及白玉锅底"
        "全月损耗较片区平均水平高566公斤，影响成本上升1.2万人民币，影响锅底毛利率下降2%；"
    )
    safe_set(ws, 12, 1, 4)
    safe_set(ws, 12, 2, "酒水毛利率中单品毛利异常")
    safe_set(ws, 12, 3,
        "(1) 问题描述：因盘点不准导致的毛利率异常问题，尤其贵重酒水盘差影响尤为明显，"
        "如8月XXX店酒水毛利率环比7月下降16.7%，经复核发现门店飞天茅台漏盘，影响酒水毛利率下降异常；\n"
        "(2) 原因：主要是门店库存盘点不准确，存在漏盘的问题"
    )
