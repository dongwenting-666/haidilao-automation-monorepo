from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STORE_ORDER = [
    "加拿大一店",
    "加拿大二店",
    "加拿大三店",
    "加拿大四店",
    "加拿大五店",
    "加拿大六店",
    "加拿大七店",
    "加拿大八店",
    "加拿大九店",
]

ALIAS_MAP = {
    "Hi Bowl": "Hi Bowl",
    "HI BOWL": "Hi Bowl",
    "HIBOWL": "Hi Bowl",
    "9452100001": "Hi Bowl",
    "9452201002": "Hi Bowl",
    "加拿大大嗨麻辣烫一店": "Hi Bowl",
    "加拿大大嗨麻辣烫一店(9452)": "Hi Bowl",
    "加拿大大嗨麻辣烫一店（9452）": "Hi Bowl",
    "加拿大一店": "加拿大一店",
    "加拿大一店夜市项目": "加拿大一店",
    "加拿大海底捞-大嗨麻辣烫一店(9451)": "加拿大一店",
    "温哥华二店": "加拿大二店",
    "加拿大二店销售公共组": "加拿大二店",
    "多伦多一店": "加拿大三店",
    "加拿大三店销售公共组": "加拿大三店",
    "多伦多二店": "加拿大四店",
    "加拿大四店销售公共组": "加拿大四店",
    "多伦多三店": "加拿大五店",
    "加拿大五店销售公共组": "加拿大五店",
    "加拿大六店": "加拿大六店",
    "加拿大六店销售公共组": "加拿大六店",
    "加拿大七店": "加拿大七店",
    "加拿大七店销售公共组": "加拿大七店",
    "加拿大麻辣烫-七店(9451)": "加拿大七店",
    "加拿大八店": "加拿大八店",
    "加拿大八店销售公共组": "加拿大八店",
    "加拿大九店": "加拿大九店",
    "加拿大九店销售公共组": "加拿大九店",
}

EXCLUDED_COST_CENTERS = {"加拿大六店维修项目"}

ROWS = [
    (1, "一、主营业务收入", "agg"),
    (2, "1、销售净收入", "leaf"),
    (3, "2、仓储服务收入", "leaf"),
    (4, "二、主营业务成本（减）", "agg"),
    (5, "1、产品销售成本", "leaf"),
    (6, "2、仓储服务成本", "leaf"),
    (7, "三、毛利率", "agg"),
    (8, "1、产品销售毛利率", "agg"),
    (9, "2、仓储服务毛利率", "agg"),
    (10, "四、费用（减）", "agg"),
    (11, "1、人工成本总额", "agg"),
    (12, "1.1、工资金额", "agg"),
    (13, "正式工工资", "leaf"),
    (14, "钟点工工资", "leaf"),
    (15, "财务补计提工资", "leaf"),
    (16, "年终奖", "leaf"),
    (17, "1.2、劳务外包", "leafmap"),
    (18, "1.3、福利费", "leafmap"),
    (19, "1.4、劳动保护费", "leafmap"),
    (20, "1.5、员工社保费", "leafmap"),
    (21, "1.6、员工餐费用", "leafmap"),
    (22, "1.7、工会经费", "leafmap"),
    (23, "1.8、宿舍费用", "agg"),
    (24, "其中：房租", "leaf"),
    (25, "水电气费", "leaf"),
    (26, "1.9、组织员工活动", "leafmap"),
    (27, "2、常规费用", "agg"),
    (28, "2.1、通讯费", "leafmap"),
    (29, "2.2、差旅费", "leafmap"),
    (30, "2.3、物料消耗", "leafmap"),
    (31, "2.4、赠送顾客费用", "leafmap"),
    (32, "2.5、市场费用", "leafmap"),
    (33, "2.6、运杂费", "leafmap"),
    (34, "2.7、仓储服务费", "leafmap"),
    (35, "2.8、燃料费", "leafmap"),
    (36, "2.9、水费", "leafmap"),
    (37, "2.10、电费", "leafmap"),
    (38, "2.11、修理费", "leafmap"),
    (39, "2.12、低值易耗品", "leafmap"),
    (40, "2.13、日常办公费", "leafmap"),
    (41, "2.14、业务招待费", "leafmap"),
    (42, "2.15、车辆费用", "leafmap"),
    (43, "2.16、其他费用", "leafmap"),
    (44, "2.17、海捞送支付门店管理费", "leafmap"),
    (45, "3、固定费用", "agg"),
    (46, "3.1、税金及附加", "leafmap"),
    (47, "3.2、财产保险费", "leafmap"),
    (48, "3.3、资产折旧费", "leafmap"),
    (49, "3.4、店面租赁费", "leafmap"),
    (50, "3.5、装修费摊销", "leafmap"),
    (51, "3.6、咨询服务费", "leafmap"),
    (52, "3.7、资产减值损失", "leafmap"),
    (53, "4、财务费用", "agg"),
    (54, "4.1、利息收入", "leafmap"),
    (55, "4.2、利息支出", "leafmap"),
    (56, "4.3、手续费", "leafmap"),
    (57, "4.4、汇兑损益", "leafmap"),
    (58, "五、营业利润合计", "agg"),
    (59, "六、其他业务收益（加）", "agg"),
    (60, "1、其他业务收入", "leaf"),
    (61, "2、其他业务成本", "leaf"),
    (62, "七、营业外收益（加）", "agg"),
    (63, "1、营业外收入", "leaf"),
    (64, "2、营业外支出", "leaf"),
    (65, "八、投资收益（加）", "agg"),
    (66, "1、公允价值变动收益", "leafmap"),
    (67, "2、投资收益", "leaf"),
    (68, '九、利润总额(亏损以"-"号表示)', "agg"),
    (69, "十、所得税费用", "leaf"),
    (70, '十一、净利润(亏损以"-"号表示)', "agg"),
]

MAP_ITEM = {
    "1.2、劳务外包": "12、劳务外包",
    "1.3、福利费": "13、福利费",
    "1.4、劳动保护费": "14、劳动保护费",
    "1.5、员工社保费": "15、员工社保费",
    "1.6、员工餐费用": "16、员工餐费用",
    "1.7、工会经费": "17、工会经费",
    "1.9、组织员工活动": "19、组织员工活动",
    "2.1、通讯费": "21、通讯费",
    "2.2、差旅费": "22、差旅费",
    "2.3、物料消耗": "23、物料消耗",
    "2.4、赠送顾客费用": "24、赠送顾客费用",
    "2.5、市场费用": "25、市场费用",
    "2.6、运杂费": "26、运杂费",
    "2.7、仓储服务费": "27、仓储服务费",
    "2.8、燃料费": "28、燃料费",
    "2.9、水费": "29、水费",
    "2.10、电费": "210、电费",
    "2.11、修理费": "211、修理费",
    "2.12、低值易耗品": "212、低值易耗品",
    "2.13、日常办公费": "213、日常办公费",
    "2.14、业务招待费": "214、业务招待费",
    "2.15、车辆费用": "215、车辆费用",
    "2.16、其他费用": "216、其他费用",
    "2.17、海捞送支付门店管理费": "217、海捞送支付门店管理费",
    "3.1、税金及附加": "31、税金及附加",
    "3.2、财产保险费": "32、财产保险费",
    "3.3、资产折旧费": "33、资产折旧费",
    "3.4、店面租赁费": "34、店面租赁费",
    "3.5、装修费摊销": "35、装修费摊销",
    "3.6、咨询服务费": "36、咨询服务费",
    "3.7、资产减值损失": "37、资产减值损失",
    "4.1、利息收入": "41、利息收入",
    "4.2、利息支出": "42、利息支出",
    "4.3、手续费": "43、手续费",
    "4.4、汇兑损益": "44、汇兑损益",
    "1、公允价值变动收益": "1、公允价值变动损益",
}


def map_store(*texts: object) -> str | None:
    for text in texts:
        if text is None:
            continue
        value = str(text).strip()
        if value in ALIAS_MAP:
            return ALIAS_MAP[value]
    for text in texts:
        if text is None:
            continue
        value = str(text).strip()
        for alias, store in ALIAS_MAP.items():
            if alias in value:
                return store
    return None


def load_account_map(mapping_path: Path) -> dict[str, tuple[str, str, float]]:
    wb = load_workbook(mapping_path, data_only=True, read_only=True)
    ws = wb["损益科目对照"]
    account_map: dict[str, tuple[str, str, float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[2] is None:
            continue
        account_map[str(row[0]).zfill(8)] = (
            str(row[2]).strip(),
            str(row[3]).strip(),
            float(row[4]),
        )
    wb.close()
    return account_map


def aggregate_direct_values(source_path: Path, account_map: dict[str, tuple[str, str, float]]):
    wb = load_workbook(source_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    direct = {store: defaultdict(float) for store in STORE_ORDER}
    excluded = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cost_element, amount = row[1], row[2]
        profit_desc, cost_desc = row[4], row[6]
        if cost_element is None or amount is None:
            continue
        code8 = str(cost_element).zfill(10)[-8:]
        mapped = account_map.get(code8)
        if not mapped:
            continue
        report_item, data_source, sign = mapped
        cost_desc_text = str(cost_desc or "").strip()
        if cost_desc_text in EXCLUDED_COST_CENTERS:
            excluded.append((cost_desc_text, code8, report_item, float(amount)))
            continue
        store = map_store(profit_desc) if data_source == "利润中心" else map_store(cost_desc, profit_desc)
        if store in STORE_ORDER:
            direct[store][report_item] += float(amount) * sign
    wb.close()
    return direct, excluded


def calculate_result(direct: dict[str, defaultdict[str, float]]) -> dict[str, dict[str, float]]:
    result = {store: {} for store in STORE_ORDER}
    for store in STORE_ORDER:
        values = result[store]
        for _, item, kind in ROWS:
            if kind == "leaf":
                values[item] = direct[store].get(item, 0.0)
            elif kind == "leafmap":
                values[item] = direct[store].get(MAP_ITEM[item], 0.0)

        values["一、主营业务收入"] = values["1、销售净收入"] + values["2、仓储服务收入"]
        values["二、主营业务成本（减）"] = values["1、产品销售成本"] + values["2、仓储服务成本"]
        values["三、毛利率"] = 0.0 if values["一、主营业务收入"] == 0 else 1 - values["二、主营业务成本（减）"] / values["一、主营业务收入"]
        values["1、产品销售毛利率"] = 0.0 if values["1、销售净收入"] == 0 else 1 - values["1、产品销售成本"] / values["1、销售净收入"]
        values["2、仓储服务毛利率"] = 0.0 if values["2、仓储服务收入"] == 0 else 1 - values["2、仓储服务成本"] / values["2、仓储服务收入"]
        values["1.1、工资金额"] = values["正式工工资"] + values["钟点工工资"] + values["财务补计提工资"] + values["年终奖"]
        values["1.8、宿舍费用"] = values["其中：房租"] + values["水电气费"]
        values["1、人工成本总额"] = (
            values["1.1、工资金额"] + values["1.2、劳务外包"] + values["1.3、福利费"] + values["1.4、劳动保护费"]
            + values["1.5、员工社保费"] + values["1.6、员工餐费用"] + values["1.7、工会经费"] + values["1.8、宿舍费用"]
            + values["1.9、组织员工活动"]
        )
        values["2、常规费用"] = (
            values["2.1、通讯费"] + values["2.2、差旅费"] + values["2.3、物料消耗"] + values["2.4、赠送顾客费用"]
            + values["2.5、市场费用"] + values["2.6、运杂费"] + values["2.7、仓储服务费"] + values["2.8、燃料费"]
            + values["2.9、水费"] + values["2.10、电费"] + values["2.11、修理费"] + values["2.12、低值易耗品"]
            + values["2.13、日常办公费"] + values["2.14、业务招待费"] + values["2.15、车辆费用"]
            + values["2.16、其他费用"] + values["2.17、海捞送支付门店管理费"]
        )
        values["3、固定费用"] = (
            values["3.1、税金及附加"] + values["3.2、财产保险费"] + values["3.3、资产折旧费"] + values["3.4、店面租赁费"]
            + values["3.5、装修费摊销"] + values["3.6、咨询服务费"] + values["3.7、资产减值损失"]
        )
        values["4、财务费用"] = values["4.1、利息收入"] + values["4.2、利息支出"] + values["4.3、手续费"] + values["4.4、汇兑损益"]
        values["四、费用（减）"] = values["1、人工成本总额"] + values["2、常规费用"] + values["3、固定费用"] + values["4、财务费用"]
        gross_profit_amount = values["一、主营业务收入"] - values["二、主营业务成本（减）"]
        values["五、营业利润合计"] = gross_profit_amount - values["四、费用（减）"]
        values["六、其他业务收益（加）"] = values["1、其他业务收入"] - values["2、其他业务成本"]
        values["七、营业外收益（加）"] = values["1、营业外收入"] - values["2、营业外支出"]
        values["八、投资收益（加）"] = values["1、公允价值变动收益"] + values["2、投资收益"]
        values['九、利润总额(亏损以"-"号表示)'] = values["五、营业利润合计"] + values["六、其他业务收益（加）"] + values["七、营业外收益（加）"] + values["八、投资收益（加）"]
        values["十、所得税费用"] = (
            0.0
            if store == "Hi Bowl"
            else (
                values['九、利润总额(亏损以"-"号表示)'] * 0.27
                if values['九、利润总额(亏损以"-"号表示)'] > 0
                else 0.0
            )
        )
        values['十一、净利润(亏损以"-"号表示)'] = values['九、利润总额(亏损以"-"号表示)'] - values["十、所得税费用"]
    return result


def write_workbook(output_path: Path, result: dict[str, dict[str, float]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "损益表"

    headers = ["序号", "项目"] + STORE_ORDER
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="FCE4D6")
    subtotal_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)

    for i, header in enumerate(headers, start=1):
        cell = ws.cell(1, i, header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (seq, item, kind) in enumerate(ROWS, start=2):
        ws.cell(row_idx, 1, seq)
        ws.cell(row_idx, 2, item)
        for col_idx, store in enumerate(STORE_ORDER, start=3):
            cell = ws.cell(row_idx, col_idx, result[store].get(item, 0.0))
            if item in {"三、毛利率", "1、产品销售毛利率", "2、仓储服务毛利率"}:
                cell.number_format = "0.00%"
            else:
                cell.number_format = "#,##0.00;[Red]-#,##0.00"

        fill = None
        if item.startswith(("一、", "二、", "三、", "四、", "五、", "六、", "七、", "八、", "九、", "十、", "十一、")):
            fill = section_fill
        elif kind == "agg":
            fill = subtotal_fill
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row_idx, col_idx).fill = fill
                ws.cell(row_idx, col_idx).font = bold

    ws.freeze_panes = "C2"
    widths = {1: 8, 2: 30}
    for idx in range(3, len(headers) + 1):
        widths[idx] = 16
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    note = wb.create_sheet("门店映射说明")
    note.append(["原始名称", "归集门店"])
    note.cell(1, 1).font = bold
    note.cell(1, 2).font = bold
    note.cell(1, 1).fill = header_fill
    note.cell(1, 2).fill = header_fill
    for alias, store in ALIAS_MAP.items():
        note.append([alias, store])
    note.append(["[排除] 加拿大六店维修项目", "不归集"])
    note.column_dimensions["A"].width = 38
    note.column_dimensions["B"].width = 16

    wb.save(output_path)


def generate_canada_pnl(source_path: Path, mapping_path: Path, output_path: Path) -> tuple[Path, int]:
    account_map = load_account_map(mapping_path)
    direct, excluded = aggregate_direct_values(source_path, account_map)
    result = calculate_result(direct)
    write_workbook(output_path, result)
    return output_path, len(excluded)
