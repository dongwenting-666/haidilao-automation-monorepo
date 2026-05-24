"""Build the 基础数据 sheet of the 毛利相关分析指标 workbook.

The 基础数据 sheet is a 126-column monthly P&L + operations table per
store, since 2018. It is the foundation of the 毛利率 derivative sheets
(细分毛利率表 / 毛利率连续对比表 / 毛利率环比 / 毛利率同比).

Column layout (1-indexed):
  1–6   Year/month/period/department identifiers
  7–76  P&L (70 line items) — sourced from canada_pnl.ROWS via ZFI0049
  77–78 (重新计算) tax + net-profit (alternate tax basis)
  79–85 Derived metrics (audit adjustment, functional fees, 火锅经营净利润, cash flow)
  86–92 Static per-store metadata — from static_meta.STORE_META
  93–126 Operations metrics — from QBI / daily-store-operation-report data
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from zfi0049_report.canada_pnl import ROWS
from zfi0049_report.static_meta import STORE_META

# 70 P&L line items, in the column order they appear (cols 7–76).
PNL_ITEMS: list[str] = [item for _, item, _ in ROWS]
assert len(PNL_ITEMS) == 70, f"Expected 70 P&L rows, got {len(PNL_ITEMS)}"


# Headers for cols 7–76 are the PNL_ITEMS values exactly.
# Headers for cols 77–85 (derived):
DERIVED_HEADERS = [
    '(重新计算)十、所得税费用',
    '(重新计算)十一、净利润(亏损以"-"号表示)',
    '审计调整',
    '职能费用',
    '火锅经营净利润',
    '经营性现金流',
    '火锅经营净利润（所得税前）',
    '火锅经营净利润率（所得税前）',
    '经营性现金流（所得税前）',
]

# Headers for cols 86–92 (static metadata):
META_HEADERS = ['地区', '国家', '城市', '开业日期',
                '门店级别', '营业状态', '门店分类']

# Headers for cols 93–126 (operations metrics):
OPS_HEADERS = [
    '营业桌数',                       # 93
    '营业桌数(考核)',                  # 94
    '所有餐位数',                      # 95
    '营业天数',                        # 96
    '08:00-13:59日均桌数',             # 97
    '14:00-16:59日均桌数',             # 98
    '17:00-21:59日均桌数',             # 99
    '22:00-(次)07:59日均桌数',         # 100
    '工作日日均桌数',                  # 101
    '节假日日均桌数',                  # 102
    '全月所有餐位数',                  # 103
    '翻台率(整体)',                    # 104
    '翻台率(考核)',                    # 105
    '全月座位数（新业态）',            # 106
    '翻坐率（新业态）',                # 107
    '人均消费(不含税优惠后)',          # 108
    '人均消费(含税优惠前)',            # 109
    '单桌消费(不含税优惠后)',          # 110
    '单桌消费(含税优惠前)',            # 111
    '打折、抹零、赠菜金额(不含税)',    # 112
    '免单金额(不含税)',                # 113
    '优惠总金额（不含税）',            # 114
    '优惠总金额（含税）',              # 115
    '合同面积',                        # 116
    '用餐人数',                        # 117
    '经营月份',                        # 118
    '营业收入(堂吃)(不含税)',          # 119
    '营业收入(堂吃)(含税)',            # 120
    '店经理',                          # 121
    '区域经理',                        # 122
    '餐位数（披露）',                  # 123
    '翻台率（披露）',                  # 124
    '门店名称',                        # 125 — duplicate of col 6's 三级部门
    '利润档位',                        # 126
]

# Full header row (126 columns).
HEADERS: list[str] = (
    ['年份', '月份', '期间', '一级部门', '二级部门', '三级部门']
    + PNL_ITEMS
    + DERIVED_HEADERS
    + META_HEADERS
    + OPS_HEADERS
)
assert len(HEADERS) == 126, f"Expected 126 cols, got {len(HEADERS)}"


@dataclass
class StoreMonthRecord:
    """One row in 基础数据 — covers a single store for a single month."""

    store: str               # e.g. '加拿大一店'
    year: int                # e.g. 2026
    month: int               # 1–12
    period_serial: int       # Excel date serial of month-end (e.g. 46082)

    # P&L: maps each item in PNL_ITEMS → float. Missing keys default to 0.
    pnl: dict[str, float] = field(default_factory=dict)

    # Derived (cols 77–85). Caller can override; otherwise computed from pnl.
    audit_adjustment: float = 0.0   # 审计调整 (col 79)
    functional_fees: float = 0.0    # 职能费用 (col 80)

    # Ops metrics — caller fills from QBI / daily-report. Missing keys → blank.
    ops: dict[str, Any] = field(default_factory=dict)


# ── Helpers ──────────────────────────────────────────────────────────────────


def _excel_date_serial(d: date) -> int:
    """Excel's 1900-based date serial. Month-end date for monthly P&L rows."""
    epoch = date(1899, 12, 30)  # Excel quirk: treats 1900 as leap (it isn't).
    return (d - epoch).days


def _month_end(year: int, month: int) -> date:
    import calendar
    last = calendar.monthrange(year, month)[1]
    return date(year, month, last)


def compute_derived(record: StoreMonthRecord) -> dict[str, float]:
    """Compute cols 77–85 from the P&L + audit/functional-fee inputs.

    Formulas derived empirically from the manual workbook (March 2026):
      - 火锅经营净利润 = 利润总额 - 审计调整 - 职能费用
      - 火锅经营净利润率 = 火锅经营净利润 / 主营业务收入
      - 经营性现金流 = TODO (need non-cash add-backs; treat as 0 until verified)
      - 重新计算 tax/profit: identical to original since current 所得税费用 = 0
        for all 加拿大 stores (Hi Bowl is the only one with tax in the existing
        canada_pnl logic, and we don't include it here).
    """
    rev = record.pnl.get("一、主营业务收入", 0.0)
    profit_total = record.pnl.get('九、利润总额(亏损以"-"号表示)', 0.0)
    net_profit_orig = record.pnl.get('十一、净利润(亏损以"-"号表示)', 0.0)
    tax_orig = record.pnl.get("十、所得税费用", 0.0)

    # 重新计算 — for now mirror originals; refine when we know the alt-tax rule.
    tax_recalc = tax_orig
    net_recalc = net_profit_orig

    hotpot_op_profit = profit_total - record.audit_adjustment - record.functional_fees
    hotpot_op_pretax = hotpot_op_profit  # tax == 0, so pre/post-tax identical
    margin_pretax = (hotpot_op_profit / rev) if rev else 0.0

    # TODO(maoli-report): cash-flow formulas — need 资产折旧费 + 装修费摊销
    # add-back logic confirmed against multiple stores before enabling.
    op_cash_flow = 0.0
    op_cash_flow_pretax = 0.0

    return {
        '(重新计算)十、所得税费用': tax_recalc,
        '(重新计算)十一、净利润(亏损以"-"号表示)': net_recalc,
        '审计调整': record.audit_adjustment,
        '职能费用': record.functional_fees,
        '火锅经营净利润': hotpot_op_profit,
        '经营性现金流': op_cash_flow,
        '火锅经营净利润（所得税前）': hotpot_op_pretax,
        '火锅经营净利润率（所得税前）': margin_pretax,
        '经营性现金流（所得税前）': op_cash_flow_pretax,
    }


def build_row(record: StoreMonthRecord) -> list[Any]:
    """Build a single 126-column row for the 基础数据 sheet."""
    meta = STORE_META.get(record.store)

    # Cols 1–6
    row: list[Any] = [
        f"{record.year}年",
        f"{record.month}月",
        record.period_serial,
        "海外门店",  # 一级部门
        "海外门店",  # 二级部门
        record.store,  # 三级部门
    ]

    # Cols 7–76: P&L
    row.extend(record.pnl.get(item, 0.0) for item in PNL_ITEMS)

    # Cols 77–85: derived
    derived = compute_derived(record)
    row.extend(derived[h] for h in DERIVED_HEADERS)

    # Cols 86–92: static metadata
    if meta is None:
        row.extend([None] * len(META_HEADERS))
    else:
        row.extend([
            meta.region, meta.country, meta.city,
            meta.open_date, meta.level, meta.status, meta.classification,
        ])

    # Cols 93–126: ops metrics (blank where unknown)
    for h in OPS_HEADERS:
        if h == '门店名称':  # col 125 dup of 三级部门
            row.append(record.store)
        else:
            row.append(record.ops.get(h))

    assert len(row) == 126, f"Expected 126, got {len(row)}"
    return row


def write_basic_data_sheet(
    ws: Worksheet,
    records: list[StoreMonthRecord],
    *,
    include_header: bool = True,
) -> int:
    """Write records to ``ws``. Returns the number of data rows written."""
    if include_header:
        ws.append(HEADERS)

    written = 0
    for rec in sorted(records, key=lambda r: (r.year, r.month, r.store)):
        ws.append(build_row(rec))
        written += 1

    # Light formatting on the header row + freeze panes.
    if include_header:
        from openpyxl.styles import Font, PatternFill
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="D9EAF7")
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(1, col_idx)
            cell.font = bold
            cell.fill = fill
        ws.freeze_panes = "G2"
        # First few cols narrow, the rest wide enough for numbers.
        widths = {1: 8, 2: 6, 3: 10, 4: 12, 5: 12, 6: 14}
        for i, w in widths.items():
            ws.column_dimensions[get_column_letter(i)].width = w
        for i in range(7, len(HEADERS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    return written
