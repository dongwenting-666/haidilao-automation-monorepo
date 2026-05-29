"""Orchestrator for the 毛利相关分析指标 workbook (9 sheets).

Pulls together the per-sheet builders into a single workbook generator:
  1. 填写说明 — static instructions
  2. 细分毛利率表 (2) — per-category gross margin breakdown
  3. 毛利率连续对比表 — 7-month rolling trend per store
  4. 毛利率环比 — MoM decomposition
  5. 毛利率同比 — YoY decomposition
  6. 表1-菜品价格变动及菜品损耗表 — per dish×spec×material
  7. 表2-原材料成本变动表 — per material per store
  8. 表3-打折优惠表 — per-store discount summary
  9. 基础数据 — 126-col P&L + ops base

Inputs are accepted as already-loaded data structures (dicts / lists)
so the same orchestrator works whether the upstream data came from live
SAP automation or from on-disk archive files. The CLI wrapper
(``gross_margin_analysis_main.py``) handles the I/O.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from zfi0049_report.basic_data import (
    StoreMonthRecord,
    write_basic_data_sheet,
)
from zfi0049_report.derivative_sheets import (
    MomRow,
    build_mom_rows,
    build_mom_sheet,
    build_trend_rows,
    build_trend_sheet,
    build_yoy_sheet,
)
from zfi0049_report.static_meta import STORE_META
from zfi0049_report.subdivided_gp_sheet import (
    build_subdivided_gp_sheet,
    compute_category_gp,
)
from zfi0049_report.table1_dish import (
    PosSale,
    Table1Row,
    WERKS_TO_STORE,
    apply_canonical_blanking,
    build_rows_for_store,
    compute_loss_impact,
    compute_revenue_impact,
    enrich_with_materials,
    enrich_with_prices,
    enrich_with_set_meals,
    load_mb5b_prices,
    load_pos_prices,
    load_pos_sales,
    load_pos_set_sales,
    load_zfi0156,
    write_sheet as write_table1_sheet,
)
from zfi0049_report.table2_material import (
    build_rows as build_table2_rows,
    write_sheet as write_table2_sheet,
)
from zfi0049_report.table3_discount import (
    build_rows as build_table3_rows,
    write_sheet as write_table3_sheet,
)

logger = logging.getLogger(__name__)


@dataclass
class GrossMarginInputs:
    """Bundle of all data needed to build the workbook.

    Most fields are optional — the orchestrator gracefully skips or stubs
    sheets when data is missing, so partial pipelines (e.g. dev with only
    one month of POS) still produce a viewable workbook.
    """

    year: int
    month: int

    # Per-store P&L for current month: store → P&L dict
    cur_pnl: dict[str, dict[str, float]] = field(default_factory=dict)
    prev_pnl: dict[str, dict[str, float]] = field(default_factory=dict)
    yoy_pnl: dict[str, dict[str, float]] = field(default_factory=dict)

    # 7-month rolling gross margin per store (most-recent-first)
    monthly_gp: dict[str, list[float]] = field(default_factory=dict)

    # POS sales per store: store_name → list of PosSale
    pos_sales: dict[str, list[PosSale]] = field(default_factory=dict)
    # Prev/YoY POS sales — used to build prev/yoy table1 rows so the
    # 细分毛利率表 prev/YoY GM columns can be computed.
    pos_prev_sales: dict[str, list[PosSale]] = field(default_factory=dict)
    pos_yoy_sales: dict[str, list[PosSale]] = field(default_factory=dict)
    # POS 菜品套餐汇总 — {(store, dish_code, spec) → Σ 应收数量}.
    pos_set_qty: dict[tuple[str, int, str], float] = field(default_factory=dict)

    # POS prices per store (current/prev/YoY): store → {(dish,spec) → (price, unit)}
    pos_prices_cur: dict[str, dict] = field(default_factory=dict)
    pos_prices_prev: dict[str, dict] = field(default_factory=dict)
    pos_prices_yoy: dict[str, dict] = field(default_factory=dict)

    # store_bom: store → list of recipe dicts (from inventory_check.db_bom)
    bom_rows: dict[str, list[dict]] = field(default_factory=dict)

    # Pre-built prev-month Table1Row list — only needed if the caller wants
    # 细分毛利率表 to show the 环比 column. Built the same way as cur table1
    # rows but using prev-month POS sales + prev-month material prices.
    prev_table1_rows: list = field(default_factory=list)

    # Full historical 基础数据 rows (one StoreMonthRecord per store-month).
    # When provided, the 基础数据 sheet mirrors these records 1:1 instead
    # of being built from cur_pnl alone — preserves the historical archive
    # (since 2018) that the manual workbook carries.
    basic_data_records: list = field(default_factory=list)

    # ZFI0156 + MB5B by werks: (werks, matnr) → value
    zfi_cur: dict = field(default_factory=dict)
    zfi_prev: dict = field(default_factory=dict)
    zfi_yoy: dict = field(default_factory=dict)
    mb5b_cur: dict = field(default_factory=dict)
    mb5b_prev: dict = field(default_factory=dict)
    mb5b_yoy: dict = field(default_factory=dict)


def _build_table1(inputs: GrossMarginInputs) -> list[Table1Row]:
    """Build the full 表1 row list across all stores."""
    all_rows: list[Table1Row] = []
    werks_by_store = {v: k for k, v in WERKS_TO_STORE.items()}
    for store, bom in inputs.bom_rows.items():
        sales = inputs.pos_sales.get(store, [])
        # Only emit recipe rows for dishes actually sold in this store —
        # matches the manual 表1 which is POS-driven per store.
        rows = build_rows_for_store(
            store, pos_sales=sales, bom_rows=bom, sold_only=bool(sales),
        )
        werks = werks_by_store.get(store)
        if werks:
            enrich_with_materials(
                rows, werks=werks,
                zfi=inputs.zfi_cur, mb5b_prices=inputs.mb5b_cur,
            )
        enrich_with_prices(
            rows,
            cur_prices=inputs.pos_prices_cur.get(store, {}),
            prev_prices=inputs.pos_prices_prev.get(store, {}),
            yoy_prices=inputs.pos_prices_yoy.get(store, {}),
        )
        compute_revenue_impact(rows)
        all_rows.extend(rows)

    # 套餐 cols 22-23 — populate from POS 菜品套餐汇总 aggregate.
    if inputs.pos_set_qty:
        enrich_with_set_meals(all_rows, inputs.pos_set_qty)
    # Loss-impact cols 32-36 — needs the full row list to compute per-
    # (store, material) aggregates. Run AFTER all stores are accumulated
    # so prev / yoy lookups work across the whole region. Canonical
    # blanking happens last so non-canonical rows are cleared at the end.
    prev_rows = inputs.prev_table1_rows or _build_prev_table1(inputs)
    compute_loss_impact(all_rows, prev_rows=prev_rows)
    apply_canonical_blanking(all_rows)
    return all_rows


def _build_prev_table1(inputs: GrossMarginInputs) -> list[Table1Row]:
    """Build a prev-month table1 row list — same shape as cur but using
    ``pos_prev_sales`` for sales and the same store_bom for recipes.

    Material prices fall back to ``mb5b_cur`` when ``mb5b_prev`` is empty
    (price drift month-over-month is small enough that the prev-GM
    figures stay representative). When neither MB5B export is available
    the per-row cost stays None and the resulting 细分毛利率表 prev cells
    blank out as None rather than producing a misleading zero.
    """
    if not inputs.pos_prev_sales:
        return []
    all_rows: list[Table1Row] = []
    werks_by_store = {v: k for k, v in WERKS_TO_STORE.items()}
    mb5b_for_prev = inputs.mb5b_prev or inputs.mb5b_cur
    zfi_for_prev = inputs.zfi_prev or inputs.zfi_cur
    for store, bom in inputs.bom_rows.items():
        sales = inputs.pos_prev_sales.get(store, [])
        if not sales:
            continue
        rows = build_rows_for_store(
            store, pos_sales=sales, bom_rows=bom, sold_only=True,
        )
        werks = werks_by_store.get(store)
        if werks:
            enrich_with_materials(
                rows, werks=werks,
                zfi=zfi_for_prev, mb5b_prices=mb5b_for_prev,
            )
        enrich_with_prices(
            rows,
            cur_prices=inputs.pos_prices_prev.get(store, {}),
        )
        all_rows.extend(rows)
    return all_rows


def _build_store_month_records(inputs: GrossMarginInputs) -> list[StoreMonthRecord]:
    """Build the cur-month 基础数据 row(s) — one per store with P&L."""
    out: list[StoreMonthRecord] = []
    # Excel month-end serial. 1900 epoch quirk handled by basic_data._excel_date_serial.
    from zfi0049_report.basic_data import _excel_date_serial, _month_end
    period = _excel_date_serial(_month_end(inputs.year, inputs.month))
    for store, pnl in inputs.cur_pnl.items():
        out.append(StoreMonthRecord(
            store=store, year=inputs.year, month=inputs.month,
            period_serial=period, pnl=pnl,
        ))
    # Add zero-record rows for stores with metadata but no P&L this month
    # (so the basic_data sheet has consistent store coverage).
    covered = {r.store for r in out}
    for store in STORE_META:
        if store in covered:
            continue
        out.append(StoreMonthRecord(
            store=store, year=inputs.year, month=inputs.month,
            period_serial=period,
        ))
    return out


def build_workbook(inputs: GrossMarginInputs, out_path: Path,
                   *, style_template: Path | None = None) -> Path:
    """Build the full 毛利相关分析指标 workbook → out_path.

    When ``style_template`` is given, copy that workbook and overwrite
    only the data cells in place — preserving the manual's exact styling
    (fonts, fills, borders, merged cells, conditional formatting, column
    widths). Otherwise build all sheets from scratch (unstyled).
    """
    if style_template is not None:
        return _build_workbook_from_template(inputs, out_path, style_template)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: instructions (static text — verbatim from the manual template).
    ws_inst = wb.create_sheet("填写说明")
    ws_inst.append(["填表说明：涉及金额的填写为本币金额"])
    ws_inst.append(["1、先更新《表1-菜品价格变动及菜品损耗表》、《表2-原材料成本变动表》、《表3-打折优惠表》"])
    ws_inst.append(["2、其次填写《细分毛利率表》、《毛利率连续对比表》"])
    ws_inst.append(["3、再填写《毛利率环比》和《毛利率同比》中贴数部分"])
    ws_inst.append(["4、通过本表数据对比分析后，需识别毛利率相关分析问题在经营分析报告上描述即可（描述内容需包含如下图示例中问题类型，内容建议等）"])
    ws_inst.append([])
    ws_inst.append([f"{inputs.year}年{inputs.month}月毛利率相关问题"])
    ws_inst.append(["序号", "问题内容", "问题描述及建议"])
    ws_inst.append([
        1, "毛利率环比下降异常",
        "(1) 问题描述：本月XX店毛利率环比下降超过3%，且毛利率低于60%；\n"
        "(2) 原因：如：本月库存盘点有误，影响成本虚增3万元，影响毛利率下降0.5%；"
        "原材料成本环比增加8万元，影响毛利率下降XXX；",
    ])
    ws_inst.append([
        2, "低毛利低点击率产品",
        "（1）问题描述：通过数据发现XXX店和XXX店销售的牛蛙（现杀）属于负毛利、低点击率产品"
        "（毛利率为-7.8%，点击率在1.5%左右），反馈至门店以及大区，由门店及大区评估是否下架或"
        "通过其他如调价等措施提升毛利率\n"
        "（2）建议：反馈至门店以及大区，由门店及大区评估是否下架或通过其他如调价等措施提升毛利率",
    ])
    ws_inst.append([
        3, "单品锅底毛利率异常",
        "(1) 问题描述：XXX店的番茄锅底及白玉锅底的毛利率较低损耗较大，该店番茄及白玉锅底"
        "全月损耗较片区平均水平高566公斤，影响成本上升1.2万人民币，影响锅底毛利率下降2%；",
    ])
    ws_inst.append([
        4, "酒水毛利率中单品毛利异常",
        "(1) 问题描述：因盘点不准导致的毛利率异常问题，尤其贵重酒水盘差影响尤为明显，"
        "如8月XXX店酒水毛利率环比7月下降16.7%，经复核发现门店飞天茅台漏盘，影响酒水毛利率下降异常；\n"
        "(2) 原因：主要是门店库存盘点不准确，存在漏盘的问题",
    ])

    # Build 表1/2/3 first so derivative sheets can consume them.
    table1_rows = _build_table1(inputs)
    prev_table1_rows = (
        inputs.prev_table1_rows or _build_prev_table1(inputs)
    )

    # Sheet 2: 细分毛利率表 (2) — per-category gross margin breakdown.
    # Categories come from POS 大类 via dish_category.map_pos_to_report_category.
    cur_gp = compute_category_gp(table1_rows)
    prev_gp = (compute_category_gp(prev_table1_rows)
               if prev_table1_rows else {})
    build_subdivided_gp_sheet(
        wb, cur_gp=cur_gp, prev_gp=prev_gp,
        year=inputs.year, month=inputs.month,
    )

    # Sheet 3: 毛利率连续对比表
    trend_rows = build_trend_rows(monthly_gp=inputs.monthly_gp)
    build_trend_sheet(wb, trend_rows)
    table2_rows = build_table2_rows(
        zfi_cur=inputs.zfi_cur,
        mb5b_cur=inputs.mb5b_cur,
        mb5b_prev=inputs.mb5b_prev,
        mb5b_yoy=inputs.mb5b_yoy,
    )
    table3_rows = build_table3_rows(
        cur_pnl=inputs.cur_pnl,
        prev_pnl=inputs.prev_pnl,
        yoy_pnl=inputs.yoy_pnl,
    )

    # Sheet 4: 毛利率环比
    mom_rows = build_mom_rows(
        cur_pnl=inputs.cur_pnl, prev_pnl=inputs.prev_pnl,
        table1_rows=table1_rows, table2_rows=table2_rows,
        table3_rows=table3_rows,
    )
    build_mom_sheet(wb, mom_rows)

    # Sheet 5: 毛利率同比
    yoy_mom_rows = build_mom_rows(
        cur_pnl=inputs.cur_pnl, prev_pnl=inputs.yoy_pnl,
        table1_rows=table1_rows, table2_rows=table2_rows,
        table3_rows=table3_rows,
    )
    build_yoy_sheet(wb, yoy_mom_rows)

    # Sheet 6: 表1
    ws1 = wb.create_sheet("表1-菜品价格变动及菜品损耗表 (模板) ")
    write_table1_sheet(ws1, table1_rows)

    # Sheet 7: 表2
    ws2 = wb.create_sheet("表2-原材料成本变动表")
    write_table2_sheet(ws2, table2_rows)

    # Sheet 8: 表3
    ws3 = wb.create_sheet("表3-打折优惠表")
    write_table3_sheet(ws3, table3_rows)

    # Sheet 9: 基础数据 — use the full historical record list when the
    # caller supplied one (mirrors the manual workbook's archive); otherwise
    # build a minimal cur-month record list from cur_pnl.
    ws_basic = wb.create_sheet("基础数据")
    records = (inputs.basic_data_records
               if inputs.basic_data_records
               else _build_store_month_records(inputs))
    write_basic_data_sheet(ws_basic, records)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info(
        "wrote 毛利分析 workbook → %s (table1=%d rows, table2=%d, table3=%d, mom=%d)",
        out_path, len(table1_rows), len(table2_rows),
        len(table3_rows), len(mom_rows),
    )
    return out_path


def _build_workbook_from_template(
    inputs: GrossMarginInputs, out_path: Path, style_template: Path,
) -> Path:
    """Fill a styled template copy with computed data (styles preserved).

    Strict mode: every data cell in our fill regions is cleared up-front,
    so anything present in the saved workbook is written by this
    function. Formula cells (=SUM, =F5*汇率, etc.) are preserved so
    Excel recomputes them on open. Static template constants (汇率,
    store-manager names, date serials, the 2023 stale-example block) are
    written from template_constants — so they're "from source" via this
    module rather than inherited from the template.
    """
    from datetime import date
    from zfi0049_report.template_fill import (
        fill_instructions_sheet, fill_positioned_rows, fill_specific_row,
        fill_table_sheet, open_template, safe_set,
    )
    from zfi0049_report.template_constants import (
        DEFAULT_FX_BY_PERIOD, STORE_STAFF,
        SUBDIVIDED_GP_2023_HEADERS, SUBDIVIDED_GP_2023_ROWS,
        SUBDIVIDED_GP_2026_NOTE,
    )
    from zfi0049_report.basic_data import _excel_date_serial, _month_end
    from zfi0049_report.table1_dish import to_row as t1_to_row
    from zfi0049_report.table2_material import to_row as t2_to_row
    from zfi0049_report.table3_discount import to_row as t3_to_row
    from zfi0049_report.basic_data import build_row as basic_build_row
    from zfi0049_report.derivative_sheets import mom_row_to_excel
    from zfi0049_report.dish_category import REPORT_CATEGORIES

    # ── Compute all data structures (same as the from-scratch path) ──
    table1_rows = _build_table1(inputs)
    prev_table1_rows = inputs.prev_table1_rows or _build_prev_table1(inputs)
    cur_gp = compute_category_gp(table1_rows)
    prev_gp = compute_category_gp(prev_table1_rows) if prev_table1_rows else {}
    table2_rows = build_table2_rows(
        zfi_cur=inputs.zfi_cur, mb5b_cur=inputs.mb5b_cur,
        mb5b_prev=inputs.mb5b_prev, mb5b_yoy=inputs.mb5b_yoy,
    )
    table3_rows = build_table3_rows(
        cur_pnl=inputs.cur_pnl, prev_pnl=inputs.prev_pnl, yoy_pnl=inputs.yoy_pnl,
    )
    mom_rows = build_mom_rows(
        cur_pnl=inputs.cur_pnl, prev_pnl=inputs.prev_pnl,
        table1_rows=table1_rows, table2_rows=table2_rows, table3_rows=table3_rows,
    )
    yoy_mom_rows = build_mom_rows(
        cur_pnl=inputs.cur_pnl, prev_pnl=inputs.yoy_pnl,
        table1_rows=table1_rows, table2_rows=table2_rows, table3_rows=table3_rows,
    )
    records = (inputs.basic_data_records
               if inputs.basic_data_records
               else _build_store_month_records(inputs))

    wb = open_template(style_template, out_path)

    # ── 填写说明 — overwrite from source (no inherited template text). ──
    fill_instructions_sheet(wb, year=inputs.year, month=inputs.month)

    # ── Static template constants written from source ──
    # Date serials + 汇率 in 表3.
    cur_serial = _excel_date_serial(_month_end(inputs.year, inputs.month))
    prev_y = inputs.year if inputs.month > 1 else inputs.year - 1
    prev_m = inputs.month - 1 if inputs.month > 1 else 12
    prev_serial = _excel_date_serial(_month_end(prev_y, prev_m))
    yoy_serial = _excel_date_serial(_month_end(inputs.year - 1, inputs.month))
    cur_fx = DEFAULT_FX_BY_PERIOD.get((inputs.year, inputs.month), 0.728597)
    prev_fx = DEFAULT_FX_BY_PERIOD.get((prev_y, prev_m), 0.728597)
    yoy_fx = DEFAULT_FX_BY_PERIOD.get((inputs.year - 1, inputs.month), 0.695265)
    if "表3-打折优惠表" in wb.sheetnames:
        ws3 = wb["表3-打折优惠表"]
        safe_set(ws3, 1, 3, cur_serial)
        safe_set(ws3, 2, 3, cur_fx)
        safe_set(ws3, 16, 3, prev_fx)
        safe_set(ws3, 17, 3, yoy_fx)
        safe_set(ws3, 3, 7, prev_serial)
        safe_set(ws3, 3, 11, yoy_serial)
    if "毛利率环比" in wb.sheetnames:
        safe_set(wb["毛利率环比"], 16, 3, prev_fx)
        safe_set(wb["毛利率环比"], 17, 3, cur_fx)
    if "毛利率同比" in wb.sheetnames:
        safe_set(wb["毛利率同比"], 20, 2, yoy_fx)
        safe_set(wb["毛利率同比"], 21, 2, cur_fx)
    if "细分毛利率表 (2)" in wb.sheetnames:
        wss = wb["细分毛利率表 (2)"]
        safe_set(wss, 1, 2, SUBDIVIDED_GP_2023_HEADERS[0])
        safe_set(wss, 2, 2, SUBDIVIDED_GP_2023_HEADERS[1])
        for i, row in enumerate(SUBDIVIDED_GP_2023_ROWS):
            r = 5 + i
            safe_set(wss, r, 2, row[0])
            for j, v in enumerate(row[1:], start=3):
                safe_set(wss, r, j, v)
        safe_set(wss, 16, 2, SUBDIVIDED_GP_2026_NOTE)

    # ── Flat tables: 表1 / 表2 / 基础数据 ──
    fill_table_sheet(wb, "表1-菜品价格变动及菜品损耗表 (模板) ",
                     [t1_to_row(r) for r in table1_rows], ncols=36)
    fill_table_sheet(wb, "表2-原材料成本变动表",
                     [t2_to_row(r) for r in table2_rows], ncols=29)
    fill_table_sheet(wb, "基础数据",
                     [basic_build_row(r) for r in records], ncols=126)

    # ── 表3: 8 store rows (store name in col 3) ──
    t3_by_store = {r.store: t3_to_row(r) for r in table3_rows}
    fill_positioned_rows(
        wb, "表3-打折优惠表", t3_by_store,
        store_col=3, start_row=5, end_row=12,
        value_cols=list(range(4, 16)),
    )

    # ── 毛利率环比 / 同比: store name in col 2, data rows 5-12 ──
    mom_by_store = {r.store: mom_row_to_excel(r) for r in mom_rows}
    fill_positioned_rows(
        wb, "毛利率环比", mom_by_store,
        store_col=2, start_row=5, end_row=12,
        value_cols=list(range(3, 24)),  # cols 3..23 (all computed data)
    )
    yoy_by_store = {r.store: mom_row_to_excel(r) for r in yoy_mom_rows}
    fill_positioned_rows(
        wb, "毛利率同比", yoy_by_store,
        store_col=2, start_row=5, end_row=12,
        value_cols=list(range(3, 24)),  # cols 3..23 (all computed data)
    )

    # ── 细分毛利率表 (2): 2026 block, store rows 20-27 ──
    # Build a full-row map per store: col2=store, then per-category cur/prev/环比.
    sub_by_store: dict[str, list] = {}
    stores_in_gp = {k[0] for k in cur_gp} | {k[0] for k in prev_gp}
    for store in stores_in_gp:
        row: list = [None] * 26
        for cat_idx, cat in enumerate(REPORT_CATEGORIES):
            col = 3 + cat_idx * 3  # 1-based
            cur = cur_gp.get((store, cat))
            prev = prev_gp.get((store, cat))
            cgm = cur[2] if cur and cur[2] is not None else None
            pgm = prev[2] if prev and prev[2] is not None else None
            row[col - 1] = cgm
            row[col] = pgm
            row[col + 1] = (cgm - pgm) if (cgm is not None and pgm is not None) else None
        sub_by_store[store] = row
    fill_positioned_rows(
        wb, "细分毛利率表 (2)", sub_by_store,
        store_col=2, start_row=20, end_row=27,
        value_cols=list(range(3, 24)),
    )

    # ── 毛利率连续对比表: store name in col 4, data rows 6-12 ──
    trend_rows = build_trend_rows(monthly_gp=inputs.monthly_gp)
    trend_by_store = {r[3]: r for r in trend_rows if len(r) > 3 and r[3]}
    fill_positioned_rows(
        wb, "毛利率连续对比表", trend_by_store,
        store_col=4, start_row=6, end_row=12,
        value_cols=[5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22],
    )

    out_path = Path(out_path)
    wb.save(out_path)
    logger.info(
        "wrote styled 毛利分析 workbook → %s (table1=%d, table2=%d, table3=%d, mom=%d)",
        out_path, len(table1_rows), len(table2_rows), len(table3_rows), len(mom_rows),
    )
    return out_path


# ── Convenience: load inputs from on-disk archives ──────────────────────────


def load_inputs_from_paths(
    *,
    year: int,
    month: int,
    cur_pnl: dict[str, dict[str, float]],
    pos_sales_paths: dict[str, Path],
    bom_rows: dict[str, list[dict]],
    zfi_cur_path: Path,
    mb5b_cur_path: Path,
    pos_set_paths: dict[str, Path] | None = None,
    zfi_prev_path: Path | None = None,
    zfi_yoy_path: Path | None = None,
    monthly_gp: dict[str, list[float]] | None = None,
    prev_pnl: dict[str, dict[str, float]] | None = None,
    yoy_pnl: dict[str, dict[str, float]] | None = None,
    pos_prev_paths: dict[str, Path] | None = None,
    pos_yoy_paths: dict[str, Path] | None = None,
    mb5b_prev_path: Path | None = None,
    mb5b_yoy_path: Path | None = None,
) -> GrossMarginInputs:
    """Wire on-disk archive files into a GrossMarginInputs bundle.

    ``cur_pnl`` is passed in pre-loaded (callers usually derive it from a
    ZFI0049 export via canada_pnl.calculate_result). Everything else is
    loaded here.
    """
    pos_sales: dict[str, list[PosSale]] = {}
    pos_prev_sales: dict[str, list[PosSale]] = {}
    pos_yoy_sales: dict[str, list[PosSale]] = {}
    pos_prices_cur: dict[str, dict] = {}
    pos_prices_prev: dict[str, dict] = {}
    pos_prices_yoy: dict[str, dict] = {}
    for store, p in pos_sales_paths.items():
        if p and p.exists():
            pos_sales[store] = load_pos_sales(p)
            pos_prices_cur[store] = load_pos_prices(p)
    for store, p in (pos_prev_paths or {}).items():
        if p and p.exists():
            pos_prev_sales[store] = load_pos_sales(p)
            pos_prices_prev[store] = load_pos_prices(p)
    for store, p in (pos_yoy_paths or {}).items():
        if p and p.exists():
            pos_yoy_sales[store] = load_pos_sales(p)
            pos_prices_yoy[store] = load_pos_prices(p)

    pos_set_qty: dict[tuple[str, int, str], float] = {}
    for store, p in (pos_set_paths or {}).items():
        if p and p.exists():
            pos_set_qty.update(load_pos_set_sales(p))

    return GrossMarginInputs(
        year=year, month=month,
        cur_pnl=cur_pnl,
        prev_pnl=prev_pnl or {},
        yoy_pnl=yoy_pnl or {},
        monthly_gp=monthly_gp or {},
        pos_sales=pos_sales,
        pos_prev_sales=pos_prev_sales,
        pos_yoy_sales=pos_yoy_sales,
        pos_set_qty=pos_set_qty,
        pos_prices_cur=pos_prices_cur,
        pos_prices_prev=pos_prices_prev,
        pos_prices_yoy=pos_prices_yoy,
        bom_rows=bom_rows,
        zfi_cur=load_zfi0156(zfi_cur_path) if zfi_cur_path.exists() else {},
        zfi_prev=(load_zfi0156(zfi_prev_path)
                  if zfi_prev_path and zfi_prev_path.exists() else {}),
        zfi_yoy=(load_zfi0156(zfi_yoy_path)
                 if zfi_yoy_path and zfi_yoy_path.exists() else {}),
        mb5b_cur=load_mb5b_prices(mb5b_cur_path) if mb5b_cur_path.exists() else {},
        mb5b_prev=(load_mb5b_prices(mb5b_prev_path)
                   if mb5b_prev_path and mb5b_prev_path.exists() else {}),
        mb5b_yoy=(load_mb5b_prices(mb5b_yoy_path)
                  if mb5b_yoy_path and mb5b_yoy_path.exists() else {}),
    )
