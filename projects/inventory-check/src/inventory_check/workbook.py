"""Multi-sheet workbook assembly — produce the full ``CA08-盘点结果``
file shaped like the manual workbook.

Why this exists
---------------
The single-sheet xlsx produced by ``report.write_report_xlsx`` mirrors
only the report column layout. The operations team's actual deliverable
is a *workbook* with ~17 sheets where the report sheet
(``CA08-本月-盘点结果.``) drives most columns via VLOOKUP formulas
into sibling data sheets:

- col 10 单价      → ``本月系统单价mb5b!$B:$P`` col 15
- col 12 上月使用   → ``上月数量需更新!A:C`` col 3
- col 14 单价差异   → ``上月盘点结果!$B:$J`` col 9 (IFERROR ⇒ 0)
- col 15 备注       → ``计算!R:Z`` col 9 (IFERROR ⇒ "")
- col 17 分类       → ``分类!B:F`` col 5

Strategy: take the previous month's manual workbook as a *template*,
preserve all the hand-curated sheets (计算, 分类, BI套餐, 红火台销售汇总,
Sheet3 POS pivot, etc.), and overwrite only the few sheets whose
contents change month-to-month:

    报告月 N's workbook ← previous month's manual workbook
        ├── replace ``CA08-本月-盘点结果.`` with N's matnrs (formulas
        │   in cols 9-17, static values in cols 1-8)
        ├── replace ``本月系统单价mb5b`` with N's MB5B export
        ├── replace ``上月数量zfi0156`` with (N-1)'s ZFI0156 export
        ├── replace ``上月数量需更新`` with (N-1)'s ZFI0156 pivoted
        ├── replace ``上月盘点结果`` with (N-1)'s report sheet
        └── keep everything else (计算, Sheet3, BI套餐, 分类, ...)

Per the workflow's design intent: 计算 is hand-curated and treated as
stable across months — the same dish→material BOM applies, only the
POS sales numbers change. We don't try to update the 计算 sheet here.

The output is openable in Excel; formulas recompute on open. Python
readers using ``data_only=True`` will see ``None`` for formula cells
until the file is opened in Excel/LibreOffice once and saved.
"""
from __future__ import annotations

import logging
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from inventory_check.dates import Month
from inventory_check.mb5b_parse import parse_mb5b_text, read_mb5b_text
from inventory_check.references import coerce_matnr
from inventory_check.report import (
    ReportRow,
    _read_zfi0156_lookup,
    build_report_rows,
)
from inventory_check.stores import Store

logger = logging.getLogger(__name__)


REPORT_SHEET = "CA08-本月-盘点结果."
MB5B_SHEET = "本月系统单价mb5b"
ZFI_RAW_SHEET = "上月数量zfi0156"
ZFI_PIVOT_SHEET = "上月数量需更新"
PREV_REPORT_SHEET = "上月盘点结果"
POS_RAW_SHEET = "红火台销售汇总"
POS_PIVOT_SHEET = "Sheet3"  # the pivot of 红火台销售汇总 that 计算!M reads from

# Reference sheets that are STORE-SPECIFIC: they carry the template-native
# store's name/data in every row, so leaking them into a different store's
# workbook produces wrong VLOOKUP results.
#  - 计算 has 门店名称 in col C; we regenerate from IPMS+POS per store.
#  - BI套餐 has 门店名称 col; without a per-store source we wipe (W → 0
#    via SUMIF over empty range, conservative default).
# The remaining template sheets — 分类, 对照表, 折算数量 — are CA-region
# master data (no per-store rows) and ARE referenced by formulas
# (分类: 426 refs from report sheet col Q). Keep them intact.
TEMPLATE_REFERENCE_SHEETS = (
    "计算", "BI套餐",
)


@dataclass(frozen=True)
class WorkbookSources:
    """Bundle of file paths the workbook assembly needs.

    ``zfi0156_path`` and ``prev_report_path`` are optional — when omitted
    the corresponding template sheets (``上月数量zfi0156`` / ``上月数量需更新``
    / ``上月盘点结果``) are inherited from the template unchanged. Useful
    for e2e validation where the template *is* the report month.
    """

    template_path: Path           # previous month's manual workbook
    mb5b_path: Path               # report month's MB5B export
    fiori_path: Path              # report month's Fiori (archive or entry)
    zfi0156_path: Path | None = None
    prev_report_path: Path | None = None
    pos_path: Path | None = None  # report month's 红火台销售汇总 (POS)
    pos_set_path: Path | None = None  # report month's 红火台套餐汇总 (POS) → BI套餐
    ipms_bom_paths: tuple[Path, ...] = ()
    """IPMS 海外菜品物料明细 exports (one per tab: 菜品 + 锅底).

    When provided, the 计算 sheet is regenerated from these BOM rows
    (manual is stale per 2026-05 user direction — IPMS is authoritative).
    When empty, the template's 计算 sheet behaviour is preserved
    (kept for native store, wiped otherwise).
    """


def _strip_pivot_tables(ws: Worksheet) -> int:
    """Remove any Excel pivot table objects bound to ``ws``.

    The manual template has real Excel pivot tables on
    ``上月数量需更新`` and ``Sheet3`` (each with a cached records file
    in xl/pivotCache/). When openpyxl saves, the pivot objects are
    re-emitted with their stale caches — which means Excel/WPS shows
    the cached values instead of the cell values we just wrote.
    Removing the pivot objects here makes the cells authoritative;
    users lose the right-click "Refresh" affordance, but the report's
    VLOOKUPs against these sheets will see our fresh data.
    """
    if not (hasattr(ws, "_pivots") and ws._pivots):
        return 0
    n = len(ws._pivots)
    ws._pivots = []
    logger.info("stripped %d pivot table object(s) from %s", n, ws.title)
    return n


def _clear_data_rows(ws: Worksheet) -> None:
    """Delete all rows below the existing header.

    Some templates have a 1-row header, some have a title+header in rows
    1+2. We don't try to be clever — caller specifies the first data
    row by leaving the header alone and calling ``ws.delete_rows`` from
    that row.
    """
    if ws.max_row <= 1:
        return
    ws.delete_rows(2, ws.max_row - 1)


def _replace_mb5b_sheet(
    ws: Worksheet, mb5b_path: Path, *, werks: str | None = None,
) -> int:
    """Overwrite ``本月系统单价mb5b`` with the new month's MB5B export.

    The MB5B Spreadsheet export is UTF-16 LE TSV. We re-parse the raw
    text directly here (vs reusing ``parse_mb5b_file``) so the column
    order matches the manual workbook's layout: ValA, 物料, dates,
    qty/amt, 计, 货币, then trailing columns (物料描述, 单价, 单位)
    that the bare 14-col SAP variant doesn't include.

    Critical: the report sheet's 单价 column is
    ``=VLOOKUP(B,本月系统单价mb5b!$B:$P,15,0)``. VLOOKUP exact-match
    returns the *first* row whose col B (matnr) matches — so when the
    raw MB5B is region-wide (CA01..CA09) the wrong plant's price wins
    if its row appears first. Filtering to ``werks`` here is what makes
    the formula correct for one store. The manual workbook's mb5b
    sheet is also pre-filtered this way (3383 CA08 rows in the March
    template, vs ~34k region-wide).

    The manual fills cols 16/17 with formulas:
        col 16 (单价)  =Mn/Hn          (ClosingAmt / ClosingQty)
        col 17 (单位)  =VLOOKUP(In, 'CA08-本月-盘点结果.'!$G:$H, 2, 0)
    We replicate those per-row.
    """
    text = read_mb5b_text(mb5b_path)
    rows = parse_mb5b_text(text)

    _clear_data_rows(ws)

    # Map our parsed dict back to the manual's 14-col raw layout.
    # The header is left intact in row 1.
    written = 0
    for r in rows:
        if not r.get("Werks") or not r.get("Matnr"):
            continue
        if werks and r.get("Werks") != werks:
            continue
        target_row = ws.max_row + 1
        line = [
            r.get("Werks", ""),                # 1 ValA
            int(r["Matnr"]) if str(r["Matnr"]).isdigit() else r["Matnr"],  # 2 物料
            r.get("DateFrom", ""),             # 3 开始日期
            r.get("DateTo", ""),               # 4 结束日期
            r.get("OpeningQty", 0),            # 5 期初库存
            r.get("ReceiptsQty", 0),           # 6 总收货数量
            r.get("IssuesQty", 0),             # 7 总发货数量
            r.get("ClosingQty", 0),            # 8 期末库存
            r.get("MeinsAlt", ""),             # 9 计 (SAP unit code)
            r.get("OpeningAmt", 0),            # 10 期初金额
            r.get("ReceiptsAmt", 0),           # 11 总收货金额
            r.get("IssuesAmt", 0),             # 12 总发货金额
            r.get("ClosingAmt", 0),            # 13 期末金额
            r.get("Currency", ""),             # 14 货币
            r.get("Matxt") or None,            # 15 物料描述 (often blank)
            f"=M{target_row}/H{target_row}",   # 16 单价 = ClosingAmt / ClosingQty
            f"=VLOOKUP(I{target_row},'{REPORT_SHEET}'!$G:$H,2,0)",  # 17 单位
            1,                                 # 18 — manual writes 1 (a row counter)
        ]
        ws.append(line)
        written += 1

    logger.info("rewrote %s with %d MB5B rows", ws.title, written)
    return written


def _replace_zfi_raw_sheet(
    ws: Worksheet, zfi_path: Path, *, werks: str | None = None,
) -> int:
    """Overwrite ``上月数量zfi0156`` with (N-1)'s ZFI0156 raw export.

    Format mirrors the SAP ZFI0156 export schema (20 cols typically). The
    SAP export is region-wide (CA01-CA09); when ``werks`` is set, we
    filter to that store's plant so each store's workbook only contains
    its own raw rows. The 工厂 column is detected by header (matches
    ``工厂`` exactly, falls back to col 3 which is its conventional
    position in the SAP layout).
    """
    src = load_workbook(zfi_path, data_only=True)
    src_ws = src.active

    werks_col = None
    if werks:
        for c in range(1, src_ws.max_column + 1):
            header = src_ws.cell(row=1, column=c).value
            if isinstance(header, str) and header.strip() == "工厂":
                werks_col = c
                break
        if werks_col is None:
            werks_col = 3  # SAP convention: 开始日期, 结束日期, 工厂, ...

    _clear_data_rows(ws)
    written = 0
    skipped = 0
    for r in range(2, src_ws.max_row + 1):
        line = [src_ws.cell(row=r, column=c).value
                for c in range(1, src_ws.max_column + 1)]
        if not any(v not in (None, "") for v in line):
            continue
        if werks and werks_col:
            row_werks = line[werks_col - 1]
            if isinstance(row_werks, str):
                row_werks = row_werks.strip()
            if row_werks != werks:
                skipped += 1
                continue
        ws.append(line)
        written += 1
    logger.info(
        "rewrote %s with %d ZFI0156 raw rows (werks=%s, skipped %d)",
        ws.title, written, werks, skipped,
    )
    return written


def _replace_zfi_pivot_sheet(
    ws: Worksheet, zfi_path: Path, *, werks: str,
) -> int:
    """Rebuild the ``上月数量需更新`` 3-column pivot from raw ZFI0156.

    Schema (from the manual file):
        row 1: ['',     '值',           '']
        row 2: ['物料', '求和项:单价1', '求和项:系统发出金额']
        rows 3+: matnr, unit_price, value

    We aggregate the raw ZFI0156 by matnr filtered to ``werks`` (the
    store's plant code). Unit price takes the first non-empty value;
    系统发出金额 is summed.
    """
    lookup = _read_zfi0156_lookup(zfi_path, werks=werks)

    # Preserve the 2-row title/header — clear only data rows starting at 3.
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    written = 0
    for matnr in sorted(lookup, key=lambda m: int(m) if m.isdigit() else m):
        rec = lookup[matnr]
        # Manual values are positive (the pivot took abs); mirror that
        # so the report sheet's col 12 reads the expected sign.
        ws.append([
            int(matnr) if matnr.isdigit() else matnr,
            rec.unit_price if rec.unit_price is not None else "",
            abs(rec.value),
        ])
        written += 1
    logger.info("rewrote %s with %d pivoted rows (werks=%s)",
                ws.title, written, werks)
    return written


def _coerce_dish_code(v: Any) -> Any:
    """Strip leading zeros from a dish code so it matches the manual.

    pos-crawler emits 菜品编码 / 菜品短编码 as zero-padded 8-char strings
    (e.g. '01010106'); the manual workbook's 计算 sheet keys them as
    ints (e.g. 1010106 — Excel strips the leading zero on int storage).
    The 计算!A lookup formula ``=C2&F2&G2&J2`` concatenates F as int →
    string → '加拿大八店1010106单锅'. If our Sheet3 stores the padded
    form ('加拿大八店01010106单锅'), the VLOOKUP misses → 实收数量
    falls back to 0 → 备注 column reads as "no usage" — silently wrong.

    Pass-through for non-numeric strings (some 编码 fields are blank
    or carry text) and ints (already in canonical form).
    """
    if v in (None, ""):
        return v
    if isinstance(v, str) and v.isdigit():
        return int(v)
    return v


def _replace_pos_sheet(ws: Worksheet, pos_path: Path) -> int:
    """Overwrite ``红火台销售汇总`` with the new month's POS export.

    The pos-crawler output is already shaped like the manual workbook's
    红火台销售汇总 sheet (12 cols, same headers): col 4 菜品编码,
    col 5 菜品短编码, col 10 实际出品数据 (= 出品 - 退菜). We copy
    verbatim except for two transforms that match the manual's storage:

    - col 4 / col 5: coerce numeric strings to int (strips zero-padding
      so '01010106' → 1010106). The 计算 sheet's lookup formula
      reads cells as ints; if our 红火台 cells stay as strings the
      Sheet3 pivot key won't match.
    - col 1 (检索): recompute as 门店&菜品编码&菜品短编码&规格 using
      the int-coerced values. The pos-crawler precomputes this with
      string concatenation, which preserves padding and breaks the
      lookup chain. Manual stores it as a formula that re-evaluates
      after concat coercion.
    """
    src = load_workbook(pos_path, data_only=True)
    src_ws = src.active

    _clear_data_rows(ws)
    written = 0
    for r in range(2, src_ws.max_row + 1):
        line = [src_ws.cell(row=r, column=c).value
                for c in range(1, src_ws.max_column + 1)]
        if not any(v not in (None, "") for v in line):
            continue
        # Pad short rows so we can safely index 1..12.
        while len(line) < 12:
            line.append(None)
        # Normalise the codes (cols 4 and 5).
        line[3] = _coerce_dish_code(line[3])
        line[4] = _coerce_dish_code(line[4])
        # Recompute col 1 (检索) from the normalised codes. Manual's
        # formula ``=B&D&E&G`` treats None as empty (Excel concat), so we
        # do the same — ``f"{None}"`` would emit the literal "None".
        store, dish_code, short_code, spec = line[1], line[3], line[4], line[6]
        line[0] = (
            f"{store}{dish_code}"
            f"{short_code if short_code is not None else ''}"
            f"{spec if spec is not None else ''}"
            if store and dish_code is not None else line[0]
        )
        ws.append(line)
        written += 1
    logger.info("rewrote %s with %d POS rows", ws.title, written)
    return written


def _rebuild_sheet3_pivot(ws: Worksheet, pos_path: Path) -> int:
    """Rebuild ``Sheet3`` as the pivot of 红火台销售汇总 by 检索.

    Manual layout (Excel pivot output):

        r1: (blank, blank)
        r2: (blank, blank)
        r3: 求和项:实际出品数据（出品数量-退菜数量）, blank
        r4: Row Labels, Grand Total
        r5: (blank), blank
        r6: blank, 0
        r7+: <lookup_key>, <sum>
        r last: Grand Total, <total sum>

    We rewrite to that exact shape so the manual's 计算!M formula
    (=IFERROR(VLOOKUP(A2,Sheet3!$A:$B,2,0),0)) finds the right key. The
    pivot key is col 1 (检索) of the POS source; the value is the sum
    of col 10 (实际出品数据).

    Aggregation is by exact-match on the lookup key; that's how Excel's
    pivot built it and how the 计算 sheet's VLOOKUP expects to query it.
    """
    src = load_workbook(pos_path, data_only=True)
    src_ws = src.active

    # Group by 检索, sum 实际出品数据 (col 10). The source file's col 1 is
    # the pos-crawler's pre-joined key with zero-padded dish codes (e.g.
    # '加拿大八店01060061四宫格'); the manual's 计算!A formula produces
    # the int-coerced form ('加拿大八店1060061单锅'). We rebuild the key
    # the same way _replace_pos_sheet does so Sheet3's keys match
    # 计算's lookup key exactly.
    pivot: dict[str, float] = {}
    for r in range(2, src_ws.max_row + 1):
        store = src_ws.cell(row=r, column=2).value
        dish_code = _coerce_dish_code(src_ws.cell(row=r, column=4).value)
        short_code = _coerce_dish_code(src_ws.cell(row=r, column=5).value)
        spec = src_ws.cell(row=r, column=7).value
        val = src_ws.cell(row=r, column=10).value
        if store in (None, "") or dish_code in (None, ""):
            continue
        key = f"{store}{dish_code}{short_code if short_code is not None else ''}{spec if spec is not None else ''}"
        if isinstance(val, (int, float)):
            pivot[key] = pivot.get(key, 0.0) + float(val)
    grand_total = sum(pivot.values())

    # Wipe everything (we own the whole sheet).
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    # Recreate the manual's exact layout.
    ws.cell(row=3, column=1, value="求和项:实际出品数据（出品数量-退菜数量）")
    ws.cell(row=4, column=1, value="Row Labels")
    ws.cell(row=4, column=2, value="Grand Total")
    ws.cell(row=5, column=1, value="(blank)")
    ws.cell(row=6, column=2, value=0)
    # Stable order — sort by key for diff-friendly output.
    target_row = 7
    for key in sorted(pivot.keys()):
        ws.cell(row=target_row, column=1, value=key)
        v = pivot[key]
        ws.cell(row=target_row, column=2, value=int(v) if v == int(v) else v)
        target_row += 1
    # Manual ends with a "(blank-shop) row 0" + "Grand Total" footer.
    ws.cell(row=target_row, column=1, value="加拿大八店")
    ws.cell(row=target_row, column=2, value=0)
    target_row += 1
    ws.cell(row=target_row, column=1, value="Grand Total")
    ws.cell(row=target_row, column=2,
            value=int(grand_total) if grand_total == int(grand_total) else grand_total)

    logger.info("rewrote %s with %d pivoted POS keys (grand_total=%s)",
                ws.title, len(pivot), grand_total)
    return len(pivot)


def _replace_prev_report_sheet(ws: Worksheet, prev_report_path: Path) -> int:
    """Overwrite ``上月盘点结果`` with last month's report sheet contents.

    The manual template's 上月盘点结果 sheet has 14 columns (no 备注 /
    回复 / 分类 — those are the recurring-input fields that aren't
    needed when the previous report is consumed only via VLOOKUP for
    单价 and 使用数量).
    """
    src = load_workbook(prev_report_path, data_only=True)
    src_ws = src.active

    _clear_data_rows(ws)
    template_cols = ws.max_column or 14

    written = 0
    for r in range(2, src_ws.max_row + 1):
        line = [src_ws.cell(row=r, column=c).value
                for c in range(1, min(src_ws.max_column, template_cols) + 1)]
        if not any(v not in (None, "") for v in line):
            continue
        # Pad to template width.
        line += [None] * (template_cols - len(line))
        ws.append(line)
        written += 1
    logger.info("rewrote %s with %d prev-report rows", ws.title, written)
    return written


# Per-row formula templates. ``{r}`` is the 1-based row number where the
# formula lives. Cols 9-17 match the manual workbook's formulas exactly.
# Cols 18-21 are the material-pivoted 差异 view added on top — one row
# per material, no double-counting because 计算's W/U/X/Z are emitted
# only on the smallest-N row of each (F, R) group (see
# inventory_check.calc_sheet.attach_formulas).
_REPORT_FORMULAS: dict[int, str] = {
    9:  "=D{r}-E{r}",                                                       # 使用数量
    10: "=VLOOKUP(B{r},本月系统单价mb5b!$B:$P,15,0)",                       # 单价
    11: "=I{r}*J{r}",                                                       # 本月使用金额
    12: "=VLOOKUP(B{r},上月数量需更新!A:C,3,0)",                            # 上月使用金额
    13: "=K{r}-L{r}",                                                       # 对比
    14: "=IFERROR(J{r}-VLOOKUP(B{r},上月盘点结果!$B:$J,9,0),0)",            # 单价差异
    15: '=IFERROR(VLOOKUP(B{r},计算!R:Z,9,FALSE),"")',                      # 备注
    17: '=IFERROR(VLOOKUP(B{r},分类!B:F,5,0),"")',                          # 分类
    # Material-pivoted material-balance view. 差异 = actual usage (col I)
    # − theoretical (sum of N*M*O across dishes using this matnr) −
    # set-meal allocation (sum of canonical W cells in 计算).
    18: '=IFERROR(SUMIF(计算!R:R,B{r},计算!Q:Q),0)',                         # 理论用量
    19: '=IFERROR(SUMIF(计算!R:R,B{r},计算!W:W),0)',                         # 套餐拼盘用量
    20: "=I{r}-R{r}-S{r}",                                                  # 差异(物料)
    21: ('=IF(T{r}=0,"",IF(T{r}>0,"多用"&TEXT(T{r},"0.00"),'
         '"少用"&TEXT(ABS(T{r}),"0.00")))'),                                 # 备注(物料)
}

# Headers for cols 18-21 — written if absent. Template's row-1 header
# only covers cols 1-17; we extend it so analysts can see the new view
# without editing the template.
_REPORT_EXT_HEADERS: dict[int, str] = {
    18: "理论用量",
    19: "套餐拼盘用量",
    20: "差异(物料)",
    21: "备注(物料)",
}


def _replace_report_sheet(
    ws: Worksheet, rows: list[ReportRow],
) -> int:
    """Overwrite ``CA08-本月-盘点结果.`` — static values cols 1-8,
    formulas 9-17 (manual layout) + 18-21 (material-pivoted 差异 view).

    ``rows`` carries our pre-computed ReportRow records, but we only
    use the static (source-derived) fields:
        col 1 行号 / col 2 物料编码 / col 3 物料名称 / col 4 库存数量 /
        col 5 盘点数量 / col 6 单位 / col 7 单位编码 / col 8 单位描述

    Cols 9-21 are written as Excel formulas so the workbook re-derives
    those values in-place — the user can edit any source sheet and the
    report updates without re-running our pipeline.
    """
    _clear_data_rows(ws)

    # Add headers for cols 18-21 if the template only carries 17 cols.
    # Idempotent: if the headers are already there (re-run on a workbook
    # we've already extended) we leave them alone.
    for col, header in _REPORT_EXT_HEADERS.items():
        if ws.cell(row=1, column=col).value != header:
            ws.cell(row=1, column=col, value=header)

    written = 0
    for row in rows:
        target_row = ws.max_row + 1
        ws.cell(row=target_row, column=1, value=row.row_no)
        ws.cell(row=target_row, column=2,
                value=int(row.matnr) if row.matnr.isdigit() else row.matnr)
        ws.cell(row=target_row, column=3, value=row.matxt or None)
        ws.cell(row=target_row, column=4,
                value=row.closing_qty if row.closing_qty != "" else None)
        ws.cell(row=target_row, column=5,
                value=row.counted_qty if row.counted_qty != "" else None)
        ws.cell(row=target_row, column=6, value=row.unit or None)
        ws.cell(row=target_row, column=7, value=row.unit_code or None)
        ws.cell(row=target_row, column=8, value=row.unit_desc or None)
        # Formulas — col 16 (回复) is left blank (manual entry column).
        for col, fmt in _REPORT_FORMULAS.items():
            ws.cell(row=target_row, column=col, value=fmt.format(r=target_row))
        written += 1

    logger.info("rewrote %s with %d report rows", ws.title, written)
    return written


def assemble_workbook(
    sources: WorkbookSources,
    *,
    store: Store,
    month: Month,
    out_path: Path,
    fiori_use_entry: bool = False,  # accepted for symmetry, unused here
) -> Path:
    """Build the full month-N workbook from the previous month's template.

    Pipeline:
      1. Copy template → out_path (preserves all 17 sheets + formatting)
      2. Replace `本月系统单价mb5b` with month-N MB5B
      3. Replace `上月数量zfi0156` with (N-1) ZFI0156 raw
      4. Replace `上月数量需更新` with (N-1) ZFI0156 pivot for ``store.werks``
      5. Replace `上月盘点结果` with (N-1) report sheet
      6. Replace `CA08-本月-盘点结果.` with month-N matnrs + formulas
      7. Save

    The 计算 sheet, Sheet3 (POS pivot), BI套餐, 红火台销售汇总, 分类,
    折算数量, 对照表, 注意事项, and miscellaneous Sheet1/2/4 / 透视表
    are inherited from the template — they're either hand-curated or
    static reference data that doesn't change month-to-month.

    `fiori_use_entry` is accepted for caller-side symmetry; the report
    rows the caller built already reflect that choice.
    """
    del fiori_use_entry  # callers may pass it; we don't need it here

    # 1. Build the report rows up front so a parse failure aborts before
    #    we copy the template (cheap fast-fail).
    rows = build_report_rows(
        store=store, month=month,
        mb5b_path=sources.mb5b_path,
        fiori_path=sources.fiori_path,
        prev_report_path=sources.prev_report_path,
        zfi0156_path=sources.zfi0156_path,
        # No calc_path — the template's 计算 sheet is the source of remarks.
    )

    # 2. Copy the template so we don't mutate the caller's reference file.
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    logger.info("copying template %s → %s", sources.template_path, out_path)
    shutil.copy2(sources.template_path, out_path)

    # 3. Open keep_links / formulas, mutate, save.
    wb = load_workbook(out_path, data_only=False)

    if MB5B_SHEET in wb.sheetnames:
        _replace_mb5b_sheet(wb[MB5B_SHEET], sources.mb5b_path, werks=store.werks)
    else:
        logger.warning("template has no %s sheet — skipping", MB5B_SHEET)

    if sources.zfi0156_path:
        if ZFI_RAW_SHEET in wb.sheetnames:
            _replace_zfi_raw_sheet(
                wb[ZFI_RAW_SHEET], sources.zfi0156_path, werks=store.werks,
            )
        if ZFI_PIVOT_SHEET in wb.sheetnames:
            _strip_pivot_tables(wb[ZFI_PIVOT_SHEET])
            _replace_zfi_pivot_sheet(
                wb[ZFI_PIVOT_SHEET], sources.zfi0156_path, werks=store.werks,
            )
    else:
        logger.info("zfi0156_path not provided — leaving 上月 sheets untouched")

    if sources.prev_report_path:
        if PREV_REPORT_SHEET in wb.sheetnames:
            _replace_prev_report_sheet(
                wb[PREV_REPORT_SHEET], sources.prev_report_path,
            )
    else:
        # No per-store prev_report supplied — wipe the template's data so
        # the previous store's numbers don't leak into this store's col 14
        # (单价差异) lookups. Header row stays.
        if PREV_REPORT_SHEET in wb.sheetnames:
            ws_prev = wb[PREV_REPORT_SHEET]
            cleared = ws_prev.max_row - 1 if ws_prev.max_row > 1 else 0
            _clear_data_rows(ws_prev)
            logger.info(
                "prev_report_path not provided — cleared %d rows from %s",
                cleared, PREV_REPORT_SHEET,
            )

    if sources.pos_path:
        if POS_RAW_SHEET in wb.sheetnames:
            _replace_pos_sheet(wb[POS_RAW_SHEET], sources.pos_path)
        if POS_PIVOT_SHEET in wb.sheetnames:
            _strip_pivot_tables(wb[POS_PIVOT_SHEET])
            _rebuild_sheet3_pivot(wb[POS_PIVOT_SHEET], sources.pos_path)
    else:
        logger.info("pos_path not provided — leaving %s and %s untouched "
                    "(report 备注 will reflect previous month's POS volume)",
                    POS_RAW_SHEET, POS_PIVOT_SHEET)

    if REPORT_SHEET in wb.sheetnames:
        _replace_report_sheet(wb[REPORT_SHEET], rows)
    else:
        raise RuntimeError(
            f"template {sources.template_path} has no {REPORT_SHEET!r} sheet"
        )

    # If the target store isn't the template's native store, wipe the
    # hand-curated reference sheets (计算/分类/折算数量/BI套餐/对照表) so
    # this workbook is self-contained — the previous owner's data
    # doesn't leak via the report sheet's VLOOKUPs.
    _wipe_template_references_if_foreign(wb, store)

    # IPMS-derived 计算 sheet — overrides whatever the template held
    # (or the wipe just left blank) with current-recipe data.
    if sources.ipms_bom_paths:
        _replace_calc_sheet(wb, store, sources.ipms_bom_paths,
                            pos_path=sources.pos_path)

    # BI套餐 — populate from POS 菜品套餐报表 if supplied (per-store data,
    # so we always rewrite when given a path; if missing, the wipe above
    # leaves W=0 which is the conservative default).
    if sources.pos_set_path:
        _replace_bi_taocan_sheet(wb, sources.pos_set_path)

    wb.save(str(out_path))
    logger.info("saved %s", out_path)
    return out_path


def _replace_bi_taocan_sheet(wb, pos_set_path: Path) -> int:
    """Rewrite the BI套餐 sheet from a POS set-meal export.

    The calc sheet's W formula reads K (菜品编码) + T (应收数量) from
    BI套餐. We copy the entire export's rows verbatim so any future
    formula that wants other cols (套餐折扣, 实收金额, etc.) still works.
    """
    if "BI套餐" not in wb.sheetnames:
        logger.warning("template has no BI套餐 sheet — skipping POS set-sale ingestion")
        return 0
    pos_set_path = Path(pos_set_path)
    if not pos_set_path.exists():
        logger.warning("POS set-sale path missing: %s — leaving BI套餐 untouched",
                       pos_set_path)
        return 0

    src_wb = load_workbook(pos_set_path, data_only=True, read_only=True)
    src_ws = src_wb[src_wb.sheetnames[0]]
    src_rows = list(src_ws.iter_rows(values_only=True))
    src_wb.close()
    if not src_rows:
        logger.info("POS set-sale xlsx is empty — leaving BI套餐 cleared")
        return 0

    target = wb["BI套餐"]
    _clear_data_rows(target)
    # Header row stays (template's). Append source data rows (skip source header).
    for r in src_rows[1:]:
        target.append(list(r))
    written = len(src_rows) - 1
    logger.info("rewrote BI套餐 with %d rows from POS 菜品套餐报表", written)
    return written


def _replace_calc_sheet(wb, store: Store, bom_paths: tuple[Path, ...],
                        pos_path: Path | None = None) -> int:
    """Regenerate the 计算 sheet from IPMS BOM exports.

    Drops all existing data rows (the manual is stale per 2026-05) and
    writes one row per IPMS BOM (dish×spec×material) tuple. The 检索
    column is keyed to ``store.pos_name`` so the report sheet's VLOOKUPs
    (col 15 备注 → 计算!R:Z) and Sheet3 lookups (计算!M ← Sheet3!A:B)
    resolve correctly per store. Formulas in M/Q/T/U/X are kept as
    formulas so Excel computes them on open.

    pos_path supplies 菜品短编码 / 大类名称 / 子类名称 for each
    (dish_code, spec) tuple. Without it, the 检索 keys won't include
    the short-code segment that Sheet3 expects, and M (实收数量) +
    Q (红火台理论量) formulas resolve to 0 for all rows.
    """
    from inventory_check.calc_sheet import (
        attach_formulas, derive_calc_rows, load_ipms_bom_rows,
        load_pos_dish_meta,
    )

    if "计算" not in wb.sheetnames:
        logger.warning("template has no 计算 sheet — skipping IPMS regen")
        return 0

    ws = wb["计算"]
    bom_rows = load_ipms_bom_rows(list(bom_paths))
    pos_meta = load_pos_dish_meta(pos_path) if pos_path else {}
    rows = derive_calc_rows(bom_rows, store_name=store.pos_name,
                            pos_meta=pos_meta)
    attach_formulas(rows, report_sheet_name=REPORT_SHEET)

    _clear_data_rows(ws)
    for r in rows:
        ws.append(r)
    logger.info("regenerated 计算 with %d rows from IPMS BOM "
                "(%d source files, pos_meta=%d entries)",
                len(rows), len(bom_paths), len(pos_meta))
    return len(rows)


def _wipe_template_references_if_foreign(wb, store: Store) -> None:
    """Clear data rows from the template's reference sheets when the
    target store doesn't own the template.

    The template's 计算 sheet has 门店名称 in col C of every data row.
    If the first data row's value matches ``store.pos_name``, the
    template *is* this store's — leave everything alone. Otherwise,
    wipe the reference sheets (header rows survive).
    """
    if "计算" not in wb.sheetnames:
        return  # nothing to wipe
    calc_ws = wb["计算"]
    template_store = (
        calc_ws.cell(row=2, column=3).value if calc_ws.max_row >= 2 else None
    )
    if template_store == store.pos_name:
        logger.info(
            "template's reference sheets belong to %s (== target) — kept",
            store.pos_name,
        )
        return
    logger.info(
        "template-native store=%r ≠ target=%r — wiping reference sheets %s",
        template_store, store.pos_name, list(TEMPLATE_REFERENCE_SHEETS),
    )
    for sheet_name in TEMPLATE_REFERENCE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        before = ws.max_row - 1 if ws.max_row > 1 else 0
        _clear_data_rows(ws)
        logger.info("  cleared %d rows from %s", before, sheet_name)
