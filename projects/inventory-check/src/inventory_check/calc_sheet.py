"""Build 计算 sheet rows from store_bom recipe data.

The 计算 sheet (26-col, hand-curated in the manual) joins POS sales with
material BOM data to compute theoretical material consumption per
dish×spec. Per the 2026-05 migration, the recipe table now lives in
Postgres (``store_bom``); ``inventory_check.db_bom.load_store_bom_rows``
is the canonical reader. The /admin/bom UI is the system of record.

Each store_bom row (one per dish×spec×material) becomes one 计算 row.
Display columns we don't have a source for (大类, 子类, 菜品单价,
菜品单位) come from POS; if POS doesn't have them they're left blank —
they don't enter the formula chain that feeds the report sheet.

Source columns from store_bom:
  dish_code        → F
  dish_short_code  → G (also fed by POS when missing in BOM)
  dish_name        → H, I
  spec             → J
  material_code    → R
  material_name    → S
  portion          → N (出品分量)
  loss_factor      → O (损耗 — already computed, no yield→loss conversion)
  unit             → Y (库存单位)
  packaging_factor → P (物料单位)

Computed columns:
  A 检索 = store_name & code_int & short_code & spec
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

# 26 columns matching the manual 计算 layout.
HEADERS: list[str | None] = [
    "检索", "区域名称", "门店名称", "大类名称", "子类名称",
    "菜品编码", "菜品短编码", "菜品名称", "菜品名称（系统）",
    "规格", "菜品单价", "菜品单位", "实收数量", "出品分量",
    "损耗", "物料单位", "红火台理论量", "物料号", "物料描述",
    "物料用量", "差异", "备注1", "套餐、拼盘用量", "备注",
    "单位", None,
]

REGION = "加拿大"


def _norm_code(c: Any) -> str:
    """Normalize numeric codes by stripping leading zeros so SAP/IPMS-padded
    strings ('01060061') and bare ints (1060061) compare equal."""
    if c is None:
        return ""
    return str(c).strip().lstrip("0")


def load_pos_dish_meta(pos_path: Path) -> dict[tuple[str, str], dict]:
    """Read POS 红火台销售汇总 → (norm_dish_code, spec) → meta dict.

    POS is the only source for 菜品短编码 (critical for the Sheet3
    lookup key), 大类名称 / 子类名称 (D/E columns), and 菜品单价 /
    菜品单位 (K/L columns — only present in xlsx written by recent
    versions of pos-crawler that capture dishPrice + unit from the
    listDishPotSale response).
    Returns empty dict if path is missing or sheet is empty.
    """
    out: dict[tuple[str, str], dict] = {}
    if not pos_path or not Path(pos_path).exists():
        return out
    wb = openpyxl.load_workbook(pos_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    iter_ = ws.iter_rows(values_only=True)
    headers = list(next(iter_))
    idx = {h: i for i, h in enumerate(headers) if h}
    code_i = idx.get("菜品编码")
    short_i = idx.get("菜品短编码")
    spec_i = idx.get("规格")
    cat_i = idx.get("大类名称")
    subcat_i = idx.get("子类名称")
    price_i = idx.get("菜品单价")
    unit_i = idx.get("菜品单位")
    if code_i is None or spec_i is None:
        wb.close()
        return out

    def _at(vals, i):
        return vals[i] if i is not None and i < len(vals) else None

    for vals in iter_:
        code = vals[code_i] if code_i < len(vals) else None
        spec = vals[spec_i] if spec_i < len(vals) else None
        if code is None or spec is None:
            continue
        key = (_norm_code(code), str(spec).strip())
        if key in out:
            continue
        out[key] = {
            "菜品短编码": _at(vals, short_i),
            "大类名称": _at(vals, cat_i),
            "子类名称": _at(vals, subcat_i),
            "菜品单价": _at(vals, price_i),
            "菜品单位": _at(vals, unit_i),
        }
    wb.close()
    return out


def derive_calc_rows(bom_rows: list[dict], *, store_name: str,
                     pos_meta: dict[tuple[str, str], dict] | None = None,
                     ) -> list[list[Any]]:
    """Build 计算 rows from store_bom data for a single store.

    The store_name parameterises both the 门店名称 column and the 检索
    key (since col A is `store + code + short + spec` joined into the
    Sheet3 lookup). All other columns come from the BOM rows.

    pos_meta supplies fields BOM may not carry: 菜品短编码 (CRITICAL —
    Sheet3's pivot keys include it, so without it the M/Q formula chain
    can't resolve), 大类名称, 子类名称, 菜品单价, 菜品单位. When a BOM
    row already has dish_short_code, it wins; POS is the fallback.
    """
    pos_meta = pos_meta or {}
    out = []
    for b in bom_rows:
        code = _norm_code(b["dish_code"])
        spec = str(b["spec"]).strip()
        matnr = _norm_code(b["material_code"])

        code_int = int(code) if code.isdigit() else code
        matnr_int = int(matnr) if matnr.isdigit() else matnr

        meta = pos_meta.get((code, spec), {})
        # Prefer the BOM-side short_code if present; fall back to POS.
        bom_short = b.get("dish_short_code")
        if bom_short in (None, ""):
            short_raw = meta.get("菜品短编码")
        else:
            short_raw = bom_short
        short_int = (int(_norm_code(short_raw))
                     if short_raw and str(_norm_code(short_raw)).isdigit()
                     else (short_raw or None))
        big_cat = meta.get("大类名称")
        sub_cat = meta.get("子类名称")
        dish_price = meta.get("菜品单价")
        dish_unit = meta.get("菜品单位")

        # 检索 = C&F&G&J — keep G as int (no leading zeros) so it
        # matches Sheet3's _rebuild_sheet3_pivot key construction.
        short_str = "" if short_int is None else str(short_int)
        key = f"{store_name}{code_int}{short_str}{spec}"

        portion = b.get("portion")
        loss = b.get("loss_factor")
        if loss in (None, ""):
            loss = 1
        unit = b.get("unit")
        packaging_factor = b.get("packaging_factor")
        name = b.get("dish_name")
        matdesc = b.get("material_name")

        # 实收数量 (M), 红火台理论量 (Q), 物料用量 (T), 差异 (U), 备注1 (V),
        # 套餐拼盘用量 (W), 备注 (X), 备注 (Z) are all formulas in the
        # manual. We emit formulas so Excel computes them on open.
        out.append([
            key,                # A 检索
            REGION,             # B 区域名称
            store_name,         # C 门店名称
            big_cat,            # D 大类名称 (from POS)
            sub_cat,            # E 子类名称 (from POS)
            code_int,           # F 菜品编码
            short_int,          # G 菜品短编码 (BOM, fallback POS)
            name,               # H 菜品名称
            name,               # I 菜品名称（系统）
            spec,               # J 规格
            dish_price,         # K 菜品单价 (from POS listDishPotSale.dishPrice)
            dish_unit,          # L 菜品单位 (from POS listDishPotSale.unit)
            None,               # M 实收数量 — formula filled below
            portion,            # N 出品分量
            loss,               # O 损耗
            packaging_factor,   # P 物料单位 — packaging factor
            None,               # Q 红火台理论量 — formula
            matnr_int,           # R 物料号
            matdesc,            # S 物料描述
            None,               # T 物料用量 — formula
            None,               # U 差异 — formula
            None,               # V 备注1
            None,               # W 套餐、拼盘用量 — SUMIF formula
            None,               # X 备注 (intermediate) — formula
            unit,               # Y 单位
            None,               # last col empty in manual
        ])
    return out


def _find_canonical_rows(rows: list[list[Any]]) -> set[int]:
    """For each (F, R) dish-material group, pick the row with smallest N.

    Returns the set of 0-based row indices that are canonical. The
    canonical row is where W/U/X/Z formulas live so the material-balance
    isn't triple-counted by SUMIF(R:R, R_i, Q:Q) firing on every spec
    row of the same dish.

    Smallest N is the per-set-meal portion (e.g. 四宫格 = 0.3, the
    quarter-pot served inside a set meal). Manual encodes this as `*N4`
    on row 2 of each dish-group; we encode it by emitting only on the
    smallest-N row so `*N{i}` produces the same number.
    """
    groups: dict[tuple[Any, Any], list[tuple[int, float]]] = {}
    for idx, r in enumerate(rows):
        f = r[5]
        rr = r[17]
        n = r[13]
        try:
            n_val = float(n) if n is not None else float("inf")
        except (TypeError, ValueError):
            n_val = float("inf")
        groups.setdefault((f, rr), []).append((idx, n_val))
    canonical: set[int] = set()
    for items in groups.values():
        # Smallest N wins; ties broken by row order (stable).
        items.sort(key=lambda t: (t[1], t[0]))
        canonical.add(items[0][0])
    return canonical


def attach_formulas(rows: list[list[Any]],
                    *, report_sheet_name: str) -> None:
    """Mutate rows in place: fill formula columns (M, Q, T, U, W, X, Z).

    Manual's pattern: per-(F, R) dish-group, only the smallest-spec row
    carries the material-balance formulas (W/U/X/Z). Q/M/T are emitted
    on every row — they're per-spec values, useful as detail and don't
    pollute the SUMIF-based balance.

    Why canonical-only for W/U/X/Z:
        SUMIF(R:R, R_i, Q:Q) sums Q across ALL rows sharing the material.
        If U were emitted on every spec row, each row would re-subtract
        the same total Q sum, triple-counting the material balance.
        Manual avoids this by leaving W/U/X/Z blank on non-canonical
        rows. We replicate.

    Formulas:
        M = VLOOKUP(A, Sheet3!$A:$B, 2, 0)
        Q = N * M * O
        T = VLOOKUP(R, '<report>'!B:I, 8, 0) [* P when packaging differs]
        W = SUMIF(BI套餐!K:K, F, BI套餐!T:T) * N * O    [canonical row only]
        U = T - SUMIF(R:R, R, Q:Q) - W                  [canonical row only]
        X = IF(U=0, "", IF(U>0, "多用"&TEXT(U,"0.00"),
                             "少用"&TEXT(ABS(U),"0.00")))
        Z = IF(X<>"", X & Y, "")
    """
    canonical = _find_canonical_rows(rows)
    for i, r in enumerate(rows, start=2):  # row 2 onward (row 1 is header)
        idx0 = i - 2
        r[12] = f"=IFERROR(VLOOKUP(A{i},Sheet3!$A:$B,2,0),0)"          # M
        r[16] = f"=N{i}*M{i}*O{i}"                                      # Q
        r[19] = (f"=IFERROR(VLOOKUP(R{i},"
                 f"'{report_sheet_name}'!B:I,8,0),0)"
                 f'*IF(P{i}="",1,P{i})')                                # T
        if idx0 in canonical:
            # W: SUMIF over BI套餐!K returns total set-meal quantity for
            # this dish across all set meals it appears in; multiply by
            # this row's N (the smallest-spec portion = per-set-meal
            # serving) and O (loss factor).
            r[22] = (f"=IFERROR(SUMIF('BI套餐'!K:K,F{i},"
                     f"'BI套餐'!T:T)*N{i}*O{i},0)")                       # W
            r[20] = f"=T{i}-SUMIF(R:R,R{i},Q:Q)-IFERROR(W{i},0)"        # U
            r[23] = (f'=IF(U{i}=0,"",IF(U{i}>0,"多用"&TEXT(U{i},"0.00"),'
                     f'"少用"&TEXT(ABS(U{i}),"0.00")))')                 # X
            r[25] = f'=IF(X{i}<>"",X{i}&Y{i},"")'                        # Z


def write_calc_workbook(rows: list[list[Any]], out_path: Path) -> None:
    """Write a standalone workbook with just the 计算 sheet.

    Used for testing/validation. Production path is to insert these rows
    into the inherited template via workbook._replace_calc_sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "计算"
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    wb.save(out_path)
