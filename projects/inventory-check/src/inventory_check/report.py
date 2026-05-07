"""Assemble the final 盘点结果 workbook from the downloaded sources.

The output mirrors the manual workbook's ``CA08-本月-盘点结果.`` sheet
schema (17 columns). Cols 9, 11, 13 are simple arithmetic; cols 10, 12,
14, 15, 17 are lookups into the other sources.

What's wired
------------
- **物料编码 / 名称 / 单位 / 单位描述** ← MB5B (one row per material for
  the requested store)
- **库存数量** ← MB5B 期末库存 (ClosingQty)
- **盘点数量** ← Fiori MengePd (the latest physical count)
- **使用数量 / 本月使用金额** ← computed in Python (no formulas)
- **单价** ← MB5B UnitPrice
- **分类** ← 分类 sheet 一级分类

What's TODO (left blank with a clear marker)
--------------------------------------------
- **上月使用金额 / 对比** — needs ZFI0156 download
- **单价差异** — needs last month's 盘点结果 file
- **备注（每月刷新）** — derived from POS sales via the 计算 sheet logic

Timing note about 盘点数量
--------------------------
The Fiori 盘点报表 entry timestamped ``YYYYMM01`` is the start-of-month
physical count. So the manual file for month N uses the Fiori entry
**from month N+1** (= end-of-month-N count). Pass the next-month export
when you have it; otherwise the current month's Fiori is used and the
log makes it explicit.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from inventory_check.dates import Month
from inventory_check.mb5b_parse import filter_by_werks, parse_mb5b_file
from inventory_check.references import coerce_matnr, material_classification_index
from inventory_check.stores import Store

logger = logging.getLogger(__name__)


# Output sheet schema — must match the manual workbook column order.
OUTPUT_COLUMNS: tuple[str, ...] = (
    "行号",
    "物料编码",
    "物料名称",
    "库存数量",
    "盘点数量",
    "单位",
    "单位编码",
    "单位描述",
    "使用数量",
    "单价",
    "本月使用金额",
    "上月使用金额",
    "对比",
    "单价差异",
    "备注（每月刷新）",
    "回复",
    "分类",
)

# Marker placed in cells we can't compute yet so the user sees the gap.
TODO_MARKER = ""


@dataclass(frozen=True)
class ReportRow:
    """One row of the output, before xlsx serialization."""

    row_no: int
    matnr: str
    matxt: str
    closing_qty: float | str
    counted_qty: float | str
    unit: str
    unit_code: str
    unit_desc: str
    usage_qty: float | str
    unit_price: float | str
    month_value: float | str
    prev_month_value: float | str  # 上月使用金额 — from ZFI0156
    delta: float | str              # 对比 = 本月 - 上月
    unit_price_diff: float | str    # 单价差异 — from last month's report
    remark: str                     # 备注 — from 计算 sheet col Z
    classification: str

    def to_xlsx_row(self) -> list[Any]:
        return [
            self.row_no,
            self.matnr,
            self.matxt,
            self.closing_qty,
            self.counted_qty,
            self.unit,
            self.unit_code,
            self.unit_desc,
            self.usage_qty,
            self.unit_price,
            self.month_value,
            self.prev_month_value,
            self.delta,
            self.unit_price_diff,
            self.remark,
            "",           # 回复 (manual entry column)
            self.classification,
        ]


def _read_prev_unit_price_by_matnr(prev_report_path: Path) -> dict[str, float]:
    """Build {Matnr → 单价} from a previous month's 盘点结果 xlsx.

    The previous report uses our own schema, so col 2 = 物料编码 and
    col 10 = 单价. Materials with non-numeric prices are skipped.
    """
    wb = load_workbook(prev_report_path, data_only=True)
    ws = wb.active
    out: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        matnr = row[1] if len(row) > 1 else None
        unit_price = row[9] if len(row) > 9 else None
        if not matnr:
            continue
        key = coerce_matnr(matnr)
        if isinstance(unit_price, (int, float)):
            out[key] = float(unit_price)
    return out


def _read_calc_remarks_by_matnr(calc_path: Path) -> dict[str, str]:
    """Read the 计算 sheet and build {Matnr → 备注 string}.

    The 计算 sheet's col R (18) is 物料号 and col Z (26) is the final
    "多用…/少用…" remark string. Multiple dishes can map to the same
    material; we keep the LAST non-empty remark we see (matching VLOOKUP's
    first-found behaviour after de-duplication).

    The sheet name varies by store ("计算" in CA08); we try the active
    sheet first, then fall back to a sheet named "计算".
    """
    wb = load_workbook(calc_path, data_only=True)
    if "计算" in wb.sheetnames:
        ws = wb["计算"]
    else:
        ws = wb.active
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        matnr = ws.cell(row=r, column=18).value  # col R
        remark = ws.cell(row=r, column=26).value  # col Z
        if matnr is None:
            continue
        key = coerce_matnr(matnr)
        text = "" if remark is None else str(remark).strip()
        if text:
            out[key] = text
    return out


@dataclass(frozen=True)
class Zfi0156Lookup:
    """Per-material aggregates from a ZFI0156 export.

    A single ZFI0156 row spans one (plant, material) pair. We aggregate
    across plants when keying by material only — value is summed,
    descriptive fields take the last-seen non-empty value (the manual
    workbook does the same via VLOOKUP first-match semantics).
    """

    value: float                # 系统发出金额, summed
    unit_price: float | None    # 系统发出单价 (last-seen), used as 单价 fallback
    matxt: str                  # 物料描述, used as 物料名称 fallback
    unit: str                   # 单位描述 (Chinese), used as 单位/单位描述 fallback
    unit_code: str              # Bun (SAP unit code, e.g. 'L'), 单位编码 fallback


def _detect_zfi_columns(ws: Any) -> tuple[int, dict[str, int]]:
    """Locate the ZFI0156 header row and return ``(header_row, name→col)``.

    Two layouts in the wild:

    1. Raw export (header row = 1): columns include 物料, 物料描述, Bun,
       单位描述, 系统发出单价, 系统发出金额, …
    2. Pre-pivoted PivotTable export: row 1 carries a generic title
       ("值"), row 2 is the header (物料, 求和项:单价1, 求和项:系统发出金额).

    We scan the first 5 rows for one with ``物料`` and pick the column
    indices we care about, falling back to plausible synonyms. Both
    layouts share enough column names that one detector handles both.
    """
    aliases: dict[str, tuple[str, ...]] = {
        "matnr":      ("物料",),
        "value":      ("系统发出金额", "求和项:系统发出金额", "求和项:发出金额", "发出金额"),
        "unit_price": ("系统发出单价", "求和项:单价1", "求和项:系统发出单价"),
        "matxt":      ("物料描述", "求和项:物料描述"),
        "unit":       ("单位描述",),
        "unit_code":  ("Bun", "单位"),
        "werks":      ("工厂",),
    }
    for r in range(1, min(6, ws.max_row + 1)):
        cells = [str(ws.cell(row=r, column=c).value or "").strip()
                 for c in range(1, ws.max_column + 1)]
        if "物料" not in cells:
            continue
        cols: dict[str, int] = {}
        for key, names in aliases.items():
            for name in names:
                if name in cells:
                    cols[key] = cells.index(name) + 1
                    break
        return r, cols
    return 0, {}


def _read_zfi0156_lookup(
    zfi_path: Path, *, werks: str | None = None,
) -> dict[str, Zfi0156Lookup]:
    """Build ``{matnr → Zfi0156Lookup}`` from a ZFI0156 export.

    By default keeps every plant in the file. Pass ``werks`` to filter
    to a single plant — required when comparing against the manual
    workbook, whose 上月使用金额 is the per-store issuance value, not the
    region aggregate.

    Pre-pivoted exports (``上月数量需更新``) usually have no 工厂
    column. In that case the values are presented as already-aggregated
    per-material totals; we keep the matnr filter and skip the werks
    filter (since there's no column to filter on).

    ``value`` is summed across rows; descriptive fields (matxt, unit,
    unit_code, unit_price) keep the last non-empty value seen, which
    for any single material are identical across plant rows.
    """
    wb = load_workbook(zfi_path, data_only=True)
    ws = wb.active
    header_row, cols = _detect_zfi_columns(ws)
    if not header_row or "matnr" not in cols or "value" not in cols:
        logger.warning(
            "ZFI0156 schema not recognised in %s — empty lookup", zfi_path,
        )
        return {}

    werks_col = cols.get("werks")
    if werks and not werks_col:
        logger.info(
            "ZFI0156 %s has no 工厂 column (likely pre-pivoted) — keeping all matnrs",
            zfi_path,
        )

    raw: dict[str, dict[str, Any]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        m = ws.cell(row=r, column=cols["matnr"]).value
        if not m:
            continue
        if werks and werks_col:
            row_werks = (ws.cell(row=r, column=werks_col).value or "")
            if str(row_werks).strip() != werks:
                continue
        key = coerce_matnr(m)
        rec = raw.setdefault(key, {"value": 0.0, "unit_price": None,
                                   "matxt": "", "unit": "", "unit_code": ""})
        v = ws.cell(row=r, column=cols["value"]).value
        if isinstance(v, (int, float)):
            rec["value"] += float(v)
        for fkey in ("unit_price",):
            if fkey in cols:
                fv = ws.cell(row=r, column=cols[fkey]).value
                if isinstance(fv, (int, float)):
                    rec[fkey] = float(fv)
        for fkey in ("matxt", "unit", "unit_code"):
            if fkey in cols:
                fv = ws.cell(row=r, column=cols[fkey]).value
                if fv not in (None, ""):
                    rec[fkey] = str(fv).strip()
    return {k: Zfi0156Lookup(**v) for k, v in raw.items()}


def _read_zfi0156_value_by_matnr(zfi_path: Path) -> dict[str, float]:
    """Backwards-compatible wrapper: just the {matnr → value} map.

    Kept for any external caller still using the old shape; the report
    builder uses ``_read_zfi0156_lookup`` directly so it can pull
    descriptive fields too.
    """
    return {k: v.value for k, v in _read_zfi0156_lookup(zfi_path).items()}


def _read_fiori_count_by_matnr(fiori_path: Path) -> dict[str, float]:
    """Build {Matnr → MengePd} from a Fiori stocktake xlsx.

    The crawler-generated workbook has columns:
    工厂 | 物料号 | 名称 | 单位 | 单位描述 | 库存数量 | 盘点数量 | …
    """
    wb = load_workbook(fiori_path, data_only=True)
    ws = wb.active
    out: dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        matnr = row[1]
        counted = row[6]
        if not matnr:
            continue
        key = str(matnr).strip()
        if isinstance(counted, (int, float)):
            out[key] = float(counted)
        elif isinstance(counted, str) and counted.strip():
            try:
                out[key] = float(counted)
            except ValueError:
                continue
    return out


def _safe_num(v: Any) -> float | str:
    """Numbers pass through; blanks render as ''."""
    if v in (None, ""):
        return ""
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str) and v.strip():
        try:
            return float(v.replace(",", ""))
        except ValueError:
            return v
    return ""


def _compact_int(v: float | str) -> float | str:
    """Render integer-valued floats as int (so xlsx doesn't show '60.0')."""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def _has_stock_activity(row: dict[str, Any]) -> bool:
    """True if the material had any quantity movement in the month.

    Includes opening/closing stock and in-month receipts/issues. The
    earlier "ClosingQty != 0" rule undercounted by ~36% on the bare
    14-col MB5B variant: in-month passthrough materials (received +
    issued, ending at zero) were dropped, but the manual workbook
    keeps them. Validated against CA08 202603: this filter captures
    415/415 of the manual's matnrs (with 7 extras whose opening,
    closing, and amounts are all zero — economically inert and
    indistinguishable from the 415 by MB5B alone; the manual curator
    drops them by judgment).
    """
    for key in ("OpeningQty", "ReceiptsQty", "IssuesQty", "ClosingQty"):
        v = row.get(key)
        if isinstance(v, (int, float)) and v != 0:
            return True
    return False


def build_report_rows(
    *,
    store: Store,
    month: Month,
    mb5b_path: Path,
    fiori_path: Path,
    prev_report_path: Path | None = None,
    zfi0156_path: Path | None = None,
    calc_path: Path | None = None,
    include_zero_activity: bool = False,
) -> list[ReportRow]:
    """Compute the report rows for one store, given the source files.

    Pure: only reads the supplied files; no network or browser.

    By default, materials with no MB5B activity (期末 == 0) are filtered
    out — that matches the manual workbook. Pass
    ``include_zero_activity=True`` to keep them.

    Optional inputs:
        prev_report_path: previous month's CA0X-盘点结果-YYYYMM.xlsx;
            used to compute 单价差异 (col 14).
        zfi0156_path: ZFI0156 export (or the pre-pivoted 上月数量需更新
            sheet); used to fill 上月使用金额 (col 12) and 对比 (col 13).
        calc_path: workbook containing the 计算 sheet (dish→material
            BOM with remark strings); used to fill 备注 (col 15). The
            sheet itself is hand-curated — we just look up by 物料号.
    """
    logger.info("loading MB5B  %s", mb5b_path)
    mb5b_rows = parse_mb5b_file(mb5b_path)
    logger.info("loading Fiori %s", fiori_path)
    counts = _read_fiori_count_by_matnr(fiori_path)
    cls_idx = material_classification_index()

    prev_prices: dict[str, float] = {}
    if prev_report_path:
        logger.info("loading prev report %s", prev_report_path)
        prev_prices = _read_prev_unit_price_by_matnr(prev_report_path)
    zfi_lookup: dict[str, Zfi0156Lookup] = {}
    if zfi0156_path:
        logger.info("loading ZFI0156 %s (filter werks=%s)", zfi0156_path, store.werks)
        zfi_lookup = _read_zfi0156_lookup(zfi0156_path, werks=store.werks)
    remarks: dict[str, str] = {}
    if calc_path:
        logger.info("loading 计算 sheet %s", calc_path)
        remarks = _read_calc_remarks_by_matnr(calc_path)

    out: list[ReportRow] = []
    n = 0
    for r in filter_by_werks(mb5b_rows, store.werks):
        matnr = coerce_matnr(r.get("Matnr", ""))
        if not matnr:
            continue
        if not include_zero_activity and not _has_stock_activity(r):
            continue
        n += 1
        zfi_rec = zfi_lookup.get(matnr)
        closing = _compact_int(_safe_num(r.get("ClosingQty")))

        # 单价: MB5B's UnitPrice column doesn't exist on the bare 15-col
        # MB5B Spreadsheet variant, so we fall back to ZFI0156's
        # 系统发出单价, and finally to ClosingAmt / ClosingQty
        # (mathematically: SAP's moving average price for this period).
        unit_price = _safe_num(r.get("UnitPrice"))
        if not isinstance(unit_price, (int, float)):
            if zfi_rec and zfi_rec.unit_price is not None:
                unit_price = zfi_rec.unit_price
            else:
                cqty = _safe_num(r.get("ClosingQty"))
                camt = _safe_num(r.get("ClosingAmt"))
                if (isinstance(cqty, (int, float)) and cqty
                        and isinstance(camt, (int, float))):
                    unit_price = round(float(camt) / float(cqty), 6)

        # 物料名称: MB5B Matxt → 分类 reference → ZFI0156 物料描述.
        matxt = (r.get("Matxt") or "").strip()
        if not matxt and matnr in cls_idx:
            matxt = cls_idx[matnr].description
        if not matxt and zfi_rec:
            matxt = zfi_rec.matxt

        # 单位 / 单位编码 / 单位描述: MB5B Meins (Chinese) is missing on
        # the 15-col layout; ZFI0156 supplies both Bun (SAP code) and
        # 单位描述 (Chinese). MeinsAlt is always present for the SAP code.
        unit_desc = (r.get("Meins") or "").strip()
        if not unit_desc and zfi_rec:
            unit_desc = zfi_rec.unit
        unit = unit_desc
        unit_code = (r.get("MeinsAlt") or "").strip()
        if not unit_code and zfi_rec:
            unit_code = zfi_rec.unit_code

        # Manual convention: when Fiori has no entry for an MB5B-active
        # material, the workbook's VLOOKUP returns #N/A and is implicitly
        # coerced to 0 — so usage = closing - 0 = closing. We mirror that
        # here so col 13 (对比) lines up. The blank-Fiori case where MB5B
        # also has no closing stock is already filtered upstream by
        # _has_stock_activity.
        counted_raw = counts.get(matnr, 0.0)
        counted: float | str = _compact_int(counted_raw)
        if isinstance(closing, (int, float)) and isinstance(counted, (int, float)):
            usage_qty: float | str = _compact_int(float(closing) - float(counted))
        else:
            usage_qty = ""
        if isinstance(usage_qty, (int, float)) and isinstance(unit_price, (int, float)):
            month_value: float | str = round(float(usage_qty) * float(unit_price), 2)
        else:
            month_value = ""

        classification = ""
        if matnr in cls_idx:
            classification = cls_idx[matnr].level1

        # 上月使用金额 from ZFI0156. The pre-pivoted 上月数量需更新 sheet
        # presents 系统发出金额 as a positive number (the manual file's
        # col 12 uses these directly). The raw ZFI0156 export has
        # negative values (outflow), so abs() normalises both shapes.
        if zfi_rec is None:
            prev_month_value: float | str = ""
            delta: float | str = ""
        else:
            prev_month_value = round(abs(float(zfi_rec.value)), 2)
            if isinstance(month_value, (int, float)):
                delta = round(float(month_value) - float(prev_month_value), 2)
            else:
                delta = ""

        # 单价差异: current 单价 - last month's 单价. Manual file uses
        # IFERROR(..., 0) so a missing prev-month price renders as 0
        # (not blank) — match that convention when a prev report was
        # provided but didn't have this material.
        prev_price = prev_prices.get(matnr)
        if not isinstance(unit_price, (int, float)):
            unit_price_diff: float | str = ""
        elif prev_price is None:
            unit_price_diff = 0 if prev_prices else ""
        else:
            unit_price_diff = round(float(unit_price) - float(prev_price), 4)

        out.append(ReportRow(
            row_no=n,
            matnr=matnr,
            matxt=matxt,
            closing_qty=closing,
            counted_qty=counted,
            unit=unit,
            unit_code=unit_code,
            unit_desc=unit_desc,
            usage_qty=usage_qty,
            unit_price=unit_price,
            month_value=month_value,
            prev_month_value=prev_month_value,
            delta=delta,
            unit_price_diff=unit_price_diff,
            remark=remarks.get(matnr, ""),
            classification=classification,
        ))
    logger.info("built %d rows for %s (%s)", n, store.werks, month.period)
    return out


def write_report_xlsx(
    rows: list[ReportRow], out_path: Path, *, sheet_name_template: str = "{werks}-本月-盘点结果."
) -> Path:
    """Write the assembled rows to an xlsx using OUTPUT_COLUMNS as the header."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Sheet title — the caller picks the template (and bakes in the plant
    # code for the multi-store extension later).
    title = (sheet_name_template.format(werks="") if rows else "盘点结果")[:31]

    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(list(OUTPUT_COLUMNS))
    for row in rows:
        ws.append(row.to_xlsx_row())
    # Generous column widths for legibility.
    widths = [8, 14, 32, 12, 12, 8, 10, 10, 12, 12, 14, 14, 14, 14, 22, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(str(out_path))
    logger.info("wrote %d rows → %s", len(rows), out_path)
    return out_path


def assemble_report(
    *,
    store: Store,
    month: Month,
    mb5b_path: Path,
    fiori_path: Path,
    out_dir: Path,
    file_name: str | None = None,
    include_zero_activity: bool = False,
    prev_report_path: Path | None = None,
    zfi0156_path: Path | None = None,
    calc_path: Path | None = None,
) -> Path:
    """Top-level: build rows and write the xlsx for one store/month."""
    rows = build_report_rows(
        store=store, month=month, mb5b_path=mb5b_path, fiori_path=fiori_path,
        prev_report_path=prev_report_path, zfi0156_path=zfi0156_path,
        calc_path=calc_path,
        include_zero_activity=include_zero_activity,
    )
    if file_name is None:
        file_name = f"{store.werks}-盘点结果-{month.period}.xlsx"
    return write_report_xlsx(rows, Path(out_dir) / file_name,
                             sheet_name_template=f"{store.werks}-本月-盘点结果.")
