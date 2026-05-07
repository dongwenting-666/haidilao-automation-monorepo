"""Unit tests for inventory_check.report — pure helpers + xlsx round-trip."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from inventory_check.dates import Month
from inventory_check.report import (
    OUTPUT_COLUMNS,
    ReportRow,
    TODO_MARKER,
    _compact_int,
    _read_calc_remarks_by_matnr,
    _read_fiori_count_by_matnr,
    _read_prev_unit_price_by_matnr,
    _read_zfi0156_value_by_matnr,
    _safe_num,
    assemble_report,
    build_report_rows,
    write_report_xlsx,
)
from inventory_check.stores import get_store


def _make_fiori_xlsx(tmp_path: Path, rows: list[tuple]) -> Path:
    """Build a minimal Fiori-shaped xlsx for tests."""
    p = tmp_path / "fiori.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "工厂", "物料号", "名称", "单位", "单位描述",
        "库存数量", "盘点数量", "状态", "盘点日期", "时间",
    ])
    for r in rows:
        ws.append(list(r))
    wb.save(str(p))
    return p


def _make_mb5b_file(tmp_path: Path, rows: list[list[str]]) -> Path:
    """Write a UTF-16 LE TSV file matching the rich MB5B export format
    (18 cols, including 物料描述 / 单价 / 单位)."""
    header = "\t".join([
        "ValA", "物料", "开始日期", "结束日期",
        "                   期初库存", "                  总收货数量",
        "                  总发货数量", "                   期末库存",
        "计", "                   期初金额", "                  总收货金额",
        "                  总发货金额", "                   期末金额",
        "货币", "物料描述", "单价", "单位", "20",
    ])
    lines = [header] + ["\t".join(r) for r in rows]
    text = "\r\n".join(lines) + "\r\n"
    p = tmp_path / "mb5b.xls"
    p.write_bytes(b"\xff\xfe" + text.encode("utf-16-le"))
    return p


def _make_bare_mb5b_file(tmp_path: Path, rows: list[list[str]]) -> Path:
    """Write the 14-column MB5B variant — no 物料描述 / 单价 / 单位.

    This is what SAP returns when the GUI export uses the default
    layout instead of the user's enriched variant. The report builder
    falls back to ZFI0156 / ClosingAmt-over-ClosingQty for unit price."""
    header = "\t".join([
        "ValA", "物料", "开始日期", "结束日期",
        "                   期初库存", "                  总收货数量",
        "                  总发货数量", "                   期末库存",
        "计", "                   期初金额", "                  总收货金额",
        "                  总发货金额", "                   期末金额",
        "货币",
    ])
    lines = [header] + ["\t".join(r) for r in rows]
    text = "\r\n".join(lines) + "\r\n"
    p = tmp_path / "mb5b_bare.xls"
    p.write_bytes(b"\xff\xfe" + text.encode("utf-16-le"))
    return p


# ── small helpers ──────────────────────────────────────────────────────


def test_safe_num_passes_through() -> None:
    assert _safe_num(60) == 60
    assert _safe_num(60.5) == 60.5
    assert _safe_num("60.5") == 60.5
    assert _safe_num("1,234.5") == 1234.5


def test_safe_num_blank() -> None:
    assert _safe_num("") == ""
    assert _safe_num(None) == ""


def test_safe_num_non_numeric_string_passes_through() -> None:
    assert _safe_num("abc") == "abc"


def test_compact_int_drops_trailing_zero() -> None:
    assert _compact_int(60.0) == 60 and isinstance(_compact_int(60.0), int)
    assert _compact_int(60.5) == 60.5
    assert _compact_int("") == ""


# ── _read_fiori_count_by_matnr ─────────────────────────────────────────


def test_read_fiori_count_basic(tmp_path: Path) -> None:
    p = _make_fiori_xlsx(tmp_path, [
        ("CA08", "1000049", "金标生抽", "L", "升", 200.9, 63.7, "已过账", "20260301", "192948"),
        ("CA08", "4509062", "烧酒", "BOT", "瓶", 60, 39, "已过账", "20260301", "192948"),
    ])
    counts = _read_fiori_count_by_matnr(p)
    assert counts["1000049"] == pytest.approx(63.7)
    assert counts["4509062"] == pytest.approx(39.0)


def test_read_fiori_count_skips_blank_matnr(tmp_path: Path) -> None:
    p = _make_fiori_xlsx(tmp_path, [
        ("CA08", "", "junk", "L", "升", 1, 2, "x", "x", "x"),
        ("CA08", "1000049", "ok", "L", "升", 1, 5, "x", "x", "x"),
    ])
    counts = _read_fiori_count_by_matnr(p)
    assert counts == {"1000049": 5.0}


def test_read_fiori_count_handles_string_numbers(tmp_path: Path) -> None:
    p = _make_fiori_xlsx(tmp_path, [
        ("CA08", "1000049", "x", "L", "升", "200", "63.7", "x", "x", "x"),
    ])
    counts = _read_fiori_count_by_matnr(p)
    assert counts["1000049"] == pytest.approx(63.7)


# ── build_report_rows (the core join) ──────────────────────────────────


def test_build_report_rows_matches_mb5b_to_fiori_count(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "63.7", "127.4", "-9.8", "181.3",
         "L", "0", "0", "0", "0",
         "CAD", "金标生抽", "3", "公升", "1"],
        ["CA08", "4509062", "2026.03.01", "2026.03.31",
         "60", "0", "0", "60",
         "BOT", "0", "0", "0", "0",
         "CAD", "烧酒", "12", "瓶", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [
        ("CA08", "1000049", "金标生抽", "L", "升-公升", 181.3, 34.3, "x", "x", "x"),
        ("CA08", "4509062", "烧酒", "BOT", "瓶", 60, 39, "x", "x", "x"),
    ])

    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert len(rows) == 2
    by_matnr = {r.matnr: r for r in rows}

    r1 = by_matnr["1000049"]
    assert r1.closing_qty == pytest.approx(181.3)
    assert r1.counted_qty == pytest.approx(34.3)
    assert r1.usage_qty == pytest.approx(147.0)  # 181.3 - 34.3
    assert r1.unit_price == 3
    assert r1.month_value == pytest.approx(441.0)  # 147 * 3

    r2 = by_matnr["4509062"]
    assert r2.closing_qty == 60 and isinstance(r2.closing_qty, int)
    assert r2.counted_qty == 39
    assert r2.usage_qty == 21
    assert r2.unit_price == 12
    assert r2.month_value == pytest.approx(252.0)  # 21 * 12


def test_build_report_rows_filters_to_store_werks(tmp_path: Path) -> None:
    """Materials from a different plant must not appear in CA08's report."""
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1", "2026.03.01", "2026.03.31",
         "10", "0", "0", "10",
         "L", "0", "0", "0", "0", "CAD", "ours", "1", "公升", "1"],
        ["CA09", "2", "2026.03.01", "2026.03.31",
         "20", "0", "0", "20",
         "L", "0", "0", "0", "0", "CAD", "theirs", "2", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert {r.matnr for r in rows} == {"1"}


def test_build_report_rows_default_count_zero_when_no_fiori_match(tmp_path: Path) -> None:
    """Manual convention: missing Fiori entry → counted=0, usage=closing.

    The workbook's VLOOKUP returns #N/A and gets coerced to 0; mirroring
    that lets col 13 (对比) reconcile against the manual report."""
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "9999", "2026.03.01", "2026.03.31",
         "100", "0", "0", "100",
         "L", "0", "0", "0", "0", "CAD", "lonely", "5", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert len(rows) == 1
    r = rows[0]
    assert r.counted_qty == 0
    assert r.usage_qty == 100
    assert r.month_value == 500.0


def test_build_report_rows_filters_zero_activity_by_default(tmp_path: Path) -> None:
    """A material with all-zero MB5B quantities is dropped by default."""
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1", "2026.03.01", "2026.03.31",
         "0", "0", "0", "0",
         "L", "0", "0", "0", "0", "CAD", "zeroes", "0", "公升", "1"],
        ["CA08", "2", "2026.03.01", "2026.03.31",
         "10", "0", "0", "5",
         "L", "0", "0", "0", "0", "CAD", "active", "1", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert {r.matnr for r in rows} == {"2"}


def test_build_report_rows_include_zero_activity_keeps_all(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1", "2026.03.01", "2026.03.31",
         "0", "0", "0", "0",
         "L", "0", "0", "0", "0", "CAD", "zeroes", "0", "公升", "1"],
        ["CA08", "2", "2026.03.01", "2026.03.31",
         "10", "0", "0", "5",
         "L", "0", "0", "0", "0", "CAD", "active", "1", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
        include_zero_activity=True,
    )
    assert {r.matnr for r in rows} == {"1", "2"}


def test_build_report_rows_unit_price_falls_back_to_zfi0156(tmp_path: Path) -> None:
    """Bare 14-col MB5B has no 单价; ZFI0156's 系统发出单价 takes over."""
    mb5b = _make_bare_mb5b_file(tmp_path, [
        ["CA08", "9999", "2026.04.01", "2026.04.30",
         "0", "10", "0", "10", "L", "0", "10", "0", "10", "CAD"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    # ZFI0156 raw with 系统发出单价 column — the only price source available.
    zfi = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "大类", "系统发出单价", "数量", "系统发出金额",
    ])
    ws.append(["46118", "46147", "CA08", "8店", "9999", "X",
               "L", "升", "费用类", 7.50, 1.0, -7.50])
    wb.save(str(zfi))

    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 4),
        mb5b_path=mb5b, fiori_path=fiori, zfi0156_path=zfi,
    )
    assert len(rows) == 1
    r = rows[0]
    assert r.unit_price == pytest.approx(7.50)
    # month_value = usage * price = (10 - 0) * 7.50 = 75.0
    assert r.month_value == pytest.approx(75.0)


def test_build_report_rows_unit_price_falls_back_to_closing_amt_over_qty(
    tmp_path: Path,
) -> None:
    """No MB5B 单价, no ZFI0156 entry — fall back to ClosingAmt / ClosingQty
    (the in-period moving average price)."""
    mb5b = _make_bare_mb5b_file(tmp_path, [
        # closing_qty=10, closing_amt=42.50 → 单价 = 4.25
        ["CA08", "9999", "2026.04.01", "2026.04.30",
         "0", "10", "0", "10", "L", "0", "42.50", "0", "42.50", "CAD"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 4),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert len(rows) == 1
    r = rows[0]
    assert r.unit_price == pytest.approx(4.25)
    # usage = 10 - 0 (no Fiori entry → counted=0), month_value = 10 * 4.25 = 42.50
    assert r.month_value == pytest.approx(42.50)


def test_build_report_rows_classification_lookup(tmp_path: Path) -> None:
    """1000049 is in the real 分类 CSV — check the level1 maps in."""
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "1", "0", "0", "1",
         "L", "0", "0", "0", "0", "CAD", "金标生抽", "1", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert rows[0].classification == "成本-小料台类"


def test_build_report_rows_fallback_matxt_from_classification(tmp_path: Path) -> None:
    """If MB5B 物料描述 is blank, fall back to 分类 sheet."""
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "1", "0", "0", "1",
         "L", "0", "0", "0", "0", "CAD", "", "1", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert "金标生抽" in rows[0].matxt


# ── _read_prev_unit_price_by_matnr ─────────────────────────────────────


def _make_prev_report_xlsx(tmp_path: Path, rows: list[tuple]) -> Path:
    """Build a minimal previous-month 盘点结果 xlsx (using our own schema)."""
    p = tmp_path / "prev.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(list(OUTPUT_COLUMNS))
    for r in rows:
        # Pad to 17 cols
        cells = list(r) + [""] * (17 - len(r))
        ws.append(cells)
    wb.save(str(p))
    return p


def test_read_prev_unit_price_basic(tmp_path: Path) -> None:
    p = _make_prev_report_xlsx(tmp_path, [
        # row_no, matnr, matxt, closing, counted, unit, ucode, udesc,
        # usage, unit_price, ...
        (1, "1000049", "生抽", 100, 50, "L", "L", "升", 50, 2.40, 120),
        (2, "4509062", "烧酒", 60, 39, "瓶", "BOT", "瓶", 21, 7.10, 149.1),
    ])
    prices = _read_prev_unit_price_by_matnr(p)
    assert prices["1000049"] == pytest.approx(2.40)
    assert prices["4509062"] == pytest.approx(7.10)


def test_read_prev_unit_price_skips_non_numeric(tmp_path: Path) -> None:
    p = _make_prev_report_xlsx(tmp_path, [
        (1, "1000049", "x", 100, 50, "L", "L", "升", 50, "n/a", ""),
        (2, "4509062", "y", 60, 39, "BOT", "BOT", "瓶", 21, 7.10, 149.1),
    ])
    prices = _read_prev_unit_price_by_matnr(p)
    assert "1000049" not in prices
    assert prices["4509062"] == pytest.approx(7.10)


def test_read_prev_unit_price_strips_trailing_dot_zero(tmp_path: Path) -> None:
    """xlsx may store matnrs as floats; we want the int-string back."""
    p = tmp_path / "prev.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(list(OUTPUT_COLUMNS))
    ws.append([1, 1000049.0, "x", 0, 0, "", "", "", 0, 5.0, 0,
               "", "", "", "", "", ""])
    wb.save(str(p))
    prices = _read_prev_unit_price_by_matnr(p)
    assert "1000049" in prices
    assert prices["1000049"] == 5.0


# ── _read_zfi0156_value_by_matnr ───────────────────────────────────────


def _make_pivoted_zfi(tmp_path: Path, rows: list[tuple]) -> Path:
    """Mimic the 上月数量需更新 sheet shape (3 cols, weird row-1 title)."""
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["", "值", ""])
    ws.append(["物料", "求和项:单价1", "求和项:系统发出金额"])
    for r in rows:
        ws.append(list(r))
    wb.save(str(p))
    return p


def test_read_zfi0156_pre_pivoted(tmp_path: Path) -> None:
    p = _make_pivoted_zfi(tmp_path, [
        ("1000049", 2.34, 333.5),
        ("4509062", 7.10, 100),
    ])
    out = _read_zfi0156_value_by_matnr(p)
    # Reader passes through whatever sign the source has; the report layer
    # takes abs() to normalise to the manual's positive convention.
    assert out["1000049"] == pytest.approx(333.5)
    assert out["4509062"] == pytest.approx(100)


def test_read_zfi0156_pivot_handles_float_matnrs(tmp_path: Path) -> None:
    p = _make_pivoted_zfi(tmp_path, [
        (1000049.0, 2.34, -333.5),
    ])
    out = _read_zfi0156_value_by_matnr(p)
    assert "1000049" in out


def test_read_zfi0156_raw_export(tmp_path: Path) -> None:
    """Raw 上月数量zfi0156 sheet — group by 物料 sum 系统发出金额."""
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "单位编码", "期初数量", "收货数量", "发货数量",
        "期末数量", "期初金额", "收货金额", "系统发出金额",
    ])
    ws.append(["46054", "46081", "CA08", "...", "1000049", "生抽",
               "L", "升", "L", 100, 50, -50, 100, 0, 0, -125])
    ws.append(["46054", "46081", "CA08", "...", "1000049", "生抽",
               "L", "升", "L", 100, 50, -50, 100, 0, 0, -75])
    ws.append(["46054", "46081", "CA08", "...", "4509062", "烧酒",
               "BOT", "瓶", "BOT", 0, 0, 0, 0, 0, 0, -50])
    wb.save(str(p))

    out = _read_zfi0156_value_by_matnr(p)
    assert out["1000049"] == pytest.approx(-200)
    assert out["4509062"] == pytest.approx(-50)


def test_read_zfi0156_raw_with_werks_filter(tmp_path: Path) -> None:
    """Raw region-wide ZFI0156 must be filterable to one plant — the
    manual workbook's 上月使用金额 is per-store, not aggregated."""
    from inventory_check.report import _read_zfi0156_lookup
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "单位编码", "期初数量", "收货数量", "发货数量",
        "期末数量", "期初金额", "收货金额", "系统发出金额",
    ])
    ws.append(["46054", "46081", "CA01", "1店", "1000049", "生抽",
               "L", "升", "L", 0, 0, 0, 0, 0, 0, -1000])
    ws.append(["46054", "46081", "CA08", "8店", "1000049", "生抽",
               "L", "升", "L", 0, 0, 0, 0, 0, 0, -125])
    ws.append(["46054", "46081", "CA08", "8店", "1000049", "生抽",
               "L", "升", "L", 0, 0, 0, 0, 0, 0, -75])
    wb.save(str(p))

    region = _read_zfi0156_lookup(p)
    assert region["1000049"].value == pytest.approx(-1200)

    ca08 = _read_zfi0156_lookup(p, werks="CA08")
    assert ca08["1000049"].value == pytest.approx(-200)
    assert ca08["1000049"].matxt == "生抽"
    assert ca08["1000049"].unit == "升"
    assert ca08["1000049"].unit_code == "L"


def test_read_zfi0156_raw_handles_18_digit_zero_padded_matnr(tmp_path: Path) -> None:
    """SAP often pads matnrs to 18 chars on raw exports — the lookup
    must reconcile that with the unpadded form MB5B/Fiori use."""
    from inventory_check.report import _read_zfi0156_lookup
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "单位编码", "期初数量", "收货数量", "发货数量",
        "期末数量", "期初金额", "收货金额", "系统发出金额",
    ])
    ws.append(["46054", "46081", "CA08", "8店", "000000000001000049", "生抽",
               "L", "升", "L", 0, 0, 0, 0, 0, 0, -100])
    wb.save(str(p))

    out = _read_zfi0156_lookup(p, werks="CA08")
    assert "1000049" in out  # zero-padded form normalised
    assert out["1000049"].value == pytest.approx(-100)


def test_read_zfi0156_pivoted_with_unit_price_column(tmp_path: Path) -> None:
    """Pre-pivoted shape: 求和项:单价1 should populate ``unit_price``."""
    from inventory_check.report import _read_zfi0156_lookup
    p = _make_pivoted_zfi(tmp_path, [
        ("1000049", 2.34, -333.5),
    ])
    out = _read_zfi0156_lookup(p)
    rec = out["1000049"]
    assert rec.unit_price == pytest.approx(2.34)
    assert rec.value == pytest.approx(-333.5)


# ── _read_calc_remarks_by_matnr ────────────────────────────────────────


def _make_calc_xlsx(tmp_path: Path, rows: list[tuple]) -> Path:
    """Build a workbook with a 计算 sheet for testing the remark lookup.

    Each tuple is (matnr, remark) — we put them in cols R (18) and Z (26)
    to match the real schema.
    """
    p = tmp_path / "calc.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "计算"
    # Header row with 26 columns; col R = 物料号, col Z = the final remark
    headers = [""] * 26
    headers[17] = "物料号"
    headers[25] = "Z"
    ws.append(headers)
    for matnr, remark in rows:
        cells = [""] * 26
        cells[17] = matnr
        cells[25] = remark
        ws.append(cells)
    wb.save(str(p))
    return p


def test_read_calc_remarks_basic(tmp_path: Path) -> None:
    p = _make_calc_xlsx(tmp_path, [
        ("1000049", "多用1.50公斤"),
        ("4509062", "少用0.20公斤"),
    ])
    out = _read_calc_remarks_by_matnr(p)
    assert out["1000049"] == "多用1.50公斤"
    assert out["4509062"] == "少用0.20公斤"


def test_read_calc_remarks_skips_empty_remark(tmp_path: Path) -> None:
    p = _make_calc_xlsx(tmp_path, [
        ("1000049", ""),
        ("4509062", "多用1.00"),
    ])
    out = _read_calc_remarks_by_matnr(p)
    assert "1000049" not in out
    assert out["4509062"] == "多用1.00"


def test_read_calc_remarks_dedups_keeping_last(tmp_path: Path) -> None:
    """A material can map to multiple dishes; the last seen value wins."""
    p = _make_calc_xlsx(tmp_path, [
        ("1000049", "first"),
        ("1000049", "second"),
    ])
    out = _read_calc_remarks_by_matnr(p)
    assert out["1000049"] == "second"


def test_read_calc_remarks_handles_float_matnrs(tmp_path: Path) -> None:
    p = tmp_path / "calc.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "计算"
    headers = [""] * 26
    headers[17] = "物料号"
    ws.append(headers)
    cells = [""] * 26
    cells[17] = 4703100.0  # float matnr
    cells[25] = "多用5.00"
    ws.append(cells)
    wb.save(str(p))
    out = _read_calc_remarks_by_matnr(p)
    assert "4703100" in out
    # Regression: the buggy rstrip would have produced "47031".
    assert "47031" not in out


def test_read_calc_remarks_falls_back_to_active_sheet_if_no_named_sheet(
    tmp_path: Path,
) -> None:
    p = tmp_path / "calc.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"  # not "计算"
    headers = [""] * 26
    headers[17] = "物料号"
    ws.append(headers)
    cells = [""] * 26
    cells[17] = "1000049"
    cells[25] = "多用1.00"
    ws.append(cells)
    wb.save(str(p))
    out = _read_calc_remarks_by_matnr(p)
    assert out["1000049"] == "多用1.00"


# ── build_report_rows with the new optional inputs ─────────────────────


def test_build_report_rows_fills_unit_price_diff(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "10", "0", "0", "10", "L", "0", "0", "0", "0", "CAD",
         "生抽", "3", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    prev = _make_prev_report_xlsx(tmp_path, [
        (1, "1000049", "生抽", 0, 0, "", "", "", 0, 2.40, 0),
    ])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori, prev_report_path=prev,
    )
    assert len(rows) == 1
    assert rows[0].unit_price_diff == pytest.approx(0.6)


def test_build_report_rows_fills_prev_month_value_and_delta(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "100", "0", "0", "50", "L", "0", "0", "0", "0", "CAD",
         "生抽", "3", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [
        ("CA08", "1000049", "生抽", "L", "升", 100, 20, "x", "x", "x"),
    ])
    # ZFI0156 says last month outflow was -333.5 (a negative = issue)
    zfi = _make_pivoted_zfi(tmp_path, [("1000049", 2.34, -333.5)])

    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori, zfi0156_path=zfi,
    )
    assert len(rows) == 1
    r = rows[0]
    # 使用 = 50 - 20 = 30; 单价 = 3; 本月使用金额 = 90
    assert r.month_value == pytest.approx(90)
    # ZFI 系统发出金额 → abs (manual convention is positive)
    assert r.prev_month_value == pytest.approx(333.5)
    # 对比 = 本月 - 上月 = 90 - 333.5 = -243.5
    assert r.delta == pytest.approx(-243.5)


def test_build_report_rows_unit_price_diff_zero_when_prev_missing_material(tmp_path: Path) -> None:
    """Manual file uses IFERROR(..., 0) — when a prev report was supplied
    but doesn't have this material, output 0 (not blank)."""
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "9999", "2026.03.01", "2026.03.31",
         "10", "0", "0", "10", "L", "0", "0", "0", "0", "CAD",
         "newbie", "5", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    # prev_report has a different material — 9999 won't be found.
    prev = _make_prev_report_xlsx(tmp_path, [
        (1, "1000049", "x", 0, 0, "", "", "", 0, 2.40, 0),
    ])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori, prev_report_path=prev,
    )
    assert rows[0].unit_price_diff == 0


def test_build_report_rows_blank_when_no_prev_or_zfi(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "10", "0", "0", "10", "L", "0", "0", "0", "0", "CAD",
         "生抽", "3", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori,
    )
    assert rows[0].prev_month_value == ""
    assert rows[0].delta == ""
    assert rows[0].unit_price_diff == ""
    assert rows[0].remark == ""


def test_build_report_rows_fills_remark_from_calc(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "10", "0", "0", "10", "L", "0", "0", "0", "0", "CAD",
         "生抽", "3", "公升", "1"],
        ["CA08", "9999", "2026.03.01", "2026.03.31",
         "5", "0", "0", "5", "L", "0", "0", "0", "0", "CAD",
         "noremark", "1", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [])
    calc = _make_calc_xlsx(tmp_path, [("1000049", "少用1.23公斤")])
    rows = build_report_rows(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori, calc_path=calc,
    )
    by_matnr = {r.matnr: r for r in rows}
    assert by_matnr["1000049"].remark == "少用1.23公斤"
    # 9999 has no entry in the 计算 sheet → blank.
    assert by_matnr["9999"].remark == ""


# ── ReportRow.to_xlsx_row ──────────────────────────────────────────────


def test_to_xlsx_row_length_matches_columns() -> None:
    r = ReportRow(
        row_no=1, matnr="x", matxt="x", closing_qty=1, counted_qty=1,
        unit="L", unit_code="L", unit_desc="升",
        usage_qty=0, unit_price=0, month_value=0,
        prev_month_value="", delta="", unit_price_diff="", remark="",
        classification="cls",
    )
    out = r.to_xlsx_row()
    assert len(out) == len(OUTPUT_COLUMNS)


def test_to_xlsx_row_emits_provided_values_for_filled_columns() -> None:
    r = ReportRow(
        row_no=1, matnr="x", matxt="x", closing_qty=1, counted_qty=1,
        unit="L", unit_code="L", unit_desc="升",
        usage_qty=0, unit_price=0, month_value=100,
        prev_month_value=80, delta=20, unit_price_diff=0.5, remark="多用2.50",
        classification="cls",
    )
    out = r.to_xlsx_row()
    assert out[10] == 100   # 本月使用金额
    assert out[11] == 80    # 上月使用金额
    assert out[12] == 20    # 对比
    assert out[13] == 0.5   # 单价差异
    assert out[14] == "多用2.50"  # 备注


def test_to_xlsx_row_passes_through_blanks_when_lookups_missed() -> None:
    r = ReportRow(
        row_no=1, matnr="x", matxt="x", closing_qty=1, counted_qty=1,
        unit="L", unit_code="L", unit_desc="升",
        usage_qty=0, unit_price=0, month_value=0,
        prev_month_value="", delta="", unit_price_diff="", remark="",
        classification="cls",
    )
    out = r.to_xlsx_row()
    assert out[11] == ""
    assert out[12] == ""
    assert out[13] == ""


# ── write_report_xlsx round-trip ───────────────────────────────────────


def test_write_report_xlsx_round_trip(tmp_path: Path) -> None:
    rows = [
        ReportRow(row_no=1, matnr="1000049", matxt="生抽",
                  closing_qty=181.3, counted_qty=34.3,
                  unit="公升", unit_code="L", unit_desc="公升",
                  usage_qty=147.0, unit_price=3.0, month_value=441.0,
                  prev_month_value=333.5, delta=107.5, unit_price_diff=0.66,
                  remark="少用1.23公斤",
                  classification="成本-小料台类"),
    ]
    out = tmp_path / "r.xlsx"
    write_report_xlsx(rows, out)
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert tuple(c.value for c in ws[1]) == OUTPUT_COLUMNS
    assert ws.cell(row=2, column=2).value == "1000049"
    assert ws.cell(row=2, column=4).value == pytest.approx(181.3)
    assert ws.cell(row=2, column=5).value == pytest.approx(34.3)
    assert ws.cell(row=2, column=9).value == pytest.approx(147.0)
    assert ws.cell(row=2, column=11).value == pytest.approx(441.0)
    assert ws.cell(row=2, column=12).value == pytest.approx(333.5)
    assert ws.cell(row=2, column=13).value == pytest.approx(107.5)
    assert ws.cell(row=2, column=14).value == pytest.approx(0.66)
    assert ws.cell(row=2, column=15).value == "少用1.23公斤"
    assert ws.cell(row=2, column=17).value == "成本-小料台类"


def test_write_report_xlsx_empty_rows_still_writes_header(tmp_path: Path) -> None:
    out = tmp_path / "r.xlsx"
    write_report_xlsx([], out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1
    assert tuple(c.value for c in ws[1]) == OUTPUT_COLUMNS


# ── assemble_report end-to-end (with synthetic inputs) ─────────────────


def test_assemble_report_writes_named_xlsx(tmp_path: Path) -> None:
    mb5b = _make_mb5b_file(tmp_path, [
        ["CA08", "1000049", "2026.03.01", "2026.03.31",
         "1", "0", "0", "1",
         "L", "0", "0", "0", "0", "CAD", "生抽", "1", "公升", "1"],
    ])
    fiori = _make_fiori_xlsx(tmp_path, [
        ("CA08", "1000049", "生抽", "L", "升", 1, 0.5, "x", "x", "x"),
    ])
    out_dir = tmp_path / "out"
    p = assemble_report(
        store=get_store("CA8DKG"), month=Month(2026, 3),
        mb5b_path=mb5b, fiori_path=fiori, out_dir=out_dir,
    )
    assert p.exists()
    assert p.name == "CA08-盘点结果-202603.xlsx"
