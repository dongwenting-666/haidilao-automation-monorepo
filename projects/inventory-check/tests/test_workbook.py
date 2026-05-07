"""Unit tests for inventory_check.workbook — the multi-sheet workbook
assembly that produces the manual-shaped output.

Coverage focus: the data-shaping helpers (_replace_pos_sheet,
_rebuild_sheet3_pivot, _replace_mb5b_sheet, _replace_zfi_pivot_sheet,
_replace_report_sheet). The full assemble_workbook flow uses
shutil.copy + load_workbook on a real template — too slow for unit
tests, covered by the e2e validation script instead.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from inventory_check.workbook import (
    POS_PIVOT_SHEET,
    POS_RAW_SHEET,
    REPORT_SHEET,
    ZFI_PIVOT_SHEET,
    ZFI_RAW_SHEET,
    _rebuild_sheet3_pivot,
    _replace_pos_sheet,
    _replace_report_sheet,
    _replace_zfi_pivot_sheet,
    _replace_zfi_raw_sheet,
)
from inventory_check.report import ReportRow


# ── _replace_pos_sheet ────────────────────────────────────────────────


def _make_pos_xlsx(tmp_path: Path, rows: list[list]) -> Path:
    """Build a POS xlsx in the pos-crawler shape (12 cols, 红火台销售汇总)."""
    p = tmp_path / "pos.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "红火台销售汇总"
    ws.append([
        "检索", "门店名称", "编码", "菜品编码", "菜品短编码", "菜品名称",
        "规格", "出品数量", "退菜数量", "实际出品数据（出品数量-退菜数量）",
        "大类名称", "子类名称",
    ])
    for r in rows:
        ws.append(list(r))
    wb.save(str(p))
    return p


def _make_target_pos_sheet() -> object:
    """A blank workbook with just a 红火台销售汇总 header — mimics what's
    in the template before _replace_pos_sheet runs."""
    wb = Workbook()
    ws = wb.active
    ws.title = POS_RAW_SHEET
    ws.append([
        "检索", "门店名称", "编码", "菜品编码", "菜品短编码", "菜品名称",
        "规格", "出品数量", "退菜数量", "实际出品数据（出品数量-退菜数量）",
        "大类名称", "子类名称",
    ])
    return ws


def test_replace_pos_sheet_basic_copy(tmp_path: Path) -> None:
    src = _make_pos_xlsx(tmp_path, [
        ["加拿大八店910045431108705整份", "加拿大八店", None, "91004543",
         "1108705", "蜂蜜照烧鸡翅", "整份", 143, 1, 142, "小吃甜品", "小吃类"],
        ["加拿大八店1002013841183整份", "加拿大八店", None, "10020138",
         "41183", "芋头", "整份", 51, 0, 51, "素菜", "根茎类"],
    ])
    ws = _make_target_pos_sheet()
    written = _replace_pos_sheet(ws, src)
    assert written == 2
    # col 1 (检索) is recomputed from cols 2/4/5/7 to match the manual's
    # int-coerced concatenation, so the input string is replaced.
    assert ws.cell(row=2, column=1).value == "加拿大八店910045431108705整份"
    assert ws.cell(row=2, column=10).value == 142
    # Dish codes (cols 4 & 5) are coerced to int when numeric — manual
    # stores them as int and the 计算 sheet's CONCAT references depend on it.
    assert ws.cell(row=3, column=4).value == 10020138


def test_replace_pos_sheet_skips_blank_rows(tmp_path: Path) -> None:
    """pos-crawler can emit blank trailing rows when a page underfills;
    the replacement must drop them so Sheet3's pivot doesn't pick up
    a None key."""
    src = _make_pos_xlsx(tmp_path, [
        ["k1", "加拿大八店", None, "1", "2", "x", "整份", 10, 0, 10, "类", "子"],
        [None, None, None, None, None, None, None, None, None, None, None, None],
        ["k2", "加拿大八店", None, "3", "4", "y", "整份", 5, 0, 5, "类", "子"],
    ])
    ws = _make_target_pos_sheet()
    written = _replace_pos_sheet(ws, src)
    assert written == 2
    assert ws.max_row == 3  # header + 2 data rows


def test_replace_pos_sheet_clears_old_data(tmp_path: Path) -> None:
    """If the template already had old POS rows, they must be wiped —
    leftover rows would feed wrong values into Sheet3's pivot."""
    src = _make_pos_xlsx(tmp_path, [
        ["new-key", "加拿大八店", None, "9", "9", "new", "整份", 1, 0, 1, "类", "子"],
    ])
    ws = _make_target_pos_sheet()
    ws.append(["stale-key-1", "加拿大八店", None, "?", "?", "old", "整份", 99, 0, 99, "类", "子"])
    ws.append(["stale-key-2", "加拿大八店", None, "?", "?", "old", "整份", 88, 0, 88, "类", "子"])
    assert ws.max_row == 3

    _replace_pos_sheet(ws, src)
    assert ws.max_row == 2  # header + 1 new row
    # col 1 is recomputed from store/code/short-code/spec — the input
    # "new-key" sentinel is overwritten with "加拿大八店99整份".
    assert ws.cell(row=2, column=1).value == "加拿大八店99整份"


# ── _rebuild_sheet3_pivot ─────────────────────────────────────────────


def test_rebuild_sheet3_pivot_groups_and_sums(tmp_path: Path) -> None:
    """Two rows with the same dish/short/spec sum into one Sheet3 entry.

    The pivot is what 计算!M reads — the 计算 sheet's lookup_key
    (=C&F&G&J) must hit a single Sheet3 row to find the right portion
    count for that dish×spec. The pivot rebuilds the key from
    store/dish_code/short_code/spec the same way _replace_pos_sheet
    does, so leading-zero / int-coercion stays consistent.
    """
    src = _make_pos_xlsx(tmp_path, [
        ["bogus-input-key-1", "加拿大八店", None, "1", "11", "x", "整份", 30, 0, 30, "c", "s"],
        ["bogus-input-key-2", "加拿大八店", None, "1", "11", "x", "整份", 70, 0, 70, "c", "s"],
        ["bogus-input-key-3", "加拿大八店", None, "2", "22", "y", "整份", 50, 0, 50, "c", "s"],
    ])
    wb = Workbook()
    ws = wb.create_sheet(POS_PIVOT_SHEET)
    keys = _rebuild_sheet3_pivot(ws, src)
    assert keys == 2

    # Find the data rows (skip the title rows in 1-6).
    pivot_data = {}
    for r in range(7, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k and k != "Grand Total" and k != "加拿大八店":
            pivot_data[k] = v
    # Keys are recomputed from store + int(dish) + int(short) + spec —
    # the input col-1 sentinels are ignored.
    assert pivot_data == {"加拿大八店111整份": 100, "加拿大八店222整份": 50}

    # Final row must be Grand Total = sum of all values.
    last = ws.max_row
    assert ws.cell(row=last, column=1).value == "Grand Total"
    assert ws.cell(row=last, column=2).value == 150


def test_rebuild_sheet3_pivot_strips_leading_zeros_from_dish_codes(
    tmp_path: Path,
) -> None:
    """The POS source emits zero-padded dish codes like '01060061'; the
    manual's 计算!A formula uses the int form ('1060061'). The Sheet3
    pivot must rebuild the key with int-coerced codes so VLOOKUP from
    计算 hits — otherwise 实收数量 falls back to 0 for every padded
    dish code (the bug we found in the April regen)."""
    src = _make_pos_xlsx(tmp_path, [
        # Padded source: 01060061 / None short / 四宫格 — must lookup as
        # '加拿大八店1060061四宫格' in Sheet3.
        ["加拿大八店010600614宫格", "加拿大八店", None, "01060061", None,
         "清油麻辣火锅", "四宫格", 100, 0, 100, "c", "s"],
    ])
    wb = Workbook()
    ws = wb.create_sheet(POS_PIVOT_SHEET)
    _rebuild_sheet3_pivot(ws, src)
    # Find the data row.
    found = False
    for r in range(7, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "加拿大八店1060061四宫格":
            assert ws.cell(row=r, column=2).value == 100
            found = True
            break
    assert found, "Sheet3 must contain the int-coerced lookup key"


def test_replace_pos_sheet_handles_none_short_code(tmp_path: Path) -> None:
    """When 菜品短编码 (col 5) is None, the recomputed 检索 must treat
    it as empty string — not embed the literal 'None'. Manual's
    ``=B&D&E&G`` treats None as empty in Excel concat."""
    src = _make_pos_xlsx(tmp_path, [
        ["should-be-overwritten", "加拿大八店", None, "70000769", None,
         "寿喜锅底", "单锅", 5, 0, 5, "c", "s"],
    ])
    ws = _make_target_pos_sheet()
    _replace_pos_sheet(ws, src)
    # int(70000769) + '' + '单锅' → 加拿大八店70000769单锅 (no 'None' substring)
    assert ws.cell(row=2, column=1).value == "加拿大八店70000769单锅"


def test_rebuild_sheet3_pivot_layout_matches_manual(tmp_path: Path) -> None:
    """Row 3-6 must match the manual workbook's pivot title layout
    exactly — 计算 sheet doesn't reference these rows but humans do
    when auditing the file, and Excel's pivot refresh would re-emit
    the same shape."""
    src = _make_pos_xlsx(tmp_path, [
        ["k", "加拿大八店", None, "1", "2", "x", "整份", 10, 0, 10, "c", "s"],
    ])
    wb = Workbook()
    ws = wb.create_sheet(POS_PIVOT_SHEET)
    _rebuild_sheet3_pivot(ws, src)

    assert ws.cell(row=3, column=1).value == "求和项:实际出品数据（出品数量-退菜数量）"
    assert ws.cell(row=4, column=1).value == "Row Labels"
    assert ws.cell(row=4, column=2).value == "Grand Total"
    assert ws.cell(row=5, column=1).value == "(blank)"
    assert ws.cell(row=6, column=2).value == 0


def test_rebuild_sheet3_pivot_skips_non_numeric(tmp_path: Path) -> None:
    """If a POS row has a non-numeric 实际出品数据 (formula-as-text or a
    blank), it shouldn't crash the pivot — just skip the value."""
    src = _make_pos_xlsx(tmp_path, [
        ["k1", "加拿大八店", None, "1", "1", "x", "整份", 10, 0, 10, "c", "s"],
        ["k2", "加拿大八店", None, "2", "2", "y", "整份", 5, 0, "blank", "c", "s"],
    ])
    wb = Workbook()
    ws = wb.create_sheet(POS_PIVOT_SHEET)
    _rebuild_sheet3_pivot(ws, src)

    # k1 should be present with value 10; k2 either absent or value 0
    keys = {}
    for r in range(7, ws.max_row + 1):
        k = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if k and k not in ("Grand Total", "加拿大八店"):
            keys[k] = v
    # Keys are recomputed from store + int(dish) + int(short) + spec.
    assert keys.get("加拿大八店11整份") == 10
    # k2 is either absent or aggregated to 0 (depending on whether the row had any int values)
    assert keys.get("加拿大八店22整份", 0) in (0, None)


# ── _replace_zfi_pivot_sheet ──────────────────────────────────────────


def _make_target_zfi_pivot_sheet() -> object:
    """A blank workbook mimicking the template's 上月数量需更新 sheet
    (row 1 = title, row 2 = header, rows 3+ = data)."""
    wb = Workbook()
    ws = wb.active
    ws.title = ZFI_PIVOT_SHEET
    ws.append([None, "值", None])
    ws.append(["物料", "求和项:单价1", "求和项:系统发出金额"])
    return ws


def test_replace_zfi_raw_filters_by_werks(tmp_path: Path) -> None:
    """Region-wide ZFI export covers CA01-CA09; per-store workbooks must
    only show their own werks rows. Without the filter, a CA01 report
    would show all 8 stores' data in 上月数量zfi0156, leaking other
    stores' numbers into auditors' view of CA01."""
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "大类", "系统发出单价", "数量", "系统发出金额",
    ])
    ws.append(["x", "x", "CA01", "1店", "1", "a", "L", "升", "类", 1, 1, -1])
    ws.append(["x", "x", "CA08", "8店", "2", "b", "L", "升", "类", 2, 2, -4])
    ws.append(["x", "x", "CA08", "8店", "3", "c", "L", "升", "类", 3, 3, -9])
    ws.append(["x", "x", "CA09", "9店", "4", "d", "L", "升", "类", 4, 4, -16])
    wb.save(str(p))

    target_wb = Workbook()
    target = target_wb.active
    target.title = ZFI_RAW_SHEET
    target.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "大类", "系统发出单价", "数量", "系统发出金额",
    ])

    written = _replace_zfi_raw_sheet(target, p, werks="CA08")
    assert written == 2  # only CA08's two rows
    assert target.max_row == 3  # header + 2 data rows
    assert target.cell(row=2, column=3).value == "CA08"
    assert target.cell(row=3, column=3).value == "CA08"


def test_replace_zfi_raw_no_filter_keeps_all_rows(tmp_path: Path) -> None:
    """When werks is None (back-compat), all rows pass through."""
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述"])
    ws.append(["x", "x", "CA01", "1店", "1", "a"])
    ws.append(["x", "x", "CA08", "8店", "2", "b"])
    wb.save(str(p))

    target_wb = Workbook()
    target = target_wb.active
    target.title = ZFI_RAW_SHEET
    target.append(["开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述"])

    written = _replace_zfi_raw_sheet(target, p)  # no werks
    assert written == 2


def test_replace_zfi_pivot_normalises_to_positive(tmp_path: Path) -> None:
    """Manual's 上月使用金额 column is positive (the manual pivot took
    abs); ours must match so col 13 (对比) sign is right."""
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "大类", "系统发出单价", "数量", "系统发出金额",
    ])
    ws.append(["x", "x", "CA08", "8店", "1000049", "生抽",
               "L", "升", "费用类", 3.0, 100, -300.5])
    wb.save(str(p))

    target = _make_target_zfi_pivot_sheet()
    written = _replace_zfi_pivot_sheet(target, p, werks="CA08")
    assert written == 1
    # Row 3 = first data row.
    assert target.cell(row=3, column=1).value == 1000049
    assert target.cell(row=3, column=3).value == 300.5  # abs(-300.5)


def test_replace_zfi_pivot_preserves_header_rows(tmp_path: Path) -> None:
    """Rows 1+2 are the manual's title/header — must survive untouched
    or the matnr-column lookup at col 1 stops working."""
    p = tmp_path / "zfi.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([
        "开始日期", "结束日期", "工厂", "工厂描述", "物料", "物料描述",
        "Bun", "单位描述", "大类", "系统发出单价", "数量", "系统发出金额",
    ])
    ws.append(["x", "x", "CA08", "8店", "1", "x", "L", "升", "x", 1, 1, -1])
    wb.save(str(p))

    target = _make_target_zfi_pivot_sheet()
    # Pre-populate row 3 with stale data to ensure clearing.
    target.append(["STALE", 999, 999])
    _replace_zfi_pivot_sheet(target, p, werks="CA08")
    assert target.cell(row=1, column=2).value == "值"
    assert target.cell(row=2, column=1).value == "物料"
    assert target.cell(row=2, column=2).value == "求和项:单价1"


# ── _replace_report_sheet ─────────────────────────────────────────────


def _make_target_report_sheet() -> object:
    """A workbook with the report header — what _replace_report_sheet expects."""
    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_SHEET
    ws.append([
        "行号", "物料编码", "物料名称", "库存数量", "盘点数量", "单位",
        "单位编码", "单位描述", "使用数量", "单价", "本月使用金额",
        "上月使用金额", "对比", "单价差异", "备注（每月刷新）", "回复", "分类",
    ])
    return ws


def _make_report_row(row_no: int, matnr: str = "1000049") -> ReportRow:
    return ReportRow(
        row_no=row_no, matnr=matnr, matxt="x",
        closing_qty=10, counted_qty=5, unit="升", unit_code="L", unit_desc="升",
        usage_qty=5, unit_price=2.0, month_value=10.0,
        prev_month_value=8.0, delta=2.0, unit_price_diff=0.5,
        remark="", classification="成本-小料台类",
    )


def test_wipe_template_references_if_foreign(tmp_path: Path) -> None:
    """When the target store isn't the template's native store, the
    reference sheets (计算/分类/etc.) must be wiped so VLOOKUPs from
    the report sheet can't pick up the previous owner's data."""
    from inventory_check.stores import Store
    from inventory_check.workbook import (
        TEMPLATE_REFERENCE_SHEETS,
        _wipe_template_references_if_foreign,
    )
    wb = Workbook()
    # 计算 sheet — col C row 2 stores the template-native store name.
    calc = wb.active
    calc.title = "计算"
    calc.append(["检索", "区域", "门店名称", "大类", "子类", "码", "短码"])
    calc.append(["k1", "加拿大", "加拿大八店", "锅底类", "锅底类", 1060061, None])
    calc.append(["k2", "加拿大", "加拿大八店", "其他", "其他", 9000001, None])
    # 分类 etc.
    for name in ("分类", "折算数量", "BI套餐", "对照表"):
        ws = wb.create_sheet(name)
        ws.append(["header"])
        ws.append(["data1"])
        ws.append(["data2"])

    foreign = Store(sap_user="CA1DKG", werks="CA01", pos_name="加拿大一店")
    _wipe_template_references_if_foreign(wb, foreign)

    # Store-specific sheets (计算, BI套餐): header survives, data wiped.
    assert calc.cell(row=1, column=1).value == "检索"
    assert calc.max_row == 1
    bi = wb["BI套餐"]
    assert bi.max_row == 1
    # Regional master sheets (分类, 折算数量, 对照表): kept intact —
    # they're CA-wide reference data, not store-specific. Wiping them
    # broke the report sheet's 分类 col VLOOKUPs.
    for name in ("分类", "折算数量", "对照表"):
        ws = wb[name]
        assert ws.max_row == 3, f"{name} should be left intact"
        assert ws.cell(row=2, column=1).value == "data1"


def test_wipe_template_references_keeps_native(tmp_path: Path) -> None:
    """When the target store IS the template's native store, references
    are left intact (the template's 计算 / 分类 are *this* store's data)."""
    from inventory_check.stores import Store
    from inventory_check.workbook import _wipe_template_references_if_foreign
    wb = Workbook()
    calc = wb.active
    calc.title = "计算"
    calc.append(["检索", "区域", "门店名称"])
    calc.append(["k1", "加拿大", "加拿大八店"])
    fenlei = wb.create_sheet("分类")
    fenlei.append(["h"])
    fenlei.append(["d"])

    native = Store(sap_user="CA8DKG", werks="CA08", pos_name="加拿大八店")
    _wipe_template_references_if_foreign(wb, native)

    assert calc.max_row == 2  # data preserved
    assert fenlei.max_row == 2


def test_replace_report_sheet_writes_static_and_formulas() -> None:
    ws = _make_target_report_sheet()
    written = _replace_report_sheet(ws, [_make_report_row(1)])
    assert written == 1
    # Static cols 1-8 carry the source values.
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == 1000049
    assert ws.cell(row=2, column=4).value == 10  # 库存
    assert ws.cell(row=2, column=5).value == 5   # 盘点
    # Formula cols 9-17 — match the manual workbook's exact templates.
    assert ws.cell(row=2, column=9).value == "=D2-E2"
    assert ws.cell(row=2, column=10).value == "=VLOOKUP(B2,本月系统单价mb5b!$B:$P,15,0)"
    assert ws.cell(row=2, column=11).value == "=I2*J2"
    assert ws.cell(row=2, column=12).value == "=VLOOKUP(B2,上月数量需更新!A:C,3,0)"
    assert ws.cell(row=2, column=13).value == "=K2-L2"
    assert ws.cell(row=2, column=14).value == "=IFERROR(J2-VLOOKUP(B2,上月盘点结果!$B:$J,9,0),0)"
    assert ws.cell(row=2, column=15).value == '=IFERROR(VLOOKUP(B2,计算!R:Z,9,FALSE),"")'
    assert ws.cell(row=2, column=17).value == '=IFERROR(VLOOKUP(B2,分类!B:F,5,0),"")'
    # col 16 (回复) stays blank — manual entry column.
    assert ws.cell(row=2, column=16).value is None
    # Material-pivoted 差异 view (cols 18-21) — appended on top of the
    # manual layout so each material row carries its own balance.
    assert ws.cell(row=1, column=18).value == "理论用量"
    assert ws.cell(row=1, column=19).value == "套餐拼盘用量"
    assert ws.cell(row=1, column=20).value == "差异(物料)"
    assert ws.cell(row=1, column=21).value == "备注(物料)"
    assert ws.cell(row=2, column=18).value == "=IFERROR(SUMIF(计算!R:R,B2,计算!Q:Q),0)"
    assert ws.cell(row=2, column=19).value == "=IFERROR(SUMIF(计算!R:R,B2,计算!W:W),0)"
    # 差异 = 使用数量(I) − 理论用量(R, col 18) − 套餐拼盘用量(S, col 19).
    assert ws.cell(row=2, column=20).value == "=I2-R2-S2"
    # 备注 derives the 多用/少用 string from the new 差异 col (T = col 20).
    note = ws.cell(row=2, column=21).value
    assert "T2" in note and "多用" in note and "少用" in note


def test_replace_report_sheet_formula_row_indexes_match_target_row() -> None:
    """Each formula's ``r`` placeholder must point at its own row, not
    a constant — VLOOKUP(B2, …) on row 5 would pull row 2's matnr."""
    ws = _make_target_report_sheet()
    rows = [_make_report_row(i, matnr=f"100{i:04d}") for i in range(1, 4)]
    _replace_report_sheet(ws, rows)
    for i, expected_r in enumerate(range(2, 5), start=1):
        assert ws.cell(row=expected_r, column=9).value == f"=D{expected_r}-E{expected_r}"
        assert (ws.cell(row=expected_r, column=10).value
                == f"=VLOOKUP(B{expected_r},本月系统单价mb5b!$B:$P,15,0)")


def test_replace_pos_sheet_strips_zero_padding_from_codes(tmp_path: Path) -> None:
    """pos-crawler emits 菜品编码 as zero-padded strings ('01010106');
    the manual stores them as ints (1010106) and the 计算 sheet's
    lookup formula uses the int form. If we don't strip here the
    Sheet3 pivot key won't match what 计算!A produces — silently
    breaking 备注 column on the report."""
    src = _make_pos_xlsx(tmp_path, [
        # Padded — must round-trip to int 1010106 for the lookup chain
        # to find this row from the 计算!A formula's int concatenation.
        ["wrong-precomputed-key", "加拿大八店", None, "01010106",
         "1108705", "x", "整份", 10, 0, 10, "c", "s"],
    ])
    ws = _make_target_pos_sheet()
    _replace_pos_sheet(ws, src)
    assert ws.cell(row=2, column=4).value == 1010106
    assert isinstance(ws.cell(row=2, column=4).value, int)
    # 检索 (col 1) must be recomputed from the int form, not the padded source.
    assert ws.cell(row=2, column=1).value == "加拿大八店10101061108705整份"


def test_strip_pivot_tables_removes_pivot_objects(tmp_path: Path) -> None:
    """The manual template has real Excel pivot tables on 上月数量需更新
    and Sheet3 — when openpyxl saves, the cached pivot data overrides
    the cells we just wrote. Stripping the pivot objects makes our
    written values authoritative."""
    from inventory_check.workbook import _strip_pivot_tables

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet3"
    # Stand-in for a TableDefinition — _strip_pivot_tables only counts
    # and clears the list, so any object works for the unit test.
    ws._pivots = [object(), object()]

    n = _strip_pivot_tables(ws)
    assert n == 2
    assert ws._pivots == []


def test_replace_report_sheet_writes_matnr_as_int_when_numeric() -> None:
    """Manual stores matnr as int (e.g. 4509062), and the VLOOKUP
    against 本月系统单价mb5b's int matnr column relies on type-equal
    matching. Strings break the lookup silently."""
    ws = _make_target_report_sheet()
    _replace_report_sheet(ws, [_make_report_row(1, matnr="1000049")])
    cell = ws.cell(row=2, column=2)
    assert cell.value == 1000049
    assert isinstance(cell.value, int)
