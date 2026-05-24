"""Unit tests for inventory_check.calc_sheet — store_bom-driven 计算 builder."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from inventory_check.calc_sheet import (
    HEADERS,
    _norm_code,
    attach_formulas,
    derive_calc_rows,
    write_calc_workbook,
)


class TestNormCode:
    def test_strips_leading_zeros(self):
        assert _norm_code("01060061") == "1060061"

    def test_int_is_stringified(self):
        assert _norm_code(1060061) == "1060061"

    def test_none_becomes_empty(self):
        assert _norm_code(None) == ""

    def test_already_unpadded(self):
        assert _norm_code("1060061") == "1060061"


class TestDeriveCalcRows:
    def test_one_row_per_bom(self):
        # store_bom rows use neutral English keys — see db_bom.load_store_bom_rows.
        bom = [
            {
                "dish_code": 1060061,
                "dish_name": "清油麻辣火锅",
                "dish_short_code": None,
                "spec": "单锅",
                "material_code": 3000759,
                "material_name": "清油底料",
                "portion": 0.42,
                "loss_factor": 1.05,
                "unit": "公斤",
                "packaging_factor": None,
            },
        ]
        rows = derive_calc_rows(bom, store_name="加拿大八店")
        assert len(rows) == 1
        r = rows[0]
        assert r[0] == "加拿大八店1060061单锅"  # 检索
        assert r[1] == "加拿大"                  # 区域
        assert r[2] == "加拿大八店"               # 门店
        assert r[5] == 1060061                  # 菜品编码 — int form
        assert r[7] == "清油麻辣火锅"             # 菜品名称
        assert r[9] == "单锅"                    # 规格
        assert r[13] == 0.42                    # 出品分量
        assert r[14] == 1.05                    # 损耗 (taken directly from BOM)
        assert r[15] is None                    # P — blank packaging_factor
        assert r[17] == 3000759                 # 物料号 — int form
        assert r[18] == "清油底料"                # 物料描述
        assert r[24] == "公斤"                   # 单位

    def test_default_loss_when_missing(self):
        bom = [{
            "dish_code": 1, "spec": "整份", "material_code": 1,
            "loss_factor": None,
        }]
        rows = derive_calc_rows(bom, store_name="加拿大一店")
        assert rows[0][14] == 1

    def test_default_loss_when_empty_string(self):
        bom = [{
            "dish_code": 1, "spec": "整份", "material_code": 1,
            "loss_factor": "",
        }]
        rows = derive_calc_rows(bom, store_name="加拿大一店")
        assert rows[0][14] == 1

    def test_packaging_factor_passes_through(self):
        bom = [{
            "dish_code": 1, "spec": "整份", "material_code": 1,
            "packaging_factor": 0.354,
        }]
        rows = derive_calc_rows(bom, store_name="加拿大一店")
        assert rows[0][15] == 0.354

    def test_search_key_uses_store_name(self):
        bom = [{"dish_code": 1060061, "spec": "单锅", "material_code": 1}]
        ca1 = derive_calc_rows(bom, store_name="加拿大一店")
        ca2 = derive_calc_rows(bom, store_name="加拿大二店")
        assert ca1[0][0].startswith("加拿大一店")
        assert ca2[0][0].startswith("加拿大二店")
        assert ca1[0][0] != ca2[0][0]

    def test_pos_meta_fills_short_code_when_bom_missing(self):
        bom = [{
            "dish_code": 1060061, "spec": "单锅", "material_code": 1,
            "dish_short_code": None,
        }]
        pos_meta = {("1060061", "单锅"): {"菜品短编码": 999}}
        rows = derive_calc_rows(bom, store_name="加拿大一店", pos_meta=pos_meta)
        assert rows[0][6] == 999
        assert rows[0][0] == "加拿大一店1060061999单锅"

    def test_bom_short_code_wins_over_pos(self):
        bom = [{
            "dish_code": 1060061, "spec": "单锅", "material_code": 1,
            "dish_short_code": 111,
        }]
        pos_meta = {("1060061", "单锅"): {"菜品短编码": 999}}
        rows = derive_calc_rows(bom, store_name="加拿大一店", pos_meta=pos_meta)
        assert rows[0][6] == 111


def _mk_calc_row(f: object = 1060061, r: object = 3000759,
                 n: object = 1.2, o: object = 1.0) -> list:
    """Build a 26-col calc row with just F, R, N, O populated.

    Mirrors the layout from derive_calc_rows: idx 5=F, 13=N, 14=O, 17=R.
    """
    row = [None] * 26
    row[5], row[13], row[14], row[17] = f, n, o, r
    return row


class TestAttachFormulas:
    def test_formula_indices_single_canonical_row(self):
        rows = [_mk_calc_row()]
        attach_formulas(rows, report_sheet_name="CA08-本月-盘点结果.")
        assert rows[0][12].startswith("=IFERROR(VLOOKUP(A2,Sheet3")
        assert rows[0][16] == "=N2*M2*O2"
        assert "CA08-本月-盘点结果." in rows[0][19]
        assert 'IF(P2=""' in rows[0][19]
        assert "BI套餐" in rows[0][22]
        assert "*N2*O2" in rows[0][22]
        assert "SUMIF" in rows[0][20]
        assert "W2" in rows[0][20]
        assert "多用" in rows[0][23] and "少用" in rows[0][23]
        assert rows[0][25] == '=IF(X2<>"",X2&Y2,"")'

    def test_canonical_row_is_smallest_n_per_dish_material_group(self):
        rows = [
            _mk_calc_row(n=1.2),
            _mk_calc_row(n=0.6),
            _mk_calc_row(n=0.3),
        ]
        attach_formulas(rows, report_sheet_name="REP")
        for r in rows:
            assert r[12].startswith("=IFERROR(VLOOKUP")
            assert r[16].startswith("=N")
            assert r[19].startswith("=IFERROR(VLOOKUP")
        assert rows[2][22] is not None and "*N4*O4" in rows[2][22]
        assert rows[2][20] is not None and "T4" in rows[2][20]
        assert rows[2][23] is not None
        assert rows[2][25] is not None
        for idx in (0, 1):
            assert rows[idx][22] is None
            assert rows[idx][20] is None
            assert rows[idx][23] is None
            assert rows[idx][25] is None

    def test_separate_groups_each_get_canonical(self):
        rows = [
            _mk_calc_row(f=1060061, r=3000759, n=1.2),
            _mk_calc_row(f=1060061, r=3000759, n=0.3),
            _mk_calc_row(f=1060062, r=3000759, n=0.6),
            _mk_calc_row(f=1060062, r=3000759, n=0.2),
        ]
        attach_formulas(rows, report_sheet_name="REP")
        assert rows[0][22] is None
        assert rows[1][22] is not None
        assert rows[2][22] is None
        assert rows[3][22] is not None


class TestWriteWorkbook:
    def test_writes_calc_sheet(self, tmp_path: Path):
        rows = [[
            "加拿大八店1060061单锅", "加拿大", "加拿大八店",
            None, None, 1060061, None, "清油麻辣火锅", "清油麻辣火锅",
            "单锅", None, None, None, 0.42, 1.05, None, None,
            3000759, "清油底料", None, None, None, None, None,
            "公斤", None,
        ]]
        out = tmp_path / "calc.xlsx"
        write_calc_workbook(rows, out)
        assert out.exists()

        wb = openpyxl.load_workbook(out, data_only=False)
        assert wb.sheetnames == ["计算"]
        ws = wb["计算"]
        assert ws.cell(1, 1).value == HEADERS[0]
        assert ws.cell(2, 1).value == "加拿大八店1060061单锅"
        assert ws.cell(2, 6).value == 1060061
