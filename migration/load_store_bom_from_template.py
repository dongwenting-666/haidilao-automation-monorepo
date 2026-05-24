"""One-off migration: load store_bom from the hand-curated 计算 sheet.

Reads the 计算 sheet of a manual inventory-check template (e.g.
``CA08-盘点结果-202603.xlsx``) and upserts one row per
(werks × dish × spec × material) into ``store_bom`` for each werks in
WERKS_CODES. The recipe table is brand-wide today (only A/C cols are
store-keyed), so we broadcast the same recipe to every store; per-store
divergence can be edited later via /admin/bom.

Usage::

    uv run --project server python migration/load_store_bom_from_template.py \\
        --template ~/Downloads/CA08-盘点结果-202603.xlsx

Idempotent — re-runnable. Uses ``upsert_bom_entry`` so existing rows are
updated in place; only NULL→value transitions overwrite.
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import openpyxl

from server.db import is_db_available, upsert_bom_entry

logger = logging.getLogger("load_store_bom")

WERKS_CODES = ("CA01", "CA02", "CA03", "CA04", "CA05", "CA06", "CA07", "CA08")

# Column index → field. Matches the 计算 sheet's hand-curated 26-col layout.
COL_DISH_CODE = 5
COL_DISH_SHORT = 6
COL_DISH_NAME = 7
COL_SPEC = 9
COL_PORTION = 13
COL_LOSS = 14
COL_PACKAGING = 15
COL_MATERIAL_CODE = 17
COL_MATERIAL_NAME = 18
COL_UNIT = 24


def _int_or_none(v):
    if v in (None, ""):
        return None
    try:
        return int(str(v).strip().lstrip("0") or "0")
    except (TypeError, ValueError):
        return None


def _float_or_none(v):
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str_or_none(v):
    if v in (None, ""):
        return None
    return str(v).strip() or None


def extract_bom_rows(template_path: Path) -> list[dict]:
    """Read 计算 sheet → list of recipe rows (store-agnostic)."""
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    if "计算" not in wb.sheetnames:
        raise ValueError(f"{template_path} has no '计算' sheet")
    ws = wb["计算"]
    out = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        dish_code = _int_or_none(vals[COL_DISH_CODE])
        material_code = _int_or_none(vals[COL_MATERIAL_CODE])
        spec = _str_or_none(vals[COL_SPEC])
        if dish_code is None or material_code is None or spec is None:
            continue
        out.append({
            "dish_code": dish_code,
            "dish_short_code": _int_or_none(vals[COL_DISH_SHORT]),
            "dish_name": _str_or_none(vals[COL_DISH_NAME]),
            "spec": spec,
            "material_code": material_code,
            "material_name": _str_or_none(vals[COL_MATERIAL_NAME]),
            "portion": _float_or_none(vals[COL_PORTION]),
            "loss_factor": _float_or_none(vals[COL_LOSS]),
            "packaging_factor": _float_or_none(vals[COL_PACKAGING]),
            "unit": _str_or_none(vals[COL_UNIT]),
        })
    wb.close()
    return out


def load_into_db(rows: list[dict], werks_list: tuple[str, ...],
                 *, created_by: str, dry_run: bool) -> tuple[int, int]:
    """Upsert each row into store_bom for each werks. Returns (written, skipped)."""
    written = 0
    skipped = 0
    for werks in werks_list:
        for r in rows:
            if dry_run:
                written += 1
                continue
            try:
                upsert_bom_entry(
                    werks=werks,
                    dish_code=r["dish_code"],
                    spec=r["spec"],
                    material_code=r["material_code"],
                    dish_name=r["dish_name"],
                    dish_short_code=r["dish_short_code"],
                    material_name=r["material_name"],
                    portion=r["portion"],
                    loss_factor=r["loss_factor"],
                    unit=r["unit"],
                    packaging_factor=r["packaging_factor"],
                    notes=None,
                    created_by=created_by,
                )
                written += 1
            except Exception as exc:
                logger.error("upsert failed (werks=%s dish=%s mat=%s spec=%s): %s",
                             werks, r["dish_code"], r["material_code"], r["spec"], exc)
                skipped += 1
    return written, skipped


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--template", required=True, type=Path,
                   help="Path to the manual workbook with a 计算 sheet")
    p.add_argument("--werks", nargs="+", default=list(WERKS_CODES),
                   help=f"Werks codes to broadcast to (default: {' '.join(WERKS_CODES)})")
    p.add_argument("--created-by", default="migration:load_store_bom_from_template",
                   help="Value for store_bom.created_by audit field")
    p.add_argument("--dry-run", action="store_true",
                   help="Parse and report row counts without writing to DB")
    args = p.parse_args()

    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(name)s] %(levelname)s %(message)s")

    template = args.template.expanduser()
    if not template.exists():
        logger.error("template not found: %s", template)
        return 2

    if not args.dry_run and not is_db_available():
        logger.error("DATABASE_URL not set or DB unreachable — refusing to write")
        return 2

    rows = extract_bom_rows(template)
    logger.info("parsed %d valid recipe rows from %s", len(rows), template)
    if not rows:
        logger.warning("nothing to load")
        return 1

    written, skipped = load_into_db(
        rows, tuple(args.werks),
        created_by=args.created_by, dry_run=args.dry_run,
    )
    verb = "would write" if args.dry_run else "wrote"
    logger.info("%s %d rows across %d werks (%d skipped)",
                verb, written, len(args.werks), skipped)
    return 0 if skipped == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
