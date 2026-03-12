"""Read data from Excel files."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def load_data_rows(
    path: Path,
    *,
    sheet_name: str | None = None,
    skip_empty_key: str | None = None,
) -> list[dict[str, Any]]:
    """Load an XLSX file into a list of row dicts (header-keyed).

    Args:
        path: Path to the XLSX file.
        sheet_name: Worksheet name to read. None = active sheet.
        skip_empty_key: If set, skip rows where this column is None
            (useful for filtering summary/blank rows).

    Returns:
        List of dicts, one per row, keyed by header values.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if skip_empty_key and data.get(skip_empty_key) is None:
            continue
        rows.append(data)
    wb.close()
    logger.info("Loaded %d rows from %s", len(rows), path.name)
    return rows


def load_mapping(
    path: Path,
    *,
    key_col: int = 0,
    value_col: int = 2,
    sheet_name: str | None = None,
    header_row: bool = True,
) -> dict[str, str]:
    """Load a two-column mapping from an XLSX file.

    Args:
        path: Path to the XLSX file.
        key_col: Column index for the key (0-based).
        value_col: Column index for the value (0-based).
        sheet_name: Worksheet name. None = active sheet.
        header_row: Whether to skip the first row as header.

    Returns:
        dict mapping key → value (both as strings).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    mapping: dict[str, str] = {}
    min_row = 2 if header_row else 1
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        key = row[key_col] if key_col < len(row) else None
        val = row[value_col] if value_col < len(row) else None
        if key is not None and val is not None:
            # Normalize numeric keys (e.g. cost element codes)
            str_key = str(int(key) if isinstance(key, (int, float)) else key)
            mapping[str_key] = str(val)
    wb.close()
    logger.info("Loaded %d mappings from %s", len(mapping), path.name)
    return mapping
