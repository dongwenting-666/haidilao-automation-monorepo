"""Workbook creation and sheet utilities."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_utils.style import BOLD_FONT, auto_size_columns, set_header_row

logger = logging.getLogger(__name__)

MAX_SHEET_NAME_LEN = 31


def truncate_sheet_name(name: str, *, prefix_to_strip: str | None = None) -> str:
    """Truncate a name to fit Excel's 31-char sheet name limit.

    Args:
        name: Raw sheet name.
        prefix_to_strip: Optional prefix to remove before truncating
            (e.g. "加拿大" to shorten store names).

    Returns:
        Name truncated to 31 characters.
    """
    if prefix_to_strip and name.startswith(prefix_to_strip):
        name = name[len(prefix_to_strip):]
    return name[:MAX_SHEET_NAME_LEN]


def create_workbook() -> Workbook:
    """Create a new workbook with the default sheet removed."""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def write_data_sheet(
    wb: Workbook,
    title: str,
    *,
    headers: list[str],
    rows: list[dict[str, Any]],
    auto_size: bool = True,
) -> Worksheet:
    """Create a new sheet and populate it with header + data rows.

    Args:
        wb: Target workbook.
        title: Sheet title (will be truncated to 31 chars).
        headers: Column header names (also used as dict keys for rows).
        rows: List of dicts with values keyed by header names.
        auto_size: Whether to auto-size columns after writing.

    Returns:
        The created worksheet.
    """
    ws = wb.create_sheet(title=title[:MAX_SHEET_NAME_LEN])
    set_header_row(ws, headers)

    for i, row in enumerate(rows, 2):
        for col, key in enumerate(headers, 1):
            ws.cell(row=i, column=col, value=row.get(key))

    if auto_size:
        auto_size_columns(ws)
    return ws


def copy_sheet_data(
    wb: Workbook,
    source_path: Path,
    *,
    title: str,
    columns: int | None = None,
    sheet_name: str | None = None,
) -> Worksheet:
    """Copy data from a source XLSX into a new sheet in the workbook.

    Args:
        wb: Target workbook.
        source_path: Path to source XLSX file.
        title: Title for the new sheet.
        columns: Number of columns to copy. None = all columns.
        sheet_name: Sheet name in source file. None = active sheet.

    Returns:
        The created worksheet.
    """
    src_wb = load_workbook(source_path, read_only=True, data_only=True)
    src_ws = src_wb[sheet_name] if sheet_name else src_wb.active

    ws = wb.create_sheet(title=title[:MAX_SHEET_NAME_LEN])
    for row in src_ws.iter_rows(values_only=True):
        if columns is not None:
            ws.append(list(row[:columns]))
        else:
            ws.append(list(row))

    src_wb.close()
    return ws
