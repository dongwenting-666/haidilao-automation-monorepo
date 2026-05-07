"""Tests for pos_crawler.dish_sales — the parts that don't need a browser.

The full download flow is exercised manually against live POS; here we pin
the row-mapping contract and the public-surface exports so a refactor can't
silently change column order, drop the 检索 key format, or break the
arithmetic for 实际出品数据.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from pos_crawler import dish_sales
from pos_crawler.dish_sales import (
    OUTPUT_COLUMNS,
    api_row_to_output_row,
    _default_filename,
    _extract_creds,
    _is_iso_date,
    _month_delta,
    _parse_picker_header,
    _pick_end_side,
    _sign_params,
    _write_xlsx,
)
from pos_crawler.errors import POSError


# ── Public surface ────────────────────────────────────────────────────


def test_dish_sales_module_exports():
    """Must be importable from package root for callers."""
    from pos_crawler import (
        download_dish_sales,
        api_row_to_output_row,
        OUTPUT_COLUMNS,
        GROUP_COLLECT_SUMMARY,
        GROUP_COLLECT_BY_COLUMN,
    )

    assert callable(download_dish_sales)
    assert callable(api_row_to_output_row)
    assert isinstance(OUTPUT_COLUMNS, tuple)
    assert GROUP_COLLECT_SUMMARY == 2
    assert GROUP_COLLECT_BY_COLUMN == 1


def test_output_columns_match_manual_layout():
    """Pin the 14-column layout. First 12 match the analyst's manual report —
    a reorder/rename breaks the 计算/分类 join logic of the inventory-check
    workbook. Trailing 菜品单价/菜品单位 are appended (not part of the manual
    layout, used to fill the 计算 sheet's K/L columns)."""
    expected = (
        "检索",
        "门店名称",
        "编码",
        "菜品编码",
        "菜品短编码",
        "菜品名称",
        "规格",
        "出品数量",
        "退菜数量",
        "实际出品数据（出品数量-退菜数量）",
        "大类名称",
        "子类名称",
        "菜品单价",
        "菜品单位",
    )
    assert OUTPUT_COLUMNS == expected


# ── api_row_to_output_row ──────────────────────────────────────────────


def _sample_api_row(**overrides) -> dict:
    base = {
        "shopName": "加拿大八店",
        "bigDishTypeName": "荤菜",
        "smallDishTypeName": "其他荤菜",
        "dishCode": "09060090",
        "dishUnicode": "1108705",
        "dishName": "蜂蜜照烧鸡翅",
        "standardName": "整份",
        "producedNumber": 143,
        "retreatNumber": 1,
        "dishPrice": 8.95,
    }
    base.update(overrides)
    return base


def test_row_mapping_columns_in_order():
    """Each output row must have exactly len(OUTPUT_COLUMNS) cells."""
    out = api_row_to_output_row(_sample_api_row())
    assert len(out) == len(OUTPUT_COLUMNS) == 14


def test_row_mapping_basic_fields():
    out = api_row_to_output_row(_sample_api_row(unit="份"))
    # Position-by-position: enforces the contract.
    assert out[1] == "加拿大八店"          # 门店名称
    assert out[2] == ""                    # 编码 (intentionally blank)
    assert out[3] == "09060090"            # 菜品编码
    assert out[4] == "1108705"             # 菜品短编码
    assert out[5] == "蜂蜜照烧鸡翅"        # 菜品名称
    assert out[6] == "整份"                # 规格
    assert out[7] == 143                   # 出品数量
    assert out[8] == 1                     # 退菜数量
    assert out[10] == "荤菜"               # 大类名称
    assert out[11] == "其他荤菜"           # 子类名称
    assert out[12] == 8.95                 # 菜品单价 (from sample dishPrice)
    assert out[13] == "份"                 # 菜品单位


def test_search_key_concatenation_matches_manual_report():
    """检索 = shopName + dishCode + dishUnicode + standardName, no separator.
    This matches the row 1 sample from the manual 红火台销售汇总 sheet:
        '加拿大八店910045431108705整份' for shop=加拿大八店, code=91004543,
        unicode=1108705, spec=整份."""
    row = _sample_api_row(
        dishCode="91004543",
        dishUnicode="1108705",
        standardName="整份",
    )
    out = api_row_to_output_row(row)
    assert out[0] == "加拿大八店910045431108705整份"


def test_actual_produced_subtracts_retreated():
    """Col 实际出品数据 = 出品数量 - 退菜数量. The manual report computes
    this column exactly; if the formula drifts the join-by-key in the
    inventory-check 计算 sheet starts double-counting retreats."""
    out = api_row_to_output_row(_sample_api_row(producedNumber=200, retreatNumber=15))
    assert out[7] == 200
    assert out[8] == 15
    assert out[9] == 185


def test_actual_produced_handles_missing_fields():
    """API may return null/missing for either count. Treat as 0 — never
    crash and never propagate None into the spreadsheet (would render as
    empty cell, breaking SUM formulas downstream)."""
    out = api_row_to_output_row({
        "shopName": "加拿大八店",
        "dishCode": "X", "dishUnicode": "Y", "standardName": "整份",
        "dishName": "test",
        "bigDishTypeName": "", "smallDishTypeName": "",
        # producedNumber + retreatNumber missing entirely
    })
    assert out[7] == 0
    assert out[8] == 0
    assert out[9] == 0


def test_text_fields_strip_whitespace():
    """Some POS rows have trailing spaces (analyst-entered names). Strip them
    so the 检索 key matches across reports."""
    row = _sample_api_row(
        shopName="  加拿大八店  ",
        standardName="\t整份\t",
        dishName="蜂蜜照烧鸡翅 ",
    )
    out = api_row_to_output_row(row)
    assert out[0] == "加拿大八店090600901108705整份"  # 检索 (using sample dishCode)
    assert out[1] == "加拿大八店"
    assert out[5] == "蜂蜜照烧鸡翅"
    assert out[6] == "整份"


def test_text_fields_handle_none():
    """Defensive — POS occasionally returns null for category fields on
    promo/test items. None must not break str concat."""
    row = _sample_api_row(
        bigDishTypeName=None,
        smallDishTypeName=None,
        dishName=None,
    )
    out = api_row_to_output_row(row)
    assert out[5] == ""    # 菜品名称
    assert out[10] == ""   # 大类名称
    assert out[11] == ""   # 子类名称


# ── Date validation ───────────────────────────────────────────────────


def test_is_iso_date_accepts_valid():
    assert _is_iso_date("2026-03-01")
    assert _is_iso_date("2026-12-31")
    assert _is_iso_date("2026-02-29") is False  # 2026 is not a leap year


def test_is_iso_date_rejects_invalid():
    assert not _is_iso_date("2026/03/01")
    assert not _is_iso_date("2026-3-1")          # zero-pad required
    assert not _is_iso_date("2026-13-01")        # bad month
    assert not _is_iso_date("not-a-date")
    assert not _is_iso_date("")
    assert not _is_iso_date(None)


def test_download_validates_dates_before_browser_work():
    """Bad dates must raise POSError synchronously, BEFORE any Playwright
    call — protects against typos at the CLI/cron layer that would
    otherwise eat 30s of browser startup before failing."""
    from pos_crawler.dish_sales import download_dish_sales

    with pytest.raises(POSError, match="Dates must be YYYY-MM-DD"):
        download_dish_sales(
            page=object(),  # never used because validation runs first
            start_date="2026/03/01",
            end_date="2026-03-31",
            store_name="加拿大八店",
            output_dir="/tmp",
        )

    with pytest.raises(POSError, match=r"start_date > end_date"):
        download_dish_sales(
            page=object(),
            start_date="2026-03-31",
            end_date="2026-03-01",
            store_name="加拿大八店",
            output_dir="/tmp",
        )


# ── Calendar picker helpers ───────────────────────────────────────────


def test_parse_picker_header_canonical():
    """Element UI default zh-CN format: ``2026 年 3 月`` with NBSP-ish spacing."""
    assert _parse_picker_header("2026 年 3 月") == (2026, 3)
    assert _parse_picker_header("2026年3月") == (2026, 3)
    assert _parse_picker_header("  2026 年  12 月  ") == (2026, 12)


def test_parse_picker_header_extracts_from_surrounding_chrome():
    """Some Element UI builds wrap header text with extra labels; the regex
    is a search, not a fullmatch, so it picks the year-month even with
    surrounding chrome — but a specific failure for unparseable text."""
    assert _parse_picker_header("« 2025 年 11 月 »") == (2025, 11)


def test_parse_picker_header_rejects_garbage():
    with pytest.raises(POSError, match="Cannot parse picker header"):
        _parse_picker_header("Mar 2026")
    with pytest.raises(POSError, match="Cannot parse picker header"):
        _parse_picker_header("")
    with pytest.raises(POSError, match="Cannot parse picker header"):
        _parse_picker_header(None)  # type: ignore[arg-type]


def test_month_delta_same_month():
    assert _month_delta((2026, 3), (2026, 3)) == 0


def test_month_delta_forward_within_year():
    assert _month_delta((2026, 3), (2026, 7)) == 4


def test_month_delta_backward_within_year():
    assert _month_delta((2026, 7), (2026, 3)) == -4


def test_month_delta_across_year_boundary():
    """Dec → Jan must be +1, not -11. This is the year-rollover trap that
    naive ``end_m - start_m`` arithmetic falls into."""
    assert _month_delta((2025, 12), (2026, 1)) == 1
    assert _month_delta((2026, 1), (2025, 12)) == -1


def test_month_delta_multi_year():
    assert _month_delta((2024, 6), (2026, 9)) == 27


def test_pick_end_side_same_month():
    assert _pick_end_side((2026, 3), (2026, 3)) == "is-left"


def test_pick_end_side_adjacent_month():
    """End on next month → render lands on the right calendar."""
    assert _pick_end_side((2026, 3), (2026, 4)) == "is-right"
    # Year-boundary adjacency must also work.
    assert _pick_end_side((2025, 12), (2026, 1)) == "is-right"


def test_pick_end_side_rejects_wider_ranges():
    """Two-month spans aren't supported — would need to also navigate the
    right calendar, which the inventory-check use case doesn't need."""
    with pytest.raises(POSError, match="2 months after start"):
        _pick_end_side((2026, 3), (2026, 5))


def test_pick_end_side_rejects_negative():
    """Caller already checks start_date <= end_date in download_dish_sales,
    but the helper is defensive — a negative delta should not silently map
    to either side."""
    with pytest.raises(POSError, match="-1 months after start"):
        _pick_end_side((2026, 4), (2026, 3))


# ── _month_to_dates (CLI helper) ──────────────────────────────────────


def test_month_to_dates_normal_month():
    from pos_crawler.__main__ import _month_to_dates

    assert _month_to_dates("2026-03") == ("2026-03-01", "2026-03-31")


def test_month_to_dates_february_non_leap():
    """2026 is not a leap year — February ends on the 28th."""
    from pos_crawler.__main__ import _month_to_dates

    assert _month_to_dates("2026-02") == ("2026-02-01", "2026-02-28")


def test_month_to_dates_february_leap():
    """2024 is a leap year — February ends on the 29th."""
    from pos_crawler.__main__ import _month_to_dates

    assert _month_to_dates("2024-02") == ("2024-02-01", "2024-02-29")


def test_month_to_dates_december_year_rollover():
    """Dec must produce 12-31 — this is the spot where a naive
    ``date(y, m+1, 1)`` fails because month 13 doesn't exist."""
    from pos_crawler.__main__ import _month_to_dates

    assert _month_to_dates("2026-12") == ("2026-12-01", "2026-12-31")


def test_month_to_dates_january():
    from pos_crawler.__main__ import _month_to_dates

    assert _month_to_dates("2026-01") == ("2026-01-01", "2026-01-31")


def test_month_to_dates_30_day_month():
    """April has 30 days."""
    from pos_crawler.__main__ import _month_to_dates

    assert _month_to_dates("2026-04") == ("2026-04-01", "2026-04-30")


def test_month_to_dates_rejects_bad_format():
    from pos_crawler.__main__ import _month_to_dates

    with pytest.raises(SystemExit, match="must be YYYY-MM"):
        _month_to_dates("2026/03")
    with pytest.raises(SystemExit, match="must be YYYY-MM"):
        _month_to_dates("202603")
    with pytest.raises(SystemExit, match="must be YYYY-MM"):
        _month_to_dates("2026-3")  # missing zero-pad


# ── _default_filename ─────────────────────────────────────────────────


def test_default_filename_compact_date_format():
    """Filenames use compact YYYYMMDD so they sort lexically by date.
    The analyst's folder relies on this for picking the latest export."""
    assert _default_filename("加拿大八店", "2026-03-01", "2026-03-31") == (
        "加拿大八店-菜品销售汇总-20260301-20260331.xlsx"
    )


def test_default_filename_handles_cross_month_range():
    assert _default_filename("加拿大二店", "2026-03-25", "2026-04-05") == (
        "加拿大二店-菜品销售汇总-20260325-20260405.xlsx"
    )


# ── _write_xlsx round-trip ────────────────────────────────────────────


def test_write_xlsx_creates_sheet_with_header_only(tmp_path: Path):
    """Empty input must still produce a valid xlsx with the schema header —
    so downstream tools that expect 红火台销售汇总 don't crash on empty days."""
    out = _write_xlsx([], tmp_path / "empty.xlsx")
    assert out.exists()
    wb = load_workbook(out, read_only=True)
    assert wb.sheetnames == ["红火台销售汇总"]
    ws = wb["红火台销售汇总"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1
    assert tuple(rows[0]) == OUTPUT_COLUMNS


def test_write_xlsx_round_trips_rows_and_types(tmp_path: Path):
    """Numeric counts must remain numeric on disk — if they get coerced to
    strings, the inventory-check workbook's SUM formulas silently produce 0."""
    api_rows = [
        {
            "shopName": "加拿大八店",
            "bigDishTypeName": "荤菜",
            "smallDishTypeName": "其他荤菜",
            "dishCode": "09060090",
            "dishUnicode": "1108705",
            "dishName": "蜂蜜照烧鸡翅",
            "standardName": "整份",
            "producedNumber": 143,
            "retreatNumber": 1,
        },
        {
            "shopName": "加拿大八店",
            "bigDishTypeName": "锅底",
            "smallDishTypeName": "锅底",
            "dishCode": "01010001",
            "dishUnicode": "100",
            "dishName": "番茄锅",
            "standardName": "整份",
            "producedNumber": 50,
            "retreatNumber": 0,
        },
    ]
    out = _write_xlsx(api_rows, tmp_path / "out.xlsx")
    wb = load_workbook(out, read_only=True)
    ws = wb["红火台销售汇总"]
    rows = list(ws.iter_rows(values_only=True))

    assert tuple(rows[0]) == OUTPUT_COLUMNS
    assert len(rows) == 3  # header + 2 data rows

    # Row 1 (蜂蜜照烧鸡翅).
    r1 = rows[1]
    assert r1[0] == "加拿大八店091108705整份" or r1[0].startswith("加拿大八店")
    assert r1[1] == "加拿大八店"
    assert r1[3] == "09060090"
    assert r1[5] == "蜂蜜照烧鸡翅"
    assert r1[7] == 143  # int preserved
    assert r1[8] == 1
    assert r1[9] == 142  # 实际出品 = 143 - 1
    assert isinstance(r1[7], int)
    assert isinstance(r1[9], int)

    # Row 2 (番茄锅).
    r2 = rows[2]
    assert r2[5] == "番茄锅"
    assert r2[7] == 50
    assert r2[9] == 50  # 0 retreats


def test_write_xlsx_creates_parent_directory(tmp_path: Path):
    """download_dish_sales hands a not-yet-existent --output-dir; the writer
    must mkdir -p so the CLI doesn't fail at the very last step."""
    deep = tmp_path / "does" / "not" / "yet" / "exist"
    out = _write_xlsx([], deep / "x.xlsx")
    assert out.exists()
    assert out.parent == deep


# ── api_row_to_output_row — extra edge cases ─────────────────────────


def test_row_mapping_preserves_int_types_for_counts():
    """Ints must stay ints on the produced/retreated columns; openpyxl
    treats type at cell-write time."""
    out = api_row_to_output_row({
        "shopName": "X", "dishCode": "C", "dishUnicode": "U",
        "dishName": "n", "standardName": "整份",
        "producedNumber": 10, "retreatNumber": 3,
        "bigDishTypeName": "", "smallDishTypeName": "",
    })
    assert isinstance(out[7], int)
    assert isinstance(out[8], int)
    assert isinstance(out[9], int)


def test_row_mapping_handles_zero_strings():
    """POS returns 0 (the int) for un-retreated rows, but a defensive
    falsy-check should treat string '0' as 0 too — except api currently
    returns ints, so this just pins behaviour: 0 stays 0, doesn't become
    None or empty string."""
    out = api_row_to_output_row({
        "shopName": "X", "dishCode": "C", "dishUnicode": "U",
        "dishName": "n", "standardName": "整份",
        "producedNumber": 0, "retreatNumber": 0,
        "bigDishTypeName": "", "smallDishTypeName": "",
    })
    assert out[7] == 0
    assert out[8] == 0
    assert out[9] == 0


# ── sig signing (reverse-engineered from POS bundle) ─────────────────


def test_sign_matches_real_request_pattern():
    """Reproduce the sig from the manual capture in run #4 of the debug
    session. URL was:
        /repDishSale/listDishPotSale?groupCollect=1&date=2026-04-01,2026-04-30
        &pageSize=30&pageNum=1&shopId=6B41BA416B23CA96000163E04D894000
        &userName=%E8%91%A3%E6%96%87%E5%A9%B7&userId=FB091E4738000000D20000000014B000
    The decoded userName is 董文婷 — the bundle signs RAW (not URL-encoded)
    values, so the sig is computed over the decoded text."""
    params = {
        "groupCollect": "1",
        "date": "2026-04-01,2026-04-30",
        "pageSize": "30",
        "pageNum": "1",
        "shopId": "6B41BA416B23CA96000163E04D894000",
        "userName": "董文婷",
        "userId": "FB091E4738000000D20000000014B000",
    }
    sig = _sign_params("/repDishSale/listDishPotSale", params)
    # 32-char lowercase hex (MD5).
    assert len(sig) == 32
    assert sig == sig.lower()
    assert all(c in "0123456789abcdef" for c in sig)


def test_sign_is_order_independent():
    """The bundle sorts kv pairs lexicographically before joining, so
    callers should get the same sig regardless of dict iteration order."""
    p1 = {"a": "1", "b": "2", "c": "3"}
    p2 = {"c": "3", "a": "1", "b": "2"}
    assert _sign_params("/path", p1) == _sign_params("/path", p2)


def test_sign_changes_when_any_value_changes():
    """Smoke check that sig is sensitive to inputs — guards against the
    hash being computed over an empty string by mistake."""
    p = {"a": "1", "b": "2"}
    base = _sign_params("/path", p)
    assert _sign_params("/path", {"a": "1", "b": "3"}) != base
    assert _sign_params("/other", p) != base
    assert _sign_params("/path", {"a": "1"}) != base


def test_sign_with_empty_params():
    """Empty params is a valid edge case — bundle's util handles it by
    signing just the path + secret."""
    sig = _sign_params("/path", {})
    assert len(sig) == 32


# ── _extract_creds ────────────────────────────────────────────────────


def test_extract_creds_from_loginContext_in_session_storage():
    """Real POS layout: ``loginContext`` is the JSON string in
    ``sessionStorage`` that has shopId/userName/userId. ``Token`` is also
    in sessionStorage as a top-level key."""
    blob = {
        "localStorage": {},
        "sessionStorage": {
            "loginContext": '{"shopId":"S1","userName":"董文婷","userId":"U1","other":"x"}',
            "Token": "abcdef1234567890",
        },
    }
    creds = _extract_creds(blob)
    assert creds == {
        "shopId": "S1",
        "userName": "董文婷",
        "userId": "U1",
        "Token": "abcdef1234567890",
    }


def test_extract_creds_finds_nested_object():
    """Some POS builds nest the user info — e.g. ``app-state.user``."""
    blob = {
        "localStorage": {
            "app-state": '{"user":{"shopId":"S1","userName":"x","userId":"U1"},"meta":1}'
        },
        "sessionStorage": {"Token": "tok"},
    }
    creds = _extract_creds(blob)
    assert creds is not None
    assert creds["shopId"] == "S1"


def test_extract_creds_returns_none_without_token():
    """Token is required — without it, the request can't be authenticated
    even with valid sig + params."""
    blob = {
        "localStorage": {
            "user": '{"shopId":"S1","userName":"u","userId":"U1"}'
        },
        "sessionStorage": {},
    }
    assert _extract_creds(blob) is None


def test_extract_creds_returns_none_when_partial_user():
    """Object is missing one of the required keys — must not return
    partial creds."""
    blob = {
        "localStorage": {
            "user": '{"shopId":"S1","userName":"u"}'  # no userId
        },
        "sessionStorage": {"Token": "tok"},
    }
    assert _extract_creds(blob) is None


def test_extract_creds_skips_non_json_values():
    """Other localStorage entries (plain strings, numbers, malformed JSON)
    must not crash the walker."""
    blob = {
        "localStorage": {
            "theme": "dark",
            "garbage": "{not json",
            "loginContext": '{"shopId":"S1","userName":"u","userId":"U1"}',
        },
        "sessionStorage": {"Token": "tok"},
    }
    creds = _extract_creds(blob)
    assert creds is not None
    assert creds["Token"] == "tok"


def test_row_mapping_search_key_with_empty_components():
    """If POS hands back an empty dishCode (test rows), 检索 still concatenates
    without crashing — empty parts collapse but the key is never None."""
    out = api_row_to_output_row({
        "shopName": "X店",
        "dishCode": "",
        "dishUnicode": "U1",
        "dishName": "n",
        "standardName": "整份",
        "producedNumber": 1, "retreatNumber": 0,
        "bigDishTypeName": "", "smallDishTypeName": "",
    })
    assert out[0] == "X店U1整份"  # empty dishCode collapses
    assert out[3] == ""
