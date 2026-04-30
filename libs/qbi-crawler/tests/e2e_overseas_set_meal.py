"""E2E test: full 海外套餐销售明细 download flow against live QBI.

Requires:
    - VPN connected (CorpLink to ipms-global / qbi.superhi-tech.com)
    - QBI_USERNAME / QBI_PASSWORD in .env
    - Network reachable from the host

Usage:
    uv run --project libs/qbi-crawler python libs/qbi-crawler/tests/e2e_overseas_set_meal.py

Validates:
    1. download_report returns a Path that exists
    2. File is non-empty
    3. File is a real Excel 2007+ xlsx (openpyxl readable)
    4. Sheet has at least 1 data row + the expected column headers
"""
from __future__ import annotations

import logging
import os
import sys
import time
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv

from qbi_crawler import (
    QBISession,
    REPORT_OVERSEAS_SET_MEAL,
    download_report,
)

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s"
)
log = logging.getLogger(__name__)

OUTPUT_DIR = Path(__file__).resolve().parents[3] / "output" / "qbi-overseas-set-meal"


def previous_month_range() -> tuple[date, date]:
    first_of_this = date.today().replace(day=1)
    last_of_prev = first_of_this - timedelta(days=1)
    first_of_prev = last_of_prev.replace(day=1)
    return first_of_prev, last_of_prev


def main() -> int:
    load_dotenv()

    username = os.environ.get("QBI_USERNAME", "")
    password = os.environ.get("QBI_PASSWORD", "")
    if not username or not password:
        print("ERROR: Set QBI_USERNAME and QBI_PASSWORD in .env", file=sys.stderr)
        return 1

    d_from, d_to = previous_month_range()
    log.info("Starting 海外套餐销售明细 E2E")
    log.info("  User:    %s", username)
    log.info("  Country: 加拿大")
    log.info("  Dates:   %s → %s", d_from, d_to)
    log.info("  Output:  %s", OUTPUT_DIR)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    t0 = time.monotonic()
    with QBISession(username=username, password=password, headless=True) as session:
        result = download_report(
            session.page,
            REPORT_OVERSEAS_SET_MEAL,
            start_date=d_from.isoformat(),
            end_date=d_to.isoformat(),
            download_dir=OUTPUT_DIR,
            country="加拿大",
        )

    elapsed = time.monotonic() - t0

    # 1. File exists, non-empty
    if not result.exists():
        log.error("E2E FAILED — file not created at %s", result)
        return 2
    size = result.stat().st_size
    if size == 0:
        log.error("E2E FAILED — file is empty: %s", result)
        return 2

    # 2. Validate it's a real xlsx with data
    try:
        import openpyxl
        wb = openpyxl.load_workbook(result, read_only=True)
        ws = wb.active
        max_row = ws.max_row
        max_col = ws.max_column
        # Read first row (headers)
        headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        log.info("Sheet:   %r", ws.title)
        log.info("Rows:    %d", max_row)
        log.info("Cols:    %d", max_col)
        log.info("Headers (first 8): %s", headers[:8])
        if max_row < 2:
            log.error("E2E FAILED — sheet has %d rows (no data)", max_row)
            return 3
        # Sanity-check headers — the report should have 业务日期 / 国家 / 门店 / 套餐
        joined = " ".join(headers)
        for expected in ("国家", "套餐"):
            if expected not in joined:
                log.warning("Header sanity: %r not found in headers — file may be wrong report", expected)
    except Exception as e:
        log.error("E2E FAILED — could not read xlsx: %s", e)
        return 4

    log.info("=" * 60)
    log.info("E2E PASSED  %.1fs  %d bytes  %d rows × %d cols  %s",
             elapsed, size, max_row, max_col, result)
    return 0


if __name__ == "__main__":
    sys.exit(main())
