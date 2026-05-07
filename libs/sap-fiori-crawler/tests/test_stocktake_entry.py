"""Unit tests for sap_fiori_crawler.entry — the in-progress 盘点录入
download path. Pure logic only: parse + map + filename. The network
POST replay is covered by the e2e suite."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from sap_fiori_crawler.entry import (
    ENTRY_OUTPUT_COLUMNS,
    _default_filename,
    entry_row_to_output_row,
    parse_invh_response,
    write_entry_xlsx,
)
from sap_fiori_crawler.errors import FioriExportError


# ── parse_invh_response ────────────────────────────────────────────────


def test_parse_invh_response_nested_results() -> None:
    payload = {"d": {"InvH_I": {"results": [{"Matnr": "1"}, {"Matnr": "2"}]}}}
    assert parse_invh_response(payload) == [{"Matnr": "1"}, {"Matnr": "2"}]


def test_parse_invh_response_flat_list() -> None:
    """Some SAP gateways flatten 'InvH_I' to a list directly."""
    payload = {"d": {"InvH_I": [{"Matnr": "1"}]}}
    assert parse_invh_response(payload) == [{"Matnr": "1"}]


def test_parse_invh_response_empty_results() -> None:
    payload = {"d": {"InvH_I": {"results": []}}}
    assert parse_invh_response(payload) == []


def test_parse_invh_response_missing_d_raises() -> None:
    with pytest.raises(FioriExportError, match="missing 'd'"):
        parse_invh_response({})


def test_parse_invh_response_missing_invh_i_raises() -> None:
    with pytest.raises(FioriExportError, match="missing 'd.InvH_I.results'"):
        parse_invh_response({"d": {"OPdtxt": "x"}})


# ── entry_row_to_output_row ────────────────────────────────────────────


def _live_invi_row() -> dict[str, object]:
    """Mirror of a real CA08 row captured during exploration."""
    return {
        "Werks": "CA08",
        "Matnr": "4509062",
        "Matxt": '"烧酒（JINRO CHANISUL FRESH，360ML/瓶）"',
        "Menge": "119",
        "MengePd": "101",
        "Msetxt": "瓶-瓶",
        "Meins": "BOT",
        "Msehi1": "BOT",
        "Status": "1",
        "Line": "00010",
    }


def test_entry_row_to_output_row_full() -> None:
    """Column order must match stocktake.OUTPUT_COLUMNS exactly."""
    out = entry_row_to_output_row(_live_invi_row())
    assert out == [
        "CA08",                                              # 工厂
        "4509062",                                           # 物料号
        '"烧酒（JINRO CHANISUL FRESH，360ML/瓶）"',         # 名称
        "BOT",                                               # 单位
        "瓶-瓶",                                             # 单位描述
        119,                                                 # 库存数量
        101,                                                 # 盘点数量 ← idx 6
        "1",                                                 # 状态
        "",                                                  # 盘点日期 (entry: blank)
        "",                                                  # 时间   (entry: blank)
    ]


def test_entry_row_to_output_row_decimal_qty() -> None:
    row = _live_invi_row() | {"Menge": "1056.500", "MengePd": "384.000"}
    out = entry_row_to_output_row(row)
    assert out[5] == 1056.5
    assert out[6] == 384  # 384.0 normalises to 384


def test_entry_row_to_output_row_blank_count() -> None:
    """Empty MengePd (count not entered yet for this row) renders blank."""
    row = _live_invi_row() | {"MengePd": ""}
    out = entry_row_to_output_row(row)
    assert out[6] == ""


def test_entry_row_to_output_row_meins_falls_back_to_msehi1() -> None:
    row = _live_invi_row() | {"Meins": ""}
    out = entry_row_to_output_row(row)
    assert out[3] == "BOT"  # falls through to Msehi1


def test_entry_row_to_output_row_strips_whitespace() -> None:
    row = _live_invi_row() | {"Werks": "  CA08  ", "Matnr": " 1234 "}
    out = entry_row_to_output_row(row)
    assert out[0] == "CA08"
    assert out[1] == "1234"


def test_entry_row_layout_matches_archive_layout() -> None:
    """Cross-check: entry xlsx column order must equal stocktake's so a
    downstream reader (report._read_fiori_count_by_matnr) works for
    either source. If this test fails, fix one or the other — do NOT
    let them drift."""
    from sap_fiori_crawler.stocktake import OUTPUT_COLUMNS as ARCHIVE_COLS
    assert ENTRY_OUTPUT_COLUMNS == ARCHIVE_COLS


# ── _default_filename ─────────────────────────────────────────────────


def test_default_filename() -> None:
    from sap_fiori_crawler.entry import _default_filename
    assert _default_filename("202604", "CA8DKG") == "SGP-CA8DKG-盘点录入-202604.xlsx"


# ── write_entry_xlsx ──────────────────────────────────────────────────


def test_write_entry_xlsx_roundtrip(tmp_path: Path) -> None:
    out = write_entry_xlsx(
        [_live_invi_row(), _live_invi_row() | {"Matnr": "9", "MengePd": "0"}],
        tmp_path / "out.xlsx",
    )
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert tuple(header) == ENTRY_OUTPUT_COLUMNS
    # col 2 = 物料号, col 7 = 盘点数量 (matching the archive layout).
    assert ws.cell(row=2, column=2).value == "4509062"
    assert ws.cell(row=2, column=7).value == 101
    assert ws.cell(row=3, column=2).value == "9"
    assert ws.cell(row=3, column=7).value == 0
