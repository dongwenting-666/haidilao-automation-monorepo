"""Unit tests for sap_fiori_crawler.stocktake (pure helpers + xlsx writer)."""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sap_fiori_crawler import (
    ENTITY_SET,
    FioriExportError,
    OUTPUT_COLUMNS,
    SERVICE_PATH,
    api_row_to_output_row,
    build_filter,
    build_period,
    build_url,
    parse_records,
    write_stocktake_xlsx,
)
from sap_fiori_crawler.stocktake import _default_filename, _to_number


# ── build_period ───────────────────────────────────────────────────────


@pytest.mark.parametrize(
    "year,month,expected",
    [
        (2026, 1, "202601"),
        (2026, 3, "202603"),
        (2026, 9, "202609"),
        (2026, 10, "202610"),
        (2026, 12, "202612"),
        (1999, 12, "199912"),
    ],
)
def test_build_period(year: int, month: int, expected: str) -> None:
    assert build_period(year, month) == expected


@pytest.mark.parametrize("month", [0, 13, -1, 100])
def test_build_period_rejects_bad_month(month: int) -> None:
    with pytest.raises(ValueError, match="month must be 1-12"):
        build_period(2026, month)


@pytest.mark.parametrize("year", [0, 1899, 10_000])
def test_build_period_rejects_bad_year(year: int) -> None:
    with pytest.raises(ValueError, match="year out of range"):
        build_period(year, 3)


# ── build_filter ───────────────────────────────────────────────────────


def test_build_filter_basic() -> None:
    assert build_filter("202603", "CA8DKG") == "ILfper eq '202603' and IUser eq 'CA8DKG'"


def test_build_filter_rejects_quote_in_user() -> None:
    """Defence against a malformed store key — never silently embed a quote."""
    with pytest.raises(ValueError, match="contains a quote"):
        build_filter("202603", "CA'INJECT")


# ── build_url ──────────────────────────────────────────────────────────


def test_build_url_contains_all_parts() -> None:
    url = build_url("202603", "CA8DKG")
    assert SERVICE_PATH in url
    assert ENTITY_SET in url
    # The filter substring is URL-encoded.
    assert "ILfper%20eq%20%27202603%27" in url
    assert "IUser%20eq%20%27CA8DKG%27" in url
    assert "format=json" in url


def test_build_url_uses_custom_base() -> None:
    url = build_url("202603", "CA8DKG", base="https://example.invalid")
    assert url.startswith("https://example.invalid" + SERVICE_PATH)


# ── _to_number ─────────────────────────────────────────────────────────


def test_to_number_integer_like_returns_int() -> None:
    """Integer-valued floats render without trailing .0 in xlsx."""
    assert _to_number("200.000") == 200
    assert isinstance(_to_number("200.000"), int)


def test_to_number_fractional_returns_float() -> None:
    assert _to_number("200.900") == pytest.approx(200.9)
    assert isinstance(_to_number("200.900"), float)


def test_to_number_blank_returns_blank() -> None:
    assert _to_number("") == ""
    assert _to_number("   ") == ""


def test_to_number_passthrough_non_numeric() -> None:
    assert _to_number("abc") == "abc"


def test_to_number_none_becomes_blank() -> None:
    """Missing fields render as blank cells, not literal None."""
    assert _to_number(None) == ""


def test_to_number_passes_through_non_string_non_none() -> None:
    assert _to_number(42) == 42


# ── api_row_to_output_row ──────────────────────────────────────────────


_SAMPLE_ROW = {
    "__metadata": {"id": "...", "uri": "...", "type": "..."},
    "IBmid": "",
    "ILfper": "000000",
    "IMatnr": "",
    "IUser": "",
    "Bmid": "99",
    "Status": "8",
    "Bukrs": "9451",
    "Werks": "CA08",
    "Matnr": "1000049",
    "Matxt": "金标生抽（海天，4.9L*2桶/件）",
    "Meins": "L",
    "Menge": "200.900",
    "MengePd": "63.700",
    "Msehi1": "L",
    "Pici": "已过账",
    "Users": "",
    "Zdate": "20260301",
    "Ztime": "192948",
    "Msetxt": "升-公升",
}


def test_api_row_to_output_row_preserves_field_order() -> None:
    """Output must align 1:1 with OUTPUT_COLUMNS."""
    out = api_row_to_output_row(_SAMPLE_ROW)
    assert len(out) == len(OUTPUT_COLUMNS)
    expected = [
        "CA08",
        "1000049",
        "金标生抽（海天，4.9L*2桶/件）",
        "L",
        "升-公升",
        pytest.approx(200.9),
        pytest.approx(63.7),
        "已过账",
        "20260301",
        "192948",
    ]
    for got, exp in zip(out, expected):
        assert got == exp


def test_api_row_to_output_row_strips_whitespace() -> None:
    row = dict(_SAMPLE_ROW)
    row["Werks"] = "  CA08  "
    row["Matxt"] = " name with spaces "
    out = api_row_to_output_row(row)
    assert out[0] == "CA08"
    assert out[2] == "name with spaces"


def test_api_row_to_output_row_handles_missing_fields() -> None:
    """Empty/missing fields must not crash; they become '' or 0."""
    out = api_row_to_output_row({})
    assert len(out) == len(OUTPUT_COLUMNS)
    assert out[0] == ""
    assert out[1] == ""
    # Numeric columns become "" when input is missing (not 0).
    assert out[5] == ""
    assert out[6] == ""


def test_api_row_to_output_row_handles_none_values() -> None:
    row = {k: None for k in _SAMPLE_ROW}
    out = api_row_to_output_row(row)
    assert all(v == "" for v in out), f"unexpected: {out}"


def test_api_row_to_output_row_integer_qty_renders_as_int() -> None:
    row = dict(_SAMPLE_ROW)
    row["Menge"] = "5.000"
    row["MengePd"] = "5.000"
    out = api_row_to_output_row(row)
    assert out[5] == 5 and isinstance(out[5], int)
    assert out[6] == 5 and isinstance(out[6], int)


# ── parse_records ──────────────────────────────────────────────────────


def test_parse_records_v2_envelope() -> None:
    payload = {"d": {"results": [{"Matnr": "1000049"}, {"Matnr": "1000050"}]}}
    rows = parse_records(payload)
    assert len(rows) == 2
    assert rows[0]["Matnr"] == "1000049"


def test_parse_records_flat_d_array() -> None:
    payload = {"d": [{"Matnr": "X"}]}
    rows = parse_records(payload)
    assert rows == [{"Matnr": "X"}]


def test_parse_records_empty_results() -> None:
    payload = {"d": {"results": []}}
    assert parse_records(payload) == []


def test_parse_records_missing_d_raises() -> None:
    with pytest.raises(FioriExportError, match="missing 'd.results' array"):
        parse_records({})


def test_parse_records_no_results_key_raises() -> None:
    with pytest.raises(FioriExportError):
        parse_records({"d": {"foo": "bar"}})


# ── _default_filename ──────────────────────────────────────────────────


def test_default_filename() -> None:
    assert _default_filename(2026, 3, "CA8DKG") == "SGP-CA8DKG-盘点-202603.xlsx"
    assert _default_filename(2026, 12, "CA9DKG") == "SGP-CA9DKG-盘点-202612.xlsx"


# ── write_stocktake_xlsx (round-trip) ──────────────────────────────────


def _make_rows(n: int = 3) -> list[dict]:
    rows = []
    for i in range(n):
        row = dict(_SAMPLE_ROW)
        row["Matnr"] = f"100004{i}"
        row["Menge"] = f"{100 + i}.000"
        row["MengePd"] = f"{50 + i}.{i}00"
        rows.append(row)
    return rows


def test_write_stocktake_xlsx_creates_file_with_header(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    write_stocktake_xlsx(_make_rows(2), out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert tuple(header) == OUTPUT_COLUMNS


def test_write_stocktake_xlsx_round_trips_values(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    rows = _make_rows(3)
    write_stocktake_xlsx(rows, out)

    wb = load_workbook(out)
    ws = wb.active
    # Header + 3 data rows
    assert ws.max_row == 4
    # First data row matches the first input.
    first = [c.value for c in ws[2]]
    expected = api_row_to_output_row(rows[0])
    assert first == expected


def test_write_stocktake_xlsx_creates_parent_dir(tmp_path: Path) -> None:
    out = tmp_path / "nested" / "deeper" / "report.xlsx"
    write_stocktake_xlsx(_make_rows(1), out)
    assert out.exists()


def test_write_stocktake_xlsx_truncates_long_sheet_name(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    long_name = "a" * 50
    write_stocktake_xlsx(_make_rows(1), out, sheet_name=long_name)
    wb = load_workbook(out)
    # Excel allows max 31 chars in sheet titles.
    assert len(wb.active.title) <= 31


def test_write_stocktake_xlsx_empty_rows_still_writes_header(tmp_path: Path) -> None:
    out = tmp_path / "empty.xlsx"
    write_stocktake_xlsx([], out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws.max_row == 1
    assert tuple(c.value for c in ws[1]) == OUTPUT_COLUMNS


# ── End-to-end against the fixture captured from the live run ──────────


_FIXTURE = Path(__file__).parent / "data" / "invhis_sample.json"


@pytest.mark.skipif(not _FIXTURE.exists(), reason="fixture not present")
def test_parse_real_payload_produces_expected_columns(tmp_path: Path) -> None:
    """Parse a real captured payload and confirm structure."""
    payload = json.loads(_FIXTURE.read_text())
    rows = parse_records(payload)
    assert len(rows) > 0
    # Every row produces a list of OUTPUT_COLUMNS length.
    for r in rows:
        out = api_row_to_output_row(r)
        assert len(out) == len(OUTPUT_COLUMNS)

    out_path = tmp_path / "real.xlsx"
    write_stocktake_xlsx(rows, out_path)
    wb = load_workbook(out_path)
    ws = wb.active
    assert ws.max_row == len(rows) + 1
