"""End-to-end test: real Fiori login + InvHisSet download for CA8DKG.

Marked ``e2e`` so the default suite skips it (login requires real
credentials and opens a browser window).

Run::

    uv run pytest libs/sap-fiori-crawler/tests/test_stocktake_e2e.py -m e2e -v
"""
from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Load .env before pytest evaluates skipif markers — the credentials live there.
try:
    from dotenv import load_dotenv

    load_dotenv("/Users/hongming-claw/haidilao-automation-monorepo/.env")
except ImportError:
    pass

from sap_fiori_crawler import (  # noqa: E402
    OUTPUT_COLUMNS,
    download_stocktake_report,
    fiori_session,
    load_store_creds,
)


pytestmark = pytest.mark.e2e


_STORE = "CA8DKG"
_YEAR = 2026
_MONTH = 3


@pytest.mark.skipif(
    not os.environ.get("SGPFIORIWEB_CREDS"),
    reason="SGPFIORIWEB_CREDS not set",
)
def test_download_stocktake_for_ca8_march(tmp_path: Path) -> None:
    """Real network call: fetch March 2026 stocktake for CA08 and verify xlsx."""
    creds = load_store_creds(_STORE)
    with fiori_session(creds, headless=False) as (browser, ctx, page):
        del browser, page
        out_path = download_stocktake_report(
            ctx, year=_YEAR, month=_MONTH, user=_STORE, out_dir=tmp_path
        )

    assert out_path.exists()
    assert out_path.suffix == ".xlsx"

    wb = load_workbook(out_path)
    ws = wb.active

    # Header row matches OUTPUT_COLUMNS
    header = [c.value for c in ws[1]]
    assert tuple(header) == OUTPUT_COLUMNS

    # Sanity check: at least 100 rows for a normal month at CA08.
    assert ws.max_row >= 100, f"only {ws.max_row - 1} data rows — looks suspicious"

    # The 工厂 column should always be CA08 for this user.
    werks_values = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert werks_values == {"CA08"}, f"unexpected 工厂 values: {werks_values}"

    # 物料号 column must be non-empty for every row.
    for r in range(2, ws.max_row + 1):
        assert ws.cell(row=r, column=2).value, f"missing material # at row {r}"
