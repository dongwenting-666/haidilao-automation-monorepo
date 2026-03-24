"""Render Excel worksheet tabs as PNG images for Lark delivery.

Reads actual cell styles (fill colors, fonts, borders, alignment, merged
cells) from the openpyxl workbook and reproduces them faithfully in a
Pillow image — the output looks like the Excel sheet itself.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.styles import Font as XlFont, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

if TYPE_CHECKING:
    from openpyxl.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

# ── Layout constants ──────────────────────────────────────────────────────────
_CELL_PAD_X = 8        # horizontal padding inside each cell
_CELL_PAD_Y = 4        # vertical padding inside each cell
_DEFAULT_ROW_HEIGHT = 28  # default row height in pixels
_GRID_COLOR = (169, 169, 169)  # Excel-style grid line color
_BORDER_COLOR = (128, 128, 128)  # Stronger border for merged cells / outer edges
_DEFAULT_BG = (255, 255, 255)
_DEFAULT_FG = (0, 0, 0)
_MAX_COL_WIDTH = 350
_MIN_COL_WIDTH = 50


# ── Font loading ──────────────────────────────────────────────────────────────
_font_cache: dict[tuple[int, bool], ImageFont.FreeTypeFont | ImageFont.ImageFont] = {}

_CJK_FONT_PATHS = [
    "/System/Library/Fonts/PingFang.ttc",
    "/System/Library/Fonts/STHeiti Light.ttc",
    "/System/Library/Fonts/Helvetica.ttc",
]


def _get_font(size: int = 12, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    """Load a CJK-capable font with caching."""
    key = (size, bold)
    if key in _font_cache:
        return _font_cache[key]

    font = None
    for path in _CJK_FONT_PATHS:
        try:
            # PingFang.ttc: index 0=Regular, index 1=Medium (semi-bold)
            idx = 1 if bold and "PingFang" in path else 0
            font = ImageFont.truetype(path, size, index=idx)
            break
        except (OSError, IOError):
            continue

    if font is None:
        font = ImageFont.load_default()

    _font_cache[key] = font
    return font


# ── Color helpers ─────────────────────────────────────────────────────────────

def _argb_to_rgb(color_str: str | None) -> tuple[int, int, int] | None:
    """Convert an ARGB hex string (e.g. 'FFFF0000') to an RGB tuple."""
    if not color_str or color_str == "00000000" or len(color_str) < 6:
        return None
    # Strip alpha channel if present
    hex_part = color_str[-6:]
    try:
        r = int(hex_part[0:2], 16)
        g = int(hex_part[2:4], 16)
        b = int(hex_part[4:6], 16)
        return (r, g, b)
    except (ValueError, IndexError):
        return None


def _cell_bg(cell: "Cell") -> tuple[int, int, int]:
    """Get the background fill color of a cell."""
    fill = cell.fill
    if fill and fill.fill_type == "solid" and fill.start_color:
        rgb = _argb_to_rgb(str(fill.start_color.rgb) if fill.start_color.rgb else None)
        if rgb and rgb != (0, 0, 0):  # skip if it resolves to black (theme fallback)
            return rgb
    return _DEFAULT_BG


def _cell_fg(cell: "Cell") -> tuple[int, int, int]:
    """Get the text color of a cell."""
    font = cell.font
    if font and font.color and font.color.rgb:
        rgb = _argb_to_rgb(str(font.color.rgb))
        if rgb:
            return rgb
    return _DEFAULT_FG


def _cell_font_size(cell: "Cell") -> int:
    """Get the font size of a cell (default 11)."""
    if cell.font and cell.font.size:
        return int(cell.font.size)
    return 11


def _cell_is_bold(cell: "Cell") -> bool:
    """Check if cell font is bold."""
    return bool(cell.font and cell.font.bold)


# ── Cell value formatting ────────────────────────────────────────────────────

def _cell_display_value(cell: "Cell") -> str:
    """Get the display string for a cell, respecting number formats."""
    if cell.value is None:
        return ""
    v = cell.value
    fmt = cell.number_format or "General"

    if isinstance(v, float):
        if "%" in fmt:
            return f"{v * 100:.2f}%"
        if "0.00" in fmt:
            return f"{v:,.2f}"
        if v == int(v):
            return f"{int(v):,}"
        return f"{v:,.2f}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v)


# ── Merged cell tracking ─────────────────────────────────────────────────────

def _build_merge_map(ws: "Worksheet") -> dict[tuple[int, int], tuple[int, int, int, int]]:
    """Build a map of (row, col) → (min_row, min_col, max_row, max_col) for merged cells."""
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
    return merge_map


# ── Main renderer ─────────────────────────────────────────────────────────────

def render_sheet(ws: "Worksheet", sheet_name: str) -> bytes:
    """Render a single worksheet to a PNG image faithfully reproducing Excel styling."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    merge_map = _build_merge_map(ws)

    # ── Phase 1: Measure column widths ────────────────────────────────────
    col_widths: list[int] = []
    for c in range(1, max_col + 1):
        # Check if worksheet has an explicit column width
        col_letter = get_column_letter(c)
        ws_col_dim = ws.column_dimensions.get(col_letter)
        if ws_col_dim and ws_col_dim.width and ws_col_dim.width > 0:
            # openpyxl width is in "characters" (~7px each)
            px = int(ws_col_dim.width * 7.5) + _CELL_PAD_X * 2
        else:
            # Auto-size from content
            px = _MIN_COL_WIDTH
            for r in range(1, max_row + 1):
                cell = ws.cell(r, c)
                # Skip cells that are part of a merge but not the top-left
                if (r, c) in merge_map:
                    mr_min_r, mr_min_c, _, _ = merge_map[(r, c)]
                    if r != mr_min_r or c != mr_min_c:
                        continue
                text = _cell_display_value(cell)
                if text:
                    font = _get_font(_cell_font_size(cell), _cell_is_bold(cell))
                    try:
                        bbox = font.getbbox(text)
                        tw = bbox[2] - bbox[0] if bbox else len(text) * 8
                    except Exception:
                        tw = len(text) * 8
                    px = max(px, tw + _CELL_PAD_X * 2)
        col_widths.append(min(max(px, _MIN_COL_WIDTH), _MAX_COL_WIDTH))

    # ── Phase 2: Measure row heights ──────────────────────────────────────
    row_heights: list[int] = []
    for r in range(1, max_row + 1):
        ws_row_dim = ws.row_dimensions.get(r)
        if ws_row_dim and ws_row_dim.height and ws_row_dim.height > 0:
            h = int(ws_row_dim.height * 1.33)  # Excel points → pixels
        else:
            h = _DEFAULT_ROW_HEIGHT
        row_heights.append(max(h, _DEFAULT_ROW_HEIGHT))

    # ── Phase 3: Compute positions ────────────────────────────────────────
    margin = 6
    table_width = sum(col_widths) + 1
    table_height = sum(row_heights) + 1

    col_x: list[int] = [margin]
    for w in col_widths:
        col_x.append(col_x[-1] + w)

    row_y: list[int] = [margin]
    for h in row_heights:
        row_y.append(row_y[-1] + h)

    img_width = table_width + margin * 2
    img_height = table_height + margin * 2

    img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    # ── Phase 4: Draw cell backgrounds ────────────────────────────────────
    drawn_merges: set[tuple[int, int, int, int]] = set()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ri = r - 1  # 0-indexed
            ci = c - 1

            # Handle merged cells
            if (r, c) in merge_map:
                mr = merge_map[(r, c)]
                if r != mr[0] or c != mr[1]:
                    continue  # not the top-left — skip
                if mr in drawn_merges:
                    continue
                drawn_merges.add(mr)

                # Merged range coordinates
                x1 = col_x[mr[1] - 1]
                y1 = row_y[mr[0] - 1]
                x2 = col_x[mr[3]] - 1  # max_col
                y2 = row_y[mr[2]] - 1  # max_row

                cell = ws.cell(mr[0], mr[1])
                bg = _cell_bg(cell)
                draw.rectangle([x1, y1, x2, y2], fill=bg)

                # Draw text in merged cell
                text = _cell_display_value(cell)
                if text:
                    fg = _cell_fg(cell)
                    font_size = min(_cell_font_size(cell) + 2, 22)  # scale up slightly for screen
                    font = _get_font(font_size, _cell_is_bold(cell))

                    # Center text in merged range
                    try:
                        bbox = font.getbbox(text)
                        tw = bbox[2] - bbox[0] if bbox else 0
                        th = bbox[3] - bbox[1] if bbox else font_size
                    except Exception:
                        tw, th = len(text) * 8, font_size

                    # Handle multi-line text (wrap_text)
                    lines = text.split("\n")
                    if len(lines) > 1:
                        total_h = th * len(lines)
                        ty = y1 + ((y2 - y1) - total_h) // 2
                        for line in lines:
                            try:
                                lbbox = font.getbbox(line)
                                lw = lbbox[2] - lbbox[0] if lbbox else 0
                            except Exception:
                                lw = len(line) * 8
                            tx = x1 + ((x2 - x1) - lw) // 2
                            draw.text((tx, ty), line, fill=fg, font=font)
                            ty += th
                    else:
                        tx = x1 + ((x2 - x1) - tw) // 2
                        ty = y1 + ((y2 - y1) - th) // 2
                        draw.text((tx, ty), text, fill=fg, font=font)
                continue

            # Regular (non-merged) cell
            x1 = col_x[ci]
            y1 = row_y[ri]
            x2 = col_x[ci + 1] - 1
            y2 = row_y[ri + 1] - 1

            cell = ws.cell(r, c)
            bg = _cell_bg(cell)
            draw.rectangle([x1, y1, x2, y2], fill=bg)

            text = _cell_display_value(cell)
            if text:
                fg = _cell_fg(cell)
                font_size = min(_cell_font_size(cell) + 2, 22)
                font = _get_font(font_size, _cell_is_bold(cell))

                try:
                    bbox = font.getbbox(text)
                    tw = bbox[2] - bbox[0] if bbox else 0
                    th = bbox[3] - bbox[1] if bbox else font_size
                except Exception:
                    tw, th = len(text) * 8, font_size

                cell_w = x2 - x1
                cell_h = y2 - y1

                # Truncate if too wide
                display = text
                while tw > cell_w - _CELL_PAD_X * 2 and len(display) > 2:
                    display = display[:-2] + "…"
                    try:
                        bbox = font.getbbox(display)
                        tw = bbox[2] - bbox[0] if bbox else 0
                    except Exception:
                        tw = len(display) * 8

                # Alignment
                alignment = cell.alignment
                h_align = alignment.horizontal if alignment else None

                if h_align == "center":
                    tx = x1 + (cell_w - tw) // 2
                elif h_align == "right":
                    tx = x2 - _CELL_PAD_X - tw
                elif h_align == "left" or h_align is None:
                    # Default: right-align numbers, left-align text
                    is_numeric = isinstance(cell.value, (int, float))
                    if is_numeric:
                        tx = x2 - _CELL_PAD_X - tw
                    else:
                        tx = x1 + _CELL_PAD_X
                else:
                    tx = x1 + _CELL_PAD_X

                ty = y1 + (cell_h - th) // 2
                draw.text((tx, ty), display, fill=fg, font=font)

    # ── Phase 5: Draw grid lines (skip lines inside merged ranges) ──────
    # Build sets of grid segments to suppress inside merged areas
    suppress_h: set[tuple[int, int]] = set()  # (row_boundary, col) — suppress horizontal line at this col
    suppress_v: set[tuple[int, int]] = set()  # (col_boundary, row) — suppress vertical line at this row
    for mr in drawn_merges:
        min_r, min_c, max_r, max_c = mr
        # Suppress internal horizontal lines
        for rb in range(min_r, max_r):  # internal row boundaries
            for c in range(min_c, max_c + 1):
                suppress_h.add((rb, c))
        # Suppress internal vertical lines
        for cb in range(min_c, max_c):  # internal col boundaries
            for r in range(min_r, max_r + 1):
                suppress_v.add((cb, r))

    # Horizontal lines — draw segments, skipping suppressed parts
    for ri in range(max_row + 1):
        seg_start = col_x[0]
        for ci in range(max_col):
            # Check if this segment (row boundary ri, column ci+1) is suppressed
            # ri=0 is the top border, ri=max_row is the bottom border
            # Suppress applies to internal row boundaries: ri corresponds to "after row ri"
            if (ri, ci + 1) in suppress_h:
                # Draw accumulated segment up to this point
                if col_x[ci] > seg_start:
                    draw.line([(seg_start, row_y[ri]), (col_x[ci], row_y[ri])], fill=_GRID_COLOR)
                seg_start = col_x[ci + 1]
        # Draw remaining segment
        if col_x[-1] > seg_start:
            draw.line([(seg_start, row_y[ri]), (col_x[-1], row_y[ri])], fill=_GRID_COLOR)

    # Vertical lines — draw segments, skipping suppressed parts
    for ci in range(max_col + 1):
        seg_start = row_y[0]
        for ri in range(max_row):
            if (ci, ri + 1) in suppress_v:
                if row_y[ri] > seg_start:
                    draw.line([(col_x[ci], seg_start), (col_x[ci], row_y[ri])], fill=_GRID_COLOR)
                seg_start = row_y[ri + 1]
        if row_y[-1] > seg_start:
            draw.line([(col_x[ci], seg_start), (col_x[ci], row_y[-1])], fill=_GRID_COLOR)

    # ── Phase 6: Draw borders for merged cells (on top of grid) ───────────
    for mr in drawn_merges:
        x1 = col_x[mr[1] - 1]
        y1 = row_y[mr[0] - 1]
        x2 = col_x[mr[3]]
        y2 = row_y[mr[2]]
        draw.rectangle([x1, y1, x2, y2], outline=_BORDER_COLOR)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


def render_all_sheets(xlsx_path: Path) -> list[tuple[str, bytes]]:
    """Render every sheet in an xlsx file to PNG images.

    Returns a list of (sheet_name, png_bytes) tuples.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    results: list[tuple[str, bytes]] = []

    for name in wb.sheetnames:
        ws = wb[name]
        try:
            png = render_sheet(ws, name)
            if png:
                results.append((name, png))
                log.info("Rendered sheet '%s' as PNG (%d bytes)", name, len(png))
        except Exception:
            log.exception("Failed to render sheet '%s'", name)

    return results
